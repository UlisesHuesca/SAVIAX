from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.forms import inlineformset_factory
from django.db.models import Sum, Q, Prefetch, Avg, FloatField, Case, When, F,DecimalField, ExpressionWrapper
from .models import Product, Subfamilia, Order, Products_Batch, Familia, Unidad, Inventario
from compras.models import Proveedor, Proveedor_Batch, Proveedor_Direcciones_Batch, Proveedor_direcciones, Estatus_proveedor, Estado
from solicitudes.models import Subproyecto, Proyecto
from requisiciones.models import Salidas, ValeSalidas
from user.models import Profile, Distrito, Banco
from .forms import ProductForm, Products_BatchForm, AddProduct_Form, Proyectos_Form, ProveedoresForm, Proyectos_Add_Form, Proveedores_BatchForm, ProveedoresDireccionesForm, Proveedores_Direcciones_BatchForm, Subproyectos_Add_Form, ProveedoresExistDireccionesForm, Add_ProveedoresDireccionesForm, DireccionComparativoForm

from .filters import ProductFilter, ProyectoFilter, ProveedorFilter, SubproyectoFilter

import csv
from django.core.paginator import Paginator
from datetime import date, datetime

import plotly.express as px
from plotly.subplots import make_subplots
import plotly.graph_objects as go
import pandas as pd
#import decimal
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt

# Create your views here.
@login_required(login_url='user-login')
def index(request):
    usuario = Profile.objects.get(staff=request.user)
    inventarios = Inventario.objects.all()
    proyectos = Proyecto.objects.all()

    # Obtener los proyectos y calcular el total
    #proyectos_total = [(proyecto, proyecto.get_projects_gastado) for proyecto in proyectos]

    # Obtener los inventarios y calcular el costo de salidas
    #inventarios_costo_salidas = [(inventario, inventario.costo_salidas) for inventario in inventarios]

    # Ordenar los inventarios por el costo de salidas en orden descendente
    #inventarios_costo_salidas_sorted = sorted(inventarios_costo_salidas, key=lambda x: x[1], reverse=True)
    # Ordenar los proyectos por el total en orden descendente
    #proyectos_total_sorted = sorted(proyectos_total, key=lambda x: x[1], reverse=True)

    # Tomar solo los primeros 50 inventarios ordenados
    #inventarios_top_50 = inventarios_costo_salidas_sorted[:50]


    # Preparar los datos para el gráfico
    #x = [proyecto.nombre for proyecto, _ in proyectos_total_sorted]
    #y = [total for _, total in proyectos_total_sorted]
    #x2 = [inventario.producto.nombre[:15] + '...' if len(inventario.producto.nombre) > 10 else inventario.producto.nombre for inventario,_ in inventarios_top_50]
    #y2 = [costo_salidas for _, costo_salidas in inventarios_top_50]



   # Crear el gráfico de barras
    #fig = make_subplots()
    #fig.add_trace(go.Bar(x=x, y=y, marker=dict(color='#3E92CC')),1,1)
    # Crear el gráfico de barras
    #fig2 = make_subplots()
    #fig2.add_trace(go.Bar(x=x2, y=y2, marker=dict(color='#3E92CC')),1,1)

    #fig.update_layout(
    #    plot_bgcolor='#9a9b9d',
    #    paper_bgcolor='white',
    #    font_color= '#3E92CC',
    #    )

    #fig2.update_layout(
    #    plot_bgcolor='#9a9b9d',
    #    paper_bgcolor='white',
    #    font_color= '#3E92CC',
    #    )

    #Convertir el gráfico en HTML para pasar a la plantilla
    #graph_proyectos = fig.to_html(full_html=False)
    #graph_inventarios = fig2.to_html(full_html=False)

    context = {
        #'graph_proyectos': graph_proyectos,
        #'graph_inventarios':graph_inventarios,
        }
    
    return render(request,'dashboard/index.html',context)

@login_required(login_url='user-login')
def proyectos(request):
    sql_salidas = """SELECT 
    solicitudes_proyecto.id AS id,
    solicitudes_proyecto.nombre AS nombre,
    SUM(requisiciones_salidas.cantidad * requisiciones_salidas.precio) AS total_salidas
    FROM
        solicitudes_proyecto
    JOIN
        dashboard_order ON solicitudes_proyecto.id = dashboard_order.proyecto_id
    JOIN
        dashboard_articulosordenados ON dashboard_order.id = dashboard_articulosordenados.orden_id
    JOIN
        dashboard_articulosparasurtir ON dashboard_articulosordenados.id = dashboard_articulosparasurtir.articulos_id
    JOIN
        requisiciones_salidas ON dashboard_articulosparasurtir.id = requisiciones_salidas.producto_id
    GROUP BY
        id, nombre
    ORDER BY
        id;
    """


    sql_gastos_pagados ="""SELECT 
    solicitudes_proyecto.id AS id,
    solicitudes_proyecto.nombre AS nombre,
    SUM(
        (gastos_articulo_gasto.cantidad * gastos_articulo_gasto.precio_unitario * 1.16) + 
			COALESCE(gastos_articulo_gasto.otros_impuestos, 0) - COALESCE(gastos_articulo_gasto.impuestos_retenidos, 0)
    ) AS total_pagado
    FROM
        solicitudes_proyecto
    JOIN
        gastos_solicitud_gasto ON solicitudes_proyecto.id = gastos_solicitud_gasto.proyecto_id
    JOIN
        gastos_articulo_gasto ON gastos_solicitud_gasto.id = gastos_articulo_gasto.gasto_id
    LEFT JOIN
        tesoreria_pago ON gastos_solicitud_gasto.id = tesoreria_pago.gasto_id
    WHERE
	    tesoreria_pago.hecho = true
    GROUP BY
	    id, nombre
    ORDER BY
        id;

    """


    sql_compras_pagos = """SELECT 
    solicitudes_proyecto.id  AS id,
    solicitudes_proyecto.nombre AS nombre,
        SUM((CASE 
                    -- Cuando la moneda es PESOS
                    WHEN cuenta_moneda.nombre = 'PESOS' THEN tesoreria_pago.monto
                    -- Cuando la moneda es DÓLARES
                  
					WHEN cuenta_moneda.nombre = 'DOLARES' THEN
						CASE 
							-- Si tiene tipo de cambio en tesoreria_pago
							WHEN tesoreria_pago.tipo_de_cambio IS NOT NULL THEN tesoreria_pago.monto * tesoreria_pago.tipo_de_cambio
							-- Usar 17 como valor predeterminado si no tiene tipo de cambio
							WHEN tesoreria_pago.tipo_de_cambio IS NULL AND compras_compra.tipo_de_cambio IS NOT NULL THEN tesoreria_pago.monto * compras_compra.tipo_de_cambio 
							ELSE tesoreria_pago.monto * 17
						END
			END
    )) AS total_pagado
    FROM
        solicitudes_proyecto
    JOIN
        dashboard_order ON solicitudes_proyecto.id = dashboard_order.proyecto_id
    JOIN 
        requisiciones_requis ON dashboard_order.id = requisiciones_requis.orden_id
    JOIN
        compras_compra ON requisiciones_requis.id = compras_compra.req_id
    JOIN
        compras_moneda ON compras_compra.moneda_id = compras_moneda.id
    LEFT JOIN 
        tesoreria_pago ON compras_compra.id = tesoreria_pago.oc_id
    LEFT JOIN
        tesoreria_cuenta ON tesoreria_pago.cuenta_id = tesoreria_cuenta.id
    LEFT JOIN
        compras_moneda AS cuenta_moneda ON tesoreria_cuenta.moneda_id = cuenta_moneda.id -- Utilizando el alias cuenta_moneda
    WHERE
        tesoreria_pago.hecho = True
    GROUP BY
        id, nombre
    ORDER BY
        proyecto_id;"""

    sql_compras = """SELECT 
	solicitudes_proyecto.id  AS id,
    solicitudes_proyecto.nombre AS nombre,
    SUM(
        CASE 
            WHEN compras_moneda.nombre = 'DOLARES' AND pagos_promedio.avg_tipo_de_cambio IS NOT NULL THEN compras_compra.costo_oc * pagos_promedio.avg_tipo_de_cambio
            WHEN compras_moneda.nombre = 'DOLARES' AND pagos_promedio.avg_tipo_de_cambio IS NULL AND compras_compra.tipo_de_cambio IS NOT NULL THEN compras_compra.costo_oc * compras_compra.tipo_de_cambio
            WHEN compras_moneda.nombre = 'DOLARES' AND pagos_promedio.avg_tipo_de_cambio IS NULL AND compras_compra.tipo_de_cambio IS NULL THEN compras_compra.costo_oc * 17
            ELSE compras_compra.costo_oc
        END
    ) AS total_costo_oc
    FROM
	    solicitudes_proyecto
    JOIN
	    dashboard_order ON solicitudes_proyecto.id = dashboard_order.proyecto_id
    JOIN 
	    requisiciones_requis ON dashboard_order.id = requisiciones_requis.orden_id
    JOIN
	    compras_compra ON requisiciones_requis.id = compras_compra.req_id
    JOIN
	    compras_moneda ON compras_compra.moneda_id = compras_moneda.id
    LEFT JOIN (
	    SELECT oc_id, AVG(tipo_de_cambio) AS avg_tipo_de_cambio
        FROM tesoreria_pago
	    group by oc_id
    ) AS pagos_promedio ON compras_compra.id = pagos_promedio.oc_id
    group by
	    id, nombre
    ORDER BY
	    id;
       """
   
     # Prefetching related data
    proyectos = Proyecto.objects.all()
    proyecto_compras_total = proyectos.raw(sql_compras)
    proyecto_pagos_total = proyectos.raw(sql_compras_pagos)
    proyecto_gastos_total = proyectos.raw(sql_gastos_pagados)
    proyectos_salidas = proyectos.raw(sql_salidas)
    dict_compras = {r.id: r.total_costo_oc for r in proyecto_compras_total}
    dict_pagos = {r.id: r.total_pagado for r in proyecto_pagos_total}
    dict_gastos = {r.id: r.total_pagado for r in proyecto_gastos_total}
    dict_salidas = {r.id: r.total_salidas for r in proyectos_salidas}

    myfilter=ProyectoFilter(request.GET, queryset=proyectos)
    proyectos = myfilter.qs

    if request.method == 'POST' and 'btnReporte' in request.POST:
        proyectos_completos = asignar_totales(proyectos, dict_compras, dict_pagos, dict_gastos, dict_salidas)
        return convert_excel_matriz_proyectos(proyectos_completos)

    #Set up pagination
    p = Paginator(proyectos, 10)
    page = request.GET.get('page')
    proyectos_list = p.get_page(page)
    
    proyectos_paginados = asignar_totales(proyectos_list, dict_compras, dict_pagos, dict_gastos, dict_salidas)


    context = {
        'proyectos':proyectos,
        'proyectos_list':proyectos_list,
        'myfilter':myfilter,
        }
    
    return render(request,'dashboard/proyectos.html',context)

def asignar_totales(proyectos_queryset, dict_compras, dict_pagos, dict_gastos, dict_salidas):
    for proyecto in proyectos_queryset:
        proyecto.total_compras = dict_compras.get(proyecto.id, 0)
        proyecto.total_pagos = dict_pagos.get(proyecto.id, 0)
        proyecto.total_gastos = dict_gastos.get(proyecto.id, 0)
        proyecto.total_salidas = dict_salidas.get(proyecto.id, 0)
    return proyectos_queryset

@login_required(login_url='user-login')
def subproyectos(request, pk):
    proyecto = Proyecto.objects.get(id=pk)
    subproyectos = Subproyecto.objects.filter(proyecto=proyecto)

    myfilter=SubproyectoFilter(request.GET, queryset=subproyectos)
    subproyectos = myfilter.qs

    #Set up pagination
    p = Paginator(subproyectos, 50)
    page = request.GET.get('page')
    subproyectos_list = p.get_page(page)

    context = {
        'proyecto':proyecto,
        'subproyectos':subproyectos,
        'subproyectos_list':subproyectos_list,
        'myfilter':myfilter,
        }

    return render(request,'dashboard/subproyectos.html',context)



@login_required(login_url='user-login')
def proyectos_edit(request, pk):

    proyecto = Proyecto.objects.get(id=pk)

    if request.method =='POST':
        form = Proyectos_Form(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request,f'Has actualizado correctamente el proyecto {proyecto.nombre}')
            return redirect('configuracion-proyectos')
    else:
        form = Proyectos_Form(instance=proyecto)


    context = {
        'form': form,
        'proyecto':proyecto,
        }
    return render(request,'dashboard/proyectos_edit.html', context)

@login_required(login_url='user-login')
def proveedor_direcciones(request, pk):
    proveedor = Proveedor.objects.get(id=pk)

    direcciones = Proveedor_direcciones.objects.filter(nombre__id=pk, completo = True)

    context = {
        'proveedor':proveedor,
        'direcciones':direcciones,
        }
    return render(request,'dashboard/direcciones_proveedor.html', context)

@login_required(login_url='user-login')
def proyectos_add(request):
    usuario = Profile.objects.get(staff=request.user)

    form = Proyectos_Add_Form()

    if request.method =='POST':
        proyecto, created = Proyecto.objects.get_or_create(distrito = usuario.distrito, complete = False)
        form = Proyectos_Add_Form(request.POST, instance = proyecto)
        if form.is_valid():
            proyecto = form.save(commit=False)
            proyecto.activo = True
            proyecto.complete = True
            proyecto.save()
            messages.success(request,'Has agregado correctamente el proyecto')
            return redirect('configuracion-proyectos')
    else:
        form = Proyectos_Add_Form()

    context = {
        'form': form,
        }

    return render(request,'dashboard/proyectos_add.html',context)

@login_required(login_url='user-login')
def subproyectos_add(request, pk):
    proyecto = Proyecto.objects.get(id=pk)
    form = Subproyectos_Add_Form()

    if request.method =='POST':
        form = Subproyectos_Add_Form(request.POST)
        if form.is_valid():
            subproyecto = form.save(commit=False)
            subproyecto.proyecto = proyecto
            subproyecto.save()
            messages.success(request,'Has agregado correctamente el subproyecto')
            return redirect('subproyectos', pk=proyecto.id)
    else:
        form = Subproyectos_Add_Form()

    context = {
        'form': form,
        'proyecto':proyecto,
        }

    return render(request,'dashboard/subproyectos_add.html',context)

@login_required(login_url='user-login')
def subproyectos_edit(request, pk):
    subproyecto = Subproyecto.objects.get(id=pk)
    proyecto = Proyecto.objects.get(id=subproyecto.proyecto.id)
    form = Subproyectos_Add_Form(instance=subproyecto)

    if request.method =='POST':
        form = Subproyectos_Add_Form(request.POST, instance=subproyecto)
        if form.is_valid():
            form.save()
            messages.success(request,'Has editado correctamente el subproyecto')
            return redirect('subproyectos', pk=subproyecto.proyecto.id)
    else:
        form = Subproyectos_Add_Form(instance=subproyecto)

    context = {
        'form': form,
        'proyecto':proyecto,
        }

    return render(request,'dashboard/subproyectos_add.html',context)


@login_required(login_url='user-login')
def staff(request):
    workers = User.objects.all()
    context= {
        'workers': workers,
        }
    return render(request,'dashboard/staff.html', context)

@login_required(login_url='user-login')
def product(request):
    usuario = Profile.objects.get(staff=request.user)
    items = Product.objects.filter(completado = True).order_by('codigo')

    myfilter=ProductFilter(request.GET, queryset=items)
    items = myfilter.qs

    #Set up pagination
    p = Paginator(items, 50)
    page = request.GET.get('page')
    items_list = p.get_page(page)

    context = {
        'usuario':usuario,
        'items': items,
        'myfilter':myfilter,
        'items_list':items_list,
        }


    return render(request,'dashboard/product.html', context)


@login_required(login_url='user-login')
def proveedores(request):
    usuario = Profile.objects.get(staff=request.user)
    # Obtén los IDs de los proveedores que cumplan con las condiciones deseadas
    proveedores_dir = Proveedor_direcciones.objects.filter(Q(estatus__nombre='NUEVO') | Q(estatus__nombre='APROBADO'))
    proveedores_ids = proveedores_dir.values_list('nombre', flat=True).distinct()
    proveedores = Proveedor.objects.filter(id__in=proveedores_ids, completo=True)

    total_prov = proveedores.count()

    myfilter=ProveedorFilter(request.GET, queryset=proveedores)
    proveedores = myfilter.qs

    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_excel_proveedores(proveedores_dir)

    #Set up pagination
    p = Paginator(proveedores, 50)
    page = request.GET.get('page')
    proveedores_list = p.get_page(page)

    context = {
        'usuario':usuario,
        'proveedores': proveedores,
        'myfilter':myfilter,
        'proveedores_list':proveedores_list,
        'total_prov':total_prov,
        }


    return render(request,'dashboard/proveedores.html', context)

@login_required(login_url='user-login')
def matriz_revision_proveedor(request):
    usuario = Profile.objects.get(staff=request.user)
    proveedores = Proveedor_direcciones.objects.filter(estatus__nombre = "REVISION")

    total_prov = proveedores.count()

    myfilter=ProveedorFilter(request.GET, queryset=proveedores)
    proveedores = myfilter.qs

    #Set up pagination
    p = Paginator(proveedores, 50)
    page = request.GET.get('page')
    proveedores_list = p.get_page(page)

    context = {
        'usuario':usuario,
        'proveedores': proveedores,
        'myfilter':myfilter,
        'proveedores_list':proveedores_list,
        'total_prov':total_prov,
        }


    return render(request,'dashboard/matriz_revision_proveedor.html', context)


@login_required(login_url='user-login')
def proveedores_update(request, pk):

    proveedores = Proveedor.objects.get(id=pk)

    if request.method =='POST':
        form = ProveedoresForm(request.POST, instance=proveedores)
        if form.is_valid():
            form.save()
            messages.success(request,f'Has actualizado correctamente el proyecto {proveedores.razon_social}')
            return redirect('dashboard-proveedores')
    else:
        form = ProveedoresForm(instance=proveedores)

    context = {
        'form': form,
        'proveedores':proveedores,
        }

    return render(request,'dashboard/proveedores_update.html', context)



@login_required(login_url='user-login')
def add_proveedores_old(request):
    usuario = Profile.objects.get(staff=request.user)
    item, created = Proveedor.objects.get_or_create(creado_por=usuario, completo = False)

    if request.method =='POST':
        form = ProveedoresForm(request.POST, request.FILES or None, instance = item)
        if form.is_valid():
            item = form.save(commit=False)
            item.completo = True
            item.save()
            messages.success(request,f'Has agregado correctamente el proveedor {item.razon_social}')
            return redirect('dashboard-proveedores')
    else:
        form = ProveedoresForm(instance=item)


    context = {
        'form': form,
        'item':item,
        }
    return render(request,'dashboard/add_proveedores.html', context)

@login_required(login_url='user-login')
def add_proveedor_direccion(request, pk):

    usuario = Profile.objects.get(staff=request.user)
    proveedor = Proveedor.objects.get(id=pk)
    form = ProveedoresDireccionesForm()

    if request.method =='POST':
        item, created = Proveedor_direcciones.objects.get_or_create(nombre = proveedor, creado_por = usuario, completo = False)
        form = ProveedoresDireccionesForm(request.POST, instance = item)
        if form.is_valid():
            item = form.save(commit=False)
            item.disitrito = usuario.distrito
            item.completo = True
            item.save()
            messages.success(request,f'Has agregado correctamente la direccion del proveedor {item.nombre.razon_social}')
            return redirect('dashboard-proveedores')
        else:
            errors = form.errors.as_text()
            messages.error(request,f'No esta validando. Errores:{errors}')
    else:
        form = ProveedoresDireccionesForm()


    context = {
        'form': form,
        #'item':item,
        'proveedor':proveedor,
        }
    return render(request,'dashboard/add_proveedor_direccion.html', context)

@login_required(login_url='user-login')
def add_proveedores2(request, pk=None):
    usuario = Profile.objects.get(staff=request.user)
    proveedor, created = Proveedor.objects.get_or_create(creado_por=usuario, completo=False)
    proveedores_dir_ids = Proveedor_direcciones.objects.filter(~Q(estatus__nombre ="REVISION"),~Q(distrito = usuario.distrito)).values_list('id', flat=True)
    
    proveedores = Proveedor.objects.filter(proveedor_direcciones__id__in=proveedores_dir_ids)
    print('proveedores:',proveedores.count())

    if usuario.tipo.proveedores == True:
        ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=Add_ProveedoresDireccionesForm, extra=1)
    else:
        ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=ProveedoresDireccionesForm, extra=1)
    

    if request.method == 'POST':
        form = ProveedoresForm(request.POST, instance = proveedor)
        formset = ProveedorDireccionesFormSet(request.POST, instance=proveedor)
        if form.is_valid() and formset.is_valid():
            proveedor = form.save(commit=False)
            proveedor.completo = True
            proveedor.save()
            direcciones = formset.save(commit=False)
            direccion = direcciones[0]
            direccion.distrito = usuario.distrito
            if usuario.tipo.proveedores == False:
                estatus = Estatus_proveedor.objects.get(nombre ="REVISION")
                direccion.estatus = estatus
            direccion.creado_por = usuario
            direccion.enviado_fecha = date.today()
            direccion.completo = True
            direccion.save()
            messages.success(request, f'Has agregado correctamente el proveedor {proveedor.razon_social} y sus direcciones')
            return redirect('dashboard-proveedores')
        else:
            print(form.errors) 
            print(formset.errors) 
            messages.success(request, 'No está validando')
    else:
        form = ProveedoresForm(instance=proveedor)
        formset = ProveedorDireccionesFormSet(instance=proveedor)

    context = {
        'form': form,
        'formset': formset,
    }
    return render(request, 'dashboard/add_proveedores_&_direccion.html', context)

@login_required(login_url='user-login')
def add_proveedores_comparativo(request, pk=None):
    usuario = Profile.objects.get(staff=request.user)
    proveedor, created = Proveedor.objects.get_or_create(creado_por=usuario, completo=False)
    #proveedores_dir_ids = Proveedor_direcciones.objects.filter(~Q(estatus__nombre ="REVISION"),~Q(distrito = usuario.distrito)).values_list('id', flat=True)
    
    #proveedores = Proveedor.objects.filter(proveedor_direcciones__id__in=proveedores_dir_ids)
    print('usuario_tipo:',usuario.tipo.proveedores)

    #if usuario.tipo.proveedores == True:
    #    ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=Add_ProveedoresDireccionesForm, extra=1)
    #else:
    ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=DireccionComparativoForm, extra=1)
    

    if request.method == 'POST':
        form = ProveedoresForm(request.POST, instance = proveedor)
        formset = ProveedorDireccionesFormSet(request.POST, instance=proveedor)
        if form.is_valid() and formset.is_valid():
            proveedor = form.save(commit=False)
            proveedor.completo = True
            proveedor.save()
            direcciones = formset.save(commit=False)
            direccion = direcciones[0]
            direccion.distrito = usuario.distrito
            if usuario.tipo.proveedores == False:
                estatus = Estatus_proveedor.objects.get(nombre ="COTIZACION")
                direccion.estatus = estatus
            direccion.creado_por = usuario
            direccion.enviado_fecha = date.today()
            direccion.completo = True
            direccion.save()
            messages.success(request, f'Has agregado correctamente el proveedor {proveedor.razon_social} y sus direcciones')
            return redirect('comparativos')
        else:
            print(form.errors) 
            print(formset.errors) 
            messages.success(request, 'No está validando')
    else:
        form = ProveedoresForm(instance=proveedor)
        formset = ProveedorDireccionesFormSet(instance=proveedor)

    context = {
        'form': form,
        'formset': formset,
    }
    return render(request, 'dashboard/add_proveedor_direccion_cotizacion.html', context)

@login_required(login_url='user-login')
def add_proveedores(request, pk=None):
    usuario = Profile.objects.get(staff=request.user)
    #proveedor, created = Proveedor.objects.get_or_create(creado_por=usuario, completo=False)
    proveedores_dir_ids = Proveedor_direcciones.objects.filter(~Q(estatus__nombre ="REVISION"),~Q(distrito = usuario.distrito)).values_list('id', flat=True)
    
    proveedores = Proveedor.objects.filter(proveedor_direcciones__id__in=proveedores_dir_ids)
    print('proveedores:',proveedores.count())

    
    form = ProveedoresExistDireccionesForm()
    form.fields['nombre'].queryset = proveedores

    if request.method == 'POST':
        form = ProveedoresExistDireccionesForm(request.POST)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.completo = True
            proveedor.save()
            #direcciones = formset.save(commit=False)
            #direccion = direcciones[0]
            #direccion.distrito = usuario.distrito
            #estatus = Estatus_proveedor.objects.get(nombre ="REVISION")
            #direccion.creado_por = usuario
            #direccion.estatus = estatus
            #direccion.completo = True
            #direccion.save()
            messages.success(request, f'Has agregado correctamente el proveedor {proveedor.razon_social} y sus direcciones')
            return redirect('dashboard-proveedores')
    
    context = {
        'proveedores':proveedores,
        'form': form,
    }
    return render(request, 'dashboard/proveedor_exist_&_direccion.html', context)

@login_required(login_url='user-login')
def edit_proveedores(request, pk):
    usuario = Profile.objects.get(staff=request.user)
    proveedor_direccion = Proveedor_direcciones.objects.get(id=pk)
    proveedor = Proveedor.objects.get(id = proveedor_direccion.nombre.id)
    #romper

    ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form =Edit_ProveedoresDireccionesForm, extra=0)
    form = ProveedoresForm(instance=proveedor)
    formset = ProveedorDireccionesFormSet(instance=proveedor)

    if request.method == 'POST':
        form = ProveedoresForm(request.POST or None, instance =proveedor)
        formset = ProveedorDireccionesFormSet(request.POST or None, instance=proveedor)
        if form.is_valid and formset.is_valid():
            
            form.save()
            direcciones = formset.save(commit=False)
            for item in direcciones:
                item.actualizado_por = usuario
                item.modificado_fecha = date.today()
                item.save()
            messages.success(request, 'Has agregado correctamente el proveedor y sus direcciones')
            return redirect('dashboard-proveedores')
        else:
            print(form.errors) 
            print(formset.errors) 
            messages.success(request, 'No está validando')

    context = {
        'form': form,
        'proveedor':proveedor,
        'formset': formset,
    }
    
    return render(request,'dashboard/add_proveedor_direccion.html', context)


@login_required(login_url='user-login')
def edit_proveedor_direccion(request, pk):

    usuario = Profile.objects.get(staff=request.user)
    direccion = Proveedor_direcciones.objects.get(id = pk)
    proveedor = Proveedor.objects.get(id = direccion.nombre.id)

    if request.method =='POST':
        form = ProveedoresDireccionesForm(request.POST, instance = direccion)
        if form.is_valid():
            direccion = form.save(commit=False)
            direccion.actualizado_por = usuario
            direccion.modificado_fecha = date.today()
            direccion.completo = True
            direccion.save()
            messages.success(request,'Has actualizado correctamente la direccion del proveedor')
            return redirect('dashboard-proveedores')
    else:
        form = ProveedoresDireccionesForm(instance = direccion)


    context = {
        'proveedor':proveedor,
        'form': form,
        'direccion':direccion,
        }
    return render(request,'dashboard/edit_direcciones_proveedores.html', context)


@login_required(login_url='user-login')
def upload_batch_proveedores(request):

    form = Proveedores_BatchForm(request.POST or None, request.FILES or None)


    if form.is_valid():
        form.save()
        form = Proveedores_BatchForm()
        proveedores_list = Proveedor_Batch.objects.get(activated = False)

        f = open(proveedores_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader)

        for row in reader:
            if not Proveedor.objects.filter(razon_social=row[0]):
                proveedor = Proveedor(razon_social=row[0], rfc=row[1])
                proveedor.save()
            else:
                messages.error(request,f'El proveedor código:{row[0]} ya existe dentro de la base de datos')

        proveedores_list.activated = True
        proveedores_list.save()
    elif request.FILES:
        messages.error(request,'El formato no es CSV')

    context = {
        'form': form,
        }

    return render(request,'dashboard/upload_batch_proveedor.html', context)

@login_required(login_url='user-login')
def upload_batch_proveedores_direcciones(request):

    form = Proveedores_Direcciones_BatchForm(request.POST or None, request.FILES or None)


    if form.is_valid():
        form.save()
        form = Proveedores_Direcciones_BatchForm()
        proveedores_list = Proveedor_Direcciones_Batch.objects.get(activated=False)

        f = open(proveedores_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader)

        for row in reader:
            if Proveedor.objects.filter(razon_social=row[0]):
                nombre = Proveedor.objects.get(razon_social=row[0])
                if Distrito.objects.filter(nombre = row[1]):
                    distrito = Distrito.objects.get(nombre = row[1])
                    if Banco.objects.filter(nombre= row[6]):
                        banco = Banco.objects.get(nombre = row[6])
                        if Estatus_proveedor.objects.filter(nombre = row[11]):
                            estatus = Estatus_proveedor.objects.get(nombre = row[11])
                            if Estado.objects.filter(nombre = row[3]):
                                estado = Estado.objects.get(nombre = row[3])
                                proveedor_direccion = Proveedor_direcciones(nombre=nombre, distrito=distrito,domicilio=row[2],estado=estado,contacto=row[4],email=row[5], banco=banco, clabe=row[7], cuenta=row[8], financiamiento=row[9],dias_credito=row[10],estatus=estatus)
                                proveedor_direccion.save()
                            else:
                                messages.error(request,f'El estado:{row[3]} no existe dentro de la base de datos')
                        else:
                             messages.error(request,f'El estatus:{row[11]} no existe dentro de la base de datos')
                    else:
                         messages.error(request,f'El banco:{row[6]} no existe dentro de la base de datos')
                else:
                    messages.error(request,f'El distrito:{row[1]} no existe dentro de la base de datos')
            else:
                messages.error(request,f'El proveedor código:{row[0]} no existe dentro de la base de datos')

        proveedores_list.activated = True
        proveedores_list.save()
    elif request.FILES:
        messages.error(request,'El formato no es CSV')

    context = {
        'form': form,
        }

    return render(request,'dashboard/upload_batch_proveedor_direcciones.html', context)



@login_required(login_url='user-login')
def upload_batch_products(request):

    form = Products_BatchForm(request.POST or None, request.FILES or None)


    if form.is_valid():
        form.save()
        form = Products_BatchForm()
        product_list = Products_Batch.objects.get(activated = False)

        f = open(product_list.file_name.path, 'r', encoding='utf-8')
        reader = csv.reader(f)
        next(reader)

        for row in reader:
            if not Product.objects.filter(codigo=row[0]):
                if Unidad.objects.filter(nombre = row[2]):
                    unidad = Unidad.objects.get(nombre = row[2])
                    if Familia.objects.filter(nombre = row[3]):
                        familia = Familia.objects.get(nombre = row[3])
                        especialista = True if row[5] == 'SI' else False
                        iva = True if row[6] == 'SI' else False
                        activo = True if row[7] == 'SI' else False
                        servicio = True if row[8] == 'SI' else False
                        if Subfamilia.objects.filter(nombre = row[4], familia = familia):
                            subfamilia = Subfamilia.objects.get(nombre = row[4], familia = familia)
                            
                            producto = Product(codigo=row[0],nombre=row[1], unidad=unidad, familia=familia, subfamilia=subfamilia,especialista=especialista,iva=iva,activo=activo,servicio=servicio,baja_item=False,completado=True)
                            producto.save()
                        else:
                            producto = Product(codigo=row[0],nombre=row[1], unidad=unidad, familia=familia,especialista=especialista,iva=iva,activo=activo,servicio=servicio,baja_item=False,completado=True)
                            producto.save()
                    else:
                        messages.error(request,f'La familia no existe dentro de la base de datos, producto:{row[0]}')
                else:
                    messages.error(request,f'La unidad no existe dentro de la base de datos, producto:{row[0]}')
            else:
                messages.error(request,f'El producto código:{row[0]} ya existe dentro de la base de datos')

        product_list.activated = True
        product_list.save()
    elif request.FILES:
        messages.error(request,'El formato no es CSV')




    context = {
        'form': form,
        }

    return render(request,'dashboard/upload_batch_products.html', context)


@login_required(login_url='user-login')
def order(request):
    orders = Order.objects.all()
    context= {
        'orders':orders,
        }

    return render(request,'dashboard/order.html', context)

@login_required(login_url='user-login')
def product_delete(request, pk):
    item = Product.objects.get(id=pk)
    if request.method == 'POST':
        item.delete()
        return redirect('dashboard-product')

    return render(request,'dashboard/product_delete.html')


@login_required(login_url='user-login')
def add_product(request):
    item, created = Product.objects.get_or_create(completado=False)

    if request.method =='POST':
        form = AddProduct_Form(request.POST, request.FILES or None, instance = item)
        #form.save(commit=False)
        item.completado = True
        if form.is_valid():
            form.save()
            item.save()
            messages.success(request,f'Has agregado correctamente el producto {item.nombre}')
            return redirect('dashboard-product')
    else:
        form = AddProduct_Form()


    context = {
        'form': form,
        'item':item,
        }
    return render(request,'dashboard/add_product.html', context)

@login_required(login_url='user-login')
def product_update(request, pk):
#def product_update_modal(request, pk):

    item = Product.objects.get(id=pk)

    if request.method =='POST':
        form = AddProduct_Form(request.POST, request.FILES or None, instance=item, )
        if form.is_valid():
            form.save()
            messages.success(request,f'Has actualizado correctamente el producto {item.nombre}')
            return redirect('dashboard-product')
    else:
        form = AddProduct_Form(instance=item)


    context = {
        'form': form,
        'item':item,
        }
    return render(request,'dashboard/product_update.html', context)



def load_subfamilias(request):

    familia_id = request.GET.get('familia_id')
    subfamilias = Subfamilia.objects.filter(familia_id = familia_id)

    return render(request, 'dashboard/subfamilia_dropdown_list_options.html',{'subfamilias': subfamilias})


@login_required(login_url='user-login')
def staff_detail(request, pk):
    worker = User.objects.get(id=pk)
    context={
        'worker': worker,
        }
    return render(request,'dashboard/staff_detail.html', context)

def convert_excel_matriz_proyectos(proyectos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Matriz_proyectos_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Proyectos')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['ID','Proyectos','Descripción','Cliente','Status de Entrega','Monto','Gastado Salidas','Suma de Compras',
              'Pagado Compras','Pagado Gastos','Creado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 40

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordtec. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30
   
    rows = [
        (
            p.id,
            p.nombre,
            p.descripcion,
            p.cliente.nombre if p.cliente else '', 
            p.status_de_entrega, 
            p.get_projects_total if p.get_projects_total is not None else 0, 
            p.suma_salidas if p.suma_salidas is not None else 0,
            p.suma_comprat if p.suma_comprat is not None else 0, 
            p.suma_pagos if p.suma_pagos is not None else 0, 
            p.suma_gastos if p.suma_gastos is not None else 0, 
            p.created_at
        ) 
        for p in proyectos
    ]

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 10:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num in [5,6,7,8,9]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_proveedores(proveedores):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Proveedores_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Proveedores')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Razón Social','RFC','Nombre Comercial','Domicilio','Teléfono','Estado','Contacto','Email','Email Opción',
               'Banco','Clabe','Cuenta','Financiamiento','Días Crédito','Estatus']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 4 or col_num == 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        if col_num == 0:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 50

    proveedores_ids = proveedores.values_list('nombre', flat=True).distinct()
    proveedores_unicos = Proveedor.objects.filter(id__in=proveedores_ids, completo=True).count()

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordtec. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.cell(row=3, column= columna_max, value="Número de proveedores:")
    ws.cell(row=3, column = columna_max + 1, value=proveedores_unicos).style = number_style

    rows = proveedores.values_list('nombre__razon_social','nombre__rfc','nombre__nombre_comercial','domicilio','telefono','estado__nombre',
                                   'contacto','email','email_opt','banco__nombre','clabe','cuenta','financiamiento','dias_credito',
                                   'estatus__nombre'
                              )

    

    #for row, subtotal, iva, total in zip(rows,subtotales, ivas, totales):
    for row in rows:
        row_num += 1
        #row_with_additional_columns = list(row) + [subtotal, iva, total]  # Agrega el subtotal a la fila existente
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 13:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = number_style
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)