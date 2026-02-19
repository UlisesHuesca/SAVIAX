from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse, FileResponse
from django.db.models.functions import Concat, Cast, Replace
from django.db.models import Value, Sum, Case, When, F, Value, Q, DecimalField, Avg, IntegerField
from django.db import models

from django.contrib import messages
 
from django.core.mail import EmailMessage, BadHeaderError
from smtplib import SMTPException
from django.core.paginator import Paginator
from django.core.files.base import ContentFile
from django.conf import settings
from io import BytesIO
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name
from django.db.models import OuterRef, Subquery
import socket

from solicitudes.models import Proyecto, Subproyecto
from dashboard.models import Inventario, Order, ArticulosparaSurtir, ArticulosOrdenados, Inventario_Batch, Product, Marca
from dashboard.forms import  Inventario_BatchForm
from user.models import Profile, User
from .models import ArticulosRequisitados, Requis, Devolucion, Devolucion_Articulos, Tipo_Devolucion
from entradas.models import Entrada, EntradaArticulo
from requisiciones.models import Salidas, ValeSalidas
from .filters import ArticulosparaSurtirFilter, SalidasFilter, EntradasFilter, DevolucionFilter, RequisFilter, Historical_Articulos_surtir_Filter
from .forms import SalidasForm, ArticulosRequisitadosForm, ValeSalidasForm, ValeSalidasProyForm, RequisForm, Rechazo_Requi_Form, DevolucionArticulosForm, DevolucionForm, ValeSalidasForm_Ext
from solicitudes.filters import SolicitudesFilter
from tesoreria.models import Pago

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt
from datetime import date, datetime

import json
import csv
import os
import io
import ast # Para leer el csr many to many
import decimal
import base64

#PDF generator

from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter, portrait
from reportlab.rl_config import defaultPageSize

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame
from bs4 import BeautifulSoup
import urllib.request, urllib.parse, urllib.error


# Create your views here.
@login_required(login_url='user-login')
def liberar_stock(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    orden = Order.objects.get(id = pk)
    productos= ArticulosparaSurtir.objects.filter(articulos__orden = orden, surtir=True)
    vale_salida, created = ValeSalidas.objects.get_or_create(almacenista = usuario,complete = False,solicitud=orden)
    salidas = Salidas.objects.filter(vale_salida = vale_salida)
    cantidad_items = salidas.count()
    proyectos = Proyecto.objects.filter(activo=True)
    subproyectos = Subproyecto.objects.all()


    formVale = ValeSalidasProyForm()
    form = SalidasForm()
    users = Profile.objects.all()

    if request.method == 'POST':
        formVale = ValeSalidasProyForm(request.POST, instance=vale_salida)
        vale = formVale.save(commit=False)
        vale.complete = True
        for producto in productos:
            if producto.cantidad == 0:
                producto.salida = True
                producto.surtir = False
                producto.save()
        if formVale.is_valid():
            formVale.save()
            messages.success(request,'La salida se ha generado de manera exitosa')
            return redirect('solicitud-autorizada')

    context= {
        'proyectos':proyectos,
        'subproyectos':subproyectos,
        'productos':productos,
        'orden':orden,
        'form':form,
        'formVale':formVale,
        'users': users,
        'vale_salida':vale_salida,
        'cantidad_items':cantidad_items,
        'salidas':salidas,
        }
    return render(request,'requisiciones/liberar_stock.html',context)



@login_required(login_url='user-login')
def solicitud_autorizada(request):
    usuario = Profile.objects.get(staff__id=request.user.id)

    if usuario.tipo.almacen == True:
        #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(surtir=True), articulos__orden__autorizar = True)
        #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(surtir=True), articulos__orden__autorizar = True, articulos__orden__tipo__tipo = "normal")
        referencia_subquery = EntradaArticulo.objects.filter(
            articulo_comprado__oc__req__orden=OuterRef('articulos__orden'),  # Relacionamos con Order a través de las relaciones intermedias
            ).values('referencia')[:1]  # Solo tomamos el primer resultado
        productos= ArticulosparaSurtir.objects.filter(Q(articulos__orden__tipo__tipo = "normal")|Q(articulos__orden__tipo__tipo = "prod_terminado"), surtir=True, articulos__orden__autorizar = True).order_by('-created_at').annotate(
            referencia=Subquery(referencia_subquery))
    #else:
        #productos = Requis.objects.filter(complete=None)
    myfilter = ArticulosparaSurtirFilter(request.GET, queryset=productos)
    productos = myfilter.qs
    #Here is where call a function to generate XLSX, using Openpyxl library

    #Set up pagination
    p = Paginator(productos, 20)
    page = request.GET.get('page')
    productos_list = p.get_page(page)


    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_solicitud_autorizada_to_xls(productos)


    context= {
        'productos':productos,
        'productos_list':productos_list,
        'myfilter':myfilter,
        'usuario':usuario,
        }
    return render(request, 'requisiciones/solicitudes_autorizadas.html',context)

@login_required(login_url='user-login')
def solicitudes_autorizadas_pendientes(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    #productos= Requis.objects.filter(complete=True, autorizar=None)
    #Aquí aparecen todas las ordenes, es decir sería el filtro para administrador, el objeto Q no tiene propiedad conmutativa
    #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(requisitar=True), articulos__orden__autorizar = True )

    #if usuario.tipo.superintendente == True:
        #productos= Requis.objects.filter(complete=True, autorizar=None, orden__superintendente=usuario)
    if usuario.tipo.almacenista == True:
        #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(surtir=True), articulos__orden__autorizar = True)
        #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(surtir=True), articulos__orden__autorizar = True, articulos__orden__tipo__tipo = "normal")
        # Subquery para obtener la referencia de EntradaArticulo
        referencia_subquery = EntradaArticulo.objects.filter(
            articulo_comprado__oc__req__orden=OuterRef('articulos__orden'),  # Relacionamos con Order a través de las relaciones intermedias
        ).values('referencia')[:1]  # Solo tomamos el primer resultado

        # Consulta principal con annotate para agregar la referencia
        productos = ArticulosparaSurtir.objects.filter(
            salida=False, 
            surtir=False, 
            articulos__orden__autorizar=True, 
            articulos__orden__tipo__tipo="normal"
        ).order_by('-created_at').annotate(
            referencia=Subquery(referencia_subquery)  # Aquí agregamos el subquery como un nuevo campo
        )
    #else:
        #productos = Requis.objects.filter(complete=None)
    myfilter = ArticulosparaSurtirFilter(request.GET, queryset=productos)
    productos = myfilter.qs

    #Set up pagination
    p = Paginator(productos, 20)
    page = request.GET.get('page')
    productos_list = p.get_page(page)

    #Here is where call a function to generate XLSX, using Openpyxl library

    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_solicitud_autorizada_to_xls(productos)


    context= {
        'productos_list':productos_list,
        'productos':productos,
        'myfilter':myfilter,
        'usuario':usuario,
        }
    return render(request, 'requisiciones/solicitudes_autorizadas_no_surtidas.html',context)

def update_devolucion(request):
    data= json.loads(request.body)
    print('creo que no entra')
    action = data["action"]
    cantidad = decimal.Decimal(data["val_cantidad"])
    devolucion = data["devolucion"]
    producto_id = data["id"]
    comentario = data["comentario"]
    devolucion = Devolucion.objects.get(id = devolucion)
    print(devolucion,comentario,producto_id,cantidad,action)
    if devolucion.tipo.nombre == "SALIDA":
        producto = Salidas.objects.get(vale_salida=devolucion.salida.vale_salida, producto__id = producto_id,)
        inv_del_producto = Inventario.objects.get(producto = producto.producto.articulos.producto.producto)
    else:
        producto = ArticulosparaSurtir.objects.get(id = producto_id)
        inv_del_producto = Inventario.objects.get(producto = producto.articulos.producto.producto)
        print(producto)    


    if action == "add":
        print('sí')
        if producto.articulos.producto.producto.critico:
            print('sí')
            if producto.articulos.producto.producto.critico.nombre == "Crítico":
                print('sí')
                cantidad_total = int(cantidad)
                cantidad = 1
                for i in range(0, cantidad_total):
                    if devolucion.tipo.nombre == "SALIDA":
                        devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto= producto.producto, vale_devolucion = devolucion, complete=False)
                    else:
                        devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto=producto, vale_devolucion = devolucion, complete=False)
                    
                    producto.seleccionado = True
                    #Se le resta a la cantidad de artículos para surtir
                    producto.cantidad = producto.cantidad - cantidad
                    #La cantidad de la devolución es igual a la cantidad que se marcó en la devolución (daaa)
                    devolucion_articulos.cantidad = cantidad
                    devolucion_articulos.comentario = comentario
                    devolucion_articulos.precio = producto.precio
                    devolucion_articulos.complete = True
                    if producto.cantidad == 0: #Si la cantidad de artículos para surtir es igual a 0, si la cantidad a devolver es 0 entonces ya no se puede surtir
                        producto.surtir = False
                    messages.success(request,'Has agregado producto para devolución de manera exitosa')
                    producto.save()
                    devolucion_articulos.save() 
            else:
                print('no')
                cantidad_total = producto.cantidad - cantidad
                if cantidad_total < 0:
                    messages.error(request,f'La cantidad que se quiere ingresar sobrepasa la cantidad disponible. {cantidad_total} mayor que {producto.cantidad}')
                else:
                    if devolucion.tipo.nombre == "SALIDA":
                        devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto= producto.producto, vale_devolucion = devolucion, complete=False)
                    else:
                        devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto=producto, vale_devolucion = devolucion, complete=False)
                    
                    producto.seleccionado = True
                    #Se le resta a la cantidad de artículos para surtir
                    producto.cantidad = producto.cantidad - cantidad
                    #La cantidad de la devolución es igual a la cantidad que se marcó en la devolución (daaa)
                    devolucion_articulos.cantidad = cantidad
                    devolucion_articulos.comentario = comentario
                    devolucion_articulos.precio = producto.precio
                    devolucion_articulos.complete = True
                    if producto.cantidad == 0: #Si la cantidad de artículos para surtir es igual a 0, si la cantidad a devolver es 0 entonces ya no se puede surtir
                        producto.surtir = False
                    messages.success(request,'Has agregado producto para devolución de manera exitosa')
                    producto.save()
                    devolucion_articulos.save()
        else:
            print('else')
            cantidad_total = producto.cantidad - cantidad
            if cantidad_total < 0:
                messages.error(request,f'La cantidad que se quiere ingresar sobrepasa la cantidad disponible. {cantidad_total} mayor que {producto.cantidad}')
            else:
                if devolucion.tipo.nombre == "SALIDA":
                    devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto= producto.producto, vale_devolucion = devolucion, complete=False)
                else:
                    devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto=producto, vale_devolucion = devolucion, complete=False)
                    
                producto.seleccionado = True
                #Se le resta a la cantidad de artículos para surtir
                producto.cantidad = producto.cantidad - cantidad
                #La cantidad de la devolución es igual a la cantidad que se marcó en la devolución (daaa)
                devolucion_articulos.cantidad = cantidad
                devolucion_articulos.comentario = comentario
                devolucion_articulos.precio = producto.precio
                devolucion_articulos.complete = True
                if producto.cantidad == 0: #Si la cantidad de artículos para surtir es igual a 0, si la cantidad a devolver es 0 entonces ya no se puede surtir
                    producto.surtir = False
                messages.success(request,'Has agregado producto para devolución de manera exitosa')
                producto.save()
                devolucion_articulos.save()
    if action == "remove":
        articulo_id = data["id_articulo"]
        if devolucion.tipo.nombre == "SALIDA":
            item = Devolucion_Articulos.objects.get(id=articulo_id, producto=producto.producto, vale_devolucion = devolucion, complete = True)
        else:
            item = Devolucion_Articulos.objects.get(id=articulo_id, producto=producto, vale_devolucion = devolucion, complete = True)
        producto.cantidad = producto.cantidad + item.cantidad
        producto.seleccionado = False
        #Esta linea la pongo debido a que es necesaria para que vuelva a aparecer el producto cuando se elimina:
        producto.surtir = True
        messages.success(request,'Has eliminado un producto de tu listado')
        producto.save()
        item.delete()

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)

def update_devolucion_salida(request):
    data= json.loads(request.body)
    action = data["action"]
    cantidad = decimal.Decimal(data["val_cantidad"])
    devolucion = data["devolucion"]
    producto_id = data["id"]
    comentario = data["comentario"]
    referencia = data["referencia"]
    devolucion = Devolucion.objects.get(id = devolucion)
    #La referencia es necesaria a diferencia del otro lado ya que con este se puede saber el producto
    if referencia == 'None':
        referencia = None
    if devolucion.tipo.nombre == "SALIDA":
        producto = Salidas.objects.get(vale_salida=devolucion.salida.vale_salida, producto__id = producto_id, referencia=referencia)
        inv_del_producto = Inventario.objects.get(producto = producto.producto.articulos.producto.producto)
    else:
        producto = ArticulosparaSurtir.objects.get(id = producto_id)
        inv_del_producto = Inventario.objects.get(producto = producto.articulos.producto.producto)
        


    if action == "add":
        cantidad_total = producto.cantidad - cantidad
        if cantidad_total < 0:
            messages.error(request,f'La cantidad que se quiere ingresar sobrepasa la cantidad disponible. {cantidad_total} mayor que {producto.cantidad}')
        else:
            if devolucion.tipo.nombre == "SALIDA":
                devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto= producto.producto, vale_devolucion = devolucion, complete=False)
            else:
                devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto=producto, vale_devolucion = devolucion, complete=False)
            
            producto.seleccionado = True
            #Se le resta a la cantidad de artículos para surtir
            producto.cantidad = producto.cantidad - cantidad
            #La cantidad de la devolución es igual a la cantidad que se marcó en la devolución (daaa)
            devolucion_articulos.cantidad = cantidad
            devolucion_articulos.comentario = comentario
            devolucion_articulos.precio = producto.precio
            devolucion_articulos.complete = True
            if producto.cantidad == 0: #Si la cantidad de artículos para surtir es igual a 0, si la cantidad a devolver es 0 entonces ya no se puede surtir
                producto.surtir = False
            messages.success(request,'Has agregado producto para devolución de manera exitosa')
            producto.save()
            devolucion_articulos.save()
    if action == "remove":
        if devolucion.tipo.nombre == "SALIDA":
            item = Devolucion_Articulos.objects.get(producto=producto.producto, vale_devolucion = devolucion, complete = True)
        else:
            item = Devolucion_Articulos.objects.get(producto=producto, vale_devolucion = devolucion, complete = True)
        producto.cantidad = producto.cantidad + item.cantidad
        producto.seleccionado = False
        messages.success(request,'Has eliminado un producto de tu listado')
        producto.save()
        item.delete()

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)

@login_required(login_url='user-login')
def autorizar_devolucion(request, pk):
    devolucion= Devolucion.objects.get(id=pk)
    productos = Devolucion_Articulos.objects.filter(vale_devolucion = devolucion)
    
    if request.method == 'POST' and 'btnAutorizar' in request.POST:
        for producto in productos:
            if devolucion.tipo.nombre == "SALIDA":
                producto_surtir = Salidas.objects.get(id=devolucion.salida.id)
                inv_del_producto = Inventario.objects.get(producto = producto.producto.articulos.producto.producto) 
                inv_del_producto._change_reason = f'Esta es una devolucion desde un salida {devolucion.id}'
            else:
                producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)
                inv_del_producto = Inventario.objects.get(producto = producto_surtir.articulos.producto.producto)
                inv_del_producto._change_reason = f'Esta es una devolucion desde un surtimiento de inventario {devolucion.id}'
                qs = EntradaArticulo.objects.filter(
                    articulo_comprado__producto__producto=producto_surtir,
                    entrada__oc__req__orden=producto_surtir.articulos.orden,
                )

                if qs.exists():
                    if qs.count() > 1:
                        entrada = qs.order_by('id').first()
                    else:
                        entrada = qs.first()

                    if entrada and entrada.cantidad_por_surtir >= producto.cantidad:
                        entrada.cantidad_por_surtir -= producto.cantidad
                        entrada.save()
                    elif entrada:
                        entrada.cantidad_por_surtir = 0
                        entrada.agotado = True
                        entrada.save()
                else:
                    messages.error(request, 'No se encontró una entrada asociada para el producto.')

                    
            inv_del_producto.cantidad = inv_del_producto.cantidad + producto.cantidad
            inv_del_producto.save()
            messages.success(request,'Has autorizado exitosamente una devolución')
        devolucion.autorizada = True
        devolucion.save()
        return redirect('matriz-autorizar-devolucion')

    context= {
        'productos':productos,
        'devolucion':devolucion,
        }

    return render(request, 'requisiciones/autorizar_devolucion.html',context)

@login_required(login_url='user-login')
def cancelar_devolucion(request, pk):
    devolucion= Devolucion.objects.get(id=pk)
    productos = Devolucion_Articulos.objects.filter(vale_devolucion = devolucion)

    if request.method == 'POST' and 'btnCancelar' in request.POST:
        for producto in productos:
            if devolucion.tipo.nombre == "SALIDA":
                producto_surtir = Salidas.objects.get(salida=devolucion.salida)
            else:
                producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)
            producto_surtir.cantidad = producto_surtir.cantidad + producto.cantidad
            producto_surtir.surtir = True
            producto_surtir.seleccionado = False
            producto_surtir.save()
            #inv_del_producto.save()
        devolucion.autorizada = False
        devolucion.save()
        return redirect('matriz-autorizar-devolucion')

    context= {
        'productos':productos,
        'devolucion':devolucion,
        }

    return render(request, 'requisiciones/cancelar_devolucion.html',context)


@login_required(login_url='user-login')
def matriz_autorizar_devolucion(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    devoluciones= Devolucion.objects.filter(complete=True, autorizada=None)
    #print(devoluciones)

    
    myfilter = DevolucionFilter(request.GET, queryset = devoluciones)
    devoluciones = myfilter.qs

    #Set up pagination
    p = Paginator(devoluciones, 20)
    page = request.GET.get('page')
    devoluciones_list = p.get_page(page)

    #Here is where call a function to generate XLSX, using Openpyxl library

    #if request.method == 'POST' and 'btnExcel' in request.POST:
    #    return convert_solicitud_autorizada_to_xls(productos)


    context= {
        'devoluciones_list':devoluciones_list,
        'devoluciones':devoluciones,
        'myfilter':myfilter,
        'usuario':usuario,
        }
    return render(request, 'requisiciones/matriz_devoluciones_autorizar.html',context)

@login_required(login_url='user-login')
def salida_material(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    orden = Order.objects.get(id = pk)
    productos= ArticulosparaSurtir.objects.filter(articulos__orden = orden, surtir=True)
    vale_salida, created = ValeSalidas.objects.get_or_create(almacenista = usuario,complete = False,solicitud=orden)
    salidas = Salidas.objects.filter(vale_salida = vale_salida)
    cantidad_items = salidas.count()
    referencias_existentes = Salidas.objects.filter(vale_salida__solicitud = orden, vale_salida__complete=True).values_list('referencia', flat=True)
    for salida in salidas:
        # Obtener las referencias asociadas a cada producto
        referencias = EntradaArticulo.objects.filter(
            articulo_comprado__oc__req__orden=salida.producto.articulos.orden,articulo_comprado__producto__producto__articulos__producto__producto=salida.producto.articulos.producto.producto,
        ).exclude(referencia__in=referencias_existentes).values_list('referencia', flat=True)

        # Agregar las referencias al producto (puedes almacenar esto en una lista o agregarlo como un atributo)
        salida.referencias = list(referencias)
        
    formVale = ValeSalidasForm()
    form = SalidasForm()
    users = Profile.objects.all()

    if request.method == 'POST':
        formVale = ValeSalidasForm(request.POST, instance=vale_salida)

        falta_referencia_critica = False 

        # Iterar sobre las salidas para asignar las referencias seleccionadas y verificar
        for salida in salidas:
            referencia_seleccionada = request.POST.get(f'referencia_{salida.id}', None)

            if referencia_seleccionada:
                salida.referencia = referencia_seleccionada
                #Para limpiar el dato si es que no se selecciona nada
            else:
                salida.referencia = None
            salida.save()
            critico = getattr(salida.producto.articulos.producto.producto, 'critico', None)
            # Verificar si el producto es crítico y no tiene referencia
            if critico and getattr(critico, 'nombre', '') == 'Crítico' and not salida.referencia:
                falta_referencia_critica = True
        # Si falta una referencia crítica, mostrar mensaje de error
        #if falta_referencia_critica:
            #messages.error(request, 'Falta una referencia de asignar')
        if formVale.is_valid():
            vale = formVale.save(commit=False)
            cantidad_salidas = 0
            cantidad_productos = productos.count()

            for producto in productos:
                producto.seleccionado = False
                if producto.cantidad == 0:
                    producto.salida = True
                    producto.surtir = False
                    cantidad_salidas = cantidad_salidas + 1
                producto.save()
            if cantidad_productos == cantidad_salidas:
                orden.requisitado == True #Esta variable creo que podría ser una variable estúpida
                orden.save()
            #vale.referencia = ref
            vale.complete = True
            vale.save()
            messages.success(request,'La salida se ha generado de manera exitosa')
            return redirect('reporte-salidas')
        if not formVale.is_valid():
            messages.error(request,'No capturaste el usuario')

    context= {
        'productos':productos,
        'form':form,
        'formVale':formVale,
        'users': users,
        #'disponible':disponible,
        'vale_salida':vale_salida,
        'cantidad_items':cantidad_items,
        'salidas':salidas,
        }

    return render(request, 'requisiciones/salida_material.html',context)


@login_required(login_url='user-login')
def salida_material_externo(request, pk):
    print('salida_externo')
    usuario = Profile.objects.get(staff__id=request.user.id)
    orden = Order.objects.get(id = pk)
    productos= ArticulosparaSurtir.objects.filter(articulos__orden = orden, surtir=True)
    vale_salida, created = ValeSalidas.objects.get_or_create(almacenista = usuario,complete = False,solicitud=orden)
    salidas = Salidas.objects.filter(vale_salida = vale_salida)
    cantidad_items = salidas.count()
    referencias_existentes = Salidas.objects.filter(vale_salida__solicitud = orden, vale_salida__complete=True).values_list('referencia', flat=True)
    for salida in salidas:
        # Obtener las referencias asociadas a cada producto
        referencias = EntradaArticulo.objects.filter(
            articulo_comprado__oc__req__orden=salida.producto.articulos.orden,articulo_comprado__producto__producto__articulos__producto__producto=salida.producto.articulos.producto.producto,
        ).exclude(referencia__in=referencias_existentes).values_list('referencia', flat=True)

        # Agregar las referencias al producto (puedes almacenar esto en una lista o agregarlo como un atributo)
        salida.referencias = list(referencias)
        
    formVale = ValeSalidasForm_Ext()
    form = SalidasForm()
    users = Profile.objects.all()

    if request.method == 'POST':
        formVale = ValeSalidasForm_Ext(request.POST, instance=vale_salida)

        falta_referencia_critica = False 

        # Iterar sobre las salidas para asignar las referencias seleccionadas y verificar
        for salida in salidas:
            referencia_seleccionada = request.POST.get(f'referencia_{salida.id}', None)

            if referencia_seleccionada:
                salida.referencia = referencia_seleccionada
                #Para limpiar el dato si es que no se selecciona nada
            else:
                salida.referencia = None
            salida.save()
            critico = getattr(salida.producto.articulos.producto.producto, 'critico', None)
            # Verificar si el producto es crítico y no tiene referencia
            if critico and getattr(critico, 'nombre', '') == 'Crítico' and not salida.referencia:
                falta_referencia_critica = True
        # Si falta una referencia crítica, mostrar mensaje de error
        #if falta_referencia_critica:
            #messages.error(request, 'Falta una referencia de asignar')
        if formVale.is_valid():
            vale = formVale.save(commit=False)
            cantidad_salidas = 0
            cantidad_productos = productos.count()

            for producto in productos:
                producto.seleccionado = False
                if producto.cantidad == 0:
                    producto.salida = True
                    producto.surtir = False
                    cantidad_salidas = cantidad_salidas + 1
                producto.save()
            if cantidad_productos == cantidad_salidas:
                orden.requisitado == True #Esta variable creo que podría ser una variable estúpida
                orden.save()
            #vale.referencia = ref
            vale.complete = True
            vale.save()
            messages.success(request,'La salida se ha generado de manera exitosa')
            return redirect('reporte-salidas')
        if not formVale.is_valid():
            messages.error(request,'No capturaste el usuario')

    context= {
        'productos':productos,
        'form':form,
        'formVale':formVale,
        'users': users,
        #'disponible':disponible,
        'vale_salida':vale_salida,
        'cantidad_items':cantidad_items,
        'salidas':salidas,
        }

    return render(request, 'requisiciones/salida_material_externo.html',context)

@login_required(login_url='user-login')
def devolucion_material(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    orden = Order.objects.get(id = pk)
    productos_sel = ArticulosparaSurtir.objects.filter(articulos__orden = orden, surtir=True)
    tipo = Tipo_Devolucion.objects.get(nombre ="APARTADO" )
    devolucion, created = Devolucion.objects.get_or_create(almacenista = usuario,complete = False,solicitud=orden, tipo=tipo)
    productos = Devolucion_Articulos.objects.filter(vale_devolucion = devolucion)
    cantidad_items = productos.count()
    form = DevolucionArticulosForm()
    form2 = DevolucionForm()

    form.fields['producto'].queryset = productos_sel

    if request.method == 'POST':
        if "agregar_devolucion" in request.POST:
            form2 = DevolucionForm(request.POST, instance=devolucion)
            if form2.is_valid():
                devolucion = form2.save(commit=False)
                devolucion.complete= True
                devolucion.hora = datetime.now().time()
                devolucion.fecha = date.today()
                devolucion.tipo.nombre = "SIN SALIDA" 
                devolucion.save()
                for producto in productos_sel:
                    producto.seleccionado = False
                    producto.save()
                messages.success(request,f'{usuario.staff.first_name}, Has hecho la devolución de manera exitosa')
                email = EmailMessage(
                    f'Cancelación de solicitud: {orden.folio}',
                    f'Estimado {orden.staff.staff.first_name} {orden.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {orden.folio} ha sido devuelta al almacén por {usuario.staff.first_name} {usuario.staff.last_name}, con el siguiente comentario {devolucion.comentario} para más información comunicarse al almacén.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                    settings.DEFAULT_FROM_EMAIL,
                    ['ulises_huesc@hotmail.com'],#orden.staff.staff.email],
                    )
                email.send()
                return redirect('solicitud-autorizada')

    context= {
        'orden':orden,
        'productos':productos,
        'form':form,
        'form2':form2,
        'devolucion': devolucion,
        'cantidad_items':cantidad_items,
        'productos_sel': productos_sel,
        }

    return render(request, 'requisiciones/devolucion_material.html',context)

@login_required(login_url='user-login')
def devolucion_material_salida(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    salidas = Salidas.objects.all()
    salida = salidas.get(id=pk)
    vale_salida = ValeSalidas.objects.get(id=salida.vale_salida.id)
    orden = Order.objects.get(id = vale_salida.solicitud.id)
    #Esta es la parte que varía de devolución de material, aquí los productos deben ser salida = True
    #productos_sel = ArticulosparaSurtir.objects.filter(articulos__orden = orden, salida=True)
    productos_sel = salidas.filter(id=pk)
    tipo = Tipo_Devolucion.objects.get(nombre ="SALIDA" )
    devolucion, created = Devolucion.objects.get_or_create(almacenista = usuario,complete = False,solicitud=orden,tipo =tipo, salida =salida)
    productos = Devolucion_Articulos.objects.filter(vale_devolucion = devolucion)
    cantidad_items = productos.count()
    form = DevolucionArticulosForm()
    form2 = DevolucionForm()

    form.fields['producto'].queryset = productos_sel

    if request.method == 'POST':
        if "agregar_devolucion" in request.POST:
            form2 = DevolucionForm(request.POST, instance=devolucion)
            if form2.is_valid():
                devolucion = form2.save(commit=False)
                devolucion.complete= True
                devolucion.hora = datetime.now().time()
                devolucion.fecha = date.today()
                devolucion.save()
                for producto in productos_sel:
                    producto.seleccionado = False
                    producto.save()
                messages.success(request,f'{usuario.staff.first_name}, Has hecho la devolución de manera exitosa')
                email = EmailMessage(
                    f'Cancelación de solicitud: {orden.folio}',
                    f'Estimado {orden.staff.staff.first_name} {orden.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {orden.folio} ha sido devuelta al almacén por {usuario.staff.first_name} {usuario.staff.last_name}, con el siguiente comentario {devolucion.comentario} para más información comunicarse al almacén.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                    'savia@vordtec.com',
                    ['ulises_huesc@hotmail.com'],#orden.staff.staff.email],
                    )
                email.send()
                return redirect('solicitud-autorizada')

    context= {
        'orden':orden,
        'productos':productos,
        'form':form,
        'form2':form2,
        'devolucion': devolucion,
        'cantidad_items':cantidad_items,
        'productos_sel': productos_sel,
        }

    return render(request, 'requisiciones/devolucion_salida.html',context)


def solicitud_autorizada_firma(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    #Aquí aparecen todas las ordenes, es decir sería el filtro para administrador
    productos= Salidas.objects.filter(producto__articulos__orden__autorizar = True, salida_firmada=False)
    myfilter = SalidasFilter(request.GET, queryset=productos)
    productos = myfilter.qs

    #Here is where XLSX is generated, using Openpyxl library | Aquí es donde se genera el XLSX
    if request.method == "POST" and 'btnExcel' in request.POST:

        return convert_solicitud_autorizada_orden_to_xls(productos)

    context= {
        'productos':productos,
        'myfilter':myfilter,
        'usuario':usuario,
        }
    return render(request, 'requisiciones/solicitudes_autorizadas_firma.html',context)


def update_salida(request):
    data= json.loads(request.body)
    action = data["action"]
    cantidad = decimal.Decimal (data["val_cantidad"])
    salida = data["salida"]
    producto_id = data["id"]
    id_salida =data["id_salida"]
    producto = ArticulosparaSurtir.objects.get(id = producto_id)
    vale_salida = ValeSalidas.objects.get(id = salida)
    inv_del_producto = Inventario.objects.get(producto = producto.articulos.producto.producto)
    entradas = EntradaArticulo.objects.filter(articulo_comprado__producto__producto = producto, agotado=False, entrada__oc__req__orden= producto.articulos.orden).aggregate(cantidad_surtir=Sum('cantidad_por_surtir'))
    suma_entradas = entradas['cantidad_surtir']

    #Si no existen entradas la suma_entradas es igual a None, lo convierto en 0 para que pueda pasar la condicional #Definitoria
    if suma_entradas == None:
        suma_entradas = 0

    if action == "add":
        #con cantidad total establezco si la "cantidad" no sobrepasa lo que tengo que surtir(producto.cantidad)     
        cantidad_total = producto.cantidad - cantidad
        if cantidad_total < 0:
            messages.error(request,f'La cantidad que se quiere ingresar sobrepasa la cantidad disponible. {cantidad_total} mayor que {producto.cantidad}')
        else:
            producto.seleccionado = True
            entradas_dir = EntradaArticulo.objects.filter(articulo_comprado__producto__producto=producto, agotado=False, entrada__oc__req__orden=producto.articulos.orden, articulo_comprado__producto__producto__articulos__orden__tipo__tipo = 'normal').order_by('id')

            try:
                EntradaArticulo.objects.filter(articulo_comprado__producto__producto__articulos__producto = inv_del_producto, articulo_comprado__producto__producto__articulos__orden__tipo__tipo = 'resurtimiento', agotado = False)
            
            except EntradaArticulo.DoesNotExist:
                entrada_res = None
            else:
                entrada_res = EntradaArticulo.objects.filter(articulo_comprado__producto__producto__articulos__producto = inv_del_producto, articulo_comprado__producto__producto__articulos__orden__tipo__tipo = 'resurtimiento', agotado = False).order_by('id')

            if entradas_dir.exists():
                print('entradas normales ??')
                entradas = EntradaArticulo.objects.filter(articulo_comprado__producto__producto = producto, agotado=False, entrada__oc__req__orden= producto.articulos.orden)
                
                for entrada in entradas:
                    if producto.cantidad > 0:
                        salida, created = Salidas.objects.get_or_create(producto=producto, vale_salida = vale_salida, complete=False)
                        salida.precio = entrada.articulo_comprado.precio_unitario
                        #Aquí siempre van a entrar los productos con criticidad ya que al tener entrada limitada a 1, entonces solo pueden caer en este >=
                        if entrada.cantidad_por_surtir >= cantidad:
                            salida.cantidad = cantidad
                            if producto.articulos.producto.producto.critico:
                                if producto.articulos.producto.producto.critico.nombre == 'Crítico':
                                    cantidad = 1
                            else:
                                cantidad = 0 #la cantidad se vuelve 0 porque si la condición se cumple indica que la cantidad por surtir es capaz de abastecer toda la cantidad
                            producto.cantidad = producto.cantidad - salida.cantidad
                            salida.entrada = entrada.id
                            entrada.cantidad_por_surtir = entrada.cantidad_por_surtir - salida.cantidad
                            salida.complete = True
                            if entrada.cantidad_por_surtir <= 0:
                                entrada.agotado = True
                            producto.save()
                            entrada.save()
                            salida.save()
                        elif entrada.cantidad_por_surtir < cantidad and cantidad > 0: #Le meto la condicional para que no se repita el proceso si la cantidad es igual o menor que 0 
                            salida.cantidad = entrada.cantidad_por_surtir #No puedo surtir mas que la cantidad que tengo disponible en la entrada
                            cantidad = cantidad - salida.cantidad #La nueva cantidad a surtir es la cantidad menos lo que ya salió
                            producto.cantidad = producto.cantidad - salida.cantidad
                            salida.entrada = entrada.id
                            salida.complete = True
                            entrada.agotado = True
                            entrada.cantidad_por_surtir = 0
                            #producto.salida =
                            #True si vuelvo la entrada de resurtimiento verdadera anulo la posibilidad de realizar más salidas
                            producto.save()
                            entrada.save()
                            salida.save()
                        inv_del_producto.cantidad_entradas = inv_del_producto.cantidad_entradas - salida.cantidad
                        #inv_del_producto.cantidad = inv_del_producto.cantidad - salida.cantidad si hago una salida que proviene de entradas voy a obtener un inv_del_producto negativo
                        inv_del_producto.save()
            elif entrada_res.exists():   #si hay resurtimiento
                for entrada in entrada_res:
                    if cantidad > 0: #Se cambia producto.cantidad, se tiene que comparar con la cantidad de la salida no contra la cantidad disponible
                        salida, created = Salidas.objects.get_or_create(producto=producto, vale_salida = vale_salida, complete=False)
                        #Que hace el código a continuación la cantidad de la salida se compara contra la cantidad por surtir de la entrada
                        #L1 si es mayor la se guarda la cantidad_ant 
                        #L2 se le resta a la cantidad lo que queda en la entrada, es decir la nueva cantidad es lo que no se pudo surtir con esa entrada(cantidad) 
                        #L3 la cantidad de la salida es igual a la cantidad original menos la cantidad que no se pudo surtir con esa entrada
                        #L4 se vacía la entrada y por lo tanto se marca como agotada.
                        # Y si no la cantidad de la salida es igual a la cantidad(puede ser modificada por el bucle anterior o no) y 
                        # entrada por surtir es igual a la cantidad por surtir menos la cantidad de la salida y la cantidad se agota 04/12/2024 
                        print('entrada_res',cantidad)
                        #Aquí siempre van a entrar los productos con criticidad ya que al tener entrada limitada a 1, entonces solo pueden caer en este >=
                        #if cantidad >= entrada.cantidad_por_surtir:
                        #    cantidad_ant = cantidad
                        #    if producto.articulos.producto.producto.critico:
                        #        if producto.articulos.producto.producto.critico.nombre == 'Crítico':
                        #            cantidad = 1
                        #    else:
                        #        cantidad = cantidad - entrada.cantidad_por_surtir
                        #    salida.cantidad = cantidad_ant - cantidad
                        #    entrada.cantidad_por_surtir = 0
                        #    entrada.agotado = True
                        if cantidad >= entrada.cantidad_por_surtir:
                            cantidad_ant = cantidad
                            cantidad = cantidad - entrada.cantidad_por_surtir
                            salida.cantidad = cantidad_ant - cantidad
                            entrada.cantidad_por_surtir = 0
                            entrada.agotado = True
                        else:
                            salida.cantidad = cantidad
                            entrada.cantidad_por_surtir = entrada.cantidad_por_surtir - salida.cantidad
                            cantidad = 0
                        producto.cantidad = producto.cantidad - salida.cantidad
                        salida.entrada = entrada.id
                        salida.complete = True
                        #if producto.cantidad_requisitar <= 0: #Esta línea se considera errónea 04/12/2024
                        #    producto.requisitar = False  #Esta línea se considera errónea 04/12/2024
                        if producto.cantidad <= 0:
                            producto.surtir = False
                        print(salida)
                        entrada.save()
                        producto.save()
                        inv_del_producto.cantidad_entradas = inv_del_producto.cantidad_entradas - salida.cantidad
                        inv_del_producto._change_reason = f'Esta es la salida de un artículo desde un resurtimiento de inventario {salida.id}'
                        salida.precio = entrada.articulo_comprado.precio_unitario
                        salida.save()
            else:    #si no hay resurtimiento
                print('ELSE')
                salida, created = Salidas.objects.get_or_create(producto=producto, vale_salida = vale_salida, complete=False)
                salida.cantidad = cantidad
                salida.entrada = 0
                salida.complete = True
                producto.cantidad = producto.cantidad - cantidad 
                if producto.cantidad_requisitar <= 0:
                    producto.requisitar = False
                if producto.articulos.orden.tipo.tipo == "prod_terminado":
                    salida.precio = producto.articulos.precio
                else:
                    salida.precio = inv_del_producto.price
                inv_del_producto._change_reason = f'Esta es la salida de inventario de un artículo'
                #inv_del_producto.cantidad = inv_del_producto.cantidad - salida.cantidad
            #inv_del_producto.cantidad_apartada = inv_del_producto.cantidad_apartada - salida.cantidad
            producto.save()
            inv_del_producto.save()
            salida.save()

       
        
    if action == "remove":
        item = Salidas.objects.get(vale_salida = vale_salida, id = id_salida)
        id_entrada = item.entrada
        if id_entrada != None:
            if id_entrada != 0:
                entrada = EntradaArticulo.objects.get(id=item.entrada)
                inv_del_producto.cantidad_entradas = inv_del_producto.cantidad_entradas + item.cantidad
                entrada.cantidad_por_surtir = entrada.cantidad_por_surtir + item.cantidad
                entrada.agotado = False
                entrada.save()
            #if entrada.entrada.oc.req.orden.tipo.tipo == "normal":
            #    inv_del_producto.cantidad_apartada = inv_del_producto.cantidad_apartada + item.cantidad
        if vale_salida.solicitud.tipo.tipo == "normal":
            inv_del_producto.cantidad_apartada = inv_del_producto.cantidad_apartada + item.cantidad
        #inv_del_producto.cantidad = inv_del_producto.cantidad + item.cantidad
        producto.seleccionado = False
        producto.salida= False
        producto.cantidad = producto.cantidad + item.cantidad
        producto.surtir = True
        producto.seleccionado_por = None 
        #producto.cantidad_requisitar = producto.cantidad_requisitar + producto.cantidad
        producto._change_reason = f'Esto es una eliminación de un artículo en una salida'
        inv_del_producto._change_reason = f'Esta es una eliminación de un artìculo en una salida {item.id}'
        producto.save()
        inv_del_producto.save()
        item.delete()

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)


@login_required(login_url='user-login')
def salida_material_usuario(request, pk):
    producto= Salidas.objects.get(id = pk)
    producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)

    if request.method == 'POST':
        producto.salida_firmada = True
        producto_surtir.salida = True
        producto_surtir.firma = True
        producto_surtir.save()
        producto.save()

        messages.success(request,f'Has realizado la salida del producto {producto.producto.articulos.producto.producto} con éxito')
        return redirect('solicitud-autorizada-firma')

    context= {
        'productos':producto,
    }

    return render(request, 'requisiciones/salida_material_usuario.html',context)

@login_required(login_url='user-login')
def matriz_salida_activos(request):
    productos = Salidas.objects.filter(validacion_activos = False, producto__articulos__producto__producto__activo = True)
    #producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)
    #activo = Activo.objects.filter(activo = productos.producto.producto)


    context= {
        'productos':productos,
    }

    return render(request, 'requisiciones/matriz_salida_activos.html',context)

@login_required(login_url='user-login')
def solicitud_autorizada_orden(request):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    #usuario = request.user.id

    perfil = Profile.objects.get(staff__id=request.user.id)
    ordenes = Order.objects.filter(requisitar = True, complete=True, autorizar=True, staff__distrito=perfil.distrito, requisitado = False).annotate(
                folio_num=Cast(Replace('folio', models.Value('PL'), models.Value('')), IntegerField())).order_by('-folio_num')
  

    if perfil.tipo.almacenista == True:
        ordenes = Order.objects.filter(requisitar = True, requisitado=False).annotate(
                folio_num=Cast(Replace('folio', models.Value('PL'), models.Value('')), IntegerField())).order_by('-folio_num')
        #ordenes = Order.objects.filter(requisitar = True, complete=True, autorizar =True)
    #perfil = Profile.objects.get(id=usuario)

    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(requisitar=True), articulos__orden__autorizar = True )
    myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    ordenes = myfilter.qs


    if request.method == "POST" and 'btnExcel' in request.POST:

        return convert_solicitud_autorizada_orden_to_xls(ordenes)

    context= {
        'ordenes':ordenes,
        'myfilter':myfilter,
        }

    return render(request, 'requisiciones/solicitudes_autorizadas_orden.html',context)


def detalle_orden(request, pk):
    orden = Order.objects.get(id=pk)
    productos = ArticulosOrdenados.objects.filter(orden=pk)

    context = {
        'productos': productos,
        'orden': orden,
     }
    return render(request,'requisiciones/orden_detail.html', context)


@login_required(login_url='user-login')
def requisicion_autorizacion(request):
    perfil = Profile.objects.get(staff__id=request.user.id)
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id

    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito

    #ordenes = Order.objects.filter(complete=True, autorizar=True, staff__distrito=perfil.distrito)
    if perfil.tipo.superintendente == True:
        requis = Requis.objects.filter(autorizar=None, orden__superintendente=perfil, complete =True)
    else:
        requis = Requis.objects.filter(complete=None)
    #requis = Requis.objects.filter(autorizar=None)


    context= {
        'requis':requis,
        }

    return render(request, 'requisiciones/requisiciones_autorizacion.html',context)

def requisicion_creada_detalle(request, pk):
    productos = ArticulosRequisitados.objects.filter(req = pk)
    requis = Requis.objects.get(id = pk)

    context = {
        'productos': productos,
        'requis': requis,
     }

    return render(request,'requisiciones/requisicion_creada_detalle.html', context)

def update_requisicion(request):
    data= json.loads(request.body)
    action = data["action"]
    producto_id = data["id"]
    pk = data["requi"]
    cantidad = decimal.Decimal(data["cantidad"])

    requi = Requis.objects.get(id=pk)
    #orden = Order.objects.get(id=requi.orden.id)
    producto = ArticulosparaSurtir.objects.get(id = producto_id)
    if action == "add":
        item, created = ArticulosRequisitados.objects.get_or_create(req=requi, producto = producto, cantidad = cantidad)
        producto.requisitar = False
        producto.seleccionado = True
        producto.save()
        item.save()
    if action == "remove":
        item = ArticulosRequisitados.objects.get(req = requi, producto = producto)
        articulo_requisitado = ArticulosparaSurtir.objects.get(id =producto_id)
        articulo_requisitado.requisitar = True
        articulo_requisitado.seleccionado = False
        articulo_requisitado.save()
        item.delete()

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)

def obtener_consecutivo(distrito, requis):
    # Obtener la última requisición del distrito basado en la fecha de creación
    ultima_requisicion = requis.filter(orden__staff__distrito=distrito, complete=True).order_by('-created_at').first()

    if not ultima_requisicion:
        # Si no hay ninguna requisición previa, devolver 1 (será el primer folio)
        return 1

    # Extraer el número de folio (después de la abreviatura del distrito)
    ultimo_numero_folio = int(ultima_requisicion.folio.replace(distrito.abreviado, ''))

    # Devolver el siguiente número
    return ultimo_numero_folio + 1



def requisicion_detalle(request, pk):
    #Vista de creación de requisición
    productos = ArticulosparaSurtir.objects.filter(articulos__orden__id = pk, requisitar= True)
    orden = Order.objects.get(id = pk)
    usuario = Profile.objects.get(staff__id=request.user.id)
    requisiciones = Requis.objects.all()
    requi, created = requisiciones.get_or_create(complete=False, orden=orden)
    requis = requisiciones.filter(orden__staff__distrito = usuario.distrito, complete = True)

    #for producto in productos:
    productos_requisitados = ArticulosRequisitados.objects.filter(req = requi)

    form = RequisForm()


    if request.method == 'POST':
        form = RequisForm(request.POST, instance=requi)
        requi.complete = True
        orden.requisitado = True
        conteo_pendientes_requisitar = productos.filter(requisitar = True).count()
        if conteo_pendientes_requisitar > 0: #cuento cuantos productos están pendientes por requisitar 
            orden.requisitado = False
        else:
            orden.requisitado = True
        for producto in productos:
            #Vuelve false para que desaparezca de la vista pero creo que debo evaluar si es la mejor manera lo mismo para orden.requisitar = False, esto me está causando problemas en la vista
            producto.seleccionado = False
            producto.save()
            #if producto.requisitar == False:
            #    orden.requisitado = False
            #    orden.save()
        if productos_requisitados:
            folio_consecutivo = obtener_consecutivo(usuario.distrito, requisiciones)
            requi.folio = str(usuario.distrito.abreviado) + str(folio_consecutivo).zfill(4)
            requi.save()
            form.save()
            orden.save()
            messages.success(request,f'Has realizado la requisición {requi.folio} con éxito')
            return redirect('solicitud-autorizada-orden')
        else:
             messages.error(request,'No se puede crear la requisición debido a que no hay productos agregados')
    #else:
        #messages.error(request,'El formulario no es válido')


    context = {
        'productos': productos,
        'productos_requisitados':productos_requisitados,
        'orden': orden,
        'requi':requi,
        'form':form,
        }

    return render(request,'requisiciones/detalle_requisitar_editar.html', context)

# Convertir la imagen a base64
def get_image_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode()


def requisicion_autorizar(request, pk):
    usuario = request.user.id
    perfil = Profile.objects.get(staff__id=usuario)
    #perfil = Profile.objects.get(id=usuario)
    requi = Requis.objects.get(id = pk)
    productos = ArticulosRequisitados.objects.filter(req = pk)
    costo_aprox = 0
    for producto in productos:
        costo_aprox = costo_aprox + producto.cantidad * producto.producto.articulos.producto.price

    if costo_aprox > 0:
        porcentaje = "{0:.2f}%".format((costo_aprox/requi.orden.subproyecto.presupuesto)*100)
    else:
        porcentaje = "0"
    resta = requi.orden.subproyecto.presupuesto - requi.orden.subproyecto.gastado - costo_aprox

    if request.method == 'POST':
        requi.requi_autorizada_por = perfil
        requi.approved_at_time = datetime.now().time()
        requi.approved_at = date.today()
        requi.autorizar = True
        requi.save()
        static_path = settings.STATIC_ROOT
        img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
        img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
        image_base64 = get_image_base64(img_path)
        logo_v_base64 = get_image_base64(img_path2)
        # Crear el mensaje HTML
        html_message = f"""
        <html>
            <head>
                <meta charset="UTF-8">
            </head>
            <body>
                <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                <p>Estimado {requi.orden.staff.staff.first_name} {requi.orden.staff.staff.last_name},</p>
                <p>Estás recibiendo este correo porque tu sol: {requi.orden.folio}| Req: {requi.folio} ha sido autorizada,</p>
                <p>por {requi.requi_autorizada_por.staff.first_name} {requi.requi_autorizada_por.staff.last_name}.</p>
                <p>El siguiente paso del sistema: Generación de OC</p>
                <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
            </body>
        </html>
        """
        try:
            email = EmailMessage(
                f'Requisición Autorizada {requi.folio}',
                body=html_message,
                from_email = settings.DEFAULT_FROM_EMAIL,
                to= ['ulises_huesc@hotmail.com', requi.orden.staff.staff.email],
                headers={'Content-Type': 'text/html'}
                )
            email.content_subtype = "html " # Importante para que se interprete como HTML
            email.send()
            messages.success(request,f'Has autorizado la requisición {requi.folio} con éxito')
        except (BadHeaderError, SMTPException, socket.gaierror) as e:
            error_message = f'Has autorizado la requisición {requi.folio} con éxito pero el correo de notificación no ha sido enviado debido a un error: {e}'
            messages.warning(request, error_message)
        return redirect('requisicion-autorizacion')

    context = {
        'productos': productos,
        'requis': requi,
        'costo_aprox': costo_aprox,
        'porcentaje': porcentaje,
        'resta': resta,
     }

    return render(request,'requisiciones/requisiciones_autorizar.html', context)

def requisicion_cancelar(request, pk):
    usuario = request.user.id
    perfil = Profile.objects.get(staff=usuario)
    requis = Requis.objects.get(id = pk)
    productos = ArticulosRequisitados.objects.filter(req = pk)

    if request.method == 'POST':
        form= Rechazo_Requi_Form(request.POST,instance=requis)
        if form.is_valid():
            requis.autorizada_por = perfil
            requis.autorizar = False
            requis.save()
            try:
                email = EmailMessage(
                    f'Requisición Rechazada {requis.folio}',
                    f'Estimado {requis.orden.staff.staff.first_name} {requis.orden.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requis.orden.folio}| Req: {requis.folio} ha sido rechazada,\n por {requis.autorizada_por.staff.first_name} {requis.autorizada_por.staff.last_name} por el siguiente motivo: \n " {requis.comentario_compras} ".\n\n Este mensaje ha sido automáticamente generado por SAVIA 2.0',
                    settings.DEFAULT_FROM_EMAIL,
                    ['ulises_huesc@hotmail.com',requis.orden.staff.staff.email],
                    )
                email.send()
                messages.error(request,f'Has cancelado la requisición {requis.folio}')
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'Has cancelado la requisición {requis.folio} con éxito, pero el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.warning(request, error_message)
            return redirect('requisicion-autorizacion')
    else:
        form = Rechazo_Requi_Form(instance=requis)


    context = {
        'productos': productos,
        'requis': requis,
        'form':form,
     }
    return render(request,'requisiciones/requisiciones_cancelar.html', context)

def render_pdf_view(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #Here ends conf.
    orden = Order.objects.get(id=pk)
    productos = ArticulosOrdenados.objects.filter(orden=pk)
    #salidas = Salidas.objects.filter(producto__articulos__orden__id=pk)


   #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760

    #Encabezado
    c.drawString(410,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'Almacén')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(490,caja_iso-10,'Subdirección Administrativa')
    c.drawString(30,caja_iso-20,'Número de documento')
    c.drawString(40,caja_iso-30,'F-ALM-N4-01.01')
    c.drawString(125,caja_iso-20,'Clasificación del documento')
    c.drawString(165,caja_iso-30,'Registro')
    c.drawString(230,caja_iso-20,'Nivel del documento')
    c.drawString(255,caja_iso-30, 'N5')
    c.drawString(315,caja_iso-20,'Revisión No.')
    c.drawString(327,caja_iso-30,'000')
    c.drawString(385,caja_iso-20,'Fecha de Emisión')
    c.drawString(395,caja_iso-30,'08/03/2024')
    c.drawString(490,caja_iso-20,'Fecha última modificación')
    c.drawString(525,caja_iso-30,'08/03/2024')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,575,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Solicitud')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,575) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor-8,585,575) #Linea 2 contorno
    c.drawInlineImage('static/images/logo vordtec_documento.png',45,747, 1.5 * cm, 0.75 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(120,caja_proveedor,'Infor')
    c.drawString(300,caja_proveedor, 'Detalles')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',11)
    c.drawString(30,caja_proveedor-20,'Vordtec de México S.A. de C.V.')
    c.setFont('Helvetica',9)
    c.drawString(30,caja_proveedor-60,'Proyecto')
    c.drawString(30,caja_proveedor-80,'Subproyecto:')
    c.drawString(30,caja_proveedor-100,'Fecha:')
    
    c.setFont('Helvetica-Bold',12)
    c.drawString(500,caja_proveedor-20,'Folio:')
    #c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_proveedor-20, orden.folio)

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(100,caja_proveedor-60, orden.proyecto.nombre)
    c.drawString(100,caja_proveedor-80, orden.subproyecto.nombre)
    c.drawString(100,caja_proveedor-100, orden.approved_at.strftime("%d/%m/%Y"))

    #Create blank list
    data =[]

    data.append(['''Código''', '''Producto''', '''Cantidad''','''Comentario'''])


    high = 540
    for producto in productos:
        data.append([producto.producto.producto.codigo, producto.producto.producto.nombre,producto.cantidad, producto.comentario])
        high = high - 18
    if high<0:
        high=0
    #print('alturaaaaaaaa')
    #print(high)

    c.setFillColor(black)
    width, height = letter
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]

    table = Table(data, colWidths=[1.2 * cm, 12 * cm, 1.5 * cm, 5.2 * cm,])
    table_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 8),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    table.setStyle(table_style)
    table_style2 = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), colors.black),
        ('FONTSIZE',(0,0),(-1,0), 6),
        #('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    table.setStyle(table_style)
    
    rows_per_page_first_page = 20
    rows_per_page_other_pages = 35

    total_rows = len(data) - 1  # Excluye el encabezado
    remaining_rows = total_rows - rows_per_page_first_page

    if remaining_rows <= 0:
        # Dibujar la tabla completa en la primera página si no hay suficientes filas para otra página
        table.wrapOn(c, c._pagesize[0], c._pagesize[1])
        table.drawOn(c, 20, high)  # Posición en la primera página
        c.setFillColor(prussian_blue)

        # Parte de los comentarios en la primera página
        c.setFillColor(black)
        c.drawString(100,high-95, orden.staff.staff.first_name +' '+ orden.staff.staff.last_name)
        c.drawString(100,high-115, orden.supervisor.staff.first_name +' '+ orden.supervisor.staff.last_name)
        c.setFillColor(prussian_blue)
        c.rect(20,30,565,30, fill=True, stroke=False)
        c.rect(20,high-100,70,15, fill=True, stroke=False)
        c.rect(20,high-120,70,15, fill=True, stroke=False)
        c.setFillColor(white)
        c.drawString(25,high-95,'Solicitado por:')
        c.drawString(25,high-115, 'Aprobado por:')
        c.setFillColor(prussian_blue)
        c.rect(20,high-40,565,25, fill=True, stroke=False)
        c.setFillColor(white)
        c.drawCentredString(320,high-40,'Observaciones')
        comentario = orden.comentario if orden.comentario else "No hay comentarios"
        options_conditions_paragraph = Paragraph(comentario, styleN)
        frame = Frame(25, 0, width, high-40, id='normal')
        frame.addFromList([options_conditions_paragraph], c)
        c.setFillColor(prussian_blue)
        c.rect(20,30,565,20, fill=True, stroke=False)
        c.setFillColor(white)

    else:
        # Dibujar las primeras 20 filas en la primera página
        first_page_data = data[:rows_per_page_first_page + 1]  # Incluye el encabezado
        first_page_table = Table(first_page_data, colWidths=[1.2 * cm, 12 * cm, 1.5 * cm, 5.2 * cm])
        first_page_table.setStyle(table_style)
        first_page_table.wrapOn(c, c._pagesize[0], c._pagesize[1])
        first_page_table.drawOn(c, 20, high + 190)

        # Parte de los comentarios en la primera página
        c.setFillColor(black)
        c.drawString(100,high+75, orden.staff.staff.first_name +' '+ orden.staff.staff.last_name)
        c.drawString(100,high+95, orden.supervisor.staff.first_name +' '+ orden.supervisor.staff.last_name)
        c.setFillColor(prussian_blue)
        c.rect(20,30,565,30, fill=True, stroke=False)
        c.rect(20,high+70,70,15, fill=True, stroke=False)
        c.rect(20,high+90,70,15, fill=True, stroke=False)
        c.setFillColor(white)
        c.drawString(25,high+75,'Solicitado por:')
        c.drawString(25,high+95, 'Aprobado por')
        c.setFillColor(prussian_blue)
        c.rect(20,high+155,565,25, fill=True, stroke=False)
        c.setFillColor(white)
        c.drawCentredString(310,high+165,'Observaciones')
        comentario = orden.comentario if orden.comentario else "No hay comentarios"
        options_conditions_paragraph = Paragraph(comentario, styleN)
        frame = Frame(25, 0, width, high+160, id='normal')
        frame.addFromList([options_conditions_paragraph], c)
        c.setFillColor(prussian_blue)
        c.rect(20,30,565,20, fill=True, stroke=False)
        c.setFillColor(white)

        # Agregar una nueva página y dibujar las filas restantes en las siguientes páginas
        remaining_data = data[rows_per_page_first_page + 1:]
        pages_needed = (len(remaining_data) + rows_per_page_other_pages - 1) // rows_per_page_other_pages
        
        for page in range(pages_needed):
            c.showPage()
            page_start = page * rows_per_page_other_pages
            page_end = page_start + rows_per_page_other_pages
            page_data = remaining_data[page_start:page_end]
            
            if page_data:
                remaining_table = Table(page_data, colWidths=[1.2 * cm, 12 * cm, 1.5 * cm, 5.2 * cm])
                remaining_table.setStyle(table_style2)
                remaining_table.wrapOn(c, c._pagesize[0], c._pagesize[1])
                remaining_table_height = len(page_data) * 18
                remaining_table_y = c._pagesize[1] - 70 - remaining_table_height - 10
                remaining_table.drawOn(c, 20, remaining_table_y)  # Posición en las páginas adicionales
            caja_proveedor = caja_iso - 65
            c.setFont('Helvetica',12)
            c.setFillColor(prussian_blue)
            # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
            c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
            c.setFillColor(white)
            c.setLineWidth(.2)
            c.setFont('Helvetica-Bold',14)
            c.drawCentredString(280,755,'Solicitud')
            c.setLineWidth(.3) #Grosor
            c.drawInlineImage('static/images/logo vordtec_documento.png',45,747, 1.5 * cm, 0.75 * cm) #Imagen vortec
            #Encabezado
            c.setFillColor(black)
            c.setLineWidth(.2)
            c.setFont('Helvetica',8)
            caja_iso = 760

            #Encabezado
            c.drawString(410,caja_iso,'Preparado por:')
            c.drawString(420,caja_iso-10,'Almacén')
            c.drawString(520,caja_iso,'Aprobación')
            c.drawString(490,caja_iso-10,'Subdirección Administrativa')
            c.drawString(30,caja_iso-20,'Número de documento')
            c.drawString(40,caja_iso-30,'F-ALM-N4-01.01')
            c.drawString(125,caja_iso-20,'Clasificación del documento')
            c.drawString(165,caja_iso-30,'Registro')
            c.drawString(230,caja_iso-20,'Nivel del documento')
            c.drawString(255,caja_iso-30, 'N5')
            c.drawString(315,caja_iso-20,'Revisión No.')
            c.drawString(327,caja_iso-30,'000')
            c.drawString(385,caja_iso-20,'Fecha de Emisión')
            c.drawString(395,caja_iso-30,'08/03/2024')
            c.drawString(490,caja_iso-20,'Fecha última modificación')
            c.drawString(525,caja_iso-30,'08/03/2024')

    c.save()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Solicitud_' + str(orden.folio) +'.pdf')

def reporte_entradas(request):
    entradas = EntradaArticulo.objects.filter(entrada__completo = True, articulo_comprado__producto__producto__articulos__producto__producto__servicio = False
                                              ).filter(Q(almacenado = True)|Q(agotado = True)|Q(cantidad_por_surtir = 0) ).order_by('-entrada__id')
    myfilter = EntradasFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs

    #Set up pagination
    p = Paginator(entradas, 50)
    page = request.GET.get('page')
    entradas_list = p.get_page(page)

    for entrada in entradas_list:
        if entrada.articulo_comprado.oc.moneda.nombre == "DOLARES":
            entrada.articulo_comprado.precio_unitario = entrada.articulo_comprado.precio_unitario * entrada.articulo_comprado.oc.tipo_de_cambio

    if request.method == "POST" and 'btnExcel' in request.POST:

        return convert_entradas_to_xls(entradas)


    context = {
        'entradas_list':entradas_list,
        'entradas':entradas,
        'myfilter':myfilter,
        }

    return render(request,'requisiciones/reporte_entradas.html', context)

def reporte_salidas(request):
    salidas = Salidas.objects.filter(producto__isnull=False).order_by('-vale_salida')
    myfilter = SalidasFilter(request.GET, queryset=salidas)
    salidas = myfilter.qs
    salidas_filtradas = salidas.filter(producto__articulos__producto__producto__servicio = False,)

    if request.method == "POST" and 'btnExcel' in request.POST:
        return convert_salidas_to_xls(salidas_filtradas)
    
     #Set up pagination
    p = Paginator(salidas, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)


    context = {
        'salidas':salidas,
        'salidas_list':salidas_list,
        'myfilter':myfilter,
        }

    return render(request,'requisiciones/reporte_salidas.html', context)


def editar_cliente(request, salida_id):
    if request.method == "POST":
        salida = get_object_or_404(Salidas, id=salida_id)
        nuevo_cliente = request.POST.get('cliente', '').strip()
        salida.cliente = nuevo_cliente
        salida.save()
        messages.success(request,f'Has actualizado los datos del cliente correctamente. Salida:{salida.id}-Vale:{salida.vale_salida.id}')
        return redirect('reporte-salidas')  # Ajusta el nombre de tu vista principal
    
@login_required(login_url='user-login')
def historico_articulos_para_surtir(request):
    registros = ArticulosparaSurtir.history.all()

    myfilter = Historical_Articulos_surtir_Filter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        }

    return render(request,'requisiciones/historicos_articulos_para_surtir.html',context)

@login_required(login_url='user-login')
def historico_salidas(request):
    registros = Salidas.history.all()

    context = {
        'registros':registros,
        }

    return render(request,'requisiciones/historico_salidas.html',context)


def convert_solicitud_autorizada_to_xls(productos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Solicitudes_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    number_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)

    columns = ['Folio','Solicitante','Proyecto','Subproyecto','Código','Artículo','Creado','Cantidad']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordtec. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style

    row_num = 1

    for producto in productos.select_related(
        "articulos__orden",
        "articulos__orden__proyecto",
        "articulos__orden__subproyecto",
        "articulos__producto__producto"
    ):
        row = [
            producto.articulos.orden.folio,
            f"{producto.articulos.orden.staff.staff.first_name} {producto.articulos.orden.staff.staff.last_name}",
            producto.articulos.orden.proyecto.nombre if producto.articulos.orden.proyecto else '',
            producto.articulos.orden.subproyecto.nombre if producto.articulos.orden.subproyecto else '',
            producto.articulos.producto.producto.codigo,
            producto.articulos.producto.producto.nombre,
            producto.articulos.orden.approved_at,   # ← tipo datetime.date
            producto.cantidad
        ]

        row_num += 1
        for col_num, value in enumerate(row):
            if col_num == 6:  # fecha
                ws.cell(row=row_num, column=col_num+1, value=value).style = date_style
            elif col_num in [7, 4]:  # números
                ws.cell(row=row_num, column=col_num+1, value=value).style = number_style
            else:  # texto
                ws.cell(row=row_num, column=col_num+1, value=str(value)).style = body_style

    # limpiar hoja inicial
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return response


def convert_solicitud_autorizada_orden_to_xls(ordenes):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Solicitudes_pend_requisicion' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)

    columns = ['Folio','Solicitante','Proyecto','Subproyecto','Creado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordtec. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style

    rows = ordenes.values_list('folio',Concat('staff__staff__first_name',Value(' '),'staff__staff__last_name'),
                            'proyecto__nombre','subproyecto__nombre','created_at')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)
#Aquí termina la implementación del XLSX

def convert_entradas_to_xls(entradas):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Entradas_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Entradas')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)

    columns = ['Folio Solicitud','Fecha','Solicitante','Proyecto','Subproyecto','Código','Articulo','Cantidad','Moneda','Tipo de Cambio','Precio']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordtec. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    
    """rows = []
    for entrada in entradas:
        # Obtén todos los pagos relacionados con esta entrada
        pagos = Pago.objects.filter(oc=entrada.entrada.oc)
        # Calcula el tipo de cambio promedio de estos pagos
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']

        # Usar el tipo de cambio de los pagos, si existe. De lo contrario, usar el tipo de cambio de la entrada
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or entrada.entrada.oc.tipo_de_cambio

        row = [
            entrada.entrada.oc.req.orden.id,
            entrada.created_at,
            f"{entrada.entrada.oc.req.orden.staff.staff.first_name} {entrada.entrada.oc.req.orden.staff.staff.last_name}",
            entrada.entrada.oc.req.orden.proyecto.nombre,
            entrada.entrada.oc.req.orden.subproyecto.nombre,
            entrada.entrada.oc.req.orden.area.nombre,
            entrada.articulo_comprado.producto.producto.articulos.producto.producto.codigo,
            entrada.articulo_comprado.producto.producto.articulos.producto.producto.nombre,
            entrada.cantidad,
            entrada.entrada.oc.moneda.nombre,
            tipo_de_cambio,
            entrada.articulo_comprado.precio_unitario,
        ]
        if row[9] == "DOLARES":
            if row[10] is None or row[10] < 15:
                row[10] = 17  # O cualquier valor predeterminado que desees
        elif row[10] is None:
                row[10] = ""

        rows.append(row)
    """
    rows = entradas.values_list(
        'entrada__oc__req__orden__id',
        'created_at',
        Concat('entrada__oc__req__orden__staff__staff__first_name',Value(' '),'entrada__oc__req__orden__staff__staff__last_name'),
        'entrada__oc__req__orden__proyecto__nombre',
        'entrada__oc__req__orden__subproyecto__nombre',
        'articulo_comprado__producto__producto__articulos__producto__producto__codigo',
        'articulo_comprado__producto__producto__articulos__producto__producto__nombre',
        'cantidad',
        'entrada__oc__moneda__nombre', #8
        Case(                                          #9
            When(entrada__oc__tipo_de_cambio__isnull=False, then = F('entrada__oc__tipo_de_cambio')),
            #When(Pago.objects.filter(oc=F('entrada.oc')).exclude(tipo_de_cambio__isnull=True).exists(),then=F('Pago__tipo_de_cambio')),
            default=1.0,  # Puedes establecer un valor predeterminado si no hay tipo de cambio.
            output_field=DecimalField(max_digits=10, decimal_places=2),  # Asegura que el campo sea decimal si es necesario.
        ),
        'articulo_comprado__precio_unitario', #10
    )

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 4:
                (ws.cell(row = row_num, column = col_num + 1, value=row[col_num])).style = date_style
            if col_num == 9:
                (ws.cell(row = row_num, column = col_num + 1, value=row[col_num])).style = money_style
            if col_num == 10:
                if row[8] == "DOLARES":
                    precio_unitario = row[10]
                    tipo_de_cambio = row[9]
                    (ws.cell(row=row_num, column=col_num + 1, value=precio_unitario * tipo_de_cambio)).style = money_style
                else:
                    (ws.cell(row=row_num, column=col_num + 1, value=row[col_num])).style = money_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)
#Aquí termina la implementación del XLSX

def convert_salidas_to_xls(salidas):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Salidas_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Salidas')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    number_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)

    columns = ['Folio Solicitud','Fecha','Solicitante','Proyecto','Subproyecto','Área','Código','Articulo','Material recibido por','Cantidad','Precio','Total']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA VORDTEC. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Grupo Vordcab S.A. de C.V.}')).style = messages_style
    

    rows = salidas.values_list(
        'producto__articulos__orden__folio',
        'created_at',
        Concat('producto__articulos__orden__staff__staff__first_name',Value(' '),'producto__articulos__orden__staff__staff__last_name'),
        'producto__articulos__orden__proyecto__nombre',
        'producto__articulos__orden__subproyecto__nombre',
        'producto__articulos__orden__area__nombre',
        'producto__articulos__producto__producto__codigo',
        'producto__articulos__producto__producto__nombre',
        Concat('vale_salida__material_recibido_por__staff__first_name',Value(' '),'vale_salida__material_recibido_por__staff__last_name'),
        'cantidad',
        Case(
            When(precio__gt = 0, then='precio'),
            When(producto__precio__gt = 0, then='producto__precio'),
            default='producto__articulos__producto__price',
        )
    )

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 1:
                value = (row[col_num]).date()
                (ws.cell(row = row_num, column = col_num+1, value = value)).style = date_style
            if col_num == 9:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = number_style
            if col_num == 10:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
        ws.cell(row=row_num, column=len(row) + 1, value=f'=J{row_num} * K{row_num}').style = money_style
    
    (ws.cell(column = columna_max , row = 3, value=f'=SUM(L2:L{row_num})')).style = money_resumen_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)
#Aquí termina la implementación del XLSX

def render_entrada_pdf(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=portrait(letter))
    #Here ends conf.
    articulo = EntradaArticulo.objects.get(id=pk)
    vale = Entrada.objects.get(id = articulo.entrada.id)
    productos = EntradaArticulo.objects.filter(entrada= vale)
    styles = getSampleStyleSheet()
    styles['BodyText'].fontSize = 6

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 770
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)


    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'SUP. ADMON')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(520,caja_iso-10,'SUB ADM')
    #Segundo renglón
    c.drawString(150,caja_iso-25,'Número de documento')
    c.drawString(160,caja_iso-35,'SEOV-ALM-N4-01-02')
    c.drawString(245,caja_iso-25,'Clasificación del documento')
    c.drawString(275,caja_iso-35,'Controlado')
    c.drawString(355,caja_iso-25,'Nivel del documento')
    c.drawString(380,caja_iso-35, 'N5')
    c.drawString(440,caja_iso-25,'Revisión No.')
    c.drawString(452,caja_iso-35,'001')
    c.drawString(510,caja_iso-25,'Fecha de Emisión')
    c.drawString(525,caja_iso-35,'24-Oct.-18')


    c.drawString(510,caja_iso-50,'Folio: ')
    #c.drawString(530,caja_iso-50, str(vale.folio))
    c.drawString(510,caja_iso-60,'Fecha:')
    c.drawString(540,caja_iso-60,vale.entrada_date.strftime("%d/%m/%Y"))

    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(530,caja_iso-50, str(vale.id))
    

    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,caja_iso-15,250,20, fill=True, stroke=False) #Barra azul superior Orden de Compra

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,caja_iso-10,'Vale de Entrada Almacén')
    c.setLineWidth(.3) #Grosor

    c.drawInlineImage('static/images/logo vordtec_documento.png',45,caja_iso-40, 3 * cm, 1.5 * cm) #Imagen vortec
   

    data =[]
    productos_data = []
    high = 670
    data.append(['''Código''','''Producto''', '''Cantidad''', '''Unidad'''])
    for producto in productos:
        producto_nombre = Paragraph(producto.articulo_comprado.producto.producto.articulos.producto.producto.nombre, styles["BodyText"])
        data.append([producto.articulo_comprado.producto.producto.articulos.producto.producto.codigo, producto_nombre, producto.cantidad, producto.articulo_comprado.producto.producto.articulos.producto.producto.unidad])
        high = high - 18
        #Lo vuelvo a captura de otra manera para el código QR
        nombre_producto = producto.articulo_comprado.producto.producto.articulos.producto.producto.nombre
        codigo_producto = producto.articulo_comprado.producto.producto.articulos.producto.producto.codigo
        producto_info = {
            'codigo': codigo_producto,
            'nombre': nombre_producto,
            'cantidad': str(producto.cantidad),
            'unidad': str(producto.articulo_comprado.producto.producto.articulos.producto.producto.unidad),
        }
        productos_data.append(producto_info)
    
    
    # Generar el código QR
    #qr = qrcode.QRCode(
    #    version=1,
    #    error_correction=qrcode.constants.ERROR_CORRECT_L,
    #    box_size=10,
    #    border=4,
    #)
    #folio = str(vale.folio)
    #fecha = vale.created_at.strftime("%d/%m/%Y")
    #qr_info = {
    #    'folio': folio,
    #    'fecha': fecha,
    #    'productos': productos_data
    #}
    #qr_data = json.dumps(qr_info)
    #qr.add_data(qr_data)
    #qr.make(fit=True)

    # Generar la imagen del QR y guardarla
    #qr_image = qr.make_image(fill_color="black", back_color="white")
    #qr_image_path = '/tmp/temp_qr.png'
    #qr_image.save(qr_image_path)
    #c.drawInlineImage(qr_image_path, 500, 440, 100, 100)  # Reemplaza x, y, width, height con tus valores


    c.setFillColor(black)
    c.setFont('Helvetica',8)
    proyecto_y = 485 if high > 500 else high - 30

    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,proyecto_y - 5 ,350,20, fill=True, stroke=False) #3ra linea azul
    c.setFillColor(black)
    c.setFont('Helvetica',7)


    c.setFillColor(white)
    c.setLineWidth(.1)
    c.setFont('Helvetica-Bold',10)
    c.drawCentredString(120,proyecto_y,'Proyecto')
    c.drawCentredString(300,proyecto_y,'Subproyecto')

    c.setFont('Helvetica',8)
    c.setFillColor(black)
    #c.drawCentredString(120,proyecto_y - 15, str(vale.solicitud.proyecto.nombre))
    #c.drawCentredString(300,proyecto_y - 15, str(vale.solicitud.subproyecto.nombre))


    c.setFillColor(black)
    c.setFont('Helvetica',8)
    #c.line(135,high-200,215, high-200) #Linea de Autorizacion
    c.drawCentredString(150,proyecto_y - 30,'Recibió')
    if vale.almacenista:
        c.drawCentredString(150,proyecto_y - 40, vale.almacenista.staff.first_name +' '+vale.almacenista.staff.last_name)

    #c.line(370,proyecto_y - 20,430, proyecto_y - 20)
    #c.drawCentredString(400,proyecto_y - 30,'Recibió')
    #c.drawCentredString(400,proyecto_y - 40, vale.material_recibido_por.staff.staff.first_name +' '+vale.material_recibido_por.staff.staff.last_name)


    #c.line(240, high-200, 310, high-200)
    c.drawCentredString(280,proyecto_y - 30,'Proveedor')
    c.drawCentredString(280,proyecto_y - 40, vale.oc.proveedor.nombre.razon_social)

    c.setFont('Helvetica',10)
    c.setFillColor(prussian_blue)
    c.setFont('Helvetica', 9)
    c.setFillColor(black)

    c.setFillColor(prussian_blue)
    c.rect(20,proyecto_y - 65,565,20, fill=True, stroke=False)
    c.setFillColor(white)

    width, height = letter
    table = Table(data, colWidths=[2 * cm, 12.5 * cm, 2.5 * cm, 2.5 * cm,])
    table.setStyle(TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 10),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, high)
    c.save()
    c.showPage()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='vale_entrada_'+str(vale.id) +'.pdf')


def render_salida_pdf(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=portrait(letter))
    #Here ends conf.
    articulo = Salidas.objects.get(id=pk)
    vale = ValeSalidas.objects.get(id = articulo.vale_salida.id)
    productos = Salidas.objects.filter(vale_salida = vale)
    
    styles = getSampleStyleSheet()
    styles['BodyText'].fontSize = 6

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 770
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)


    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(430,caja_iso-10,'Almacén')
    c.drawString(510,caja_iso,'Aprobación')
    c.drawString(480,caja_iso-10,'Subdirección Admninistrativa')
    
    c.drawString(50,caja_iso-30,'Número de documento')
    c.drawString(60,caja_iso-40,'F-ALM-N4-01.02')
    c.drawString(145,caja_iso-30,'Clasificación del documento')
    c.drawString(175,caja_iso-40,'Registro')
    c.drawString(255,caja_iso-30,'Nivel del documento')
    c.drawString(280,caja_iso-40, 'N5')
    c.drawString(340,caja_iso-30,'Revisión No.')
    c.drawString(352,caja_iso-40,'001')
    c.drawString(400,caja_iso-30,'Fecha de Emisión')
    c.drawString(415,caja_iso-40,'01/09/2018')
    c.drawString(490,caja_iso-30,'Fecha última modificación')
    c.drawString(522,caja_iso-40,'08/03/2024')


    c.drawString(510,caja_iso-60,'Folio: ')
    c.drawString(530,caja_iso-60, str(vale.id))
    c.drawString(510,caja_iso-70,'Fecha:')
    c.drawString(540,caja_iso-70,vale.created_at.strftime("%d/%m/%Y"))


    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,caja_iso-15,250,20, fill=True, stroke=False) #Barra azul superior Orden de Compra

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,caja_iso-10,'Vale de Salida Almacén')
    c.setLineWidth(.3) #Grosor

    c.drawInlineImage('static/images/logo vordtec_documento.png',30,caja_iso-22, 2 * cm, 1 * cm) #Imagen vortec


    data =[]
    #high = 670
    data.append(['''Código''','''Producto''','''Referencia''','''Cliente''','''Cantidad''', '''Unidad''','''P.Unitario''', '''Importe'''])
    for producto in productos:
        producto_nombre = Paragraph(producto.producto.articulos.producto.producto.nombre, styles["BodyText"])
        data.append([producto.producto.articulos.producto.producto.codigo, producto_nombre, producto.referencia, producto.cliente, producto.cantidad, producto.producto.articulos.producto.producto.unidad, producto.precio, producto.precio * producto.cantidad])
        #high = high - 18
   
    c.setFillColor(black)
    c.setFont('Helvetica',8)
    proyecto_y = 760

    
    width, height = letter
    table = Table(data, colWidths=[1.5 * cm, 6.5 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm, 2.0 * cm])
    table.setStyle(TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 10),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ]))
    # 1) Decidimos el "top" donde quieremos que INICIE la tabla (debajo del bloque azul o debajo del encabezado)
    top_table_y = proyecto_y - 85   # ajusta este margen a tu gusto
   
    # 2) Calculamos alto real
    tw, th = table.wrap(width - 40, height)
    # 3) Dibuja usando esquina inferior = top - alto
    table.drawOn(c, 20, top_table_y - th)

    gap_after_table = 25
    proyecto_y = top_table_y - th - gap_after_table  # aquí anclas tu bloque 

    # ---- tu bloque original usando proyecto_y ----
    c.setFillColor(prussian_blue)
    c.rect(20, proyecto_y - 5, 250, 20, fill=True, stroke=False)

    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 10)
    c.drawCentredString(70,  proyecto_y, 'Proyecto')
    c.drawCentredString(165, proyecto_y, 'Subproyecto')

    c.setFillColor(colors.black)
    c.setFont('Helvetica', 8)
    c.drawCentredString(70,  proyecto_y - 15, str(vale.solicitud.proyecto.nombre))
    c.drawCentredString(165, proyecto_y - 15, str(vale.solicitud.subproyecto.nombre))

    # Firmas (igual que ya lo tienes, pero basadas en proyecto_y)
    c.setFont('Helvetica', 8)
    c.drawCentredString(100, proyecto_y - 30, 'Entregó')
    c.drawCentredString(100, proyecto_y - 40, vale.almacenista.staff.first_name +' '+vale.almacenista.staff.last_name)

    #c.line(370, proyecto_y - 20, 430, proyecto_y - 20)
    c.drawCentredString(500, proyecto_y - 30, 'Recibió')
    if vale.material_recibido_por:
        c.drawCentredString(500, proyecto_y - 40,vale.material_recibido_por.staff.first_name +' '+vale.material_recibido_por.staff.last_name)
    elif vale.cliente:
        c.drawCentredString(500, proyecto_y - 40, str(vale.cliente+ ' ' + vale.destino))

    c.drawCentredString(300, proyecto_y - 30, 'Autorizó')
    c.drawCentredString(300, proyecto_y - 40, vale.solicitud.staff.staff.first_name + '--' + vale.solicitud.staff.staff.last_name)

    # Franja azul final debajo de firmas
    c.setFillColor(prussian_blue)
    c.rect(20, proyecto_y - 65, 565, 20, fill=True, stroke=False)




   
    c.save()
    c.showPage()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='vale_salida_'+str(vale.id) +'.pdf')

@login_required(login_url='user-login')
def requisiciones_status(request):
    #pk = request.session.get('selected_profile_id')
    #perfil = Profile.objects.get(id = pk)
    usuario = Profile.objects.get(staff__id=request.user.id)
    
    if usuario.tipo.nombre == "PROVEEDORES" or usuario.tipo.nombre == "VIS_ADQ":
        requis = Requis.objects.filter(complete = True).order_by('-folio')
    else:
        requis = Requis.objects.filter(orden__distrito = usuario.distrito, complete = True).order_by('-folio')
   
    #requis = Requis.objects.filter(autorizar=True, colocada=False)

    myfilter = RequisFilter(request.GET, queryset=requis)
    requis = myfilter.qs

     #Set up pagination
    p = Paginator(requis, 50)
    page = request.GET.get('page')
    requis_list = p.get_page(page)

    #if request.method == 'POST' and 'btnExcel' in request.POST:
    #    return convert_excel_matriz_requis(requis)

    context= {
        'myfilter': myfilter,
        'requis':requis,
        'requis_list':requis_list,
        }

    return render(request, 'requisiciones/requisiciones.html',context)

def render_requisicion_pdf_view(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Here ends conf.
    requisicion = Requis.objects.get(id=pk)
    productos = ArticulosRequisitados.objects.filter(req=pk)
    #salidas = Salidas.objects.filter(producto__articulos__orden__id=pk)
   #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)

    if requisicion.colocada:
        estado = 'Colocada'
        color = prussian_blue
    elif requisicion.autorizar:
        estado = 'Autorizada'
        color = Color(0.0, 0.39, 0.0)
    elif requisicion.autorizar == False:
        estado = 'Cancelada'
        color = rojo
    else:
        estado = 'No Autorizado Aún'
        color = prussian_blue

    c.drawInlineImage('static/images/logo vordtec_documento.png',45,730, 3 * cm, 1.5 * cm) #Imagen vortec
    #Encabezado
    c.drawString(415,caja_iso-5,'Preparado por:')
    c.drawString(515,caja_iso-5,'Aprobación')
    c.drawString(20,caja_iso-35,'Número de documento')
    c.drawString(125,caja_iso-35,'Clasificación del documento')
    c.drawString(245,caja_iso-35,'Nivel del documento')
    c.drawString(335,caja_iso-35,'Revisión No.')
    c.drawString(400,caja_iso-35,'Fecha de Emisión')
    c.drawString(485,caja_iso-35,'Fecha última modificación')
    c.setFont('Helvetica',8)
    c.drawString(485,caja_iso-15,'Subdirección administrativa')
    c.drawString(30,caja_iso-45,'F-ADQ-N4-01.01')
    c.drawString(158,caja_iso-45,'Registro')
    c.drawString(415,caja_iso-15,'Adquisiciones')
    c.drawString(273,caja_iso-45, 'N5')
    c.drawString(347,caja_iso-45,'001')
    c.drawString(415,caja_iso-45,'22-Nov.-17')
    c.drawString(510,caja_iso-45,'22-Nov.-17')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,575,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Requisición')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,575) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor-8,585,575) #Linea 2 contorno
    

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    c.drawCentredString(296,caja_proveedor, 'Detalles')

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawRightString(140, caja_proveedor - 20, 'Folio de solicitud:')
    c.drawString(142, caja_proveedor - 20, requisicion.orden.folio)
    c.drawRightString(140,caja_proveedor-35,'Folio de requisición:')
    c.drawString(142, caja_proveedor - 35, requisicion.folio)
    c.drawRightString(140,caja_proveedor-50,'Proyecto/Orden de Trabajo:')
    c.drawString(142, caja_proveedor - 50, requisicion.orden.proyecto.nombre)
    c.drawRightString(140,caja_proveedor-65,'Subproyecto:')
    c.drawString(142, caja_proveedor - 65, requisicion.orden.subproyecto.nombre)
    c.drawRightString(140,caja_proveedor-80,'Área o Departamento:')
    c.drawString(142, caja_proveedor - 80, requisicion.orden.area.nombre)
    c.drawRightString(140,caja_proveedor-95,'Requisitó:')
    if requisicion.orden.staff is None:
        c.drawString(142, caja_proveedor - 95, 'N/A')
    else:
        c.drawString(142, caja_proveedor - 95, requisicion.orden.staff.staff.first_name + ' ' + requisicion.orden.staff.staff.last_name)
    c.drawRightString(140,caja_proveedor-110,'Autorizó:')
    if requisicion.requi_autorizada_por is None:
        c.drawString(142, caja_proveedor - 110, 'N/A')
    else:
        c.drawString(142, caja_proveedor - 110, requisicion.requi_autorizada_por.staff.first_name + ' ' + requisicion.requi_autorizada_por.staff.last_name)
    #str(preevaluacion.creado_at.strftime("%d/%m/%Y %H:%M"))
    c.drawString(270,caja_proveedor-20,'Fecha de elaboración:')
    c.drawString(365,caja_proveedor-20,str(requisicion.created_at.strftime("%d/%m/%Y %H:%M")))
    c.drawString(270,caja_proveedor-35,'Almacen que recibe:')
    c.drawString(355,caja_proveedor-35,requisicion.orden.subproyecto.nombre)
    c.drawString(270,caja_proveedor-50,'Archivo adjunto:')
    try:
        soporte_name = requisicion.orden.soporte.name
        if soporte_name:  # Verifica si soporte_name no es None ni una cadena vacía
            file_name = os.path.basename(soporte_name)
        else:
            file_name = None
    except AttributeError:
        file_name = None
        c.drawString(340,caja_proveedor-50,'N/A')
    else:
        c.drawString(340,caja_proveedor-50,str(file_name))
    
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',11)
    almacenista = Profile.objects.filter(tipo__almacenista = True).first()
    #if requisicion.orden.staff:
    #    c.drawString(130,caja_proveedor-20, requisicion.orden.staff.staff.first_name+' '+ requisicion.orden.staff.staff.last_name)
    #else:    
    #    c.drawString(130,caja_proveedor-20, almacenista.staff.last_name+' '+ almacenista.staff.last_name)
    #c.drawString(130,caja_proveedor-40, requisicion.orden.staff.distrito.nombre)
    #c.drawString(130,caja_proveedor-60, requisicion.orden.proyecto.nombre)
    #c.drawString(130,caja_proveedor-80, requisicion.orden.subproyecto.nombre)
    c.drawCentredString(300,caja_proveedor-80, 'Estado:')
    if requisicion.approved_at:
        c.setFillColor(color)
        c.drawString(325,caja_proveedor-80, estado +' '+ requisicion.approved_at.strftime("%d/%m/%Y %H:%M"))
    else:
        c.setFillColor(color)
        c.drawString(325,caja_proveedor-80, estado)
        # Datos de la tabla

    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=6,
        leading=10,
        alignment=1,  # Centrar el texto
        wordWrap='CJK',
        textColor=colors.white,
    )
    data = [
        ['', '', '', '', '', '', 'Detalle de productos y servicios', '', ''],
    ]

    data.append([
        Paragraph('#', custom_style),
        Paragraph('Cantidad', custom_style),
        Paragraph('Codigo producto', custom_style),
        Paragraph('Producto o Servicio', custom_style),
        Paragraph('Unidad de medida', custom_style),
        Paragraph('Criticidad', custom_style),
        Paragraph('Descripción general', custom_style),
        Paragraph('Descripción especifica', custom_style),
        Paragraph('Descripción técnica', custom_style),
        Paragraph('Criterios aceptación', custom_style),
    ])

    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6,
        leading=10,
        alignment=1,  # Centrar el texto
        wordWrap='CJK'
    )
    number=1
    for producto in productos:
        if producto.producto.articulos.producto.producto.servicio == True:
            tipo = 'Servicio'
        else:
            tipo = 'Producto'
        if producto.producto.articulos.producto.producto.unidad.nombre is None:
            name = 'N/A'
        else:
            name = producto.producto.articulos.producto.producto.unidad.nombre
        if producto.producto.articulos.producto.producto.critico is None:
            critico = 'N/A'
        else:
            critico = producto.producto.articulos.producto.producto.critico.nombre
        if producto.producto.articulos.producto.producto.descripcion_especifica is None:
            especifico = 'N/A'
        else:
            especifico = producto.producto.articulos.producto.producto.descripcion_especifica
        if producto.producto.articulos.producto.producto.criterios_aceptacion is None:
            criterios = 'N/A'
        else:
            criterios = producto.producto.articulos.producto.producto.criterios_aceptacion
        if producto.producto.articulos.producto.producto.especs is None:
            especs = 'N/A'
        else:
            especs = producto.producto.articulos.producto.producto.especs
        data.append([
            Paragraph(str(number), custom_style),
            Paragraph(str(producto.cantidad), custom_style),
            Paragraph(producto.producto.articulos.producto.producto.codigo, custom_style),
            Paragraph(tipo, custom_style),
            Paragraph(name, custom_style),
            Paragraph(critico, custom_style),
            Paragraph(producto.producto.articulos.producto.producto.nombre, custom_style),
            Paragraph(especifico, custom_style),
            Paragraph(especs, custom_style),
            Paragraph(criterios, custom_style)
        ])
        number +=1


    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    #Primer renglón
    c.drawCentredString(70,48,'Clasificación:')
    c.drawCentredString(140,48,'Nivel:')
    c.drawCentredString(240,48,'Preparado por:')
    c.drawCentredString(350,48,'Aprobado:')
    c.drawCentredString(450,48,'Fecha emisión:')
    c.drawCentredString(550,48,'Rev:')
    #Segundo renglón
    c.drawCentredString(70,34,'Controlado')
    c.drawCentredString(140,34,'N5')
    c.drawCentredString(240,34,'SEOV-ALM-N4-01-01')
    c.drawCentredString(350,34,'SUB ADM')
    c.drawCentredString(450,34,'24/Oct/2018')
    c.drawCentredString(550,34,'001')

    c.setFillColor(black)
    width, height = letter
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]

    if requisicion.comentario_compras is not None:
        comentario = requisicion.comentario_compras
    else:
        comentario = "No hay comentarios"

    c.setFillColor(prussian_blue)
    c.rect(40,135,530,15, fill=True, stroke=False)
    c.setFillColor(white)
    c.drawCentredString(296,140,'Observaciones')
    options_conditions_paragraph = Paragraph(comentario, styleN)
    # Crear un marco (frame) en la posición específica
    frame = Frame(40, -60, 530, 200, id='normal')
    # Agregar el párrafo al marco
    frame.addFromList([options_conditions_paragraph], c)
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)

    c.setFillColor(black)
    #if requisicion.orden:
    #    c.drawCentredString(180,high-240, requisicion.orden.staff.staff.first_name +' '+ requisicion.orden.staff.staff.last_name)
    #else:    
    #     c.drawCentredString(180,high-240, almacenista.staff.last_name+' '+ almacenista.staff.last_name)
    #c.drawCentredString(180,high-240, requisicion.created_by.staff.staff.first_name +' '+ requisicion.created_by.staff.staff.last_name)
    ##c.line(140,high-241,220,high-241)
    #c.drawCentredString(180,high-250, 'Solicitado')
    #if requisicion.autorizar == False:
    #    c.setFillColor(rojo)
    #    c.drawCentredString(410, high-240, '{Esta requisicion ha sido Cancelada}')
    #    c.setFont('Helvetica-Bold',14)
    #    c.drawString(370,670, 'CANCELADA')
    #elif requisicion.autorizar:
    #    c.setFillColor(prussian_blue)
    #    c.drawCentredString(410,high-240, requisicion.orden.superintendente.staff.first_name+' '+ requisicion.orden.superintendente.staff.last_name)
    #    c.setFont('Helvetica-Bold',14)
    #    c.drawString(370,670, 'APROBADA')
    #else:
    #    c.setFillColor(rojo)
    #    c.drawCentredString(410,high-240, requisicion.orden.superintendente.staff.first_name+' '+ requisicion.orden.superintendente.staff.last_name)
    #    c.setFont('Helvetica-Bold',22)
        #c.saveState()
        # Trasladar el origen del canvas al punto (370, 670)
        #c.translate(200, 300)
        # Rotar el canvas 45 grados en sentido horario
        #c.rotate(45)
        #c.drawString(30,0, 'NO AUTORIZADA AÚN')
        #c.restoreState()
    #c.setFillColor(black)
    #c.setFont('Helvetica',12)
    #c.line(360,high-241,460,high-241)
    #c.drawCentredString(410,high-250,'Superintendente')

    # Definición de la tabla        1       2           3       4       5           6          7        8        9      10
    table = Table(data, colWidths=[0.8 * cm, 1.5 * cm, 1.6 * cm, 1.5 * cm, 1.7 * cm, 1.5 * cm, 3 * cm, 2.8 * cm, 2.9 * cm, 2.8 * cm])
    # Estilos de la tabla
    table_style = TableStyle([
        # Estilo general
        ('INNERGRID', (0, 2), (-1, -1), 0.25, colors.black),  # Celdas de los datos
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Encabezado principal (primera fila)
        #('SPAN', (0, 0), (-1, 0)),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003153')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        # Encabezado de columnas (segunda fila)
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#003153')),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.white),
        ('FONTSIZE', (0, 1), (-1, 1), 8),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('ALIGN', (0, 1), (-1, 1), 'CENTER'),  # Centrar el texto del encabezado de columnas
        # Cuerpo de la tabla
        ('TEXTCOLOR', (0, 2), (-1, -1), colors.black),
        ('FONTSIZE', (0, 2), (-1, -1), 6),
    ])

    table.setStyle(table_style)
    width, height = letter
    frame = Frame(20, -120, width - 40, height - 100, id='normal')
    frame.addFromList([table], c)

    #c.showPage()
    c.save()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Requisición_' + str(requisicion.folio) +'.pdf')

@login_required
def reporte_devoluciones(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    if usuario.tipo.almacen == True:
        entradas = Devolucion.objects.all().order_by('-fecha').select_related('solicitud')
    else:
        entradas = Devolucion.objects.none()
    myfilter = DevolucionFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs

    if request.method == "POST":
        
        if 'btnExcel' in request.POST:
            return convert_devoluciones_to_xls2(entradas)
        if 'btnproductosxl' in request.POST:
            return convert_devoluciones_productos_to_xls(entradas)       
    
    #Set up pagination
    p = Paginator(entradas, 50)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    context = {
        'ordenes_list':ordenes_list,
        'entradas':entradas,
        'myfilter':myfilter,
        }
    
    return render(request,'requisiciones/reporte_devoluciones.html', context)



def convert_devoluciones_to_xls2(entradas):
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y añade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Matriz_Devoluciones")

     
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})

    #columns = ['Folio Solicitud', 'Solicitante', 'Almacenista','Proyecto', 'Subproyecto', 'Fecha creación','Productos','Tipo','Autorizada','Fecha autorización','Comentario']
    columns = ['Folio Solicitud', 'Solicitante', 'Almacenista','Proyecto', 'Subproyecto', 'Fecha creación','Tipo','Autorizada','Fecha autorización','Comentario']

    columna_max = len(columns)+2

    worksheet.write(0, columna_max - 1, 'Reporte Creado Automáticamente por SAVIA 2.0 Vordcab. UH', messages_style)
    worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas

    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    row_num = 0
    for dev in entradas:
        if dev.tipo:
            tipo = dev.tipo.nombre
        else:
            tipo = ''
        if dev.autorizada is True:
            autorizado = 'Autorizado'
        elif dev.autorizada is False:
            autorizado = 'No Autorizado'
        else:
            autorizado = 'Pendiente'
        row_num += 1
        # Crear la lista de productos con nombre y cantidad
        #productos_lista = [
        #    f"{producto['producto__producto__nombre']} (Cantidad: {producto['cantidad']})"
        #    for producto in dev.solicitud.productos.values('producto__producto__nombre', 'cantidad')
        #]
        # Unir la lista en una cadena
        #productos_str = ", ".join(productos_lista)

        row = [
            dev.solicitud.folio,
            f"{dev.solicitud.staff.staff.first_name} {dev.solicitud.staff.staff.last_name}",
            f"{dev.almacenista.staff.first_name} {dev.almacenista.staff.last_name}",
            dev.solicitud.proyecto.nombre,
            dev.solicitud.subproyecto.nombre,
            str(dev.created_at),
            #productos_str,  # Productos concatenados
            tipo,
            autorizado,
            str(dev.fecha),
            dev.comentario,
        ]
        
        for col_num, cell_value in enumerate(row):
        # Define el formato por defecto
            cell_format = body_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

      
        #worksheet.write_formula(row_num, 19, f'=IF(ISBLANK(R{row_num+1}), L{row_num+1}, L{row_num+1}*R{row_num+1})', money_style)
    
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Matriz_devoluciones_{dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response


def convert_devoluciones_productos_to_xls(entradas_qs):
    dev_ids = list(entradas_qs.values_list('id', flat=True))
    renglones = (
        Devolucion_Articulos.objects
        .filter(vale_devolucion_id__in=dev_ids)
        .select_related(
            'vale_devolucion__solicitud__staff__staff',
            'vale_devolucion__solicitud__proyecto',
            'vale_devolucion__solicitud__subproyecto',
            'vale_devolucion__tipo',
            'vale_devolucion__almacenista__staff',
            'producto__articulos__producto__producto',
        )
        .order_by('-vale_devolucion__fecha', '-created_at', 'id')
    )

    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet("Matriz_Devoluciones_Prod")

    head_style  = wb.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial',  'font_size': 11})
    body_style  = wb.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = wb.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style  = wb.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    dt_style    = wb.add_format({'num_format': 'dd/mm/yyyy hh:mm', 'font_name': 'Calibri', 'font_size': 10})

    columns = [
        'Folio Solicitud', 'Solicitante', 'Almacenista', 'Proyecto', 'Subproyecto',
        'Fecha creación', 'Tipo', 'Autorizada', 'Fecha autorización', 'Comentario Devolución',
        'ID Devolución', 'Código Producto', 'Nombre Producto', 'Cantidad', 'Precio', 'Importe',
        'Comentario Producto'
    ]
    for i, c in enumerate(columns):
        ws.write(0, i, c, head_style)
        ws.set_column(i, i, 18)
    leyenda_fmt = wb.add_format({'font_name':'Arial Narrow', 'font_size':11})
    ws.write(0, len(columns)+1, 'Reporte Creado Automáticamente por SAVIA 2.0 Vordtec. UH', leyenda_fmt)
    ws.write(1, len(columns)+1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.',   leyenda_fmt)
    ws.set_column(len(columns)+1, len(columns)+1, 32)

    row = 0
    for r in renglones:
        row += 1
        d = r.vale_devolucion

        folio   = d.solicitud.folio if d and d.solicitud else ''
        solicit = f"{d.solicitud.staff.staff.first_name} {d.solicitud.staff.staff.last_name}".strip() if (d and d.solicitud and d.solicitud.staff and d.solicitud.staff.staff) else ''
        almac   = f"{d.almacenista.staff.first_name} {d.almacenista.staff.last_name}".strip() if (d and d.almacenista and d.almacenista.staff) else ''
        proy    = d.solicitud.proyecto.nombre if (d and d.solicitud and d.solicitud.proyecto) else ''
        subproy = d.solicitud.subproyecto.nombre if (d and d.solicitud and d.solicitud.subproyecto) else ''
        # ⬇️ FECHAS como objetos datetime/date (SIN str)
        f_crea  = d.created_at if d else None                    # DateTimeField
        f_aut   = d.fecha if d else None                         # DateField (o None)
        tipo    = d.tipo.nombre if (d and d.tipo) else ''
        if d and d.autorizada is True:   autorizada = 'Autorizado'
        elif d and d.autorizada is False: autorizada = 'No Autorizado'
        else:                             autorizada = 'Pendiente'
        cdev    = d.comentario or '' if d else ''

        try:
            base_prod = r.producto.articulos.producto.producto
            codigo = base_prod.codigo or ''
            nombre = base_prod.nombre or ''
        except Exception:
            codigo, nombre = '', ''

        cantidad = float(r.cantidad or 0)
        precio   = float(r.precio   or 0)
        importe  = cantidad * precio
        cren     = r.comentario or ''
        #refren   = r.referencia or ''
        #f_reng   = r.created_at  # ⬅️ DateTimeField (mantener datetime)

        fila = [
            folio, solicit, almac, proy, subproy, f_crea, tipo, autorizada, f_aut, cdev,
            r.id, codigo, nombre, cantidad, precio, importe, cren
        ]

        for col, val in enumerate(fila):
            # Columnas con fechas: creación(5), autorización(8), renglón(18)
            if col in (5, 8, 18) and val:
                # Si viene con timezone, la limpiamos
                if hasattr(val, 'tzinfo'):
                    val = val.replace(tzinfo=None)
                # Convertir a datetime solo si es date
                if isinstance(val, dt.date) and not isinstance(val, dt.datetime):
                    val = dt.datetime.combine(val, dt.time.min)
                ws.write_datetime(row, col, val, date_style)
            elif col in (13, 14, 15):  # Cantidad, Precio, Importe
                ws.write_number(row, col, float(val), money_style if col in (14, 15) else body_style)
            else:
                ws.write(row, col, val, body_style)


    wb.close()
    output.seek(0)
    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename=Matriz_devoluciones_productos_{dt.date.today()}.xlsx'
    resp.set_cookie('descarga_iniciada', 'true', max_age=20)
    output.close()
    return resp



@login_required
def terminado_salida_surtir(request, pk):
    entrada = get_object_or_404(EntradaArticulo, id=pk)
    perfil = Profile.objects.get(staff__id=request.user.id)
    vale, created = ValeSalidas.objects.get_or_create(solicitud_terminado = entrada.producto_terminado.solicitud,almacenista=perfil,proyecto=entrada.producto_terminado.solicitud.proyecto,subproyecto=entrada.producto_terminado.solicitud.subproyecto,complete=True)
    vale.save()
    salida, created = Salidas.objects.get_or_create(vale_salida=vale,producto_terminado=entrada.producto_terminado,cantidad=entrada.cantidad,complete =True)
    salida.save()
    entrada.liberado = True
    entrada.save()
    #Modificar el inventario para la salida
    inventario = entrada.producto_terminado.producto
    inventario.cantidad -= entrada.cantidad
    inventario.comentario = 'Salida de producto terminado'
    inventario.save()
    messages.success(request,f'Salida creada para la entrada: {entrada.id}, producto {entrada.producto_terminado.producto.producto.nombre}') 
    return redirect('producto-terminado-salida')