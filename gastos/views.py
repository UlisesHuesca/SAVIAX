from django.shortcuts import render, redirect
from datetime import date, datetime
from django.contrib import messages
from django.core.mail import EmailMessage, BadHeaderError
import socket
import traceback
from smtplib import SMTPException
from dashboard.models import Inventario, Order, ArticulosparaSurtir, ArticulosOrdenados, Tipo_Orden, Product
from inventoryproject.settings import EMAIL_HOST_USER
from solicitudes.models import Proyecto, Subproyecto, Operacion
from tesoreria.models import Pago, Cuenta
from .models import Solicitud_Gasto, Articulo_Gasto, Entrada_Gasto_Ajuste, Conceptos_Entradas, Factura
from .forms import Solicitud_GastoForm, Articulo_GastoForm, Articulo_Gasto_Edit_Form, Pago_Gasto_Form, Articulo_Gasto_Factura_Form, Entrada_Gasto_AjusteForm, Conceptos_EntradasForm, FacturaForm, Autorizacion_Gasto_Form
from tesoreria.forms import Facturas_Gastos_Form 
from tesoreria.utils import extraer_texto_de_pdf, encontrar_variables
from compras.views import attach_oc_pdf
from .filters import Solicitud_Gasto_Filter
from user.models import Profile
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, FileResponse
from django.core.paginator import Paginator
from django.db.models import Sum
from django.urls import reverse
import os
import base64
import json
import io
import xml.etree.ElementTree as ET
import decimal
from django.db.models import Q
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import pytz
#PDF generator
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter, portrait
from reportlab.rl_config import defaultPageSize 
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame
from bs4 import BeautifulSoup
from django.conf import settings

#Excel stuff
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import xlsxwriter
from io import BytesIO

# Create your views here.
@login_required(login_url='user-login')
def crear_gasto(request):
    colaborador = Profile.objects.all()
    articulos_gasto = Articulo_Gasto.objects.all()
    inventario = Inventario.objects.all()
    usuario = colaborador.get(staff__id=request.user.id)
    superintendentes = colaborador.filter(tipo__superintendente=True, staff__is_active=True,).exclude(tipo__nombre="Admin")
    proyectos = Proyecto.objects.filter(activo=True)
    #subproyectos = Subproyecto.objects.all()
    #colaborador = Profile.objects.all()
    #Tengo que revisar primero si ya existe una orden pendiente del usuario
    gasto, created = Solicitud_Gasto.objects.get_or_create(complete= False, staff=usuario)
    
    articulo, created = articulos_gasto.get_or_create(completo = False, staff=usuario)

    productos = articulos_gasto.filter(gasto=gasto, completo = True)
    articulos_gasto = Inventario.objects.filter(producto__gasto = True, producto__baja_item = False) #Cuando Alberto envie los conceptos se implementa
    
    proyectos_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.nombre) + ' |' + str(item.descripcion)
        } for item in proyectos
    ]



    productos_para_select2 = [
        {
            'id': item.id,
            'text': str(item.producto.nombre),
            'iva': str(item.producto.iva)
        } for item in articulos_gasto
    ]

  
    articulos = inventario.filter(producto__gasto = False)
    facturas = Factura.objects.filter(solicitud_gasto = gasto)
    form_product = Articulo_GastoForm()
    form = Solicitud_GastoForm()
    factura_form = FacturaForm()

    if request.method =='POST':
        if "btn_agregar" in request.POST:
            form = Solicitud_GastoForm(request.POST, instance=gasto)
            #abrev= usuario.distrito.abreviado
            if form.is_valid():
                gasto = form.save(commit=False)
                gasto.complete = True
                gasto.created_at = date.today()
                gasto.created_at_time = datetime.now().time()
                gasto.staff =  usuario
                gasto.save()
                form.save()
                messages.success(request, f'La solicitud {gasto.id} ha sido creada')
                return redirect('solicitudes-gasto')
        if "btn_producto" in request.POST:
            form_product = Articulo_GastoForm(request.POST, request.FILES or None, instance=articulo)
            if form_product.is_valid():
                articulo = form_product.save(commit=False)
                articulo.gasto = gasto
                articulo.completo = True
                articulo.save()
                messages.success(request, 'La solicitud de creacion de articulo funciona')
                return redirect('crear-gasto')
        if "btn_factura" in request.POST:
            factura_form = FacturaForm(request.POST, request.FILES)
            if factura_form.is_valid():
                factura = factura_form.save(commit=False)
                factura.solicitud_gasto = gasto  # Asume que ya tienes una instancia de Solicitud_Gasto en 'gasto'
                factura.fecha_subida = datetime.now()
                factura.save()
                messages.success(request, 'Factura agregada correctamente.')
                return redirect('crear-gasto')


    context= {
        'facturas':facturas,
        'productos':productos,
        'productos_para_select2':productos_para_select2,
        'colaborador':colaborador,
        'form':form,
        'form_product': form_product,
        'gasto':gasto,
        'superintendentes':superintendentes,
        'proyectos_para_select2':proyectos_para_select2,
        #'subproyectos':subproyectos,
        'factura_form': factura_form,
    }
    return render(request, 'gasto/crear_gasto.html', context)

def delete_gasto(request, pk):
    articulo = Articulo_Gasto.objects.get(id=pk)
    messages.success(request,f'El articulo {articulo.producto} ha sido eliminado exitosamente')
    articulo.delete()

    return redirect('crear-gasto')

def eliminar_factura(request, pk):
    articulo = Factura.objects.get(id=pk)
    messages.success(request,f'La factura {articulo.id} ha sido eliminada exitosamente')
    articulo.delete()

    return redirect('crear-gasto')

def editar_gasto(request, pk):
    producto = Articulo_Gasto.objects.get(id=pk)

    form = Articulo_Gasto_Edit_Form(instance=producto)

    if request.method =='POST':
        form = Articulo_Gasto_Edit_Form(request.POST, instance=producto)

        if form.is_valid():
            form.save()

            messages.success(request,f'Se ha guardado el artículo {producto} correctamente')
            return HttpResponse(status=204)
        #else:
            #messages.error(request,'Se lo llevo SPM')


    context= {
        'producto': producto,
        'form': form,
        }

    return render(request, 'gasto/editar_gasto.html', context)

@login_required(login_url='user-login')
def solicitudes_gasto(request):

    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    perfil = Profile.objects.get(staff__id=request.user.id)


   
    if perfil.tipo.nombre == "Admin" or perfil.tipo.nombre == "Gerente" or perfil.tipo.superintendente:
        solicitudes = Solicitud_Gasto.objects.filter(complete = True).order_by('-created_at')
    else:
        solicitudes = Solicitud_Gasto.objects.filter(complete=True, staff = perfil).order_by('-created_at')

    myfilter=Solicitud_Gasto_Filter(request.GET, queryset=solicitudes)
    solicitudes = myfilter.qs

    #Set up pagination
    p = Paginator(solicitudes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:
       return convert_excel_gasto_matriz(solicitudes)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'gasto/solicitudes_gasto.html',context)

@login_required(login_url='user-login')
def detalle_gastos(request, pk):
    productos = Articulo_Gasto.objects.filter(gasto__id=pk)
    facturas = Factura.objects.filter(solicitud_gasto__id = pk)


    context= {
        'productos':productos,
        'facturas':facturas,
        'pk':pk,
        }

    return render(request, 'gasto/detalle_gasto.html', context)

@login_required(login_url='user-login')
def gastos_pendientes_autorizar(request):

    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    perfil = Profile.objects.get(staff__id=request.user.id)

    if perfil.tipo.nombre == "Admin":
        solicitudes = Solicitud_Gasto.objects.filter(complete=True, autorizar = None).order_by('-folio')
    else:
        solicitudes = Solicitud_Gasto.objects.filter(complete=True, autorizar = None, superintendente = perfil).order_by('-folio')

    ids_solicitudes_validadas = [solicitud.id for solicitud in solicitudes if solicitud.get_validado]

    solicitudes = Solicitud_Gasto.objects.filter(id__in=ids_solicitudes_validadas)

    myfilter=Solicitud_Gasto_Filter(request.GET, queryset=solicitudes)
    solicitudes = myfilter.qs

    #Set up pagination
    p = Paginator(solicitudes, 10)
    page = request.GET.get('page')
    solicitudes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)

    context= {
        'ordenes_list':solicitudes_list,
        'myfilter':myfilter,
        }

    return render(request, 'gasto/pendientes_autorizar_gasto.html', context)

@login_required(login_url='user-login')
def gastos_pendientes_autorizar2(request):

    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    perfil = Profile.objects.get(staff__id=request.user.id)

    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    #if perfil.tipo.superintendente == True:
    #    solicitudes = Solicitud_Gasto.objects.filter(complete=True, staff__distrito=perfil.distrito).order_by('-folio')
    #elif perfil.tipo.supervisor == True:
    #    solicitudes = Solicitud_Gasto.objects.filter(complete=True, staff__distrito=perfil.distrito, supervisor=perfil).order_by('-folio')
    #else:
    solicitudes = Solicitud_Gasto.objects.filter(complete=True, autorizar = True, autorizar2 = None).order_by('-folio')

    myfilter=Solicitud_Gasto_Filter(request.GET, queryset=solicitudes)
    solicitudes = myfilter.qs

    #Set up pagination
    p = Paginator(solicitudes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'gasto/pendientes_autorizar_gasto2.html', context)

@login_required(login_url='user-login')
def autorizar_gasto(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    gasto = Solicitud_Gasto.objects.get(id = pk)
    productos = Articulo_Gasto.objects.filter(gasto = gasto)
    facturas = Factura.objects.filter(solicitud_gasto__id = pk)


    if request.method =='POST' and 'btn_autorizar' in request.POST:
        gasto.autorizar = True
        gasto.approved_at = date.today()
        gasto.approved_at_time = datetime.now().time()
        gasto.sol_autorizada_por = Profile.objects.get(staff__id=request.user.id)
        gasto.save()
        messages.success(request, f'{perfil.staff.first_name} {perfil.staff.last_name} has autorizado la solicitud {gasto.id}')
        return redirect ('gastos-pendientes-autorizar')


    context = {
        'facturas':facturas,
        'gasto': gasto,
        'productos': productos,
    }

    return render(request,'gasto/autorizar_gasto.html', context)


@login_required(login_url='user-login')
def cancelar_gasto(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    gasto = Solicitud_Gasto.objects.get(id = pk)
    productos = Articulo_Gasto.objects.filter(gasto = gasto)

    if request.method =='POST' and 'btn_cancelar' in request.POST:
        gasto.autorizar = False
        gasto.approved_at = date.today()
        gasto.approved_at_time = datetime.now().time()
        gasto.sol_autorizada_por = Profile.objects.get(staff__id=request.user.id)
        gasto.save()
        messages.info(request, f'{perfil.staff.first_name} {perfil.staff.last_name} has cancelado la solicitud {gasto.id}')
        return redirect ('gastos-pendientes-autorizar')

    context = {
        'gasto': gasto,
        'productos': productos,
    }

    return render(request,'gasto/cancelar_gasto.html', context)

@login_required(login_url='user-login')
def autorizar_gasto2(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    gasto = Solicitud_Gasto.objects.get(id = pk)
    productos = Articulo_Gasto.objects.filter(gasto = gasto)

    if request.method =='POST' and 'btn_autorizar' in request.POST:
        gasto.autorizar2 = True
        gasto.autorizado_por2 = perfil
        gasto.approbado_fecha2 = date.today()
        gasto.approved_at_time2 = datetime.now().time()
        gasto.save()
        messages.success(request, f'{perfil.staff.first_name} {perfil.staff.last_name} has autorizado el gasto {gasto.id}')
        return redirect ('gastos-pendientes-autorizar2')


    context = {
        'gasto': gasto,
        'productos': productos,
    }

    return render(request,'gasto/autorizar_gasto2.html', context)


@login_required(login_url='user-login')
def cancelar_gasto2(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    gasto = Solicitud_Gasto.objects.get(id = pk)
    productos = Articulo_Gasto.objects.filter(gasto = gasto)
    form = Autorizacion_Gasto_Form(instance = gasto)

    if request.method =='POST':
        form = Autorizacion_Gasto_Form(request.POST, instance = gasto)

        if form.is_valid():
            gasto = form.save(commit = False)
            gasto.autorizar2 = False
            gasto.approbado_fecha2 = date.today()
            gasto.approved_at_time2 = datetime.now().time()
            gasto.autorizado_por2 = perfil
            gasto.save()
            messages.info(request, f'{perfil.staff.first_name} {perfil.staff.last_name} has cancelado la solicitud {gasto.id}')
            return HttpResponse(status=204)

    
    context = {
        'form': form,
        'gasto': gasto,
        'productos': productos,
    }

    return render(request,'gasto/cancelar_gasto2.html', context)




# Create your views here.
@login_required(login_url='user-login')
def pago_gastos_autorizados(request):
    usuario = Profile.objects.get(staff__id=request.user.id)

    if usuario.tipo.tesoreria == True:
        gastos = Solicitud_Gasto.objects.filter(autorizar=True, pagada=False, autorizar2=True).order_by('-folio')


    myfilter = Solicitud_Gasto_Filter(request.GET, queryset=gastos)
    gastos = myfilter.qs


    context= {
        'gastos':gastos,
        'myfilter':myfilter,
        }

    return render(request, 'gasto/pago_gastos_autorizados.html',context)

@login_required(login_url='user-login')
def pago_gasto(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    gasto = Solicitud_Gasto.objects.get(id=pk)
    pagos_alt = Pago.objects.filter(gasto=gasto, hecho=True)
    cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')

   
    form = Pago_Gasto_Form()
    remanente = gasto.get_total_solicitud - gasto.monto_pagado


    if request.method == 'POST':
        pago, created = Pago.objects.get_or_create(tesorero = usuario, distrito = usuario.distrito, hecho=False, gasto=gasto)
        form = Pago_Gasto_Form(request.POST or None, request.FILES or None, instance = pago)
        if form.is_valid():
            pago = form.save(commit = False)
            #pago.gasto = gasto
            pago.pagado_date = date.today()
            pago.pagado_hora = datetime.now().time()
            pago.hecho = True
            total_pagado = round(gasto.monto_pagado  + pago.monto,2)
            total_sol = round(gasto.get_total_solicitud,2)
            #El bloque a continuación se generó para resolver los problemas de redondeo, se comparan las dos cantidades redondeadas en una variable y se activa una bandera (flag) que indica si son iguales o no!
            if total_sol == total_pagado:
                flag = True
            else:
                flag = False
            if total_pagado > gasto.get_total_solicitud:
                messages.error(request,f'{usuario.staff.first_name}, el monto introducido más los pagos anteriores superan el monto total del viático')
            else:
                if flag:
                    gasto.pagada = True
                    gasto.save()
                pago.save()
                pagos = Pago.objects.filter(gasto=gasto, hecho=True)
                #archivo_oc = attach_oc_pdf(request, gasto.id)
                try:
                    email = EmailMessage(
                        f'Gasto Autorizado {gasto.id}',
                        f'Estimado(a) {gasto.staff.staff.first_name} {gasto.staff.staff.last_name}:\n\nEstás recibiendo este correo porque ha sido pagado el gasto con folio: {gasto.id}.\n\n\nVordtec de México S.A. de C.V.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                        'savia@vordtec.com',
                        ['ulises_huesc@hotmail.com',gasto.staff.staff.email],
                        )
                    #email.attach(f'OC_folio_{gasto.id}.pdf',archivo_oc,'application/pdf')
                    email.attach('Pago.pdf',request.FILES['comprobante_pago'].read(),'application/pdf')
                    if pagos.count() > 0:
                        for item in pagos:
                            email.attach(f'Gasto{gasto.id}_P{item.id}.pdf',item.comprobante_pago.read(),'application/pdf')
                    email.send()
                    messages.success(request,f'Gracias por registrar tu pago, {usuario.staff.first_name}')
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'{usuario.staff.first_name}, Gracias por registrar tu pago, pero el correo de notificación no ha sido enviado debido a un error: {e}'
                    messages.warning(request, error_message)
                return redirect('pago-gastos-autorizados')
        else:
            form = Pago_Gasto_Form()
            messages.error(request,f'{usuario.staff.first_name}, No se pudo subir tu documento')

    context= {
        'gasto':gasto,
        #'pago':pago,
        'form':form,
        'pagos_alt':pagos_alt,
        'cuentas':cuentas,
        'remanente':remanente,
    }

    return render(request,'gasto/pago_gasto.html',context)


def prellenar_formulario_gastos(request):
    #print('prellenar_formulario_gastos')
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        pdf_content = request.FILES.get('comprobante_pago')
        
        if not pdf_content:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        pdf_content = pdf_content.read()
        
        
        texto_extraido = extraer_texto_de_pdf(pdf_content)
        datos_extraidos = encontrar_variables(texto_extraido)
            #divisa_cuenta_extraida = datos_extraidos.get('divisa_cuenta', '').strip()

     
        divisa_cuenta_extraida = datos_extraidos.get('divisa_cuenta', '').strip()

        
        fecha_str = datos_extraidos.get('fecha', '').strip()
        #print(fecha_str)
        fecha_formato_correcto = None  # Valor por defecto en caso de que no se pueda procesar la fecha
        
        if fecha_str:
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y')
                fecha_formato_correcto = fecha_obj.strftime('%Y-%m-%d')
            except ValueError:
                # Opcional: Agregar alguna forma de logging o notificación de que la fecha no es válida
                print('Se lo llevó madres')
                pass
        
        numero_cuenta_extraido = datos_extraidos.get('cuenta_retiro', '').strip().lstrip('0')
       
        cuenta_objeto = None
       
        #print('numero_cuenta_extraido',numero_cuenta_extraido)
        if numero_cuenta_extraido:
            try:
                
                cuenta_objeto = Cuenta.objects.get(cuenta__contains=numero_cuenta_extraido)
                print('cuenta_objeto:', cuenta_objeto)
            except Cuenta.DoesNotExist:
                print('Cuenta retiro no encontrada:', numero_cuenta_extraido)
                return JsonResponse({'error': 'Cuenta retiro no encontrada'}, status=404)
            except Exception as e:
                print('Error inesperado al buscar cuenta retiro:', e)
                print(traceback.format_exc())
                return JsonResponse({'error': 'Error interno'}, status=500)
            
        
        
        #print("destino_cuenta",datos_extraidos.get('cuenta_deposito', '').strip().lstrip('0') or None) 
        datos_para_formulario = {
            'monto': datos_extraidos.get('importe_operacion', '').replace('MXP', '').replace(',', '').strip() or None,
            'pagado_real': fecha_formato_correcto,  # Valor procesado o None
            'cuenta': cuenta_objeto.id if cuenta_objeto else None,
            'divisa_cuenta': divisa_cuenta_extraida or None,
            #'hora_operacion': datos_extraidos.get('hora_operacion', '') or None,
        }
        
        return JsonResponse(datos_para_formulario)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required(login_url='user-login')
def matriz_facturas_gasto(request, pk):
    gasto = Solicitud_Gasto.objects.get(id = pk)
    articulos_gasto = Articulo_Gasto.objects.filter(gasto = gasto)
    facturas = Factura.objects.filter(solicitud_gasto = gasto)
    form =  Facturas_Gastos_Form(instance=gasto)
    factura_form = FacturaForm()
    next_url = request.GET.get('next','matriz-pagos')
    if request.method == 'POST':
        form = Facturas_Gastos_Form(request.POST, instance=gasto)
        if "btn_factura_completa" in request.POST:
            if form.is_valid():
                form.save()
                messages.success(request,'Haz cambiado el status de facturas completas')
                return redirect(next_url)
            else:
                messages.error(request,'No está validando')
        if "btn_factura" in request.POST:
            factura_form = FacturaForm(request.POST, request.FILES)
            if factura_form.is_valid():
                factura = factura_form.save(commit=False)
                factura.solicitud_gasto = gasto  # Asume que ya tienes una instancia de Solicitud_Gasto en 'gasto'
                factura.fecha_subida = datetime.now()
                factura.save()
                messages.success(request, 'Factura agregada correctamente.')
                return redirect(next_url)


    context={
        'form':form,
        'factura_form':factura_form,
        'articulos_gasto':articulos_gasto,
        'gasto':gasto,
        'facturas':facturas,
        }

    return render(request, 'gasto/matriz_factura_gasto.html', context)

def get_image_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode()
    
@login_required(login_url='user-login')
def eliminar_factura_gasto(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    factura = Factura.objects.get(id=pk)
    gasto = factura.solicitud_gasto
    comentario = request.POST.get('comentario')
    # Construir la URL usando reverse
    next_url = request.GET.get('next', None)
    print('URLLLLLLLLLLLLLLLLLLL2')
    print(next_url)
    # Construir la URL usando reverse
    matriz_url = reverse('matriz-facturas-gasto', args=[gasto.id])
    static_path = settings.STATIC_ROOT
    img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
    img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
    image_base64 = get_image_base64(img_path)
    logo_v_base64 = get_image_base64(img_path2)
    # Crear el mensaje HTML
    html_message = f"""
    <html>
        <head>
            <meta charset="UTF-8">
        </head>
        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                <tr>
                    <td align="center">
                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                            <tr>
                            </tr>
                            <tr>
                                <td style="padding: 20px;">
                                    <p style="font-size: 18px; text-align: justify;">
                                        <p>Estimado {factura.solicitud_gasto.staff.staff.first_name} {factura.solicitud_gasto.staff.staff.last_name},</p>
                                    </p>
                                    <p style="font-size: 16px; text-align: justify;">
                                        Estás recibiendo este correo porque tu factura subida el: <strong>{factura.fecha_subida.date()}</strong> en el gasto <strong>{gasto.id}</strong> ha sido eliminada.</p>
                                    <p>Comentario:</p>
                                    {comentario}
                               
                                    </p>
                                <p style="font-size: 16px; text-align: justify;">
                                    Att: {perfil.staff.first_name} {perfil.staff.last_name}
                                </p>
                                    <p style="text-align: center; margin: 20px 0;">
                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                    </p>
                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                        Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
    </html>
    """
    try:
        email = EmailMessage(
            f'Factura eliminada',
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[gasto.staff.staff.email],
            headers={'Content-Type': 'text/html'}
            )
        email.content_subtype = "html " # Importante para que se interprete como HTML
        if factura.archivo_pdf:
            pdf_path = factura.archivo_pdf.path
            if os.path.exists(pdf_path):  # Verificar si el archivo realmente existe
                with open(pdf_path, 'rb') as pdf_file:
                    email.attach(factura.archivo_pdf.name, pdf_file.read(), 'application/pdf')
            else:
                print(f"El archivo PDF no se encuentra en la ruta: {pdf_path}")

        if factura.archivo_xml:
            xml_path = factura.archivo_xml.path
            if os.path.exists(xml_path):  # Verificar si el archivo realmente existe
                with open(xml_path, 'rb') as xml_file:
                    email.attach(factura.archivo_xml.name, xml_file.read(), 'application/xml')
            else:
                print(f"El archivo XML no se encuentra en la ruta: {xml_path}")

        email.send()
        messages.success(request, f'La factura {factura.id} ha sido eliminada exitosamente')
    except (BadHeaderError, SMTPException, socket.gaierror) as e:
        error_message = f'La factura {factura.id} ha sido eliminada, pero el correo no ha sido enviado debido a un error: {e}'
        messages.success(request, error_message)
    factura.delete()
    # Si next_url existe, redirigir agregando el parámetro `next`
    if next_url:
        return redirect(f'{matriz_url}?next={next_url}')
    else:
        return redirect(matriz_url)

def facturas_gasto(request, pk):
    articulo = Articulo_Gasto.objects.get(id = pk)
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)
    #factura, created = Facturas.objects.get_or_create(pago=pago, hecho=False)
    form = Articulo_Gasto_Factura_Form(instance= articulo)

    if request.method == 'POST':
        form = Articulo_Gasto_Factura_Form(request.POST or None, request.FILES or None, instance = articulo)
        if form.is_valid():
            form.save()
            messages.success(request,'Las facturas se subieron de manera exitosa')
            return redirect('matriz-compras')
        else:
            form = Articulo_Gasto_Factura_Form()
            messages.error(request,'No se pudo subir tu documento')

    context={
        'articulo':articulo,
        'form':form,
        }

    return render(request, 'gasto/facturas_gasto.html', context)


def matriz_gasto_entrada(request):
    #articulos_gasto = Articulo_Gasto.objects.filter(gasto = gasto)

    #articulos_gasto = Articulo_Gasto.objects.all()
    articulos_gasto = Articulo_Gasto.objects.filter(
        Q(producto__producto__nombre = "MATERIALES")|Q(producto__producto__nombre = "HERRAMIENTA"), 
        completo = True, 
        validacion = False, 
        gasto__autorizar = None, 
        gasto__tipo__tipo='REEMBOLSO'
        )

    context={
        'articulos_gasto':articulos_gasto,
        #'form':form,
    }

    return render(request, 'gasto/matriz_entrada_almacen.html', context)

def gasto_entrada(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    #Tengo que revisar primero si ya existe una orden pendiente del usuario
    articulo_gasto = Articulo_Gasto.objects.get(id=pk)
    entrada, created = Entrada_Gasto_Ajuste.objects.get_or_create(completo= False, almacenista=usuario, gasto = articulo_gasto)
    articulo, created = Conceptos_Entradas.objects.get_or_create(completo = False, entrada = entrada)
    last_order = Order.objects.filter(staff__distrito = usuario.distrito).order_by('-last_folio_number').first()
    productos = Conceptos_Entradas.objects.filter(entrada=entrada, completo = True)
    articulos = Inventario.objects.filter(producto__gasto = False)
    form_product = Conceptos_EntradasForm()
    form = Entrada_Gasto_AjusteForm()

    if request.method =='POST':
        if "btn_agregar" in request.POST:
            form = Entrada_Gasto_AjusteForm(request.POST, instance = entrada)
            if form.is_valid():
                entrada = form.save(commit=False)
                entrada.completo = True
                entrada.completado_fecha = date.today()
                entrada.completado_hora = datetime.now().time()
                entrada.save()
                articulo_gasto.validacion = True
                articulo_gasto.save()
                messages.success(request, f'La entrada del gasto {entrada.id} ha sido creada')
               
                abrev= usuario.distrito.abreviado
                if last_order == None:
                    #No hay órdenes para este distrito todavía
                    folio_number = 1
                else:
                    folio_number = last_order.last_folio_number + 1
                last_folio_number = folio_number
                tipo = Tipo_Orden.objects.get(tipo ='normal')
                folio = str(abrev) + str(folio_number).zfill(4)  
                orden_producto, created = Order.objects.get_or_create(staff = articulo_gasto.staff, complete = None, distrito = articulo_gasto.staff.distrito)
                orden_producto.folio =folio
                orden_producto.tipo = tipo
                orden_producto.last_folio_number = last_folio_number
                orden_producto.created_at = date.today()
                orden_producto.approved_at = date.today()
                orden_producto.created_at_time = datetime.now().time()
                orden_producto.approved_at_time = datetime.now().time()
                orden_producto.autorizar = True
                orden_producto.supervisor = articulo_gasto.staff
                orden_producto.superintendente = articulo_gasto.gasto.superintendente
                orden_producto.proyecto = articulo_gasto.proyecto
                orden_producto.subproyecto = articulo_gasto.subproyecto
                area = Operacion.objects.get(nombre="GASTO")
                orden_producto.area = area
                orden_producto.complete = True
                orden_producto.save()
                destinatarios = list(Profile.objects.filter(tipo__almacen=True).values_list('staff__email', flat=True))

                destinatarios.extend(['ulises_huesc@hotmail.com', articulo_gasto.staff.staff.email])
                destinatarios = [e.strip() for e in destinatarios if e and e.strip()]
                destinatarios = list(dict.fromkeys(destinatarios))
                for item_producto in productos:
                    producto_inventario = Inventario.objects.get(producto= item_producto.concepto_material.producto)
                    #productos_por_surtir = ArticulosparaSurtir.objects.filter(articulos__producto=producto_inventario, requisitar = True)
                    articulo_ordenado = ArticulosOrdenados.objects.create(producto=producto_inventario, orden = orden_producto, cantidad=item_producto.cantidad)
                    productos_por_surtir = ArticulosparaSurtir.objects.create(
                        articulos = articulo_ordenado,
                        cantidad=item_producto.cantidad,
                        precio = item_producto.precio_unitario,
                        surtir=True,
                        comentario=f"Esta solicitud es proveniente del gasto {articulo_gasto.gasto.folio}",
                        created_at=date.today(),
                        created_at_time=datetime.now().time(),
                    )
                    #Calculo el precio 
                    producto_inventario.price = ((item_producto.precio_unitario * item_producto.cantidad)+ ((producto_inventario.cantidad_apartada + producto_inventario.cantidad) * producto_inventario.price))/(producto_inventario.cantidad + item_producto.cantidad + producto_inventario.cantidad_apartada)
                    #La cantidad en inventario + la cantidad del producto en la entrada
                    producto_inventario.cantidad_apartada = producto_inventario.cantidad_apartada + item_producto.cantidad
                    #producto_inventario.save()
                    producto_inventario._change_reason = f'Esta es una entrada desde un gasto {item_producto.id}'
                    producto_inventario.save()
               
                try:
                    email = EmailMessage(
                        subject=f'Entrada de producto por gasto: {articulo_gasto.producto.producto.nombre} |Gasto: {articulo_gasto.gasto.folio}|Solicitud:{orden_producto.folio}',
                        body=(
                            f'Estimado {articulo_gasto.staff.staff.first_name} {articulo_gasto.staff.staff.last_name},\n'
                            f'Estás recibiendo este correo porque tu producto: {articulo_gasto.producto.producto.nombre}'
                            f'ha sido validado por el almacenista {usuario.staff.first_name} {usuario.staff.last_name}, '
                            f'favor de pasar a firmar el vale de salida para terminar con este proceso.\n\n'
                            f' Este mensaje ha sido automáticamente generado por SAVIA VORDTEC'
                        ),
                        from_email = EMAIL_HOST_USER,
                        to = destinatarios,
                        )
                    email.send()
                    
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'{usuario.staff.first_name}, El correo notificación no ha sido enviado debido a un error: {e}'
                    messages.warning(request, error_message)
                return redirect('matriz-gasto-entrada')
        if "btn_producto" in request.POST:
            form_product = Conceptos_EntradasForm(request.POST, instance=articulo)
            if form_product.is_valid():
                articulo = form_product.save(commit=False)
                articulo.completo = True
                articulo.save()
                messages.success(request, 'Has guardado exitosamente un artículo')
                return redirect('gasto-entrada',pk= pk)

    context= {
        'articulo_gasto':articulo_gasto,
        'productos':productos,
        'form':form,
        'form_product': form_product,
        'articulos':articulos,
        'entrada':entrada,
    }

    return render(request, 'gasto/crear_entrada.html', context)

def delete_articulo_entrada(request, pk):
   
    articulo = Conceptos_Entradas.objects.get(id=pk)
    gasto = articulo.entrada.gasto.id
    messages.success(request,f'El articulo {articulo.concepto_material} ha sido eliminado exitosamente')
    articulo.delete()

    return redirect('gasto-entrada',pk= gasto)


def render_pdf_gasto(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #Here ends conf.
    gasto = Solicitud_Gasto.objects.get(id=pk)
    productos = Articulo_Gasto.objects.filter(gasto=gasto)
    facturas = Factura.objects.filter(solicitud_gasto = gasto)

   #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)



    #Encabezado
    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'SUP. ADMON')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(520,caja_iso-10,'SUB ADM')
    c.drawString(150,caja_iso-20,'Número de documento')
    #c.drawString(160,caja_iso-30,'F-ADQ-N4-01.02')
    c.drawString(245,caja_iso-20,'Clasificación del documento')
    #c.drawString(275,caja_iso-30,'Controlado')
    c.drawString(355,caja_iso-20,'Nivel del documento')
    c.drawString(380,caja_iso-30, 'N5')
    c.drawString(440,caja_iso-20,'Revisión No.')
    c.drawString(452,caja_iso-30,'000')
    c.drawString(510,caja_iso-20,'Fecha de Emisión')
    c.drawString(525,caja_iso-30,'01/2024')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,575,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Comprobación de Gastos')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,575) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor-8,585,575) #Linea 2 contorno
    c.drawInlineImage('static/images/logo vordtec_documento.png',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(120,caja_proveedor,'Infor')
    c.drawString(300,caja_proveedor, 'Detalles')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(30,caja_proveedor-20,'Solicitó:')
    c.drawString(30,caja_proveedor-40,'Distrito:')
    c.drawString(30,caja_proveedor-60,'Clase')
    c.drawString(30,caja_proveedor-80,'Banco:')
    c.drawString(30,caja_proveedor-100,'Fecha:')
    # Segunda columna del encabezado
    c.drawString(280,caja_proveedor-60,'Depositar a:')
    c.drawString(280,caja_proveedor-20,'Cuenta:')
    c.drawString(280,caja_proveedor-40,'Clabe:')


    
    c.setFont('Helvetica-Bold',12)
    c.drawString(500,caja_proveedor-20,'FOLIO:')
    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_proveedor-20, str(gasto.id))

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(100,caja_proveedor-20, gasto.staff.staff.first_name+' '+ gasto.staff.staff.last_name)
    c.drawString(100,caja_proveedor-40, gasto.staff.distrito.nombre)
    c.drawString(100,caja_proveedor-60, gasto.tipo.tipo)
    if gasto.staff.banco:
        c.drawString(100,caja_proveedor-80, gasto.staff.banco.nombre)
    else:
        c.drawString(100,caja_proveedor-80, "Sin registro")
    if gasto.approved_at:
        c.drawString(100,caja_proveedor-100, gasto.approved_at.strftime("%d/%m/%Y"))
    else:
        c.drawString(100,caja_proveedor-100, 'Sin fecha')
    # Segunda Columna del encabezado
    if gasto.colaborador:
        c.drawString(350,caja_proveedor-60,gasto.colaborador.staff.first_name+' '+ gasto.colaborador.staff.last_name)
        if gasto.staff.cuenta_bancaria:
            c.drawString(350,caja_proveedor-20,str(gasto.colaborador.cuenta_bancaria))
        else:
            c.drawString(350,caja_proveedor-20, "Sin registro")
        if gasto.staff.clabe:
            c.drawString(350,caja_proveedor-40,str(gasto.colaborador.clabe))
        else:
            c.drawString(350,caja_proveedor-40, "Sin registro")
    else:
        c.drawString(350,caja_proveedor-60,gasto.staff.staff.first_name+' '+ gasto.staff.staff.last_name)
        if gasto.staff.cuenta_bancaria:
            c.drawString(350,caja_proveedor-20,str(gasto.staff.cuenta_bancaria))
        else:
            c.drawString(350,caja_proveedor-20, "Sin registro")
        if gasto.staff.clabe:
            c.drawString(350,caja_proveedor-40,str(gasto.staff.clabe))
        else:
            c.drawString(350,caja_proveedor-40, "Sin registro")

    #Create blank list
    data =[]

    data.append(['''Código''', '''Nombre''', '''Cantidad''','''Precio''', '''Subtotal''', '''Total''','''Comentario'''])


    high = 540
    for producto in productos:
         # Convert to Decimal and round to two decimal places
        cantidad = producto.cantidad if producto.cantidad is not None else 0
        cantidad_redondeada = Decimal(cantidad).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        precio = producto.precio_unitario if producto.cantidad is not None else 0
        precio_unitario_redondeado = Decimal(precio).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        subtotal = Decimal(cantidad_redondeada * precio_unitario_redondeado).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if producto.otros_impuestos is None:
            producto.otros_impuestos = 0
        if subtotal is None:
            subtotal = 0
        total = Decimal(subtotal) + Decimal(producto.otros_impuestos)
        data.append([
            producto.producto.producto.codigo, 
            producto.producto.producto.nombre,
            cantidad_redondeada, 
            precio_unitario_redondeado,
            subtotal, 
            total,
            #producto.comentario,
            ])
        high = high - 18


    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    #Primer renglón
    c.drawCentredString(70,48,'Clasificación:')
    c.drawCentredString(140,48,'Nivel:')
    c.drawCentredString(240,48,'Preparado por:')
    c.drawCentredString(350,48,'Aprobado:')
    c.drawCentredString(450,48,'Fecha emisión:')
    c.drawCentredString(550,48,'Rev:')
    #Segundo renglón
    c.drawCentredString(70,34,'Controlado')
    c.drawCentredString(140,34,'N5')
    c.drawCentredString(240,34,'SEOV-ALM-N4-01-01')
    c.drawCentredString(350,34,'SUB ADM')
    c.drawCentredString(450,34,'24/Oct/2018')
    c.drawCentredString(550,34,'001')

    c.setFillColor(black)
    width, height = letter
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]

    #if gasto.comentario is not None:
    #    comentario = gasto.comentario
    #else:
    #    comentario = "No hay comentarios"

   
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    # Personalizar el estilo de los párrafos
    custom_style = ParagraphStyle(
    'CustomStyle',
        parent=styles['BodyText'],
        fontSize=6,  # Reducir el tamaño de la fuente a 6
        leading=8,   # Aumentar el espacio entre líneas para asegurar que el texto no se superponga
        alignment=TA_LEFT,  # Alineación del texto
        # Puedes añadir más ajustes si es necesario
    )
    for i, row in enumerate(data):
        for j, item in enumerate(row):
            if i!=0 and j == 6:
                text = '' if item is None else str(item)
                data[i][j] = Paragraph(text, custom_style)

    table = Table(data, colWidths=[1.2 * cm, 6 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm, 1.5* cm, 6 * cm,])
    table_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 8),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    table.setStyle(table_style)

    #pdf size
    pos_table1 = high-20
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, pos_table1)
    # Crear una lista de datos para la tabla secundaria
    data_secundaria = []
    data_secundaria.append(['Proyecto', 'Subproyecto'])  # Encabezados de la tabla secundaria

    # Añadir filas de proyectos y subproyectos
    for producto in productos:
        data_secundaria.append([producto.proyecto.nombre, producto.subproyecto.nombre])

    # Crear la tabla secundaria
    table_secundaria = Table(data_secundaria, colWidths=[7 * cm, 7 * cm])  # Ajusta las medidas según necesites

    # Estilo para la tabla secundaria
    table_secundaria_style = TableStyle([
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BOX', (0,0), (-1,-1), 0.25, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (1,0), colors.grey),  # Fondo gris para los encabezados
        ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),  # Texto blanco para los encabezados
        ('TEXTCOLOR', (0,1), (-1,-1), colors.black),  # Texto negro para el cuerpo
        ('FONTSIZE', (0,0), (-1,-1), 8),  # Tamaño de fuente para toda la tabla
        # Añade aquí más estilos si lo necesitas
    ])

    table_secundaria.setStyle(table_secundaria_style)

    # Posición de la tabla secundaria en el PDF
    x_pos = 20  # Ajusta la posición X como sea necesario
    y_pos = pos_table1 - (len(data) * 18) - 20  # Ajusta la posición Y según el espacio ocupado por la primera tabla y cualquier otro contenido

    # Dibujar la tabla secundaria en el canvas
    table_secundaria.wrapOn(c, width, height)
    table_secundaria.drawOn(c, x_pos, y_pos)

    # 1. Preparar los datos para la tabla de facturas
    facturas = Factura.objects.filter(solicitud_gasto=gasto)
    data_facturas = [['Datos de XML', 'Nombre', 'Monto']]  # Encabezados de la tabla de facturas

    suma_total = Decimal('0.00')
    for factura in facturas:
        
        if factura.archivo_xml:
            emisor = factura.emisor  # Aquí emisor es un diccionario
            # Convertir el total a Decimal (o float) antes de formatear
            descripciones = [tupla[0] for tupla in emisor['resultados']]
            descripciones_str = ', '.join(descripciones)
           
            try:
                total_factura = Decimal(emisor['total'])
            except (InvalidOperation, ValueError):
                total_factura = Decimal('0.00')  # Si no es convertible, usa 0.00

            suma_total += total_factura  # Suma al total acumulado
            data_facturas.append([
                descripciones_str, 
                emisor['nombre'],
                f"${total_factura:,.2f}",  # Formatea el total como una cadena de texto
            ])

    for i, row in enumerate(data_facturas):
        for j, item in enumerate(row):
            if i!=0 and j == 0:
                 # Proporcionar un valor predeterminado si 'item' es None
                text = '' if item is None else str(item)
                data_facturas[i][j] = Paragraph(text, custom_style)
    # Crear un marco (frame) en la posición específica
    
    # 2. Crear la tabla de facturas
    table_factura = Table(data_facturas, colWidths=[11 * cm, 6 * cm, 2 * cm,])

   # Estilo para la tabla secundaria
    table_facturas_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 8),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    

    table_factura.setStyle(table_facturas_style)
    #Parrafó de totales
    data_totales = []
    diferencia_totales = suma_total - Decimal(gasto.get_total_solicitud)
    if diferencia_totales > 0:
        color_diferencia = colors.green
    elif diferencia_totales < 0:
        color_diferencia = colors.red
    else:
        color_diferencia = colors.black 
    total_str = "${:,.2f}".format(suma_total)  # Convierte Decimal a string y formatea
    # 4. Posición de la tabla de facturas en el PDF
    # Asumiendo que 'y_pos' es la posición Y después de dibujar la tabla secundaria y cualquier otro contenido
    

    data_totales = [
    ['Total solicitado', 'Total comprobado', 'Saldo A cargo/Favor en Pesos'],  # Encabezados
    ['$' + str(gasto.get_total_solicitud), f"${suma_total:,.2f}", Paragraph(f'${diferencia_totales:,.2f}', ParagraphStyle('CustomStyle', textColor=color_diferencia))]
]

    #data_totales.append(['Total solicitado', 'Total comprobado', 'Saldo A cargo/Favor en Pesos'])  # Encabezados de la tabla secundaria
    #data_totales.append(['$' + str(gasto.get_total_solicitud), total_str, '$' + str(diferencia_totales) ])
    table_totales = Table(data_totales, colWidths=[5 * cm, 5 * cm, 5 * cm])  # Ajusta las medidas según necesites
    table_totales.setStyle(table_secundaria_style)
    # Añadir filas de proyectos y subproyectos
   
    table_totales.wrapOn(c, width, height)
    y_totales_pos = y_pos - (len(data_totales) * 15 + 30) 
    table_totales.drawOn(c, 20, y_totales_pos)

    c.setFillColor(prussian_blue)
    c.rect(20, y_totales_pos-50,565,25, fill=True, stroke=False)
    c.setFillColor(white)
    c.drawCentredString(320, y_totales_pos-45,'Observaciones')
    c.setFillColor(black)
    #options_conditions_paragraph = Paragraph(comentario, styleN)
    # Crear un marco (frame) en la posición específica

    frame = Frame(50, 0, width, y_totales_pos-45, id='normal')

    # Agregar el párrafo al marco
    #frame.addFromList([options_conditions_paragraph], c)
    c.drawCentredString(230, y_totales_pos-190, gasto.staff.staff.first_name +' '+ gasto.staff.staff.last_name)
    c.drawCentredString(230,  y_totales_pos-205, 'Solicitado')
   
    c.setFillColor(black)
    c.drawCentredString(410, y_totales_pos-190, gasto.superintendente.staff.first_name +' '+ gasto.superintendente.staff.last_name)
    c.line(360,  y_totales_pos-195,460,  y_totales_pos-195)
    c.drawCentredString(410, y_totales_pos-205,'Aprobado por')


    c.showPage()
    y_facturas_pos =height - (len(data_facturas) * 18) - 220  # Ajusta según sea necesario
    
    
    #total_paragraph = Paragraph(total_str, styleN)
    #frame = Frame(50, 0, width, y_facturas_pos-100, id='normal')
    #frame.addFromList([total_paragraph], c)
    # Dibujar la tabla de facturas en el canvas
    table_factura.wrapOn(c, width, height)
    table_factura.drawOn(c, 20, y_facturas_pos)

    c.save()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Comprobación_Gasto_' + str(gasto.id) +'.pdf')

def convert_excel_gasto_matriz(gastos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Gastos_' + str(date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Gastos')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Folio','Fecha Autorización','Proyectos','Subproyectos','Comentarios','Colaborador','Solicitado para',
               'Importe','Fecha Creación','Status','Autorizado por','Facturas','Status de Pago']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5: #Columna del proveedor
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 20

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de Gastos").style = head_style
    ws.cell(row=4, column = columna_max, value="Sumatoria de Pagos Pendientes").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=4, column=columna_max + 1, value=f"=SUM(I:I)").style = money_resumen_style
  
    for gasto in gastos:
        row_num = row_num + 1    

        if gasto.pagada:
            pagada = "Tiene Pago"
        else: 
            pagada ="No tiene pago"

        if gasto.facturas.exists():
            facturas = "Con Facturas"
        else:
            facturas = "Sin Facturas"
        
        if gasto.autorizar2:
            status = "Autorizado"
            
            if gasto.autorizado_por2:
                autorizado_por = str(gasto.autorizado_por2.staff.first_name) + ' ' + str(gasto.autorizado_por2.staff.last_name)
            else:
                autorizado_por ="NR"
        elif gasto.autorizar2 == False:
            status = "Cancelado"
            autorizado_por =   str(gasto.autorizado_por2.staff.first_name) + ' ' + str(gasto.autorizado_por2.staff.last_name)
        elif gasto.autorizar:
            autorizado_por =str(gasto.superintendente.staff.first_name) + ' ' + str(gasto.superintendente.staff.last_name)
            status = "Autorizado | Falta una autorización"
        elif gasto.autorizar == False:
            status = "Cancelado"
            autorizado_por = str(gasto.superintendente.staff.last_name)
        else:
            autorizado_por = "Faltan autorizaciones"
            status = "Faltan autorizaciones"
        if gasto.approbado_fecha2 is None:
            gasto.approbado_fecha2 = ''
        proyectos = set()
        subproyectos = set()
        comentarios = set()
        articulos_gasto = Articulo_Gasto.objects.filter(gasto=gasto)
        for articulo in articulos_gasto:
            if articulo.proyecto:
                proyectos.add(str(articulo.proyecto.nombre))
            if articulo.subproyecto:
                subproyectos.add(str(articulo.subproyecto.nombre))
            if articulo.comentario:
                comentarios.add(str(articulo.comentario))

        proyectos_str = ', '.join(proyectos)
        subproyectos_str = ', '.join(subproyectos)
        comentarios_str = ', '.join(comentarios)

        row = [
            gasto.id,
            gasto.approbado_fecha2,
            proyectos_str,
            subproyectos_str,
            comentarios_str,
            gasto.staff.staff.first_name + ' ' + gasto.staff.staff.last_name,
            gasto.colaborador.staff.first_name + ' '  + gasto.colaborador.staff.last_name if gasto.colaborador else '',
            gasto.get_total_solicitud,
            gasto.created_at,
            status,
            autorizado_por,
            facturas,
            pagada,
            #f'=IF(I{row_num}="",G{row_num},I{row_num}*G{row_num})',  # Calcula total en pesos usando la fórmula de Excel
            #created_at_naive,
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num ==1 or col_num == 9:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
       
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)