from django.shortcuts import render, redirect
from dashboard.models import Inventario, Order, ArticulosOrdenados, ArticulosparaSurtir, Inventario_Batch, Marca, Product, Tipo_Orden, Plantilla, ArticuloPlantilla
from requisiciones.models import Requis, ArticulosRequisitados, ValeSalidas
from compras.models import Compra
from tesoreria.models import Pago
from solicitudes.models import Subproyecto, Operacion, Proyecto
from entradas.models import EntradaArticulo, Entrada
from requisiciones.views import get_image_base64
from gastos.models import Entrada_Gasto_Ajuste, Conceptos_Entradas
from .forms import InventarioForm, OrderForm, Inv_UpdateForm, Inv_UpdateForm_almacenista, ArticulosOrdenadosForm, Conceptos_EntradasForm, Entrada_Gasto_AjusteForm, Order_Resurtimiento_Form, ArticulosOrdenadosComentForm, Plantilla_Form, ArticuloPlantilla_Form
from dashboard.forms import Inventario_BatchForm
from user.models import Profile, Distrito, Almacen
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.conf import settings
import json
from django.db.models import Sum, Value, F, Sum, When, Case, DecimalField
from .filters import InventoryFilter, SolicitudesFilter, SolicitudesProdFilter, InventarioFilter, HistoricalInventarioFilter, HistoricalProductoFilter
from django.contrib import messages
import decimal
# Import Pagination Stuff
from django.core.paginator import Paginator
from datetime import date, datetime
# Import Excel Stuff
from django.db.models.functions import Concat
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt
import csv
import ast
import os
from django.core.mail import EmailMessage, BadHeaderError
from smtplib import SMTPException
# Create your views here.


def updateItem(request):
    data= json.loads(request.body)
    productId = data['productId']
    action = data['action']

    usuario = Profile.objects.get(staff__id=request.user.id)
    producto = Inventario.objects.get(id=productId)
    tipo = Tipo_Orden.objects.get(tipo ='normal')
    order, created = Order.objects.get_or_create(staff=usuario, complete=False, tipo = tipo, distrito = usuario.distrito)

    orderItem, created = ArticulosOrdenados.objects.get_or_create(orden = order, producto= producto)

    if action == 'add':
        orderItem.cantidad = (orderItem.cantidad + 1)
        message = f"Item was added: {orderItem}"
        orderItem.save()
    elif action == 'remove':
        orderItem.delete()
        message = f"Item was removed: {orderItem}"

    return JsonResponse(message, safe=False)

def updateItemRes(request):
    data= json.loads(request.body)
    productId = data['productId']
    action = data['action']

    usuario = Profile.objects.get(staff__id=request.user.id)
    producto = Inventario.objects.get(id=productId)
    tipo = Tipo_Orden.objects.get(tipo ='resurtimiento')
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distrito)
    orderItem, created = ArticulosOrdenados.objects.get_or_create(orden = order, producto= producto)

    if action == 'add':
        orderItem.cantidad = (orderItem.cantidad + 1)
        message = f"Item was added: {orderItem}"
        orderItem.save()
    elif action == 'remove':
        orderItem.delete()
        message = f"Item was removed: {orderItem}"

    return JsonResponse(message, safe=False)

#Vista de seleccion de productos, requiere login
@login_required(login_url='user-login')
def product_selection_resurtimiento(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    tipo = Tipo_Orden.objects.get(tipo ='resurtimiento')
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distrito)
    productos = Inventario.objects.filter(cantidad__lt =F('minimo'))
    cartItems = order.get_cart_quantity
    myfilter=InventoryFilter(request.GET, queryset=productos)
    productos = myfilter.qs


    context= {
        'myfilter': myfilter,
        'productos':productos,
        'productosordenadosres':cartItems,
        }
    return render(request, 'solicitud/product_selection_resurtimiento.html', context)

@login_required(login_url='user-login')
def crear_plantilla(request):
    usuarios = Profile.objects.all()
    creador = usuarios.get(staff__id=request.user.id)
    productos = Inventario.objects.filter(distrito = creador.distrito)
    
    plantilla, created = Plantilla.objects.get_or_create(creador = creador, complete = False)
    productos_plantilla = ArticuloPlantilla.objects.filter(plantilla = plantilla)
    form = Plantilla_Form()
    form_producto = ArticuloPlantilla_Form()

    if request.method =='POST' and "CrearBtn" in request.POST:
        form = Plantilla_Form(request.POST, instance=plantilla)
        if form.is_valid():
            plantilla = form.save(commit=False)
            plantilla.complete = True
            plantilla.save()
            messages.success(request, 'Has creado exitósamente la plantilla')
            return redirect('matriz-plantillas')
        else:
            messages.error(request, 'No está entrando')
    else:
        errores = form.errors.as_data()  # Esto te da un diccionario ordenado con los errores
        for campo, errores_campo in errores.items():
            for error in errores_campo:
                # Aquí, error es una instancia de ValidationError
                # Puedes acceder al mensaje de error con error.message
                print(f"Error en el campo {campo}: {error.message}")
                # Si quisieras hacer algo más con cada error, este es el lugar
           


    context = {
        'plantilla':plantilla,
        'productos':productos,
        'form':form,
        'form_producto':form_producto,
        'productos_plantilla':productos_plantilla,
    }

    return render(request, 'solicitud/crear_plantilla.html', context)

def update_item_plantilla(request):
    data= json.loads(request.body)
    plantilla_id = int(data['plantilla_id'])
    id_producto = int(data['id_producto'])
    cantidad = decimal.Decimal(data['cantidad'])
    comentario = str(data['comentario'])
    comentario_plantilla = str(data['comentario_plantilla'])
    print(comentario)
    action = data['action']

    usuario = Profile.objects.get(staff__id=request.user.id)
    producto = Inventario.objects.get(id=id_producto)
    
    plantilla = Plantilla.objects.get(id=plantilla_id)

    item, created = ArticuloPlantilla.objects.get_or_create(plantilla = plantilla, producto= producto)

    if action == 'add':
        item.cantidad = cantidad
        item.modified_at = date.today()
        item.modified_by = usuario
        item.comentario_articulo = comentario
        item.comentario_plantilla = comentario_plantilla
        messages.success(request, f'El producto {item.producto.producto.nombre} ha sido creado')
        item.save()
    elif action == 'remove':
        messages.success(request, f'El producto {item.producto.producto.nombre} ha sido eliminado')
        item.delete()
        

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)


#Vista de seleccion de productos, requiere login
@login_required(login_url='user-login')
def product_selection(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    tipo = Tipo_Orden.objects.get(tipo ='normal')
    #order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo = tipo)
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distrito)
    productos = Inventario.objects.filter(complete=True)
    cartItems = order.get_cart_quantity
    myfilter=InventoryFilter(request.GET, queryset=productos)
    productos = myfilter.qs

    #Set up pagination
    p = Paginator(productos, 30)
    page = request.GET.get('page')
    productos_list = p.get_page(page)


    context= {
        'myfilter': myfilter,
        'productos_list':productos_list,
        'productos':productos,
        'productosordenados':cartItems,
        }
    return render(request, 'solicitud/product_selection.html', context)

#Vista para crear solicitud
@login_required(login_url='user-login')
def checkout(request):
    usuarios = Profile.objects.all()
    ordenes = Order.objects.all()
    usuario = usuarios.get(staff=request.user)
    
    #Tengo que revisar primero si ya existe una orden pendiente del usuario
    orders = ordenes.filter(staff__distrito = usuario.distrito)
    last_order = orders.order_by('-last_folio_number').first()
    #consecutivo = orders.count() + 1
    proyectos = Proyecto.objects.filter(activo=True)
    subproyectos = Subproyecto.objects.all()
    tipo = Tipo_Orden.objects.get(tipo ='normal')

    order, created = ordenes.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distrito)

    if usuario.tipo.supervisor:
        supervisores = usuarios.filter(staff=request.user)
        order.supervisor = usuario
    else:
        supervisores = usuarios.filter(tipo__supervisor = True)

    if usuario.tipo.superintendente:
        superintendentes = usuarios.filter(staff=request.user)
        order.superintendente = usuario
    else:
        superintendentes = usuarios.filter(tipo__superintendente = True, staff__is_active = True).exclude(tipo__nombre="Admin")


    form = OrderForm(instance = order)
    form.fields['area'].queryset = Operacion.objects.exclude(nombre='GASTO')



    if order.staff != usuario:
        productos = None
        cartItems = 0
    else:
        productos = order.articulosordenados_set.all()
        cartItems = order.get_cart_quantity

    if request.method =='POST' and 'agregar' in request.POST:
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            order = form.save(commit=False)
            order.created_at = date.today()
            order.created_at_time = datetime.now().time()
            abrev= usuario.distrito.abreviado
            if last_order == None:
                #No hay órdenes para este distrito todavía
                folio_number = 1
            else:
                folio_number = last_order.last_folio_number + 1
                order.last_folio_number = folio_number
            order.folio = str(abrev) + str(folio_number).zfill(4)
            productos_html = '<ul>'
            if usuario.tipo.supervisor == True:
                for producto in productos:
                    productos_html += f'<li>{producto.producto.producto.nombre}: {producto.cantidad}.</li>'
                    # We fetch inventory product corresponding to product (that's why we use product.id)
                    # We create a new product line in a new database to control the ArticlestoDeliver (ArticulosparaSurtir)
                    prod_inventario = Inventario.objects.get(id = producto.producto.id)
                    ordensurtir , created = ArticulosparaSurtir.objects.get_or_create(articulos = producto)
                    #cond:1 evalua si la cantidad en inventario es mayor que lo solicitado
                    if prod_inventario.cantidad >= producto.cantidad and order.tipo.tipo == "normal":  #si la cantidad solicitada es mayor que la cantidad en inventario
                        prod_inventario.cantidad = prod_inventario.cantidad - producto.cantidad
                        prod_inventario.cantidad_apartada = producto.cantidad + prod_inventario.cantidad_apartada
                        prod_inventario._change_reason = f'Se modifica el inventario en view: autorizada_sol:{order.id} cond:1'
                        ordensurtir.cantidad = producto.cantidad
                        ordensurtir.precio = prod_inventario.price
                        ordensurtir.surtir = True
                        ordensurtir.requisitar = False
                        ordensurtir.save()
                        prod_inventario.save()
                    elif producto.cantidad >= prod_inventario.cantidad and producto.cantidad > 0 and order.tipo.tipo == "normal" and producto.producto.producto.servicio == False: #si la cantidad solicitada es mayor que la cantidad en inventario
                        ordensurtir.cantidad = prod_inventario.cantidad #lo que puedes surtir es igual a lo que tienes en el inventario
                        ordensurtir.precio = prod_inventario.price
                        ordensurtir.cantidad_requisitar = producto.cantidad - ordensurtir.cantidad #lo que falta por surtir
                        prod_inventario.cantidad_apartada = prod_inventario.cantidad_apartada + prod_inventario.cantidad
                        prod_inventario.cantidad = 0
                        if ordensurtir.cantidad > 0: #si lo que se puede surtir es mayor que 0
                            ordensurtir.surtir = True
                        ordensurtir.requisitar = True
                        order.requisitar = True
                        prod_inventario.save()
                        ordensurtir.save()
                    elif prod_inventario.cantidad + prod_inventario.cantidad_entradas == 0 or producto.producto.producto.servicio == True:
                        ordensurtir.requisitar = True
                        ordensurtir.cantidad_requisitar = producto.cantidad
                        order.requisitar = True
                        print(producto.producto.producto.servicio)
                        if producto.producto.producto.servicio == True:
                            requi, created = Requis.objects.get_or_create(complete = True, orden = order)
                            requitem, created = ArticulosRequisitados.objects.get_or_create(req = requi, producto= ordensurtir, cantidad = producto.cantidad, almacenista = usuario)
                            requi.folio = str(abrev) + str(folio_number).zfill(4)
                            if productos.count() == 1: 
                                order.requisitar=False
                                order.requisitado = True
                            ordensurtir.requisitar=False
                            requi.save()
                            requitem.save()
                        ordensurtir.save()
                        order.save()
                order.autorizar = True
                order.approved_at = date.today()
                order.approved_at_time = datetime.now().time()
                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
                productos_html += '</ul>'
                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)
                # Crear el mensaje HTML
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body>
                        <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                        <p>Estimado {order.staff.staff.first_name} {order.staff.staff.last_name},</p>
                        <p>Estás recibiendo este correo porque tu solicitud folio:{order.folio}  se ha generado</p>
                        <p>Con los productos siguientes</p>
                        {productos_html}
                        <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                        <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                    </body>
                </html>
                """
                try:
                    email = EmailMessage(
                        f'Solicitud Autorizada {order.folio}',
                        body=html_message,
                        from_email= settings.DEFAULT_FROM_EMAIL,
                        to=[order.staff.staff.email, 'ulises_huesc@hotmail.com'],
                        headers={'Content-Type': 'text/html'}
                        )
                    email.content_subtype = "html " # Importante para que se interprete como HTML
                    email.send()
                    messages.success(request, f'La solicitud {order.folio} ha sido creada')
                except (BadHeaderError, SMTPException) as e:
                    error_message = f'La solicitud {order.folio} ha sido creada, pero el correo no ha sido enviado debido a un error: >>> {e}'
                    messages.warning(request, error_message)
                order.sol_autorizada_por = Profile.objects.get(staff__id=request.user.id)    
                cartItems = '0'
            else:
                for producto in productos:
                    productos_html += f'<li>{producto.producto.producto.nombre}: {producto.cantidad}.</li>'
                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
                productos_html += '</ul>'
                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)
                # Crear el mensaje HTML
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body>
                        <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                        <p>Estimado {order.staff.staff.first_name} {order.staff.staff.last_name},</p>
                        <p>Estás recibiendo este correo porque tu solicitud folio:{order.folio}  se ha generado</p>
                        <p>Con los productos siguientes</p>
                        {productos_html}
                        <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                        <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                    </body>
                </html>
                """
                try:
                    email = EmailMessage(
                        f'Solicitud Autorizada {order.folio}',
                        body=html_message,
                        from_email= settings.DEFAULT_FROM_EMAIL,
                        to=[order.staff.staff.email, 'ulises_huesc@hotmail.com'],
                        headers={'Content-Type': 'text/html'}
                        )
                    email.content_subtype = "html " # Importante para que se interprete como HTML
                    email.send()
                    messages.success(request, f'La solicitud {order.folio} ha sido creada')
                except (BadHeaderError, SMTPException) as e:
                    error_message = f'La solicitud {order.folio} ha sido creada, pero el correo no ha sido enviado debido a un error: {e}'
                    messages.warning(request, error_message)
            order.complete = True
            order.save()
            return redirect('solicitud-matriz')

    


    context= {
        'form':form,
        'productos':productos,
        'orden':order,
        'productosordenados':cartItems,
        'supervisores':supervisores,
        'superintendentes':superintendentes,
        'proyectos':proyectos,
        'subproyectos':subproyectos,
    }
    return render(request, 'solicitud/checkout.html', context)

def product_quantity_edit(request, pk):
    item = ArticulosOrdenados.objects.get(id= pk)
    form= ArticulosOrdenadosForm(instance = item)

    if request.method == 'POST':
        form = ArticulosOrdenadosForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return HttpResponse(status=204)

    context = {
        'form': form,
        'item':item,
        }

    return render(request, 'solicitud/product_quantity_edit.html', context)

def product_comment_add(request, pk):
    item = ArticulosOrdenados.objects.get(id= pk)
    form= ArticulosOrdenadosComentForm(instance = item)

    if request.method == 'POST':
        form = ArticulosOrdenadosComentForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return HttpResponse(status=204)

    context = {
        'form': form,
        'item':item,
        }

    return render(request, 'solicitud/product_comment_add.html', context)

#Vista para crear solicitud de resurtimiento
@login_required(login_url='user-login')
def checkout_resurtimiento(request):
    usuario = Profile.objects.get(staff=request.user)
    #Tengo que revisar primero si ya existe una orden pendiente del usuario
    superintendentes = Profile.objects.filter(tipo__superintendente = True, staff__is_active = True).exclude(tipo__nombre="Admin")
    proyectos = Proyecto.objects.filter(activo=True)
    subproyectos = Subproyecto.objects.all()
    orders = Order.objects.filter(staff__distrito = usuario.distrito)
    last_order = orders.order_by('-last_folio_number').first()
    #consecutivo = orders.count()+1



    tipo = Tipo_Orden.objects.get(tipo ='resurtimiento')
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distrito)
    almacen = Operacion.objects.get(nombre = "ALMACEN")

    if order.staff != usuario:
        productos = None
        cartItems = 0
    else:
        productos = order.articulosordenados_set.all()
        cartItems = order.get_cart_quantity

    form = OrderForm(instance = order)


    if request.method =='POST':
        form = Order_Resurtimiento_Form(request.POST, instance=order)
        if form.is_valid():
            order = form.save(commit=False)
            order.supervisor = usuario
            order.created_at = date.today()
            order.created_at_time = datetime.now().time()
            order.complete = True
            order.area = almacen
            abrev= usuario.distrito.abreviado
            last_order = orders.order_by('-last_folio_number').first()
            if last_order == None:
                #No hay órdenes para este distrito todavía
                folio_number = 1
            else:
                folio_number = last_order.last_folio_number + 1
                order.last_folio_number = folio_number
            order.folio = str(abrev) + str(folio_number).zfill(4)

            requi, created = Requis.objects.get_or_create(complete = True, orden = order)
            requi.folio = str(abrev) + str(requi.id).zfill(4)
            requi.save()
            for producto in productos:
                ordensurtir , created = ArticulosparaSurtir.objects.get_or_create(articulos = producto)
                requitem, created = ArticulosRequisitados.objects.get_or_create(req = requi, producto= ordensurtir, cantidad = producto.cantidad)
                ordensurtir.requisitar = True
                ordensurtir.cantidad_requisitar = producto.cantidad
                ordensurtir.save()
                requitem.save()
            order.requisitado = True
            order.autorizar = True
            order.approved_at = date.today()
            order.approved_at_time = datetime.now().time()
            requi.save()
            order.save()
        #abrev= usuario.distrito.abreviado
        #order.folio = str(abrev) + str(order.id).zfill(4)
            messages.success(request, f'La solicitud {order.folio} junto con la requisición {requi.folio} ha sido creada')
            cartItems = '0'
            return redirect('solicitud-matriz')
    else:
        form = OrderForm(request.POST)

    context= {
        'proyectos':proyectos,
        'form':form,
        'productos':productos,
        'orden':order,
        'productosordenadosres':cartItems,
        'superintendentes':superintendentes,
        'subproyectos':subproyectos,
    }
    return render(request, 'solicitud/checkout_resurtimiento.html', context)


#Vista para crear solicitud
@login_required(login_url='user-login')
def checkout_editar(request, pk):

    order = Order.objects.get(id=pk)

    usuario = Profile.objects.get(id=request.user.id)

    productos = order.articulosordenados_set.all()
    cartItems = order.get_cart_quantity
    form = OrderForm(instance=order, distrito = usuario.distrito)


    if request.method =='POST':
        form = OrderForm(request.POST, instance=order, distrito = usuario.distrito)
        order.complete = True
        if form.is_valid():
            form.save()
            cartItems = '0'
            return redirect('solicitud-matriz')

    context= {
        'form':form,
        'productos':productos,
        'orden':order,
        'productosordenados':cartItems,
    }
    return render(request, 'solicitud/checkout.html', context)

@login_required(login_url='user-login')
def solicitud_pendiente(request):

    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    perfil = Profile.objects.get(staff__id=request.user.id)


    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.superintendente == True or perfil.tipo.nombre == "Control":
        ordenes = Order.objects.filter(complete=True, staff__distrito=perfil.distrito).order_by('-folio')
    elif perfil.tipo.supervisor == True:
        ordenes = Order.objects.filter(complete=True, staff__distrito=perfil.distrito, supervisor=perfil).order_by('-folio')
    else:
        ordenes = Order.objects.filter(complete=True, staff = perfil).order_by('-folio')

    myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    ordenes = myfilter.qs

    #Set up pagination
    p = Paginator(ordenes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:

        return convert_excel_solicitud_matriz(ordenes)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'solicitud/solicitudes_pendientes.html',context)

@login_required(login_url='user-login')
def solicitud_matriz(request):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    perfil = Profile.objects.get(staff__id=request.user.id)


     #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.superintendente == True or perfil.tipo.nombre == "Control":
        ordenes = Order.objects.filter(complete=True, staff__distrito=perfil.distrito).order_by('-folio')
    elif perfil.tipo.supervisor == True:
        ordenes = Order.objects.filter(complete=True, staff__distrito=perfil.distrito, supervisor=perfil).order_by('-folio')
    else:
        ordenes = Order.objects.filter(complete=True, staff = perfil).order_by('-folio')

    myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    ordenes = myfilter.qs

    #Set up pagination
    p = Paginator(ordenes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:

        return convert_excel_solicitud_matriz(ordenes)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'solicitud/solicitudes_creadas.html',context)

@login_required(login_url='user-login')
def matriz_plantillas(request):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    perfil = Profile.objects.get(staff__id=request.user.id)
    plantillas_list= Plantilla.objects.filter(complete=True)

    #myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    #ordenes = myfilter.qs

    #Set up pagination
    #p = Paginator(ordenes, 10)
    #page = request.GET.get('page')
    #ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

    #    return convert_excel_solicitud_matriz(ordenes)

    context= {
        'plantillas_list':plantillas_list,
        #'myfilter':myfilter,
        }

    return render(request, 'solicitud/matriz_plantillas.html',context)

@login_required(login_url='user-login')
def productos_plantilla(request, pk):
    plantilla = Plantilla.objects.get(id=pk)
    productos = ArticuloPlantilla.objects.filter(plantilla=plantilla)

    context = {
        'productos':productos,
    }

    return render(request,'solicitud/productos_plantilla.html', context)

@login_required(login_url='user-login')
def editar_plantilla(request, pk):
    plantilla = Plantilla.objects.get(id=pk)
    usuarios = Profile.objects.all()
    usuario = usuarios.get(staff__id=request.user.id)
    productos = Inventario.objects.filter(distrito = usuario.distrito)
    
    productos_plantilla = ArticuloPlantilla.objects.filter(plantilla = plantilla)
    form = Plantilla_Form(instance = plantilla)
    form_producto = ArticuloPlantilla_Form()

    if request.method =='POST' and "CrearBtn" in request.POST:
        form = Plantilla_Form(request.POST, instance=plantilla)
        if form.is_valid():
            plantilla = form.save(commit=False)
            plantilla.complete = True
            plantilla.modified_at = date.today()
            plantilla.modified_by = usuario
            plantilla.save()
            messages.success(request, 'Has modificado exitósamente la plantilla')
            return redirect('matriz-plantillas')
        else:
            messages.error(request, 'No está entrando')
    else:
        errores = form.errors.as_data()  # Esto te da un diccionario ordenado con los errores
        for campo, errores_campo in errores.items():
            for error in errores_campo:
                # Aquí, error es una instancia de ValidationError
                # Puedes acceder al mensaje de error con error.message
                print(f"Error en el campo {campo}: {error.message}")
                # Si quisieras hacer algo más con cada error, este es el lugar
           


    context = {
        'plantilla':plantilla,
        'productos':productos,
        'form':form,
        'form_producto':form_producto,
        'productos_plantilla':productos_plantilla,
    }
    
    return render(request, 'solicitud/editar_plantilla.html', context)

@login_required(login_url='user-login')
def crear_solicitud_plantilla(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    tipo = Tipo_Orden.objects.get(tipo ='normal')
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distrito)
    
    # Obtiene la plantilla por su ID
    plantilla = Plantilla.objects.get(id=pk)
    
    # Añade productos de la plantilla a la orden
    for articulo in plantilla.articuloplantilla_set.all():
        # Aquí asumo que tienes un modelo que conecta un producto con una orden (quizás se llame "ArticulosOrdenados" o algo similar).
        # Si ese modelo no existe, deberás adaptar este código.
        articulo_orden, created = ArticulosOrdenados.objects.get_or_create(orden=order, producto=articulo.producto)
        articulo_orden.cantidad += articulo.cantidad  # Aumenta la cantidad basada en la plantilla
         # Copia el comentario del artículo de la plantilla al artículo ordenado.
        articulo_orden.comentario = articulo.comentario_articulo

        articulo_orden.save()

    return redirect('solicitud-checkout')  # Redirige al usuario a la selección de productos, donde ahora verá los productos de la plantilla añadidos


@login_required(login_url='user-login')
def solicitud_matriz_productos(request):

    perfil = Profile.objects.get(staff__id=request.user.id)

     #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.superintendente == True or perfil.tipo.nombre == "Control":
        productos = ArticulosOrdenados.objects.filter(orden__complete=True, orden__staff__distrito=perfil.distrito).order_by('-orden__folio')
    elif perfil.tipo.supervisor == True:
        productos = ArticulosOrdenados.objects.filter(orden__complete=True, orden__staff__distrito=perfil.distrito, orden__supervisor=perfil).order_by('-orden__folio')
    else:
        productos = ArticulosOrdenados.objects.filter(orden__complete=True, orden__staff = perfil).order_by('-orden__folio')

    myfilter=SolicitudesProdFilter(request.GET, queryset=productos)
    productos = myfilter.qs
    perfil = Profile.objects.get(staff__id=request.user.id)


    #Set up pagination
    p = Paginator(productos, 15)
    page = request.GET.get('page')
    productos_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:

        return convert_excel_solicitud_matriz_productos(productos)

    context= {
        'productos':productos_list,
        'myfilter':myfilter,
        }
    return render(request, 'solicitud/solicitudes_creadas_productos.html',context)



@login_required(login_url='user-login')
def inventario(request):
    sql_apartadas = """SELECT 
    dashboard_inventario.id AS id,
    SUM(dashboard_articulosparasurtir.cantidad) AS total_cantidad_por_surtir
    FROM
        dashboard_inventario
    JOIN
        dashboard_articulosordenados ON dashboard_inventario.id = dashboard_articulosordenados.producto_id
    JOIN
        dashboard_articulosparasurtir ON dashboard_articulosordenados.id = dashboard_articulosparasurtir.articulos_id
    WHERE
        dashboard_articulosparasurtir.surtir = TRUE
    GROUP BY
        dashboard_inventario.id; """

    sql_entradas = """SELECT 
    dashboard_inventario.id AS id,
        SUM(entradas_entradaarticulo.cantidad_por_surtir) AS total_entradas_por_surtir
    FROM
        dashboard_inventario
    JOIN
        dashboard_product ON dashboard_inventario.producto_id = dashboard_product.id
    JOIN
        dashboard_articulosordenados ON dashboard_inventario.id = dashboard_articulosordenados.producto_id
    JOIN
        dashboard_articulosparasurtir ON dashboard_articulosordenados.id = dashboard_articulosparasurtir.articulos_id
    JOIN
        requisiciones_articulosrequisitados ON dashboard_articulosparasurtir.id = requisiciones_articulosrequisitados.producto_id
    JOIN
        compras_articulocomprado ON requisiciones_articulosrequisitados.id = compras_articulocomprado.producto_id
    JOIN
        entradas_entradaarticulo ON compras_articulocomprado.id = entradas_entradaarticulo.articulo_comprado_id
    WHERE 
        dashboard_inventario.complete = TRUE AND
        dashboard_product.servicio = FALSE AND
        dashboard_product.gasto = FALSE
    GROUP BY
        dashboard_inventario.id;
    """

    resultados_sql_apartadas = Inventario.objects.raw(sql_apartadas)
    resultados_sql_entradas = Inventario.objects.raw(sql_entradas)
    dict_resultados = {r.id: r.total_cantidad_por_surtir for r in resultados_sql_apartadas}
    dict_entradas = {r.id: r.total_entradas_por_surtir for r in resultados_sql_entradas}
    perfil = Profile.objects.get(staff=request.user)
    existencia = Inventario.objects.filter(
        complete=True,
        producto__servicio = False, 
        producto__gasto = False
        ).order_by('producto__codigo')

    if perfil.tipo.nombre == 'Admin' or perfil.tipo.nombre == 'SuperAdm':
        perfil_flag = True
    else:
        perfil_flag = False

    valor_inv = 0
    for inv in existencia:
        inv.total_entradas = dict_entradas.get(inv.id,0)
        inv.total_apartado = dict_resultados.get(inv.id,0) #2 ciclos for uno para calcular el valor del inventario
        valor_inv += (inv.cantidad + inv.total_apartado) * inv.price # y otro para calcular los apartados
   

    myfilter = InventarioFilter(request.GET, queryset=existencia)
    existencia = myfilter.qs

    #Set up pagination
    p = Paginator(existencia, 50)
    page = request.GET.get('page')
    existencia_list = p.get_page(page)

    cuenta_productos = existencia.count()

    if request.method =='POST' and 'btnExcel' in request.POST:
        return convert_excel_inventario(existencia, valor_inv, dict_entradas, dict_resultados)

    context = {
        'cuenta_productos':cuenta_productos,
        'perfil_flag':perfil_flag,
        'existencia': existencia,
        'myfilter': myfilter,
        'existencia_list':existencia_list,
        #'entradas':entradas,
        'valor_inv': valor_inv,
        }

    return render(request,'dashboard/inventario.html', context)

@login_required(login_url='user-login')
def ajuste_inventario(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    productos_sel = Inventario.objects.filter(complete=True, producto__servicio = False, producto__gasto = False)
    ajuste, created = Entrada_Gasto_Ajuste.objects.get_or_create(almacenista = usuario, completo = False, gasto = None)
    productos_ajuste = Conceptos_Entradas.objects.filter(entrada = ajuste)
    cantidad_items = productos_ajuste.count()
    form = Conceptos_EntradasForm()
    form2 = Entrada_Gasto_AjusteForm()

    form.fields['concepto_material'].queryset = productos_sel

    if request.method == 'POST':
        if "agregar_ajuste" in request.POST:
            form2 = Entrada_Gasto_AjusteForm(request.POST, instance=ajuste)
            if form2.is_valid():
                ajuste.completo= True
                ajuste.completado_hora = datetime.now().time()
                ajuste.completado_fecha = date.today()
                messages.success(request,f'{usuario.staff.first_name},Has hecho un ajuste de manera exitosa')
                #ajuste.save()
                for item_producto in productos_ajuste:
                    producto_inventario = Inventario.objects.get(producto= item_producto.concepto_material.producto)
                    productos_por_surtir = ArticulosparaSurtir.objects.filter(articulos__producto=producto_inventario, requisitar = True)
                    #Calculo el precio 
                    producto_inventario.price = ((item_producto.precio_unitario * item_producto.cantidad)+ ((producto_inventario.cantidad_apartada + producto_inventario.cantidad) * producto_inventario.price))/(producto_inventario.cantidad + item_producto.cantidad + producto_inventario.cantidad_apartada)
                    #La cantidad en inventario + la cantidad del producto en la entrada
                    producto_inventario.cantidad = producto_inventario.cantidad + item_producto.cantidad
                    for item in productos_por_surtir:
                        orden_producto = Order.objects.get(id = item.articulos.orden.id)                
                        #Si la cantidad en inventario es mayor que la cantidad requisitada
                        if producto_inventario.cantidad >= item.cantidad_requisitar:
                            cantidad = item.cantidad_requisitar
                        else:
                            cantidad = producto_inventario.cantidad
                        item.requisitar = False
                        item.cantidad = item.cantidad + cantidad
                        item.cantidad_requisitar = item.cantidad_requisitar - cantidad
                        if item.cantidad_requisitar == 0:
                            item.surtir = True
                        #Se reduce la cantidad de inventario y se aumenta la apartada
                        producto_inventario.cantidad = producto_inventario.cantidad - cantidad
                        #producto_inventario.cantidad_apartada = producto_inventario.cantidad_apartada + cantidad
                        producto_inventario.save()
                        item.save()
                        articulos_por_surtir = ArticulosparaSurtir.objects.filter(articulos__orden=orden_producto)
                        #Se cuentan los articulos por surtir de esa orden, se cuentan los articulos que ya no requieren requisición
                        numero_articulos = articulos_por_surtir.count()
                        numero_articulos_requisitados = articulos_por_surtir.filter(requisitar = False).count()
                        #si el numero total de articulos por surtir ya no requieren requisición
                        if numero_articulos == numero_articulos_requisitados:
                            orden_producto.requisitar = False   # entonces ya no se requiere que la Orden se requisite
                            orden_producto.save()
                    producto_inventario._change_reason = f'Esta es una ajuste desde un ajuste de inventario {ajuste.id}'
                    producto_inventario.save()
                ajuste.save()
                #email = EmailMessage(
                #    f'Ajuste de producto: {ajuste.id}',
                #    f'Estimado {usuario.staff.first_name} {usuario.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {orden.folio} ha sido devuelta al almacén por {usuario.staff.first_name} {usuario.staff.last_name}, con el siguiente comentario {devolucion.comentario} para más información comunicarse al almacén.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                #    'savia@vordtec.com',
                #    ['ulises_huesc@hotmail.com'],#orden.staff.staff.email],
                #    )
                #email.send()

                return redirect('solicitud-inventario')

    context= {
        'productos_ajuste':productos_ajuste,
        'form':form,
        'form2':form2,
        'ajuste': ajuste,
        'cantidad_items':cantidad_items,
        'productos_sel': productos_sel,
        }

    return render(request, 'dashboard/ajuste_inventario.html',context)

def update_ajuste(request):
    data= json.loads(request.body)
    action = data["action"]
    cantidad = decimal.Decimal(data["cantidad"])
    ajuste = data["ajuste"]
    producto_id = int(data["id"])
    precio = decimal.Decimal(data["precio"])
    producto = Inventario.objects.get(id=producto_id)
    ajuste = Entrada_Gasto_Ajuste.objects.get(id = ajuste)
    if action == "add":
        articulo, created = Conceptos_Entradas.objects.get_or_create(concepto_material=producto, entrada = ajuste)
        articulo.precio_unitario = precio
        articulo.cantidad = cantidad
        articulo.save()
        messages.success(request,'Has agregado producto de manera exitosa')
        ajuste.save()
    if action == "remove":
        articulo = Conceptos_Entradas.objects.get(concepto_material = producto, entrada = ajuste)
        messages.success(request,'Has eliminado un producto de tu listado')
        articulo.delete()
    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)


@login_required(login_url='user-login')
def upload_batch_inventario(request):

    form = Inventario_BatchForm(request.POST or None, request.FILES or None)


    if form.is_valid():
        form.save()
        form = Inventario_BatchForm()
        inventario_list = Inventario_Batch.objects.get(activated = False)

        f = open(inventario_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader)

        for row in reader:
            if Product.objects.filter(codigo=row[0]):
                producto = Product.objects.get(codigo=row[0])
                if Distrito.objects.filter(nombre = row[1]):
                    distrito = Distrito.objects.get(nombre = row[1])
                    if Almacen.objects.filter(nombre = row[2]):
                        almacen = Almacen.objects.get(nombre = row[2])
                        inventario = Inventario(producto=producto,distrito=distrito, almacen=almacen, ubicacion=row[3], estante=row[4], cantidad=row[5], price=row[6], minimo=row[7],comentario=row[8],complete=True)
                        inventario.save()
                        #marcas_str = row[2]
                        #marcas_names = ast.literal_eval(marcas_str)
                        #marcas_names = map(str.lower, marcas_names) # normalize them, all lowercase
                        #marcas_names = list(set(marcas_names)) # remove duplicates

                        #for marca in marcas_names:
                        #    marca, _ = Marca.objects.get_or_create(nombre=marca)
                        #    inventario.marca.add(marca)
                        #    inventario.save()
                    else:
                        messages.error(request,'El almacén no existe en la base de datos')
                else:
                     messages.error(request,'El distrito no existe en la base de datos')
            else:
                messages.error(request,f'El producto código:{row[0]} ya existe dentro de la base de datos')

        inventario_list.activated = True
        inventario_list.save()
    elif request.FILES:
        messages.error(request,'El formato no es CSV')




    context = {
        'form': form,
        }

    return render(request,'dashboard/upload_batch_inventario.html', context)




def inventario_add(request):
    #usuario = request.user.id
    perfil = Profile.objects.get(staff__id=request.user.id)

    #productos.exclude(id__in=existing)
    form = InventarioForm()
    #form.fields['producto'].queryset = productos


    if request.method =='POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.complete = True
            item._change_reason = 'Se agrega producto el inventario en view: inventario_add'
            item.distrito = perfil.distrito
            item.save()
            messages.success(request, f'El artículo {item.producto.codigo}:{item.producto.nombre} se ha agregado exitosamente')
            return HttpResponse(status=204)
    #else:
        #form = InventarioForm()

    context = {
        'form': form,
        #'productos':productos,
        }

    return render(request,'dashboard/inventario_add.html',context)

@login_required(login_url='user-login')
def inventario_update_modal(request, pk):
    perfil = Profile.objects.get(staff=request.user)
    item = Inventario.objects.get(id=pk)



    if perfil.tipo.nombre == 'SuperAdm' or perfil.tipo.nombre == 'Admin':
        flag_perfil = True
    else:
        flag_perfil = False


    if request.method =='POST':
        if perfil.tipo.nombre == 'SuperAdm' or perfil.tipo.nombre == 'Admin':
            form = Inv_UpdateForm(request.POST, instance=item)
        else:
            form = Inv_UpdateForm_almacenista(request.POST, instance= item)
        if request.POST['comentario'] and 'btnUpdate' in request.POST:
            if form.is_valid():
                item = form.save(commit=False)
                item._change_reason = item.comentario +'. Se modifica inventario en view: inventario_update_modal'
                item.save()
                messages.success(request, f'El artículo {item.producto.codigo}:{item.producto.nombre} se ha actualizado exitosamente')
                return HttpResponse(status=204)
        else:
            messages.error(request, 'Debes agregar un comentario con respecto al cambio realizado')
    else:
        if perfil.tipo.nombre == 'SuperAdm' or perfil.tipo.nombre == 'Admin' or perfil.tipo.nombre == "Almacen":
            form = Inv_UpdateForm(instance=item)
        else:
            form = Inv_UpdateForm_almacenista(instance= item)


    context = {
        'flag_perfil':flag_perfil,
        'form': form,
        'item':item,
        }

    return render(request,'dashboard/inventario_update_modal.html',context)


@login_required(login_url='user-login')
def historico_inventario(request):
    registros = Inventario.history.all()

    myfilter = HistoricalInventarioFilter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        }

    return render(request,'dashboard/historico_inventario.html',context)


@login_required(login_url='user-login')
def historico_producto(request):
    registros = Product.history.all()

    myfilter = HistoricalProductoFilter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        }

    return render(request,'dashboard/historico_producto.html',context)


@login_required(login_url='user-login')
def inventario_delete(request, pk):
    item = Inventario.objects.get(id=pk)

    if request.method == 'POST':
        item.delete()
        return redirect('solicitud-inventario')

    return render(request,'dashboard/inventario_delete.html')

@login_required(login_url='user-login')
def solicitud_autorizacion(request):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    #usuario = request.user.id
    perfil = Profile.objects.get(staff__id=request.user.id)
    #perfil = Profile.objects.get(id=usuario)

    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    
    ordenes = Order.objects.filter(complete=True, autorizar=None, staff__distrito=perfil.distrito).order_by('-folio')
    if perfil.tipo.nombre == "Admin":
        ordenes = ordenes
    else:
        ordenes = ordenes.filter(supervisor=perfil)
    myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    ordenes = myfilter.qs


    context= {
        'myfilter':myfilter,
        'ordenes':ordenes,
        #'perfil':perfil,
        }

    return render(request, 'autorizacion/solicitudes_pendientes_autorizacion.html',context)

def detalle_autorizar(request, pk):
    productos = ArticulosOrdenados.objects.filter(orden=pk)

    context = {
        'productos': productos,
     }
    return render(request,'autorizacion/detail.html', context)

@login_required(login_url='user-login')
def autorizada_sol(request, pk):
    usuario = request.user.id
    perfil = Profile.objects.get(staff__id=request.user.id)
    #perfil = Profile.objects.get(id=usuario)
    order = Order.objects.get(id = pk)
    productos = ArticulosOrdenados.objects.filter(orden = pk)
    requis = Requis.objects.filter(orden__staff__distrito = perfil.distrito)
    consecutivo = requis.count() + 1

    if request.method =='POST':
        #We go trough all the products one by one contained in the order
        for producto in productos:
            # We fetch inventory product corresponding to product (that's why we use product.id)
            # We create a new product line in a new database to control the ArticlestoDeliver (ArticulosparaSurtir)
            prod_inventario = Inventario.objects.get(id = producto.producto.id)
            ordensurtir , created = ArticulosparaSurtir.objects.get_or_create(articulos = producto)
            #cond:1 evalua si la cantidad en inventario es mayor que lo solicitado
            if prod_inventario.cantidad >= producto.cantidad and order.tipo.tipo == "normal":
                prod_inventario.cantidad = prod_inventario.cantidad - producto.cantidad
                prod_inventario.cantidad_apartada = prod_inventario.apartada
                prod_inventario._change_reason = f'Se modifica el inventario en view: autorizada_sol:{order.id} cond:1'
                ordensurtir.cantidad = producto.cantidad
                ordensurtir.precio = prod_inventario.price
                ordensurtir.surtir = True
                ordensurtir.requisitar = False
                ordensurtir.save()
                prod_inventario.save()
            elif producto.cantidad >= prod_inventario.cantidad and producto.cantidad > 0 and order.tipo.tipo == "normal" and producto.producto.producto.servicio == False: #si la cantidad solicitada es mayor que la cantidad en inventario 
                ordensurtir.cantidad = prod_inventario.cantidad #lo que puedes surtir es igual a lo que tienes en el inventario
                ordensurtir.precio = prod_inventario.price
                ordensurtir.cantidad_requisitar = producto.cantidad - ordensurtir.cantidad #lo que falta por surtir
                prod_inventario.cantidad_apartada = prod_inventario.apartada
                prod_inventario.cantidad = 0
                if ordensurtir.cantidad > 0: #si lo que se puede surtir es mayor que 0
                    ordensurtir.surtir = True
                ordensurtir.requisitar = True
                order.requisitar = True
                prod_inventario.save()
                ordensurtir.save()
            elif prod_inventario.cantidad + prod_inventario.cantidad_entradas == 0 or order.tipo.tipo == "resurtimiento" or  producto.producto.producto.servicio == True:
                ordensurtir.requisitar = True
                ordensurtir.cantidad_requisitar = producto.cantidad
                order.requisitar = True
                if producto.producto.producto.servicio == True:
                    requi, created = Requis.objects.get_or_create(complete = True, orden = order)
                    requitem, created = ArticulosRequisitados.objects.get_or_create(req = requi, producto= ordensurtir, cantidad = producto.cantidad)
                    requi.folio = str(perfil.distrito.abreviado)+str(requi.id).zfill(4)
                    order.requisitar=False
                    ordensurtir.requisitar=False
                    requi.save()
                    requitem.save()
                ordensurtir.save()
                order.save()
        order.autorizar = True
        order.approved_at = date.today()
        order.approved_at_time = datetime.now().time()
        #send_mail(
        #    f'Solicitud Autorizada {order.folio}',
        #    f'{order.staff.staff.first_name}, la solicitud {order.folio} ha sido autorizada. Este mensaje ha sido automáticamente generado por SAVIA X',
        #    'saviax.vordcab@gmail.com',
        #    [order.staff.staff.email],
        #    )
        order.sol_autorizada_por = Profile.objects.get(staff__id=request.user.id)
        order.save()

        messages.success(request, f'{perfil.staff.first_name} has autorizado la solicitud {order.folio}')
        return redirect('solicitud-pendientes-autorizacion')


    context = {
        'orden': order,
        'productos': productos,
    }

    return render(request,'autorizacion/autorizada.html', context)

def cancelada_sol(request, pk):
    order = Order.objects.get(id = pk)
    productos = ArticulosOrdenados.objects.filter(orden = pk)

    if request.method =='POST':
        order.autorizar = False
        order.save()
        messages.error(request, f'La orden {order} ha sido cancelada')
        return redirect('solicitud-pendientes-autorizacion')


    context = {
        'orden': order,
        'productos': productos,
    }

    return render(request,'autorizacion/cancelada.html', context)


def status_sol(request, pk):
    solicitud = Order.objects.get(id = pk)
    product_solicitudes = ArticulosOrdenados.objects.filter(orden=pk)
    product_surtir = ArticulosparaSurtir.objects.filter(articulos__orden = pk)
    listo_surtir = False
    for item in product_surtir:
        if item.surtir == True:
            listo_surtir = True

    num_prod_sol= product_solicitudes.count
    context = {
        'listo_surtir':listo_surtir,
        'solicitud': solicitud,
        'product_solicitudes': product_solicitudes,
        'num_prod_sol': num_prod_sol,
    }

    try:
        requi = Requis.objects.get(orden = solicitud, complete = True )
    except Requis.DoesNotExist:
        requi = False

    salidas = ValeSalidas.objects.filter(solicitud = solicitud)
    if salidas and not requi:
        exist_salida = True
        context.update({
            'exist_salida':exist_salida,
            'salidas':salidas,
        })

    if requi:
        exist_req = True
        prod_req = ArticulosRequisitados.objects.filter(req__id = requi.id)
        num_prod_req = prod_req.count()
        compras = Compra.objects.filter(req = requi, complete = True)

        context.update({
            'requi': requi,
            'exist_req': exist_req,
            'num_prod_req': num_prod_req,
            'prod_req':prod_req,
            'compras':compras,
        })

        if compras:
            pagos = Pago.objects.filter(oc__req = requi)
            exist_oc = True
            context.update({
                'exist_oc': exist_oc,
                'pagos':pagos,
            })

            if pagos:
                exist_pago = True
                entradas = Entrada.objects.filter(oc__req = requi, completo = True)
                exist_entradas = bool(entradas)
                salidas = ValeSalidas.objects.filter(solicitud = solicitud)
                exist_salidas = bool(salidas)

                context.update({
                    'exist_pago': exist_pago,
                    'exist_entradas': exist_entradas,
                    'entradas': entradas,
                    'salidas':salidas,
                    'exist_salidas': exist_salidas,
                })

                if entradas:
                    articulos_entradas = EntradaArticulo.objects.filter(entrada__oc__req = requi)
                    context.update({
                        'articulos_entradas': articulos_entradas,
                    })

    return render(request,'solicitud/detalle.html', context)


# AJAX
def load_subproyectos(request):

    proyecto_id = request.GET.get('proyecto_id')
    subproyectos = Subproyecto.objects.filter(proyecto_id = proyecto_id)

    return render(request, 'solicitud/subproyecto_dropdown_list_options.html',{'subproyectos': subproyectos})

def convert_excel_inventario(existencia, valor_inventario, dict_entradas, dict_resultados):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Inventario_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Inventario')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Código','Producto','Distrito','Unidad','Cantidad','Cantidad Apartada','Cantidad Entradas','Minimos','Ubicación','Estante','Precio','Total']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 0:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10
        if col_num== 1:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+3

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordtec. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Inventario Costo Total:')).style = messages_style
    (ws.cell(column = columna_max +1, row=3, value = valor_inventario)).style = money_resumen_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    #rows = existencia.values_list('producto__codigo','producto__nombre','distrito__nombre','producto__unidad__nombre','cantidad','cantidad_apartada','ubicacion','estante','price')
    rows = existencia.all()

    for inventario in rows:
        row_num += 1

        inventario.total_entradas = dict_entradas.get(inventario.id, 0)
        inventario.total_apartado = dict_resultados.get(inventario.id, 0)       
        # Aquí estás creando una lista manualmente con los valores que necesitas
        row = [
            inventario.producto.codigo,
            inventario.producto.nombre,
            inventario.distrito.nombre,
            inventario.producto.unidad.nombre,
            inventario.cantidad,
            #inventario.apartada,  # Aquí utilizas la propiedad apartada
            inventario.total_apartado,
            inventario.total_entradas,
            inventario.minimo,
            inventario.ubicacion,
            inventario.estante,
            inventario.price
        ]

        for col_num in range(len(row)):
            cell = ws.cell(row=row_num, column=col_num +1, value=row[col_num])
            if col_num > 2 and col_num != 8:
                cell.style = body_style #(ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            elif col_num == 8:
                cell.style = money_style #(ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
            else:
                cell.style = body_style#(ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style

        total_value = inventario.cantidad * inventario.price + inventario.apartada * inventario.price
        total_cell = ws.cell(row=row_num, column=len(row)+1, value=total_value)
        total_cell.style = money_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_solicitud_matriz_productos(productos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Solicitudes_por_producto_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Folio','Solicitante','Proyecto','Subproyecto','Operación','Cantidad','Código', 'Producto','Creado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 4 or col_num == 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25



    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia X. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20

    rows = productos.values_list('orden__id',Concat('orden__staff__staff__first_name',Value(' '),'orden__staff__staff__last_name'),'orden__proyecto__nombre','orden__subproyecto__nombre',
                                'orden__operacion__nombre','cantidad','producto__producto__codigo','producto__producto__nombre','orden__created_at')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_solicitud_matriz(ordenes):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Solicitudes_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Folio','Solicitante','Proyecto','Subproyecto','Operación','Creado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia X. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20

    rows = ordenes.values_list('folio',Concat('staff__staff__first_name',Value(' '),'staff__staff__last_name'),'proyecto__nombre','subproyecto__nombre',
                                'area__nombre','created_at')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)