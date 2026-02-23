from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.http import JsonResponse, HttpResponse, FileResponse
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import F, Avg, Value, ExpressionWrapper, fields, Sum, Q, Case, When, DecimalField, Max, Prefetch, OuterRef, Subquery, Count, FloatField
from django.db.models.functions import Concat, Coalesce
from django.conf import settings
from django.core.mail import EmailMessage, BadHeaderError
from django.core.files.base import ContentFile
from dashboard.models import Inventario, Order, ArticulosOrdenados, ArticulosparaSurtir, Producto_Calidad
from requisiciones.models import Requis, ArticulosRequisitados
from user.models import Profile
from tesoreria.models import Pago
from requisiciones.views import get_image_base64
from .filters import CompraFilter, ArticulosRequisitadosFilter,  ArticuloCompradoFilter, HistoricalArticuloCompradoFilter, ComparativoFilter, Item_ComparativoFilter
from .models import ArticuloComprado, Compra, Proveedor, Proveedor_direcciones, Comparativo, Item_Comparativo, Preevaluacion, Estatus_proveedor, Evidencia
from tesoreria.models import Facturas
from .forms import CompraForm, ArticuloCompradoForm, ComparativoForm, Item_ComparativoForm, Compra_ComentarioForm, PreevaluacionForm, Compra_Comment_Form, UploadFileForm
from requisiciones.forms import Articulo_Cancelado_Form
from requisiciones.filters import RequisFilter
from tesoreria.forms import Facturas_Form
from entradas.models import Entrada, No_Conformidad, NC_Articulo
import json
from datetime import date, datetime, timedelta
from num2words import num2words
import socket

import decimal

#PDF generator
import io
import os
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter
from reportlab.rl_config import defaultPageSize

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame
from bs4 import BeautifulSoup
import urllib.request, urllib.parse, urllib.error
from io import BytesIO

from smtplib import SMTPException
# Import Excel Stuff
from openpyxl import Workbook #,save_virtual_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
#from openpyxl.writer.excel import save_virtual_workbook
#from openpyxl import
import datetime as dt
#from urllib.parse import (
#    ParseResult,
#    SplitResult,
#    _coerce_args,
#    _splitnetloc,
#    _splitparams,
#    scheme_chars,
#)
#from urllib.parse import urlencode as original_urlencode
#from urllib.parse import uses_params
import ssl
# Create your views here.

@login_required(login_url='user-login')
def requisiciones_autorizadas(request):
    perfil = Profile.objects.get(staff__id=request.user.id)
    if perfil.tipo.compras == True:
        requis = Requis.objects.filter(autorizar=True, colocada=False).order_by('-approved_at')
    else:
        requis = Requis.objects.filter(complete=None)
    #requis = Requis.objects.filter(autorizar=True, colocada=False)
    myfilter = RequisFilter(request.GET, queryset=requis)
    requis = myfilter.qs



    tag = dof()
     #Set up pagination
    p = Paginator(requis, 50)
    page = request.GET.get('page')
    requis = p.get_page(page)

    context= {
        'myfilter':myfilter,
        'requis':requis,
        'tags':tag,
        }

    return render(request, 'compras/requisiciones_autorizadas.html',context)

@login_required(login_url='user-login')
def productos_pendientes(request):
    perfil = Profile.objects.get(staff__id=request.user.id)
    
    if perfil.tipo.compras == True:
        requis = Requis.objects.filter(autorizar=True, colocada=False)
    else:
        requis = Requis.objects.filter(complete=None)

    articulos = ArticulosRequisitados.objects.filter(req__autorizar = True, req__colocada=False, cantidad_comprada__lt = F("cantidad"), cancelado = False)
    myfilter = ArticulosRequisitadosFilter(request.GET, queryset=articulos)
    articulos = myfilter.qs

    if request.method == 'POST' and 'btnReporte' in request.POST:
        return convert_excel_productos_requisitados(articulos)
      
   
    #else:
        #messages.error(request,'Nada')


    context= {
        'requis':requis,
        'articulos':articulos,
        'myfilter':myfilter,
        #'productos_calidad': productos_calidad,
        }

    return render(request, 'compras/productos_pendientes.html',context)

def eliminar_articulos(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    productos = ArticulosRequisitados.objects.filter(req = pk, cantidad_comprada__lt = F("cantidad"), cancelado=False)
    requis = Requis.objects.get(id = pk)

    form = Articulo_Cancelado_Form()

    if request.method == 'POST' and "btn_eliminar" in request.POST:
        pk = request.POST.get('id')
        producto = ArticulosRequisitados.objects.get(id=pk)
        form = Articulo_Cancelado_Form(request.POST,instance=producto)
        if form.is_valid():
            articulo = form.save()
            productos = ArticulosRequisitados.objects.filter(req = producto.req)
            productos_cancelados = productos.filter(cancelado = True).count()
            productos_requisitados = productos.count() - productos_cancelados
            productos_comprados = productos.filter(cantidad_comprada__gte = F("cantidad")).count()
            if productos_requisitados == productos_comprados:
                requis.colocada = True
                requis.save()
            email = EmailMessage(
                f'Producto Eliminado {producto.producto.articulos.producto.producto.nombre}',
                f'Estimado(a) {producto.req.orden.staff.staff.first_name}:\n\nEstás recibiendo este correo porque el producto: {producto.producto.articulos.producto.producto.nombre} de la solicitud: {producto.req.orden.folio} ha sido eliminado, por la siguiente razón: {producto.comentario_cancelacion} \n\n Atte.{perfil.staff.first_name}{perfil.staff.last_name}  \nVORDTEC DE MÉXICO S.A. de C.V.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                'savia@vordtec.com',
                ['ulises_huesc@hotmail.com',producto.req.orden.staff.staff.email,],
                )
            email.send()
            messages.success(request,f' Has eliminado el {producto.producto.articulos.producto} correctamente')
            return redirect('requisicion-autorizada')




    context = {
        'form':form,
        'productos': productos,
        'requis': requis,
        }

    return render(request,'compras/eliminar_articulos.html', context)

def articulos_restantes(request, pk):
    productos = ArticulosRequisitados.objects.filter(req = pk, cantidad_comprada__lt = F("cantidad"), cancelado=False)
    #productos = ArticulosRequisitados.objects.filter(req = pk, cantidad_comprada__lt = F("cantidad"))
    requis = Requis.objects.get(id = pk)
    next_url = request.GET.get('next','requisiciones-status')

    context = {
        'productos': productos,
        'requis': requis,
        'next_url': next_url,
        }

    return render(request,'compras/articulos_restantes.html', context)

def dof():
#Trying to fetch DOF
    try:
        # Configurar el tiempo máximo de espera (en segundos)
        timeout = 2  # Ajusta el tiempo de espera según tus necesidades
        socket.setdefaulttimeout(timeout)
    
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        url = 'https://www.dof.gob.mx/#gsc.tab=0'
        html = urllib.request.urlopen(url, context=ctx).read()
        soup = BeautifulSoup(html,'html.parser')
        #tags = soup.find_all('p')

        tags = []
        for tag in soup.find_all('p'):
        #for anchor in tag.find_all('span'):
            tags.append(tag.contents)

        #substr = 'DOLAR'
        #if any(substr in str for str in tags):
        #   tag = tags[str][1]


        tag = tags[4][3]

        return tag
    except socket.timeout:
        return "Error: El tiempo de espera para la consulta ha sido superado."
    except Exception as e:
        # Manejo de la excepción - log, mensaje de error, etc.
        return f"Error al obtener datos: {e}"

def oc(request, pk):
    productos = ArticulosRequisitados.objects.filter(req = pk)
    req = Requis.objects.get(id = pk)
    usuario = Profile.objects.get(id=request.user.id)
    oc, created = Compra.objects.get_or_create(complete = False, req = req, creada_por = usuario)
    form_product = ArticuloCompradoForm()
    form = CompraForm(instance=oc)



    context= {
        'req':req,
        'form':form,
        'oc':oc,
        'productos':productos,
        'form_product':form_product,
        }

    return render(request, 'compras/oc.html',context)

def compras_devueltas(request):
    #productos = ArticulosRequisitados.objects.filter(req = pk)
    #req = Requis.objects.get(id = pk)
    usuario = Profile.objects.get(staff__id=request.user.id)
    compras = Compra.objects.filter(regresar_oc = True)
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs

    #form_product = ArticuloCompradoForm()
    #form = CompraForm(instance=oc)



    context= {
        'myfilter':myfilter,
        'compras_list':compras,
        }

    return render(request, 'compras/compras_devueltas.html',context)

def compra_edicion(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    oc = Compra.objects.get(id =pk)
    colaborador_sel = Profile.objects.all()
    productos_comp = ArticuloComprado.objects.filter(oc = oc)
    productos = ArticulosRequisitados.objects.filter(req = oc.req, sel_comp = False)
    req = Requis.objects.get(id = oc.req.id)
    comparativos = Comparativo.objects.filter(completo =True)
    proveedores = Proveedor_direcciones.objects.filter(id = oc.proveedor.id)
    form_product = ArticuloCompradoForm()
    form = CompraForm(instance=oc)

    proveedor_para_select2 = [
        {
            'id': proveedor.id, 
            'text': proveedor.nombre.razon_social,
            #'distrito': proveedor.
        } for proveedor in proveedores
    ]


    productos_para_select2 = [
        {
            'id': producto.id,
            'text': str(producto), 
            'cantidad': str(producto.cantidad), 
            'cantidad_pendiente': str(producto.cantidad_comprada),
            'precioref': str(producto.producto.articulos.producto.producto.preciomax),
            'porcentaje': str(producto.producto.articulos.producto.producto.porcentaje)
        } for producto in productos
    ]
    
        
    productos_comp_to_function = [
        {
            'id': producto.id,
            'precio': str(producto.precio_unitario),
            'precio_ref': str(producto.producto.producto.articulos.producto.producto.preciomax),
            'porcentaje': str(producto.producto.producto.articulos.producto.producto.porcentaje)
        } for producto in productos_comp
    ] 

    comparativos_para_select2 = [
        {
            'id': comparativo.id, 
            'text': str(comparativo.nombre)
        } for comparativo in comparativos
    ]


    tag = dof()
    subtotal = 0
    iva = 0
    total = 0
    dif_cant = 0
    #form.fields['deposito_comprador'].queryset = colaborador_sel
    for item in productos_comp:
        subtotal = decimal.Decimal(subtotal + item.cantidad * item.precio_unitario)
        if item.producto.producto.articulos.producto.producto.iva == True:
            iva = round(subtotal * decimal.Decimal(0.16),2)
        total = decimal.Decimal(subtotal + decimal.Decimal(iva))

    if request.method == 'POST' and  "crear" in request.POST:
        form = CompraForm(request.POST, instance=oc)
        costo_oc = 0
        costo_iva = 0
        articulos = ArticuloComprado.objects.filter(oc=oc)
        requisitados = ArticulosRequisitados.objects.filter(req = oc.req)
        cuenta_art_comprados = requisitados.filter(art_surtido = True).count()
        cuenta_art_totales = requisitados.count()
        if cuenta_art_totales == cuenta_art_comprados and cuenta_art_comprados > 0:
            req.colocada = True
        else:
            req.colocada = False
        for articulo in articulos:
            costo_oc = costo_oc + articulo.precio_unitario * articulo.cantidad
            if articulo.producto.producto.articulos.producto.producto.iva == True:
                costo_iva = decimal.Decimal(costo_oc * decimal.Decimal(0.16))
        for producto in requisitados:
            dif_cant = dif_cant + producto.cantidad - producto.cantidad_comprada
            if producto.art_surtido == False:
                producto.sel_comp = False
                producto.save()
        oc.complete = True
        if oc.tipo_de_cambio != None and oc.tipo_de_cambio > 0:
            oc.costo_iva = decimal.Decimal(costo_iva)
            oc.costo_oc = decimal.Decimal(costo_oc + costo_iva)
        else:
            oc.costo_iva = decimal.Decimal(costo_iva)
            oc.costo_oc = decimal.Decimal(costo_oc + costo_iva)
        if form.is_valid():
            abrev= usuario.distrito.abreviado
            #oc.folio = str(abrev) + str(consecutivo).zfill(4)
            oc.regresar_oc = False
            form.save()
            oc.save()
            req.save()
            messages.success(request,f'{usuario.staff.first_name}, Has modificado la OC {oc.get_folio} correctamente')
            return redirect('compras-devueltas')
    else:
        for field, errors in form.errors.items():
            error_messages[field] = errors.as_text()


    context= {
        'comparativos_para_select2': comparativos_para_select2,
        'proveedor_para_select2': proveedor_para_select2,
        'productos_comp_to_function': productos_comp_to_function,
        'productos_para_select2':productos_para_select2,
        #'proveedores':proveedores,
        'productos':productos,
        'form':form,
        'oc':oc,
        'productos_comp':productos_comp,
        'form_product':form_product,
        'subtotal':subtotal,
        'iva':iva,
        'total':total,
        }

    return render(request, 'compras/compra_edicion.html',context)



def update_oc(request):
    data= json.loads(request.body)
    action = data["action"]
    cantidad = data["val_cantidad"]
    producto_id = data["id"]
    productos = ArticulosRequisitados.objects.get(id=producto_id)
    pk = data["oc"]
    precio = data["val_precio"]
    oc = Compra.objects.get(id=pk)
    if action == "add":
        cantidad_total = productos.cantidad_comprada + decimal.Decimal(cantidad)
        if cantidad_total > productos.cantidad:
            messages.error(request,f'La cantidad que se quiere comprar sobrepasa la cantidad requisitada {cantidad_total} mayor que {productos.cantidad}')
        else:
            comp_item, created = ArticuloComprado.objects.get_or_create(oc=oc, producto=productos)
            productos.cantidad_comprada = productos.cantidad_comprada + decimal.Decimal(cantidad)
            messages.success(request,f'Estos son los productos comprados ahora {productos.cantidad_comprada}')
            if productos.cantidad_comprada == productos.cantidad:
                productos.art_surtido = True
            if comp_item.cantidad == None:
                comp_item.cantidad = 0
            comp_item.cantidad = comp_item.cantidad + decimal.Decimal(cantidad)
            comp_item.precio_unitario = precio
            productos.sel_comp = True
            comp_item.save()
            productos.save()
    if action == "remove":
        comp_item = ArticuloComprado.objects.get(oc = oc, producto = productos)
        productos.art_surtido = False
        productos.sel_comp = False
        productos.cantidad_comprada = productos.cantidad_comprada - comp_item.cantidad
        productos.save()
        comp_item.delete()

    return JsonResponse('Item updated, action executed: '+ action, safe=False)

def oc_modal(request, pk):
    #productos = ArticulosRequisitados.objects.filter(req = pk, sel_comp = False)
    productos = ArticulosRequisitados.objects.filter(req = pk, cantidad_comprada__lt = F("cantidad"), cancelado=False)
    req = Requis.objects.get(id = pk)
    proveedores = Proveedor_direcciones.objects.filter(
        Q(estatus__nombre='NUEVO') | Q(estatus__nombre='APROBADO')| Q(estatus__nombre='PREAPROBADO'))
    usuario = Profile.objects.get(staff__id=request.user.id)
    colaborador_sel = Profile.objects.all()
    compras = Compra.objects.all()
    comparativos = Comparativo.objects.filter(completo =True)
    oc, created = Compra.objects.get_or_create(complete = False, req = req, creada_por = usuario, regresar_oc = False)
    #consecutivo = compras.count() + 1
    productos_comp = ArticuloComprado.objects.filter(oc=oc)
    form_product = ArticuloCompradoForm()
    form = CompraForm(instance=oc)
    tag = dof()
    subtotal = 0
    iva = 0
    total = 0
    dif_cant = 0
    #form.fields['deposito_comprador'].queryset = colaborador_sel
    error_messages = {}
    productos_para_select2 = [
        {
            'id': producto.id,
            'text': str(producto.producto.articulos.producto), 
            'cantidad': str(producto.cantidad), 
            'cantidad_pendiente': str(producto.cantidad_comprada),
            'preciomax': str(producto.producto.articulos.producto.producto.preciomax),
            'porcentaje': str(producto.producto.articulos.producto.producto.porcentaje)
        } for producto in productos
    ]
        
    productos_comp_to_function = [
        {
            'id': producto.id,
            'precio': str(producto.precio_unitario),
            'precio_max': str(producto.producto.producto.articulos.producto.producto.preciomax),
            'porcentaje': str(producto.producto.producto.articulos.producto.producto.porcentaje)
        } for producto in productos_comp
    ] 

    comparativos_para_select2 = [
        {
            'id': comparativo.id, 
            'text': str(comparativo.nombre)
        } for comparativo in comparativos
    ]





    for item in productos_comp:
        subtotal = decimal.Decimal(subtotal + item.cantidad * item.precio_unitario)
        if item.producto.producto.articulos.producto.producto.iva == True:
            iva = round(subtotal * decimal.Decimal(0.16),2)
        total = decimal.Decimal(subtotal + decimal.Decimal(iva))

    if request.method == 'POST' and  "crear" in request.POST:
        form = CompraForm(request.POST, instance=oc)
        
        if form.is_valid():
            costo_oc = 0
            costo_iva = 0
            articulos = ArticuloComprado.objects.filter(oc=oc)
            requisitados = ArticulosRequisitados.objects.filter(req = oc.req)
            cuenta_art_comprados = requisitados.filter(art_surtido = True).count()
            cuenta_art_totales = requisitados.count()
            if cuenta_art_totales == cuenta_art_comprados and cuenta_art_comprados > 0: #Compara los artículos comprados vs artículos requisitados
                req.colocada = True
            else:
                req.colocada = False
            for articulo in articulos:
                costo_oc = costo_oc + articulo.precio_unitario * articulo.cantidad
                if articulo.producto.producto.articulos.producto.producto.iva == True:
                    costo_iva = decimal.Decimal(costo_oc * decimal.Decimal(0.16))
            for producto in requisitados:
                dif_cant = dif_cant + producto.cantidad - producto.cantidad_comprada
                if producto.art_surtido == False:
                    producto.sel_comp = False
                    producto.save()
            oc.complete = True
            if oc.tipo_de_cambio != None and oc.tipo_de_cambio > 0:
                oc.costo_iva = decimal.Decimal(costo_iva)
                oc.costo_oc = decimal.Decimal(costo_oc + costo_iva)
            else:
                oc.costo_iva = decimal.Decimal(costo_iva)
                oc.costo_oc = decimal.Decimal(costo_oc + costo_iva)
            abrev= usuario.distrito.abreviado
            #oc.folio = str(abrev) + str(consecutivo).zfill(4)
            form.save()
            oc.save()
            req.save()
            static_path = settings.STATIC_ROOT
            img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
            img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
            image_base64 = get_image_base64(img_path)
            logo_v_base64 = get_image_base64(img_path2)
            # Crear el mensaje HTML
            html_message = f"""
            <html>
                <head>
                    <meta charset="UTF-8">
                </head>
                <body>
                    <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                    <p>Estimado {oc.req.orden.staff.staff.first_name} {oc.req.orden.staff.staff.last_name},</p>
                    <p>Estás recibiendo este correo porque tu solicitud: {oc.req.orden.folio}| Req: {oc.req.folio} se ha convertido en la OC: {oc.get_folio},</p>
                    <p>creada por {oc.creada_por.staff.first_name} {oc.creada_por.staff.last_name}.</p>
                    <p>El siguiente paso del sistema: Autorización de OC por Superintedencia Administrativa</p>
                    <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                    <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                </body>
            </html>
            """
            try:
                email = EmailMessage(
                    f'OC Elaborada {oc.get_folio}',
                    body=html_message,
                    #f'Estimado {requi.orden.staff.staff.staff.first_name} {requi.orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requi.orden.folio}| Req: {requi.folio} ha sido autorizada,\n por {requi.requi_autorizada_por.staff.staff.first_name} {requi.requi_autorizada_por.staff.staff.last_name}.\n El siguiente paso del sistema: Generación de OC \n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                    from_email = settings.DEFAULT_FROM_EMAIL,
                    to= ['ulises_huesc@hotmail.com',oc.req.orden.staff.staff.email],
                    headers={'Content-Type': 'text/html'}
                    )
                email.content_subtype = "html " # Importante para que se interprete como HTML
                email.send()
                messages.success(request,f'{usuario.staff.first_name}, Has generado la OC {oc.get_folio} correctamente')
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'{usuario.staff.first_name}, Has generado la OC {oc.get_folio} correctamente pero el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.warning(request, error_message)
            return redirect('requisicion-autorizada')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()

    context= {
        'comparativos_para_select2': comparativos_para_select2,
        'productos_comp_to_function': productos_comp_to_function,
        'error_messages': error_messages,
        'proveedores':proveedores,
        'productos_para_select2':productos_para_select2,
        'req':req,
        'form':form,
        'oc':oc,
        'productos':productos,
        'form_product':form_product,
        'tag':tag,
        'productos_comp':productos_comp,
        'subtotal':subtotal,
        'iva':iva,
        'total':total,
        #'colaborador_sel':colaborador_sel,
        }
    return render(request, 'compras/oc.html',context)

@login_required(login_url='user-login')
def mostrar_comparativo(request, pk):
    comparativo = Comparativo.objects.get(id=pk)
    productos = Item_Comparativo.objects.filter(comparativo = comparativo)
    
    context= {
        'comparativo':comparativo,
        'productos':productos,
        }

    return render(request, 'compras/mostrar_comparativo.html',context)

@login_required(login_url='user-login')
def preevaluaciones(request, pk):
    #Carga el template para ver todas las preevaluaciones de un proveedor
    proveedor = Proveedor.objects.get(id=pk)
    preevaluaciones= Preevaluacion.objects.filter(nombre = proveedor, completo = True)

    context = {
        'proveedor':proveedor,
        'preevaluaciones':preevaluaciones,
    }

    return render(request, 'compras/preevaluaciones.html',context)

@login_required(login_url='user-login')
def preevaluacion(request, pk):
    #Crea preevaluación para un proveedor
    usuario = Profile.objects.get(staff__id=request.user.id)
    proveedor = Proveedor.objects.get(id=pk)
    preevaluacion, created = Preevaluacion.objects.get_or_create(nombre = proveedor, completo= False)
    form = PreevaluacionForm(instance = preevaluacion)
    error_messages = {}

    if request.method == 'POST':
        form = PreevaluacionForm(request.POST, instance = preevaluacion)
        #print(request.POST)
        if form.is_valid():
            print(form.cleaned_data)
            preevaluacion = form.save(commit=False)
            preevaluacion.completo = True
            preevaluacion.creado_por = usuario
            preevaluacion.modified_at = datetime.now()
            preevaluacion.save()
            messages.success(request,f'Has creado la preevaluación con éxito')
            return redirect('dashboard-proveedores')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()

    
    context= {
        'proveedor': proveedor,
        'error_messages':error_messages,
        'form': form,
        }

    return render(request, 'compras/preevaluacion.html',context)

def autorizacion_preevaluacion(request):
    preevaluaciones = Preevaluacion.objects.filter(completo = True, resultado = None)

    context = {
        'preevaluaciones':preevaluaciones,
    }

    return render(request, 'compras/matriz_autorizacion_preevaluacion.html', context)

def autorizar_preevaluacion(request, pk):
    preevaluacion = Preevaluacion.objects.get(id = pk)
    #proveedor = Proveedor.objects.get(id=preevaluacion.proveedor.id)
    direcciones_prov = Proveedor_direcciones.objects.filter(nombre = preevaluacion.nombre.id, completo=True)

    if request.method == 'POST' and 'btn_autorizar' in request.POST:
        preevaluacion.resultado = True
        for prov in direcciones_prov:
            if not prov.estatus or prov.estatus.id == 2:
                status = Estatus_proveedor.objects.get(id=5)
                prov.estatus = status
                prov.save()
        preevaluacion.save()
        messages.success(request,f'La preevaluacion {preevaluacion.id} ha sido autorizada')
        return redirect('autorizacion-preevaluacion')
    #else:
    #    messages.success(request,'Nada')


    context = {
        'preevaluacion':preevaluacion,
    }

    return render(request, 'compras/autorizar_preevaluacion.html', context)

def cancelar_preevaluacion(request, pk):
    preevaluacion = Preevaluacion.objects.get(id = pk)

    if request.method == 'POST' and 'btn_autorizar' in request.POST:
        preevaluacion.resultado = False
        preevaluacion.save()
        messages.success(request,f'La preevaluacion {preevaluacion.id} ha sido cancelada')
        return redirect('autorizacion-preevaluacion')
    #else:
    #    messages.success(request,'Nada')


    context = {
        'preevaluacion':preevaluacion,
    }

    return render(request, 'compras/cancelar_preevaluacion.html', context)



@login_required(login_url='user-login')
def matriz_oc(request):
    compras = Compra.objects.filter(complete=True).order_by('-id')
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    # Calcular el total de órdenes de compra
    total_de_oc = compras.count()
     # Calcular el número de OC que cumplen el criterio (created_at - approved_at <= 3)
    time_difference = ExpressionWrapper(F('created_at') - F('req__approved_at'), output_field=fields.DurationField())
    compras_con_criterio = compras.annotate(time_difference=time_difference).filter(time_difference__lte=timedelta(days=3))
    oc_cumplen = compras_con_criterio.count()

     # Calcular el indicador de cumplimiento (oc_cumplen / total_de_oc)
    if total_de_oc > 0:
        cumplimiento = (oc_cumplen / total_de_oc)*100
    else:
        cumplimiento = 0

     #Set up pagination
    p = Paginator(compras, 50)
    page = request.GET.get('page')
    compras_list = p.get_page(page)

    if request.method == 'POST' and 'btnReporte' in request.POST:
        return convert_excel_matriz_compras(compras)

    

    context= {
        'compras_list':compras_list,
        'compras':compras,
        'myfilter':myfilter,
        'cumplimiento': cumplimiento,
        }

    return render(request, 'compras/matriz_compras.html',context)

@login_required(login_url='user-login')
def matriz_oc_productos(request):
    compras = Compra.objects.filter(complete=True)
    articulos = ArticuloComprado.objects.filter(oc__complete = True).order_by('-oc__created_at')
   
    myfilter = ArticuloCompradoFilter(request.GET, queryset=articulos)
    articulos = myfilter.qs

    productos_optimized = articulos.select_related(
        'oc__req__orden__staff__staff',
        'oc__req__orden',
        'oc__req__orden__proyecto',
        'oc__req__orden__subproyecto',
        'oc__req__orden__area',
        'oc__proveedor__nombre',
        'producto__producto__articulos__producto__producto'
    ).only(
        'oc__folio',
        'oc__req__folio',
        'oc__req__orden__folio',
        'oc__req__orden__staff__staff__first_name',
        'oc__req__orden__staff__staff__last_name',
        'oc__req__orden__proyecto__nombre',
        'oc__req__orden__subproyecto__nombre',
        'created_at',
        'oc__proveedor__nombre__razon_social',
        'oc__req__orden__area__nombre',
        'cantidad',
        'producto__producto__articulos__producto__producto__codigo',
        'producto__producto__articulos__producto__producto__nombre',
        'precio_unitario',
        #'subtotal_parcial',
        #'iva_parcial',
        #'total'
    )

    #Set up pagination
    p = Paginator(articulos, 50)
    page = request.GET.get('page')
    articulos_list = p.get_page(page)

    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_excel_solicitud_matriz_productos(productos_optimized)

    context= {
        'articulos_list':articulos_list,
        'articulos':articulos,
        'compras':compras,
        'myfilter':myfilter,
        }

    return render(request, 'compras/matriz_oc_productos.html',context)

@login_required(login_url='user-login')
def productos_oc(request, pk):
    compra = Compra.objects.get(id=pk)
    productos = ArticuloComprado.objects.filter(oc=compra)


    context = {
        'compra':compra,
        'productos':productos,
    }

    return render(request,'compras/oc_producto.html',context)

@login_required(login_url='user-login')
def upload_facturas(request, pk):
    pago = Pago.objects.get(id = pk)
    facturas = Facturas.objects.filter(pago = pago, hecho=True)
    factura, created = Facturas.objects.get_or_create(pago=pago, hecho=False)
    form = Facturas_Form()

    if request.method == 'POST':
        form = Facturas_Form(request.POST or None, request.FILES or None, instance = factura)
        factura = form.save(commit=False)
        factura.fecha_subido = date.today()
        factura.hora_subido = datetime.now().time()
        factura.hecho = True
        if form.is_valid():
            form.save()
            factura.save()
            messages.success(request,'Las facturas se subieron de manera exitosa')
            return redirect('matriz-compras')
        else:
            form = Facturas_Form()
            messages.error(request,'No se pudo subir tu documento')

    context={
        'facturas':facturas,
        'form':form,
        }

    return render(request, 'compras/upload.html', context)

@login_required(login_url='user-login')
def upload_xml(request, pk):
    compra = Compra.objects.get(id = pk)
    form = CompraFactForm()

    if request.method == 'POST':
        form = CompraFactForm(request.POST or None, request.FILES or None, instance = compra)
        if form.is_valid():
            form.save()
            return redirect('matriz-compras')
        else:
            form = CompraFactForm()
            messages.error(request,'No se pudo subir tu documento')

    context={
        'compra':compra,
        'form': form,
        }

    return render(request, 'compras/upload_xml.html', context)

@login_required(login_url='user-login')
def autorizacion_oc1(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    if usuario.tipo.oc_superintendencia == True:
        compras = Compra.objects.filter(complete=True, autorizado1= None).order_by('-created_at')
    else:
        compras = Compra.objects.filter(flete=True,costo_fletes='1').order_by('-created_at')
    #compras = Compra.objects.filter(complete=True, autorizado1= None).order_by('-folio')
     #Set up pagination
    p = Paginator(compras, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)


    context= {
        'compras':compras,
        'salidas_list':salidas_list,
        }

    return render(request, 'compras/autorizacion_oc1.html',context)

def cancelar_oc1(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc = pk)

    if compra.costo_fletes == None:
        costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes != None:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio

    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
        else:
            costo_fletes = 0
    costo_total = costo_fletes + costo_oc
    resta = compra.req.orden.subproyecto.presupuesto - costo_total - compra.req.orden.subproyecto.gastado
    porcentaje = "{0:.2f}%".format((costo_oc/compra.req.orden.subproyecto.presupuesto)*100)

    if request.method == 'POST':
        compra.oc_autorizada_por = usuario
        compra.autorizado1 = False
        compra.autorizado_date1 = date.today()
        compra.autorizado_hora1 = datetime.now().time()
        compra.save()
        messages.error(request,f'Has cancelado la compra con FOLIO: {compra.get_folio}')
        return redirect('autorizacion-oc1')

    context = {
        'compra':compra,
        'productos': productos,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
     }
    return render(request,'compras/cancelar_oc1.html', context)

def cancelar_oc2(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc = pk)
    form = Compra_Comment_Form(instance = compra)
    
    
    if compra.costo_fletes == None:
        costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
        else:
            costo_fletes = 0
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
        else:
            costo_fletes = 0
    costo_total = costo_fletes + costo_oc
    resta = compra.req.orden.subproyecto.presupuesto - costo_total - compra.req.orden.subproyecto.gastado
    porcentaje = "{0:.2f}%".format((costo_oc/compra.req.orden.subproyecto.presupuesto)*100)


    if request.method == 'POST':
        form = Compra_Comment_Form(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save(commit = False)
            compra.oc_autorizada_por2 = usuario
            compra.autorizado2 = False
            compra.autorizado_date2 = date.today()
            compra.autorizado_hora2 = datetime.now().time()
            compra.save()
            messages.success(request,f'Has cancelado la compra con FOLIO: {compra.get_folio}')
            return HttpResponse(status=204)

    context = {
        'form':form,
        'compra':compra,
        'productos': productos,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
     }
    return render(request,'compras/cancelar_oc2.html', context)

def back_oc(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc = pk)
    #Traigo la requisición para poderla activar de nuevo
    requi = Requis.objects.get(id=compra.req.id)
    costo_fletes = 0
    if compra.costo_fletes == None:
        costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
    costo_total = costo_fletes + costo_oc
    resta = compra.req.orden.subproyecto.presupuesto - costo_total - compra.req.orden.subproyecto.gastado
    porcentaje = "{0:.2f}%".format((costo_oc/compra.req.orden.subproyecto.presupuesto)*100)

    form = Compra_ComentarioForm()

    if request.method == 'POST':
        form = Compra_ComentarioForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save(commit = False)
            if not compra.autorizado1:
                compra.oc_autorizada_por = perfil
                compra.autorizado1 = None
                compra.complete = False
                compra.autorizado_date1 = date.today()
                compra.autorizado_hora1 = datetime.now().time()
                compra.regresar_oc = True
            else:
                compra.oc_autorizada_por2 = perfil
                compra.autorizado2 = None
                compra.autorizado1 = None
                compra.complete = False
                compra.autorizado_date2 = date.today()
                compra.autorizado_hora2 = datetime.now().time()
                compra.regresar_oc = True
            #Esta línea es la que activa a la requi
            #requi.colocada = False
            compra.save()
            #requi.save()
            messages.success(request,f'Has regresado la compra con FOLIO: {compra.get_folio} y ahora podrás encontrar esos productos en el apartado devolución')
            return redirect('compras-devueltas')

    context = {
        'form':form,
        'compra':compra,
        'productos': productos,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
     }

    return render(request,'compras/back_oc.html', context)




def autorizar_oc1(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc=pk)
    # Esto asume que ya has obtenido una instancia específica de Compra con 'compra = Compra.objects.get(id=pk)'
    proyecto_id = compra.req.orden.proyecto.id
    #print('proyecto',proyecto_id)
    compras_por_sumar = Compra.objects.filter(req__orden__proyecto__id=proyecto_id, complete = True, pagada = True)
    # Ahora, sumamos los 'costo_oc' para todas las Compras que están bajo el mismo proyecto.
    total_costo_oc = 0
    total_costo_autorizado = 0
    total_costo_pagado = 0
    # Recorremos cada compra para calcular los totales.
    for comprax in compras_por_sumar:
        # Ajustamos el costo_oc si la moneda es DOLARES.
        if comprax.moneda:
            if comprax.moneda.nombre == "DOLARES":
                costo_oc = comprax.costo_oc
                tc = comprax.tipo_de_cambio or 17            
                costo_oc_ajustado = comprax.costo_oc * tc

            else:
                costo_oc_ajustado = comprax.costo_oc
        else:
            costo_oc_ajustado = comprax.costo_oc
        
        # Sumamos al total general.
        total_costo_oc += costo_oc_ajustado
        
        # Si la compra está autorizada, la sumamos al total autorizado.
        if comprax.autorizado2:
            total_costo_autorizado += costo_oc_ajustado
        
        # Si la compra está pagada, la sumamos al total pagado.
        if comprax.pagada:
            total_costo_pagado += costo_oc_ajustado

    # Ahora tenemos los totales calculados.
    #print('Total costo OC:', total_costo_oc)
    #print('Total costo OC (autorizado):', total_costo_autorizado)
    #print('Total costo OC (pagado):', total_costo_pagado)

    costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes

    #print(costo_oc)
    #print(compra.req.orden.subproyecto.gastado)
    
    costo_total = costo_fletes + costo_oc 
    suma_presupuesto = compra.req.orden.proyecto.subproyectos.aggregate(total_presupuesto=Sum('presupuesto'))['total_presupuesto']
    suma_presupuesto = suma_presupuesto if suma_presupuesto is not None else 0
    resta = suma_presupuesto - total_costo_pagado - costo_total
    if costo_oc > 0 and compra.req.orden.subproyecto.presupuesto > 0:
        porcentaje = "{0:.2f}%".format((costo_oc/compra.req.orden.subproyecto.presupuesto)*100)
    else:
        porcentaje = ""


    if request.method == 'POST':
        compra.autorizado1 = True
        compra.oc_autorizada_por = usuario
        compra.autorizado_date1 = date.today()
        compra.autorizado_hora1 = datetime.now().time()
        compra.save()
        static_path = settings.STATIC_ROOT
        img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
        img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
        image_base64 = get_image_base64(img_path)
        logo_v_base64 = get_image_base64(img_path2)
        html_message = f"""
            <html>
                <head>
                    <meta charset="UTF-8">
                </head>
                <body>
                    <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                    <p>Estimado {compra.req.orden.staff.staff.first_name} {compra.req.orden.staff.staff.last_name},</p>
                    <p>Estás recibiendo este correo porque tu OC {compra.get_folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido autorizada por {compra.oc_autorizada_por.staff.first_name} {compra.oc_autorizada_por.staff.last_name},</p>
                    <p>El siguiente paso del sistema: Autorización de OC por Gerencia de Planta</p>
                    <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                    <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                </body>
            </html>
        """
        try:
            email = EmailMessage(
                f'OC Autorizada {compra.get_folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                body=html_message,
                from_email = settings.DEFAULT_FROM_EMAIL,
                to= ['ulises_huesc@hotmail.com',compra.req.orden.staff.staff.email],
                headers={'Content-Type': 'text/html'}
                )
            email.content_subtype = "html " # Importante para que se interprete como HTML
            email.send()
            messages.success(request, f'{usuario.staff.first_name} has autorizado la solicitud {compra.get_folio}')
        except (BadHeaderError, SMTPException, socket.gaierror) as e:
            error_message = f'{usuario.staff.first_name} has autorizado la compra {compra.get_folio} pero el correo de notificación no ha sido enviado debido a un error: {e}'
            messages.success(request, error_message)  
        return redirect('autorizacion-oc1')

    context={
        'total_costo_pagado':total_costo_pagado,
        'suma_presupuesto':suma_presupuesto,
        'compra':compra,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
        }

    return render(request, 'compras/autorizar_oc1.html',context)

@login_required(login_url='user-login')
def autorizacion_oc2(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    #if usuario.tipo.oc_gerencia == True:
    #    compras = Compra.objects.filter(complete = True, autorizado1 = True, autorizado2= None).order_by('-folio')
    #else:
    #    compras = Compra.objects.filter(flete=True,costo_fletes='1')
    compras = Compra.objects.filter(complete = True, autorizado1 = True, autorizado2= None).order_by('-folio')

    p = Paginator(compras, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    context= {
        'compras':compras,
        'salidas_list':salidas_list,
        }

    return render(request, 'compras/autorizacion_oc2.html',context)


def autorizar_oc2(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc=pk)
    proyecto_id = compra.req.orden.proyecto.id
    #print('proyecto',proyecto_id)
    compras_por_sumar = Compra.objects.filter(req__orden__proyecto__id=proyecto_id, complete = True, pagada = True)
    # Ahora, sumamos los 'costo_oc' para todas las Compras que están bajo el mismo proyecto.
    total_costo_oc = 0
    total_costo_autorizado = 0
    total_costo_pagado = 0
    # Recorremos cada compra para calcular los totales.
    for comprax in compras_por_sumar:
        # Ajustamos el costo_oc si la moneda es DOLARES.
        if comprax.moneda:
            if comprax.moneda.nombre == "DOLARES":
                costo_oc = comprax.costo_oc
                tc = comprax.tipo_de_cambio or 17            
                costo_oc_ajustado = comprax.costo_oc * tc

            else:
                costo_oc_ajustado = comprax.costo_oc
        else:
            costo_oc_ajustado = comprax.costo_oc
        
        # Sumamos al total general.
        total_costo_oc += costo_oc_ajustado
        
        # Si la compra está autorizada, la sumamos al total autorizado.
        if comprax.autorizado2:
            total_costo_autorizado += costo_oc_ajustado
        
        # Si la compra está pagada, la sumamos al total pagado.
        if comprax.pagada:
            total_costo_pagado += costo_oc_ajustado

    # Ahora tenemos los totales calculados.
    #print('Total costo OC:', total_costo_oc)
    #print('Total costo OC (autorizado):', total_costo_autorizado)
    #print('Total costo OC (pagado):', total_costo_pagado)

    costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
    costo_total = costo_fletes + costo_oc
    suma_presupuesto = compra.req.orden.proyecto.subproyectos.aggregate(total_presupuesto=Sum('presupuesto'))['total_presupuesto']
    suma_presupuesto = suma_presupuesto if suma_presupuesto is not None else 0
    resta = suma_presupuesto - total_costo_pagado - costo_total
    if costo_oc > 0 and compra.req.orden.subproyecto.presupuesto > 0:
        porcentaje = "{0:.2f}%".format((costo_oc/compra.req.orden.subproyecto.presupuesto)*100)
    else:
        porcentaje = ""

    if request.method == 'POST':
        compra.autorizado2 = True
        compra.oc_autorizada_por2 = usuario
        compra.autorizado_date2 = date.today()
        compra.autorizado_hora2 = datetime.now().time()
        compra.save()
        archivo_oc = attach_oc_pdf(request, compra.id)
        static_path = settings.STATIC_ROOT
        img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
        img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
        image_base64 = get_image_base64(img_path)
        logo_v_base64 = get_image_base64(img_path2)
        if compra.cond_de_pago.nombre == "CREDITO":
            html_message2 = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body>
                        <p>Estimado(a) {compra.proveedor.contacto}| Proveedor {compra.proveedor.nombre}:,</p>
                        <p>Estás recibiendo este correo porque has sido seleccionado para surtirnos la OC adjunta con folio: {compra.folio}.<p>
                        <p>&nbsp;</p>
                        <p> Atte. {compra.creada_por.staff.first_name} {compra.creada_por.staff.last_name}</p> 
                        <p>GRUPO VORDCAB S.A. de C.V.</p>
                        <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                        <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                        <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                    </body>
                </html>
            """
            try:
                email = EmailMessage(
                f'Compra Autorizada {compra.get_folio}|SAVIA',
                body=html_message2,
                from_email =settings.DEFAULT_FROM_EMAIL,
                to= ['ulises_huesc@hotmail.com', compra.creada_por.staff.email, compra.proveedor.email, 'lizeth.ojeda@vordtec.com'],
                headers={'Content-Type': 'text/html'}
                )
                email.content_subtype = "html " # Importante para que se interprete como HTML
                email.attach(f'folio:{compra.get_folio}.pdf',archivo_oc,'application/pdf')
                email.send()
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'correo de notificación no ha sido enviado debido a un error: {e}'
            html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body>
                        <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                        <p>Estimado {compra.req.orden.staff.staff.first_name} {compra.req.orden.staff.staff.last_name},</p>
                        <p>Estás recibiendo este correo porque tu OC {compra.get_folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido autorizada por {compra.oc_autorizada_por2.staff.first_name} {compra.oc_autorizada_por2.staff.last_name},</p>
                        <p>El siguiente paso del sistema: Recepción por parte de Almacén |Compra a crédito</p>
                        <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                        <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                    </body>
                </html>
            """
            try:
                email = EmailMessage(
                    f'OC Autorizada Gerencia {compra.get_folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                    body=html_message,
                    #f'Estimado {requi.orden.staff.staff.staff.first_name} {requi.orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requi.orden.folio}| Req: {requi.folio} ha sido autorizada,\n por {requi.requi_autorizada_por.staff.staff.first_name} {requi.requi_autorizada_por.staff.staff.last_name}.\n El siguiente paso del sistema: Generación de OC \n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                    from_email = settings.DEFAULT_FROM_EMAIL,
                    to= ['ulises_huesc@hotmail.com',],#[requi.orden.staff.staff.staff.email],
                    headers={'Content-Type': 'text/html'}
                    )
                email.content_subtype = "html " # Importante para que se interprete como HTML
                email.send()
                for producto in productos:
                    if producto.producto.producto.articulos.producto.producto.especialista == True:
                        archivo_oc = attach_oc_pdf(request, compra.id)
                        email = EmailMessage(
                            f'Compra Autorizada {compra.get_folio}',
                            f'Estimado proveedor,\n Estás recibiendo este correo porque ha sido aprobada una OC que contiene el producto código:{producto.producto.producto.articulos.producto.producto.codigo} descripción:{producto.producto.producto.articulos.producto.producto.nombre} el cual requiere la liberación de calidad\n Este mensaje ha sido automáticamente generado por SAVIA X',
                            settings.DEFAULT_FROM_EMAIL,
                            ['ulises_huesc@hotmail.com',],
                            )
                        email.attach(f'folio:{compra.get_folio}.pdf',archivo_oc,'application/pdf')
                        email.send()
                messages.success(request, f'{usuario.staff.first_name} has autorizado la solicitud {compra.get_folio}')
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'{usuario.staff.first_name} has autorizado la compra {compra.folio} pero el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.warning(request, error_message)   
        else:
            html_message = f"""
                    <html>
                        <head>
                            <meta charset="UTF-8">
                        </head>
                        <body>
                            <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                            <p>Estimado {compra.req.orden.staff.staff.first_name} {compra.req.orden.staff.staff.last_name},</p>
                            <p>Estás recibiendo este correo porque tu OC {compra.get_folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido autorizada por {compra.oc_autorizada_por2.staff.first_name} {compra.oc_autorizada_por2.staff.last_name},</p>
                            <p>El siguiente paso del sistema: Pago por parte de tesorería</p>
                            <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                            <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                        </body>
                    </html>
                """
            try:
                email = EmailMessage(
                    f'OC Autorizada Gerencia {compra.get_folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                    body=html_message,
                    #f'Estimado {requi.orden.staff.staff.staff.first_name} {requi.orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requi.orden.folio}| Req: {requi.folio} ha sido autorizada,\n por {requi.requi_autorizada_por.staff.staff.first_name} {requi.requi_autorizada_por.staff.staff.last_name}.\n El siguiente paso del sistema: Generación de OC \n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                    from_email = settings.DEFAULT_FROM_EMAIL,
                    to= ['ulises_huesc@hotmail.com'],#[requi.orden.staff.staff.staff.email],
                    headers={'Content-Type': 'text/html'}
                    )
                email.content_subtype = "html " # Importante para que se interprete como HTML
                email.send()
                messages.success(request, f'{usuario.staff.first_name} has autorizado la compra {compra.get_folio}')
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'{usuario.staff.first_name} has autorizado la compra {compra.get_folio} pero el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.warning(request, error_message)   
        return redirect('autorizacion-oc2')

    context={
        'suma_presupuesto':suma_presupuesto,
        'compra':compra,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
        }

    return render(request, 'compras/autorizar_oc2.html',context)

@login_required(login_url='user-login')
def comparativos(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    

    comparativos = Comparativo.objects.filter(completo = True)
    myfilter = ComparativoFilter(request.GET, queryset=comparativos)
    comparativos = myfilter.qs

     #Set up pagination
    p = Paginator(comparativos, 50)
    page = request.GET.get('page')
    comparativos_list = p.get_page(page)
    
    context= {
        'myfilter':myfilter,
        'comparativos':comparativos,
        'comparativos_list':comparativos_list,
    }
    return render(request,'compras/comparativos.html', context)

@login_required(login_url='user-login')
def crear_comparativo(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    #Tengo que revisar primero si ya existe una orden pendiente del usuario
    
    comparativo, created = Comparativo.objects.get_or_create(completo= False, creada_por=usuario)
    productos = Item_Comparativo.objects.filter(comparativo = comparativo, completo = True)

    #proveedores = Proveedor_direcciones.objects.all()
    articulos = Inventario.objects.all()
    form_item = Item_ComparativoForm()
    form = ComparativoForm()

    if request.method =='POST':
        if "btn_creacion" in request.POST:
            form = ComparativoForm(request.POST, request.FILES or None, instance=comparativo)
            #abrev= usuario.distrito.abreviado
            if form.is_valid():
                comparativo = form.save(commit=False)
                comparativo.completo = True
                comparativo.created_at = date.today()
                #comparativo.created_at_time = datetime.now().time()
                comparativo.creado_por =  usuario
                comparativo.save()
                #form.save()
                messages.success(request, f'El comparativo {comparativo.id} ha sido creado')
                return redirect('comparativos')
        if "btn_producto" in request.POST:
            articulo, created = Item_Comparativo.objects.get_or_create(completo = False, comparativo = comparativo)
            form_item = Item_ComparativoForm(request.POST, instance=articulo)
            if form_item.is_valid():
                articulo = form_item.save(commit=False)
                articulo.completo = True
                articulo.save()
                messages.success(request, 'Se ha agregado el artículo exitosamente')
                return redirect('crear_comparativo')
        
    context= {
        'productos':productos,
        'form':form,
        'form_item':form_item,
        'articulos':articulos,
        'comparativo':comparativo,
        #'proveedores':proveedores,
    }

    return render(request, 'compras/crear_comparativo.html', context)

#Ajax Select2
def carga_proveedor(request):
    #pk_perfil = request.session.get('selected_profile_id')
    #colaborador_sel = Profile.objects.all()
    #usuario = colaborador_sel.get(id = pk_perfil)
    term = request.GET.get('term')      

    proveedores = Proveedor_direcciones.objects.filter(
        Q(estatus__nombre="NUEVO") | Q(estatus__nombre="APROBADO")| Q(estatus__nombre='PREAPROBADO'),
        nombre__razon_social__icontains = term
    ).values(
        'id','nombre__razon_social','domicilio','estado__nombre','estatus__nombre',
        'financiamiento','dias_credito'
    )
    data = list(proveedores)
    #data = []
    for prov in proveedores:
        # Aquí agregamos la lógica para extraer los items comparativos
        try:
            proveedor_id = Proveedor.objects.get(razon_social= prov['nombre__razon_social'])
        except Proveedor.DoesNotExist:
            prov['items_comparativos'] = []
            data.append(prov)
            continue

        items = []
        
        preevaluaciones = Preevaluacion.objects.filter(nombre_id=proveedor_id)
        #print('prov',prov)
    
        if not preevaluaciones.exists():
            prov['items_comparativos'] = items
            data.append(prov)
            continue
            #print('Preevaluación:', preevaluacion)

        for preevaluacion in preevaluaciones:  
            if preevaluacion.comparativo_model is not None:  
                comparativo_items = list(preevaluacion.comparativo_model.items_comparativos.values('producto__producto__id', 'producto__producto__nombre'))
                items.extend(comparativo_items)
                #print('Items del comparativo:', comparativo_items)
            else:
                pass
        
        prov['items_comparativos'] = items
        #print(comparativo_items)
        data.append(prov)
    #print(proveedores)
    return JsonResponse(data, safe=False)


def carga_proveedor_comparativo(request):
    pk_perfil = request.session.get('selected_profile_id')
    
    #print(pk_perfil)
    colaborador_sel = Profile.objects.all()
    #usuario = colaborador_sel.get(id = pk_perfil)
    term = request.GET.get('term')
    proveedores = (
        Proveedor_direcciones.objects
        .filter(nombre__razon_social__icontains=term)
        .annotate(text=F('nombre__razon_social'))
        .values('id', 'text')
        .distinct()
    )

    data = list(proveedores)
    print(data)    
    return JsonResponse(data, safe=False)

#Ajax Select2
def carga_productos(request):
    #pk_perfil = request.session.get('selected_profile_id')
    #colaborador_sel = Profile.objects.all()
    #usuario = colaborador_sel.get(id = pk_perfil)
    term = request.GET.get('term')
    articulos =(Inventario.objects.filter( Q(producto__nombre__icontains=term) |
            Q(producto__codigo__icontains=term)
        )
        .values(
            'id',
            'producto__codigo',
            'producto__nombre'
        )[:20]  # límite recomendable
    )

    
    data = [
        {
            "id": item['id'], 
            "text": f"{item['producto__codigo']} - {item['producto__nombre']}" 
        } for item in articulos
    ]
    #data = list(articulos)
        
    return JsonResponse(data, safe=False)

@login_required(login_url='user-login')
def editar_comparativo(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    comparativo =Comparativo.objects.get(id = pk)
    productos = Item_Comparativo.objects.filter(comparativo = comparativo, completo = True)
    proveedores = Proveedor_direcciones.objects.all()
    articulos = Inventario.objects.all()
    form_item = Item_ComparativoForm()
    form = ComparativoForm(instance = comparativo)

    if request.method =='POST':
        if "btn_agregar" in request.POST:
            form = ComparativoForm(request.POST, request.FILES, instance = comparativo)
            #abrev= usuario.distrito.abreviado
            if form.is_valid():
                comparativo = form.save(commit=False)
                comparativo.completo = True
                comparativo.created_at = date.today()
                #comparativo.created_at_time = datetime.now().time()
                comparativo.creado_por =  usuario
                comparativo.save()
                #form.save()
                messages.success(request, f'El comparativo {comparativo.id} ha sido modificado')
                return redirect('comparativos')
        if "btn_producto" in request.POST:
            articulo, created = Item_Comparativo.objects.get_or_create(completo = False, comparativo = comparativo)
            form_item = Item_ComparativoForm(request.POST, instance=articulo)
            if form_item.is_valid():
                articulo = form_item.save(commit=False)
                articulo.completo = True
                articulo.save()
                messages.success(request, 'Se ha agregado el artículo exitosamente')
                return redirect('editar-comparativo')
        
    context= {
        'productos':productos,
        'form':form,
        'form_item':form_item,
        'articulos':articulos,
        'comparativo':comparativo,
        'proveedores':proveedores,
    }

    return render(request, 'compras/actualizar_comparativo.html', context)

def articulos_comparativo(request, pk):
    comparativo =Comparativo.objects.get(id = pk)
    articulos = Item_Comparativo.objects.filter(comparativo__id = pk , completo = True)

    context= {
        'comparativo':comparativo,
        'articulos':articulos,
    }
    return render(request, 'compras/articulos_comparativo.html', context)

def articulo_comparativo_delete(request, pk):
   
    articulo = Item_Comparativo.objects.get(id=pk)
    comparativo = articulo.comparativo.id
   
    messages.success(request,f'El articulo ha sido eliminado exitosamente')
    articulo.delete()

    return redirect('crear_comparativo')


def evidencias_proveedor(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    evidencias = Evidencia.objects.filter(oc = compra, hecho=True)
    next_url = request.GET.get('next')
    print('next_url:', next_url)
    

    context={
        'next':next_url,
        #'form':form,
        'compra':compra,
        'evidencias':evidencias,
        'usuario':usuario,
        }

    return render(request, 'compras/evidencias_proveedor.html', context)

def subir_evidencias(request, pk):

    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    form = UploadFileForm()
    

    if request.method == 'POST':
        if 'btn_registrar' in request.POST:
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                
                files_evidencia = request.FILES.getlist('evidencia_file')
                print(request.FILES)
                if not files_evidencia:
                    messages.error(request, 'Debes subir al menos un archivo.')
                    return HttpResponse(status=204)
                for archivo_evidencia in files_evidencia:
                    evidencia = Evidencia.objects.create(
                        oc=compra,
                        file = archivo_evidencia,
                        hecho = True,
                        uploaded = datetime.now(),
                        subido_por = usuario
                    )
                    evidencia.save()
                messages.success(request, 'Las evidencias se registraron de manera exitosa')

            else:
                messages.error(request,'No se pudo subir tu documento')


    context={
        'form': form, 
        'compra': compra,
        'next':next_url,
    }

    return render(request, 'compras/subir_evidencias.html', context)

@login_required(login_url='user-login')
def historico_articulos_compras(request):
    registros = ArticuloComprado.history.all()

    myfilter = HistoricalArticuloCompradoFilter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        }

    return render(request,'compras/historico_articulos_comprados.html',context)

def descargar_pdf(request, pk):
    compra = get_object_or_404(Compra, id=pk)
    buf = generar_pdf(compra)
    return FileResponse(buf, as_attachment=True, filename='oc_' + str(compra.id) + '.pdf')

@xframe_options_sameorigin
def ver_oc_pdf(request, compra_id):
    compra = get_object_or_404(Compra, pk=compra_id)
    buf = generar_pdf(compra)  # tu función
    filename = f"OC_{compra.get_folio}.pdf"

    resp = FileResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


def preevaluacion_pdf(request, pk):
    preevaluacion = get_object_or_404(Preevaluacion, id=pk)
    if preevaluacion.tipo_evaluacion.nombre == 'Inicial':
        buf = generar_preevaluacion_inicial_pdf(preevaluacion)
    else:
        buf = generar_preevaluacion_pdf(preevaluacion)
    return FileResponse(buf, as_attachment=True, filename='Preevaluacion_' + str(preevaluacion.id) + '.pdf')

def attach_oc_pdf(request, pk):
    compra = get_object_or_404(Compra, id=pk)
    buf = generar_pdf(compra)

    # Si en algún lugar más de tu código necesitas hacer más cosas antes de retornar buf.getvalue(),
    # entonces aquí es el lugar para hacerlo. Por ahora, sólo retornaremos el valor.

    return buf.getvalue()


def get_paragraph_height(text, style, max_width):
    p = Paragraph(text, style)
    width, height = p.wrap(max_width, 0)
    return height

def dibujar_encabezado(c, caja_iso):
    c.drawString(410, caja_iso + 10, 'Preparado por:')
    c.drawString(410, caja_iso, 'Adquisiciones')
    c.drawString(500, caja_iso + 10, 'Aprobación')
    c.drawString(475, caja_iso, 'Subdirección Administrativa')
    c.drawString(20, caja_iso - 20, 'Número de documento')
    c.drawString(30, caja_iso - 30, 'F-ADQ-N4-01.02')
    c.drawString(145, caja_iso - 20, 'Clasificación del documento')
    c.drawString(175, caja_iso - 30, 'Registro')
    c.drawString(255, caja_iso - 20, 'Nivel del documento')
    c.drawString(280, caja_iso - 30, 'N5')
    c.drawString(340, caja_iso - 20, 'Revisión No.')
    c.drawString(352, caja_iso - 30, '001')
    c.drawString(410, caja_iso - 20, 'Fecha de Emisión')
    c.drawString(425, caja_iso - 30, '')
    c.drawString(500, caja_iso - 20, 'Fecha de Modificación')
    c.drawString(525, caja_iso - 30, '')

def generar_pdf(compra):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    productos = ArticuloComprado.objects.filter(oc=compra.id).order_by('id')
    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    #Elaborar caja
    caja_iso = 760
    # Dibujar encabezado
    dibujar_encabezado(c, caja_iso)
    caja_proveedor = caja_iso - 50
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,30, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.rect(20,caja_proveedor,565,10, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,570,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,760,'Orden de compra')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor,20,570) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor,585,570) #Linea 2 contorno
    c.drawInlineImage('static/images/logo vordtec_documento.png',40,755, 1.5 * cm, 0.75 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',9)
    c.drawString(120,caja_proveedor+1,'Autorización')
    c.drawString(400,caja_proveedor+1, 'Datos de Proveedor')
    inicio_central = 300
    # Definir el estilo de párrafo
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleN.fontName = 'Helvetica'
    styleN.fontSize = 8
    styleN.leading = 10
    max_width = 580 - inicio_central - 115
    c.line(inicio_central,caja_proveedor,inicio_central,570) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica',8)
    c.drawRightString(130,caja_proveedor-8,'Folio de solicitud:')
    c.drawRightString(130,caja_proveedor-18,'Folio de Requisición:')
    c.drawRightString(130,caja_proveedor-28,'Folio de orden de compra:')
    c.drawRightString(130,caja_proveedor-38,'Proyecto/Orden de Trabajo:')
    c.drawRightString(130,caja_proveedor-48,'Subproyecto:')
    c.drawRightString(130,caja_proveedor-58,'Elaboró:')
    c.drawRightString(130,caja_proveedor-68,'Autorizó:')
    c.drawRightString(130,caja_proveedor-78,'Fecha:')

    c.drawString(135,caja_proveedor-8, compra.req.orden.folio)
    c.drawString(135,caja_proveedor-18, compra.req.folio)
    c.drawString(135,caja_proveedor-28, compra.get_folio) #podría ser folio también
    c.drawString(135,caja_proveedor-38, compra.req.orden.proyecto.nombre)
    c.drawString(135,caja_proveedor-48, compra.req.orden.subproyecto.nombre)
    c.drawString(135,caja_proveedor-58, compra.req.orden.staff.staff.first_name+' '+compra.req.orden.staff.staff.last_name)
    if compra.oc_autorizada_por2:
        c.drawString(135,caja_proveedor-68, compra.oc_autorizada_por2.staff.first_name+' '+ compra.oc_autorizada_por2.staff.last_name)
    c.drawString(135,caja_proveedor-78, str(compra.autorizado_date2))

    c.setFillColor(black)
    c.setFont('Helvetica',8)
    c.drawRightString(inicio_central + 110,caja_proveedor-8,'Nombre:')
    c.drawRightString(inicio_central + 110,caja_proveedor-28,'RFC:')
    c.drawRightString(inicio_central + 110,caja_proveedor-38,'Número de Cuenta Bancaria:')
    c.drawRightString(inicio_central + 110,caja_proveedor-48,'Nombre del Banco:')
    c.drawRightString(inicio_central + 110,caja_proveedor-58,'CLABE:')
    c.drawRightString(inicio_central + 110,caja_proveedor-68,'SWIFT:')
    c.drawRightString(inicio_central + 110,caja_proveedor-78,'Estatus:')
    
    name = Paragraph(compra.proveedor.nombre.razon_social, styleN)
    width, height = name.wrap(max_width, 100)
    name.drawOn(c, inicio_central + 115, caja_proveedor - height)
    #c.drawString(inicio_central + 115,caja_proveedor-8, name)
    c.drawString(inicio_central + 115,caja_proveedor-28, compra.proveedor.nombre.rfc)
    if compra.proveedor.cuenta:
        c.drawString(inicio_central + 115,caja_proveedor-38, compra.proveedor.cuenta)
        c.drawString(inicio_central + 115,caja_proveedor-48, compra.proveedor.banco.nombre)
        c.drawString(inicio_central + 115,caja_proveedor-58, compra.proveedor.clabe)
    else:
        c.drawString(inicio_central + 115,caja_proveedor-38, 'No definido')
        c.drawString(inicio_central + 115,caja_proveedor-48, 'No definido')
        c.drawString(inicio_central + 115,caja_proveedor-58, 'No definido')
    if compra.proveedor.swift:
        c.drawString(inicio_central + 115,caja_proveedor-68, compra.proveedor.swift)
    c.drawString(inicio_central + 115,caja_proveedor-78, compra.proveedor.estatus.nombre)

    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    c.rect(20,caja_proveedor-90,565,10, fill=True, stroke=False)

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',9)
    c.drawString(90,caja_proveedor-89,'Condiciones Comerciales')
    c.drawString(370,caja_proveedor-89, 'Datos de Facturación')

    c.setFillColor(black)
    c.setFont('Helvetica',8)
    c.drawRightString(130,caja_proveedor-100,'Tiempo de entrega:')
    c.drawRightString(130,caja_proveedor-110,'Política de Garantía:')
    c.drawRightString(130,caja_proveedor-120,'Condición de pago:')
    c.drawRightString(130,caja_proveedor-130,'Vigencia de cotización:')

    
    c.drawString(135,caja_proveedor-100, str(compra.dias_de_entrega))
    #c.drawString(135,caja_proveedor-110, compra.uso_del_cfdi.descripcion)
    c.drawString(135,caja_proveedor-120, compra.cond_de_pago.nombre )
    #c.drawString(135,caja_proveedor-130, compra.uso_del_cfdi.descripcion)


    c.drawRightString(inicio_central + 110,caja_proveedor-100,'Moneda:')
    c.drawRightString(inicio_central + 110,caja_proveedor-110,'Uso del CFDI:')
    c.drawRightString(inicio_central + 110,caja_proveedor-120,'Enviar factura al correo:')
    c.drawRightString(inicio_central + 110,caja_proveedor-130,'Regimen Fiscal:')
   
    c.drawString(inicio_central + 115,caja_proveedor-100, compra.moneda.nombre)
    c.drawString(inicio_central + 115,caja_proveedor-110, compra.uso_del_cfdi.descripcion)
    c.drawString(inicio_central + 115,caja_proveedor-120, compra.creada_por.staff.email)
    c.drawString(inicio_central + 115,caja_proveedor-130, '601 - General de Ley Personas Morales')
    # Definir el estilo de párrafo
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleN.fontName = 'Helvetica'
    styleN.fontSize = 6
    styleN.leading = 8

    data =[]
    data_c = []
    high = 510
    item = 0
    #TABLAAAA 1
    data.append(['''Partida''','''Código''','''Descripción General''', '''Cantidad''', '''Unidad''', '''P.Unitario''', '''Descuento''', '''Importe'''])
    # Contador para limitar la reducción de high
    count = 0
    max_count = 12
    for producto in productos:
        item += 1
        importe = producto.precio_unitario * producto.cantidad
        importe_rounded = round(importe, 4)
        numero = Paragraph(str(item), styleN)
        codigo = Paragraph(producto.producto.producto.articulos.producto.producto.codigo, styleN)
        descripcion = Paragraph(producto.producto.producto.articulos.producto.producto.nombre, styleN)
        cantidad = Paragraph(str(producto.cantidad), styleN)
        unidad = Paragraph(str(producto.producto.producto.articulos.producto.producto.unidad), styleN)
        unitario = Paragraph(str(producto.precio_unitario), styleN)
        importe_rounded = Paragraph(str(importe_rounded), styleN)
        data.append([
            numero,
            codigo,
            descripcion,
            cantidad,
            unidad,
            unitario,
            '',
            importe_rounded
        ])
        # Reducir high solo 12 veces
        if count < max_count:
            high -= 18
            count += 1
    
    
    c.setFillColor(black)
    c.setFont('Helvetica',8)

    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,210,390,10, fill=True, stroke=False) #2ra linea azul, donde esta el proyecto y el subproyecto, se coloca altura de 150
    c.setFillColor(black)
    c.setFillColor(white)
    c.setLineWidth(.1)
    c.setFont('Helvetica-Bold',10)
    c.drawString(200,211,'Total con letra')

    c.setLineWidth(.3)
    c.line(410,220,410,160) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(410,160,580,160)

    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)

    montos_align = 480
    c.drawRightString(montos_align,210,'Sub Total:')
    c.drawRightString(montos_align,200,'IVA 16%:')
    c.drawRightString(montos_align,190,'Importe Neto:')
    #c.drawRightString(montos_align,180,'Costo fletes:')
    c.setFillColor(prussian_blue)
    c.setFillColor(black)
    c.drawString(20,130,'Opciones y condiciones:')
    c.setFont('Helvetica',8)

    c.drawCentredString(175,70,'Autorización')
    c.drawCentredString(390,70,'Autorización')

    c.drawCentredString(175,80,'Superintendente Administrativo')
    c.drawCentredString(390,80,'Gerencia Zona')
    c.setFont('Helvetica-Bold',8)
    if compra.autorizado1:
        c.drawCentredString(175,90,compra.oc_autorizada_por.staff.first_name + ' ' +compra.oc_autorizada_por.staff.last_name)
    if compra.autorizado2:
        c.drawCentredString(390,90,compra.oc_autorizada_por2.staff.first_name + ' ' + compra.oc_autorizada_por2.staff.last_name)

    c.setFont('Helvetica',10)
    subtotal = compra.costo_oc - compra.costo_iva
    
    importe_neto = compra.costo_oc
    if compra.impuestos_adicionales:
        subtotal = subtotal - compra.impuestos_adicionales
        c.setFillColor(black)
        c.setFont('Helvetica-Bold',9)
        #c.drawRightString(montos_align,170,'Impuestos Adicionales:') <<< se borra porque se encima con el total
        c.setFont('Helvetica',10)
        costo_impuestos = format(float(compra.impuestos), ',.2f')
        c.drawRightString(montos_align + 90, 180, '$' + str(costo_impuestos))
        c.drawRightString(montos_align, 180, 'Impuestos:')

    costo_subtotal = format(float(subtotal), ',.2f')
    c.drawRightString(montos_align + 90,210,'$ ' + str(costo_subtotal))
    costo_con_iva = format(float(compra.costo_iva), ',.2f')
    c.drawRightString(montos_align + 90,200,'$ ' + str(costo_con_iva))
    costo_oc =  format(float(compra.costo_oc), ',.2f')
    c.drawRightString(montos_align + 90,190,'$ ' + str(costo_oc))

    #c.drawRightString(montos_align + 90,180,'$ ' + str(compra.costo_fletes))
    c.setFillColor(prussian_blue)

    total =  format(float(compra.costo_plus_adicionales), ',.2f')
    if compra.costo_fletes:
        importe_neto = importe_neto + compra.costo_fletes
        c.drawRightString(montos_align,160,'Total:')
        c.drawRightString(montos_align,180,'Costo fletes:')
        c.drawRightString(montos_align + 90,180,'$ ' + str(compra.costo_fletes))
        c.drawRightString(montos_align + 90,160,'$ ' + str(total))
    else:
        c.drawRightString(montos_align,170,'Total:')
        
        c.drawRightString(montos_align + 90,170,'$ ' + str(total))
    
    
    c.setFillColor(prussian_blue)
    c.setFont('Helvetica', 9)

    if compra.moneda.nombre == "PESOS":
        c.drawString(40,201, num2words(compra.costo_plus_adicionales, lang='es', to='currency', currency='MXN'))
    if compra.moneda.nombre == "DOLARES":
        c.drawString(40,201, num2words(compra.costo_plus_adicionales, lang='es', to='currency',currency='USD'))
        
    c.setFillColor(black)
    width, height = letter
    styles = getSampleStyleSheet()

    if compra.opciones_condiciones is not None:
        options_conditions = compra.opciones_condiciones
    else:
        options_conditions = "NA"

    options_conditions_paragraph = Paragraph(options_conditions, styleN)


    # Crear un marco (frame) en la posición específica
    frame = Frame(135, 0, width-145, height-648, id='normal')

    # Agregar el párrafo al marco
    frame.addFromList([options_conditions_paragraph], c)
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    #TABLA 1 GENERACION OTRA HOJA
    table = Table(data, colWidths=[1.2 * cm, 1.5 * cm, 10 * cm, 1.5 * cm, 1.3 * cm, 1.5 * cm,1.5 * cm, 1.5 * cm,])
    table_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 7),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    # Definir el número de filas por página
    rows_per_page_first = 12
    rows_per_page_subsequent = 24
    # Dibuja la primera página
    first_page_data = data[:rows_per_page_first + 1]  # Incluye encabezado
    table = Table(first_page_data, colWidths=[1.2 * cm, 1.5 * cm, 10 * cm, 1.5 * cm, 1.3 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm])
    table.setStyle(table_style)
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, high)

    high = 630
    remaining_data = data[rows_per_page_first + 1:]

    while remaining_data:
        c.showPage()

        # Selecciona la cantidad de filas para la página actual
        current_page_data = [data[0]] + remaining_data[:rows_per_page_subsequent]  # Incluye encabezado
        remaining_data = remaining_data[rows_per_page_subsequent + 1:]

        # Calcula la altura basada en el número de filas actuales
        row_height = 18  # Altura de cada fila
        num_rows_current_page = len(current_page_data)
        total_table_height = row_height * num_rows_current_page

        # Ajusta la posición de la tabla en función de la altura total
        high = height - total_table_height - 120  # Margen superior ajustado
        if high < 50:  # Si la tabla se sale de la página, ajusta la altura
            high = 50

        # Dibuja la tabla
        table = Table(current_page_data, colWidths=[1.2 * cm, 1.5 * cm, 10 * cm, 1.5 * cm, 1.3 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm])
        table.setStyle(table_style)
        table.wrapOn(c, width, height)
        table.drawOn(c, 20, high)
        
        c.setFillColor(black)
        c.setLineWidth(.2)
        c.setFont('Helvetica',8)
        # Dibujar encabezado
        dibujar_encabezado(c, caja_iso)

        caja_proveedor = caja_iso - 65
        c.setFont('Helvetica', 12)
        c.setFillColor(prussian_blue)
        c.rect(150,750,250,30, fill=True, stroke=False) #Barra azul superior Orden de Compra
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 14)
        c.drawCentredString(280, 760, 'Orden de compra')
        c.drawInlineImage('static/images/logo vordtec_documento.png',40,755, 1.5 * cm, 0.75 * cm) #Imagen vortec
        #################################
    c.showPage()
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    # Dibujar encabezado
    dibujar_encabezado(c, caja_iso)

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica', 12)
    c.setFillColor(prussian_blue)
    c.rect(150,750,250,30, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(280, 760, 'Orden de compra')
    c.drawInlineImage('static/images/logo vordtec_documento.png',40,755, 1.5 * cm, 0.75 * cm) #Imagen vortec
    # Define el número de filas por página
    rows_per_page_first = 20
    rows_per_page_subsequent = 20
    high = 620
    # Estilos para los párrafos
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name='HeaderStyle',
        parent=styles['Normal'],
        fontSize=7,
        textColor='white',
        alignment=1,  # 1 es para centrar el texto
    )
    data_c = [
        [Paragraph('#', header_style),
        Paragraph('Cantidad', header_style),
        Paragraph('Código', header_style),
        Paragraph('Producto/Servicio', header_style),
        Paragraph('Criticidad', header_style),
        Paragraph('Descripción General', header_style),
        Paragraph('Especificaciones Técnicas', header_style),
        Paragraph('Criterios de aceptación', header_style)]
    ]
    item = 0
    count = 0
    max_count = 20 
    for producto in productos:
        item = item + 1
        try:
            producto_calidad = str(producto.producto.producto.articulos.producto.producto.producto_calidad.requisitos)
        except AttributeError:
            producto_calidad = ' '  # Mostrar espacio en lugar de None
        
        data_c.append([
            Paragraph(str(item), styleN),
            Paragraph(str(producto.cantidad) if producto.cantidad else ' ', styleN),
            Paragraph(producto.producto.producto.articulos.producto.producto.codigo if producto.producto.producto.articulos.producto.producto.codigo else ' ', styleN),
            Paragraph('Servicio' if producto.producto.producto.articulos.producto.producto.servicio else 'Producto', styleN),
            Paragraph(producto.producto.producto.articulos.producto.producto.critico.nombre if producto.producto.producto.articulos.producto.producto.critico else 'ND', styleN),
            Paragraph(producto.producto.producto.articulos.producto.producto.nombre if producto.producto.producto.articulos.producto.producto.nombre else ' ', styleN),
            #Paragraph('AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA', styleN),
            Paragraph(producto.producto.producto.articulos.producto.producto.especs if producto.producto.producto.articulos.producto.producto.especs else ' ', styleN),
            Paragraph(producto_calidad if producto_calidad else ' ', styleN)
        ])
        if count < max_count:
            high -= 21
            count += 1
    # Dibuja la primera página
    first_page_data = data_c[:rows_per_page_first + 1]  # Incluye encabezado
    table = Table(first_page_data, colWidths=[0.5 * cm, 1 * cm, 2 * cm, 1.5 * cm, 1.5 * cm, 4 * cm, 4 * cm, 6* cm])
    TABLE_LEFT = 20
    TABLE_TOP  = 690   # empieza con 690 y bájalo si tu header invade
    BOTTOM     = 40
    GAP        = 8
    avail_w = width - (TABLE_LEFT * 2)
    avail_h = (TABLE_TOP - GAP) - BOTTOM
    w, h = table.wrap(avail_w, avail_h)

    table.drawOn(c, TABLE_LEFT, (TABLE_TOP - GAP) - h)

    remaining_data = data_c[rows_per_page_first + 1:]
    # Actualiza la posición y los datos restantes
    #high = height - 100  # Ajustar para la primera página

    while remaining_data:
        c.showPage()
        high = 630
        c.setFillColor(black)
        c.setLineWidth(.2)
        c.setFont('Helvetica',8)
        # Dibujar encabezado
        dibujar_encabezado(c, caja_iso)

        caja_proveedor = caja_iso - 65
        c.setFont('Helvetica', 12)
        c.setFillColor(prussian_blue)
        c.rect(150,750,250,30, fill=True, stroke=False) #Barra azul superior Orden de Compra
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 14)
        c.drawCentredString(280, 760, 'Orden de compra')
        c.drawInlineImage('static/images/logo vordtec_documento.png',40,755, 1.5 * cm, 0.75 * cm) #Imagen vortec
        if len(remaining_data) < 15:
            high = high-100
        # Selecciona la cantidad de filas para la página actual
        current_page_data = [data_c[0]] + remaining_data[:rows_per_page_subsequent]  # Incluye encabezado
        remaining_data = remaining_data[rows_per_page_subsequent:]
        # Calcula la altura total de la tabla en función de las filas
        num_rows_current_page = len(current_page_data) - 1  # Excluye el encabezado de la cuenta
        total_table_height = 21 * num_rows_current_page  # Altura de cada fila

        # Dibuja la tabla
        table = Table(current_page_data, colWidths=[1.2 * cm, 1.5 * cm, 2 * cm, 1.5 * cm, 1.3 * cm, 9.5 * cm, 1.5 * cm, 1.5 * cm])
        table.setStyle(table_style)
        TABLE_LEFT = 20
        TABLE_TOP = 500   # <-- AJUSTA ESTO (prueba 660, 650, 640 según tu header)
        BOTTOM = 40
        GAP = 8
        
        avail_w = width - (TABLE_LEFT * 2)
        avail_h = (TABLE_TOP - GAP) - BOTTOM
        w, h = table.wrap(avail_w, avail_h)

        table.drawOn(c, TABLE_LEFT, (TABLE_TOP - GAP) - h)
    c.showPage()

    # Barra azul superior
    c.setFillColor(prussian_blue)
    c.rect(150, height - 50, 320, 30, fill=True, stroke=False)  # Ajusta las coordenadas según sea necesario

    # Título en la barra azul
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(310, height - 40, 'Objetivos De Calidad')

    # Imagen del logo
    c.drawInlineImage('static/images/logo vordtec_documento.png', 40, height - 60, 2.5 * cm, 1.25 * cm)  # Ajusta las coordenadas y tamaño según sea necesario

    # Puntos de los objetivos
    c.setFont('Helvetica', 12)
    c.setFillColor(colors.black)
    # Define el estilo para el párrafo
    objective_style = ParagraphStyle(
        name='ObjectiveStyle',
        fontName='Helvetica',
        fontSize=12,
        textColor=colors.black,
        spaceBefore=10,
        spaceAfter=10,
        leftIndent=10,
        rightIndent=10,
        alignment=0  # Alineación justificada
    )

    # Texto del párrafo con formato HTML
    text = '''
    <para>
    • Mantener el Sistema de Gestión de la Calidad ISO 9001:2015.<br/>
    <br/>
    • Comunicar los objetivos de calidad al 100 % del personal de VORDTEC DE MEXICO S.A de C.V. Así como a sus proveedores y contratistas para su cumplimiento.<br/>
    <br/>
    • Mantener por debajo del 4% anual el número de rechazos por producto no conforme.<br/>
    <br/>
    • Obtener un mínimo de 90% de satisfacción al cliente en las encuestas anuales.<br/>
    <br/>
    • Iniciar el proceso de implementación del departamento de Ingeniería en Vordtec de México S. A. de C. V.
    </para>
    '''
    paragraph = Paragraph(text, objective_style)
    max_width = width - 60  # Margen de 20 puntos en cada lado
    paragraph_width, paragraph_height = paragraph.wrap(max_width, height)

    # barra zul larga en la parte superior de la página
    c.setFillColor(prussian_blue)
    c.rect(40, height - 100, 530, 20, fill=True, stroke=False)

    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 12)
    c.drawCentredString(75, height - 95, 'Objetivos')
    #párrafo centrado en la página con márgenes
    paragraph_y = height - 40 - paragraph_height - 70 
    paragraph.drawOn(c, 30, paragraph_y)
    # azul corta en la parte inferior del texto
    c.setFillColor(prussian_blue)
    c.rect(40, height - 260, 530, 5, fill=True, stroke=False)

    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)

    c.showPage()
    c.save()
    buf.seek(0)
    return buf

def dias_laborables(inicio, fin):
    # Definimos los días festivos directamente dentro de la función
    festivos = [
        date(2024, 3, 18),  # Festivo específico 1
        date(2024, 3, 26),  # Festivo específico 2
    ]
    
    if not isinstance(inicio, date):
        inicio = date(2023, 3, 1)
    # Comenzamos el conteo desde el día siguiente al 'inicio' para no incluir el día inicial en el conteo
    dia_actual = inicio + timedelta(days=1)
    dias_habiles = 0
    if fin == "No existe":
        fin = date.today()

    

    while dia_actual < fin:  # Cambiamos a < para no incluir el día 'fin' en el conteo
        if dia_actual.weekday() < 5 and dia_actual not in festivos:
            dias_habiles += 1
        dia_actual += timedelta(days=1)
    
    # Verificamos si el día 'fin' debe contarse como un día hábil
    if fin.weekday() < 5 and fin not in festivos:
        dias_habiles += 1
    
    return dias_habiles

def convert_excel_matriz_compras(compras):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename= Matriz_compras_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Compra','Requisición','Solicitud','Proyecto','Subproyecto','Área','Solicitante','Creado','Req. Autorizada','Proveedor',
               'Crédito/Contado','Costo','Monto_Pagado','Status Pago','Status Autorización','Tipo Item','Días de entrega','Moneda',
               'Tipo de cambio','Entrada','Fecha Entrada','Fecha Inicio','Diferencia de Fechas','Status Entrega','No Conformidades','Total en pesos']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Grupo Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de OC's").style = head_style
    #ws.cell(row=4, column = columna_max, value="OC dentro de tiempo").style = head_style
    #ws.cell(row=5, column = columna_max, value="% de cumplimiento").style = head_style
    ws.cell(row=6, column = columna_max, value="Monto total de OC's").style = head_style

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    #ws.cell(row=4, column=columna_max + 1, value=f"=COUNTIF({get_column_letter(len(columns)-1)}:{get_column_letter(len(columns)-1)}, \"<=3\")").style = body_style
    #ws.cell(row=5, column=columna_max + 1, value=f"={get_column_letter(columna_max+1)}4/{get_column_letter(columna_max+1)}3").style = percent_style
    ws.cell(row=6, column=columna_max + 1, value=f"=SUM({get_column_letter(len(columns))}:{get_column_letter(len(columns))})").style = money_resumen_style

    #KPIS
    ws.cell(column = columna_max, row = 8, value='7.1 Porcentaje de órdenes de compra entregadas a tiempo').style = messages_style
    ws.cell(row=9, column = columna_max, value="Total de OC's con fecha Inicio").style = head_style
    ws.cell(row=10, column = columna_max, value="OC fuera de tiempo de entrega").style = head_style
    ws.cell(row=11, column = columna_max, value="% de cumplimiento").style = head_style

     # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=9, column=columna_max + 1, value=f"=COUNTIFS(U:U, \"<>No Existe\", U:U, \"<>\")").style = body_style
    ws.cell(row=10, column=columna_max + 1, value=f"=COUNTIF(X:X, \"Fuera de tiempo\")").style = body_style
    ws.cell(row=11, column=columna_max + 1, value=f"=1-({get_column_letter(columna_max+1)}10/{get_column_letter(columna_max+1)}9)").style = percent_style

    ws.cell(column = columna_max, row = 13, value='7.2.Porcentaje de productos o servicios recibidos sin no conformidades').style = messages_style
    ws.cell(row=14, column = columna_max, value="Total de OC's recibidas").style = head_style
    ws.cell(row=15, column = columna_max, value="Total de no conformidades").style = head_style
    ws.cell(row=16, column = columna_max, value="% de cumplimiento").style = head_style


    ws.cell(row=14, column=columna_max + 1, value=f"=COUNTIF(S:S, \"Entregado\")").style = body_style
    ws.cell(row=15, column=columna_max + 1, value=f"=COUNTIFS(Y:Y, \"<>No Existe\", Y:Y, \">0\")").style = body_style
    ws.cell(row=16, column=columna_max + 1, value=f"=1-IFERROR(({get_column_letter(columna_max+1)}15/{get_column_letter(columna_max+1)}14), 0)").style = percent_style

    rows = []
    for compra in compras:
        no_conformidades_count = 0
        # Obtén todos los pagos relacionados con esta compra
        pagos = Pago.objects.filter(oc=compra)
       
        # Calcula el tipo de cambio promedio de estos pagos
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg'] 

        # Usar el tipo de cambio de los pagos, si existe. De lo contrario, usar el tipo de cambio de la compra
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or compra.tipo_de_cambio or ""
        autorizado_text = 'Autorizado' if compra.autorizado2 else 'No Autorizado' if compra.autorizado2 == False or compra.autorizado1 == False else 'Pendiente Autorización'
        pagado_text = 'Pagada' if compra.pagada else 'No Pagada'
        entrada_text = 'Entregado' if compra.entrada_completa else 'No Entregado'
        condicion_fecha_ultima_entrada = True
        # Definimos la fecha de referencia
        fecha_referencia = datetime.strptime("27/03/2024","%d/%m/%Y").date()
        ultima_fecha_recepcion = compra.vale_entrada.aggregate(
            ultima_fecha=Max('articulos__fecha_recepcion')
        )['ultima_fecha']
       
        if ultima_fecha_recepcion:
        # Comparamos la fecha de vale_entrada con la fecha de referencia
            if ultima_fecha_recepcion.date() >= fecha_referencia:
                if compra.recepcion_completa:
                    # Si la recepción está completa, usamos la fecha de recepción del vale
                    fecha_ultima_entrada = ultima_fecha_recepcion.date()
        elif compra.entrada_completa:  # Verificamos si entrada es True para esta compra
            entradas = Entrada.objects.filter(oc=compra)
            ultima_entrada = entradas.order_by('-entrada_date').first()
           
            if ultima_entrada:  # Verificamos si existe al menos una entrada
                fecha_ultima_entrada = ultima_entrada.entrada_date
                # Contabilizar no_conformidades ligadas a las entradas de esta compra
                no_conformidades_count = No_Conformidad.objects.filter(oc=compra).count()
            else:
                # No hay entradas para esta compra
                fecha_ultima_entrada = "No existe"
                condicion_fecha_ultima_entrada = False
                no_conformidades_count = "No existe"
        else:
        # El atributo 'entrada' en Compra no es True
            fecha_ultima_entrada = "No existe"
            condicion_fecha_ultima_entrada = False
            no_conformidades_count = "No existe"
        
        ultimo_pago = None
        if compra.pagada:
            ultimo_pago = pagos.order_by('-pagado_date').first()

        
        if compra.cond_de_pago.nombre == "CONTADO" and compra.autorizado2 and compra.pagada:
            fecha_inicio = ultimo_pago.pagado_date
        elif compra.cond_de_pago.nombre == "CREDITO" and compra.autorizado2:
            fecha_inicio = compra.autorizado_date2
        else:
            fecha_inicio = "No existe"

        if condicion_fecha_ultima_entrada != False and fecha_inicio != "No existe" or fecha_inicio is not None:
            if fecha_ultima_entrada != None: 
                diferencia_fechas = dias_laborables(fecha_inicio, fecha_ultima_entrada)
        elif fecha_inicio != "No existe" and fecha_inicio is not None:
            diferencia_fechas = dias_laborables(fecha_inicio, date.today())
        else:
            diferencia_fechas = 0

        if compra.dias_de_entrega == None:
            compra.dias_de_entrega = 0
        if fecha_inicio == "No existe":
            cumplimiento_entrada = "No Evaluable"
        elif compra.dias_de_entrega >= diferencia_fechas:
            cumplimiento_entrada = "En tiempo"
        else:
            cumplimiento_entrada = "Fuera de tiempo"

        articulos = compra.articulocomprado_set.all()
        todos_servicios = all(articulo.producto.producto.articulos.producto.producto.servicio for articulo in articulos)
        ningun_servicio = all(not articulo.producto.producto.articulos.producto.producto.servicio for articulo in articulos)

        if todos_servicios:
            tipo_producto = "SERVICIOS"
        elif ningun_servicio:
            tipo_producto = "PRODUCTOS"
        else:
            tipo_producto = "PRODUCTO/SERVICIOS"


        row = [
        compra.id,
        compra.req.folio,
        compra.req.orden.folio,
        compra.req.orden.proyecto.nombre,
        compra.req.orden.subproyecto.nombre,
        compra.req.orden.area.nombre,
        f"{compra.req.orden.staff.staff.first_name} {compra.req.orden.staff.staff.last_name}",
        compra.created_at,
        compra.req.approved_at,
        compra.proveedor.nombre.razon_social,
        compra.cond_de_pago.nombre,
        compra.costo_oc,
        compra.monto_pagado,
        pagado_text,
        autorizado_text,
        tipo_producto,
        compra.dias_de_entrega,
        compra.moneda.nombre,
        tipo_de_cambio,
        entrada_text,
        fecha_ultima_entrada,
        fecha_inicio,
        diferencia_fechas,
        cumplimiento_entrada,
        no_conformidades_count
    ]
        if row[16] == "DOLARES":
            if row[17] is None or row[17] < 15:
                row[17] = 17  # o compra.pago_oc.tipo_de_cambio si así es como obtienes el valor correcto de tipo_de_cambio
        elif row[17] is None:  # por si acaso, aún manejar el caso donde 'tipo_de_cambio' es None
            row[17] = ""

        rows.append(row)

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 8 or col_num == 7 or col_num == 20 or col_num ==21    :
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 10 or col_num == 11 or col_num == 12 or col_num == 17 or col_num == 18:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
        # Agregamos la fórmula DATEDIF. Asumiendo que las columnas 'Creado' y 'Req. Autorizada'
        # están en las posiciones 8 y 9 respectivamente (empezando desde 0), las posiciones en Excel serán 9 y 10 (empezando desde 1).
        #ws.cell(row=row_num, column=len(columns)-1, value=f"=NETWORKDAYS(I{row_num}, H{row_num})").style = body_style
        # Agregar la fórmula de "Total en pesos"
        ws.cell(row=row_num, column = len(columns), value=f"=IF(ISBLANK(S{row_num}), L{row_num}, L{row_num}*S{row_num})").style = money_style
    
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    return(response)

def convert_excel_solicitud_matriz_productos(productos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = OC_por_producto_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Compras_Producto')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    number_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['OC','RQ','Sol','Solicitante','Proyecto','Subproyecto','Fecha','Proveedor','Estatus Proveedor','Área','Cantidad','Código', 
               'Producto','Criticidad','P.U.','Moneda','Tipo de Cambio','Subtotal','IVA','Total','Estatus','Pagada','NC','Status Entrega']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 4 or col_num == 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        if col_num == 9:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30



    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA Vordtec. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    # Calcular el número total de OCs únicas
    total_ocs = productos.values('oc').distinct().count()

    #ocs_por_proveedor = (
    #productos.values('oc__proveedor__nombre__razon_social')  # Agrupar por proveedor
    #.annotate(total_ocs=Count('oc', distinct=True))  # Contar OCs únicas
    #.order_by('-total_ocs')  # Ordenar por el mayor número de OCs
    #)
    ocs_por_proveedor = (
    productos.values('oc__proveedor__nombre__razon_social', 'oc__proveedor__estatus__nombre')  # Agrupar por proveedor y estatus
    .annotate(
        total_ocs=Count('oc', distinct=True),  # Contar OCs únicas
        total_costo_oc=Sum(
            Case(
                When(producto__producto__articulos__producto__producto__iva=True, then=F('precio_unitario') * F('cantidad') * Value(1.16)),  # Con IVA
                default=F('precio_unitario') * F('cantidad'),  # Sin IVA
                output_field=FloatField()
            )
        )
    )
    .order_by('-total_ocs')  # Ordenar por el mayor número de OCs
    )
    # Agregar los encabezados de los proveedores al costado de la tabla principal
   
    (ws.cell(column=columna_max, row=1, value='Proveedor')).style = head_style
    (ws.cell(column=columna_max + 1, row=1, value='Total de OCs')).style = head_style
    (ws.cell(column=columna_max + 2, row=1, value='Estatus del Proveedor')).style = head_style
    (ws.cell(column=columna_max + 3, row=1, value='Costo Total de OCs')).style = head_style

    # Agregar datos por proveedor
    fila_inicio = 2
    for index, proveedor_data in enumerate(ocs_por_proveedor, start=fila_inicio):
        proveedor_nombre = proveedor_data['oc__proveedor__nombre__razon_social']
        total_ocs = proveedor_data['total_ocs']
        estatus_proveedor = proveedor_data['oc__proveedor__estatus__nombre']
        total_costo_oc = proveedor_data['total_costo_oc'] or 0  # Manejar valores nulos en la suma

        ws.cell(column=columna_max, row=index, value=proveedor_nombre).style = body_style
        ws.cell(column=columna_max + 1, row=index, value=total_ocs).style = body_style
        ws.cell(column=columna_max + 2, row=index, value=estatus_proveedor).style = body_style
        ws.cell(column=columna_max + 3, row=index, value=total_costo_oc).style = money_style

        ws.column_dimensions[get_column_letter(columna_max)].width = 30
        ws.column_dimensions[get_column_letter(columna_max + 1)].width = 15
        ws.column_dimensions[get_column_letter(columna_max + 2)].width = 20
        ws.column_dimensions[get_column_letter(columna_max + 3)].width = 18

    fila_total = row_num + 2
    columna_total = columna_max + 6
    ws.cell(row=fila_total, column=columna_total, value="Total de OCs").style = head_style
    ws.cell(
        row=fila_total,
        column=columna_total + 1,
        value=f"=SUM({get_column_letter(columna_max + 1)}:{get_column_letter(columna_max + 1)})"
    ).style = body_style

    rows = []
    #cumplimiento_por_oc = {} 
    for producto in productos:
        # Extract the needed attributes
        compra_id = producto.oc.id
        req_folio = producto.oc.req.folio
        orden_folio = producto.oc.req.orden.folio
        staff_name = f"{producto.oc.req.orden.staff.staff.first_name} {producto.oc.req.orden.staff.staff.last_name}"
        proyecto_nombre = producto.oc.req.orden.proyecto.nombre
        subproyecto_nombre = producto.oc.req.orden.subproyecto.nombre
        created_at = producto.oc.created_at
        proveedor_nombre = producto.oc.proveedor.nombre.razon_social
        status_proveedor = producto.oc.proveedor.estatus.nombre 
        area_nombre = producto.oc.req.orden.area.nombre
        cantidad = producto.cantidad
        codigo = producto.producto.producto.articulos.producto.producto.codigo
        producto_nombre = producto.producto.producto.articulos.producto.producto.nombre
        criticidad = producto.producto.producto.articulos.producto.producto.critico
        precio_unitario = producto.precio_unitario
        moneda_nombre = producto.oc.moneda.nombre
        if producto.oc.autorizado2:
            estatus = 'Autorizada'
        elif producto.oc.autorizado1 == False or producto.oc.autorizado2 == False:
            estatus = 'Cancelada'  
        else:
            estatus = 'No autorizada aún'
        pagada = 'SI' if producto.oc.pagada == True else "NO"
        # Calculate total, subtotal, and IVA using attributes from producto
        subtotal = producto.subtotal_parcial
        iva = producto.iva_parcial
        total = subtotal + iva

        tiene_nc = NC_Articulo.objects.filter(
            Q(articulo_comprado=producto) |
            Q(entrada_articulo__articulo_comprado=producto)
        ).exists()

        nc_txt = "SI" if tiene_nc else "NO"
       

        # Handling the currency conversion logic
        pagos = Pago.objects.filter(oc_id=compra_id)
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or producto.oc.tipo_de_cambio
        condicion_fecha_ultima_entrada = True
        fecha_referencia = datetime.strptime("27/03/2024","%d/%m/%Y").date()
        ultima_fecha_recepcion = producto.oc.vale_entrada.aggregate(
            ultima_fecha=Max('articulos__fecha_recepcion')
        )['ultima_fecha']
       
        if ultima_fecha_recepcion:
        # Comparamos la fecha de vale_entrada con la fecha de referencia
            if ultima_fecha_recepcion.date() >= fecha_referencia:
                if producto.oc.recepcion_completa:
                    # Si la recepción está completa, usamos la fecha de recepción del vale
                    fecha_ultima_entrada = ultima_fecha_recepcion.date()
        elif producto.oc.entrada_completa:  # Verificamos si entrada es True para esta compra
            entradas = Entrada.objects.filter(oc=producto.oc)
            ultima_entrada = entradas.order_by('-entrada_date').first()
           
            if ultima_entrada:  # Verificamos si existe al menos una entrada
                fecha_ultima_entrada = ultima_entrada.entrada_date
              
            else:
                # No hay entradas para esta compra
                fecha_ultima_entrada = "No existe"
                condicion_fecha_ultima_entrada = False
        else:
        # El atributo 'entrada' en Compra no es True
            fecha_ultima_entrada = "No existe"
            condicion_fecha_ultima_entrada = False
           
        ultimo_pago = None
        if producto.oc.pagada:
            ultimo_pago = pagos.order_by('-pagado_date').first()

        if producto.oc.cond_de_pago.nombre == "CONTADO" and producto.oc.autorizado2 and producto.oc.pagada:
            fecha_inicio = ultimo_pago.pagado_date
        elif producto.oc.cond_de_pago.nombre == "CREDITO" and producto.oc.autorizado2:
            fecha_inicio = producto.oc.autorizado_date2
        else:
            fecha_inicio = "No existe"

        if condicion_fecha_ultima_entrada != False and fecha_inicio != "No existe" or fecha_inicio is not None:
            if fecha_ultima_entrada != None: 
                diferencia_fechas = dias_laborables(fecha_inicio, fecha_ultima_entrada)
        elif fecha_inicio != "No existe" and fecha_inicio is not None:
            diferencia_fechas = dias_laborables(fecha_inicio, date.today())
        else:
            diferencia_fechas = 0

        if producto.oc.dias_de_entrega == None:
            producto.oc.dias_de_entrega = 0
        if fecha_inicio == "No existe":
            cumplimiento_entrada = "No Evaluable"
        elif producto.oc.dias_de_entrega >= diferencia_fechas:
            cumplimiento_entrada = "En tiempo"
        else:
            cumplimiento_entrada = "Fuera de tiempo"

        #if moneda_nombre == "DOLARES" and tipo_de_cambio:
        #    total = total * tipo_de_cambio
        if criticidad is None:
            criticidad = ''
        else:
            criticidad = str(criticidad)
        if tipo_de_cambio is None:
            tipo_de_cambio = ''
        # Constructing the row
        row = [
            compra_id, #0
            req_folio, #1
            orden_folio, #2
            staff_name, #3
            proyecto_nombre, #4 
            subproyecto_nombre, #5 
            created_at, #6
            proveedor_nombre, #7
            status_proveedor, #8  
            area_nombre, #9
            cantidad, #10
            codigo,  #11
            producto_nombre, #12
            criticidad, #13
            precio_unitario, #14
            moneda_nombre, #15
            tipo_de_cambio, #16
            subtotal, #17
            iva, #18
            total, #19
            estatus, #20
            pagada, #21
            nc_txt, #22
            cumplimiento_entrada #23

        ]
        rows.append(row)

    # Building the Excel sheet with rows
    for row in rows:
        row_num += 1
        for col_num, cell_value in enumerate(row):
            ws.cell(row=row_num, column=col_num + 1, value=str(cell_value)).style = body_style
            if col_num == 5:
                ws.cell(row=row_num, column=col_num + 1, value=cell_value).style = body_style
            if col_num == 6:
                ws.cell(row=row_num, column=col_num + 1, value=cell_value).style = date_style
            if col_num == 10:
                ws.cell(row=row_num, column=col_num + 1, value=cell_value).style = number_style
            if col_num in [14, 16,17, 18, 19] :
                ws.cell(row=row_num, column=col_num + 1, value=cell_value).style = money_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_productos_requisitados(articulos):
    
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Requisiciones_por_producto_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Productos_Requisitados')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    number_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['RQ','Sol','Solicitante','Proyecto','Subproyecto','Fecha','Área','Cantidad','Código', 'Producto']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 4 or col_num == 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        if col_num == 9:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30



    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordtec. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Grupo Vordcab S.A. de C.V.}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20

    rows = []

    for producto in articulos:
        # Extract the needed attributes
        req_folio = producto.req.folio
        orden_folio = producto.req.orden.folio
        staff_name = f"{producto.req.orden.staff.staff.first_name} {producto.req.orden.staff.staff.last_name}"
        proyecto_nombre = producto.req.orden.proyecto.nombre
        subproyecto_nombre = producto.req.orden.subproyecto.nombre
        created_at = producto.req.created_at.replace(tzinfo=None)
        area_nombre = producto.req.orden.area.nombre
        cantidad = producto.cantidad
        codigo =producto.producto.articulos.producto.producto.codigo
        producto_nombre = producto.producto.articulos.producto.producto.nombre
      
    # Constructing the row
        row = [
            req_folio, 
            orden_folio, 
            staff_name, 
            proyecto_nombre, 
            subproyecto_nombre, 
            created_at,
            area_nombre,
            cantidad, 
            codigo, 
            producto_nombre, 
        ]
        rows.append(row)

    # Building the Excel sheet with rows
    for row in rows:
        row_num += 1
        for col_num, cell_value in enumerate(row):
            ws.cell(row=row_num, column=col_num + 1, value=str(cell_value)).style = body_style
            if col_num == 5:
                ws.cell(row=row_num, column=col_num + 1, value=cell_value).style = date_style
            if col_num == 7:
                ws.cell(row=row_num, column=col_num + 1, value=cell_value).style = number_style
           

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def generar_preevaluacion_inicial_pdf(preevaluacion):
    gerente = Profile.objects.get(tipo__nombre='Gerente')

    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',8)
    caja_iso = 760
    c.drawString(410,caja_iso + 10,'Preparado por:')
    c.drawString(510,caja_iso + 10,'Aprobación')
    c.drawString(20,caja_iso-20,'Número de documento')
    c.drawString(125,caja_iso-20,'Clasificación del documento')
    c.drawString(250,caja_iso-20,'Nivel del documento')
    c.drawString(345,caja_iso-20,'Revisión No.')
    c.drawString(410,caja_iso-20,'Fecha de Emisión')
    c.drawString(500,caja_iso-20,'Fecha de Modificación')
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)
    c.setFont('Helvetica',8)
    c.drawString(415,caja_iso,'Adquisiciones')
    c.drawString(485,caja_iso,'Subdirección Administrativa')
    c.drawString(30,caja_iso-30,'F-ADQ-N4-01.03')
    c.drawString(160,caja_iso-30,'Registro')
    c.drawString(280,caja_iso-30, 'N5')
    c.drawString(355,caja_iso-30,'000')
    c.drawString(425,caja_iso-30,'02/01/2024')
    c.drawString(520,caja_iso-30,'02/01/2024')

    caja_proveedor = caja_iso - 50
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(100,750,300,30, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.rect(20,caja_proveedor,565,10, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(250,760,'Reporte de evaluación inicial')
    c.setLineWidth(.3) #Grosor

    c.drawInlineImage('static/images/logo vordtec_documento.png',40,755, 1.5 * cm, 0.75 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',9)
    c.drawString(250,caja_proveedor+1,'Información General')

    c.setFillColor(black)
    c.setFont('Helvetica-Bold',8)
    c.drawString(40,caja_proveedor-15,'Fecha de Evaluación:')
    #c.line(133,caja_proveedor-16,330,caja_proveedor-16)
    c.drawString(40,caja_proveedor-35,'Nombre del Proveedor:')
    #c.line(133,caja_proveedor-36,330,caja_proveedor-36)
    c.drawString(40,caja_proveedor-55,'Número de identificación del proveedor:')
    #c.line(198,caja_proveedor-56,330,caja_proveedor-56)

    c.setFont('Helvetica',8)
    c.drawString(135,caja_proveedor-15, str(preevaluacion.creado_at.strftime("%d/%m/%Y %H:%M")))
    c.drawString(135,caja_proveedor-35, preevaluacion.nombre.razon_social)
    c.drawString(200,caja_proveedor-55, str(preevaluacion.nombre.id))

    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    #c.rect(20,caja_proveedor-90,565,10, fill=True, stroke=False) #Segunda caja

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',9)
    #c.drawString(90,caja_proveedor-89,'')


    data =[]
    data1 =[] 
    high = 530
    item = 0
    

    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent = styles['Normal'],
        fontName = 'Helvetica',
        fontSize=6,
        leaging = 10,
        wordWrap='CJK',
        alignment=TA_CENTER,  # Centrar el texto
        )
    
    custom_style_tight = ParagraphStyle(
        'CustomTight',
        parent=styles['Normal'],
        leading=8,  # Ajusta este valor según el espaciado que desees
        fontSize=5,
        textColor=colors.black,
    )

    # Definición del texto para los párrafos
    text_1 = "Nota: Para los proveedores de productos y servicios criticos, se requiere el cumplimiento total de los criterios para obtener la pre aprobación."
    text_2 = "La documentación de respaldo debe adjuntarse con la evaluación, incluyendo un detalle especifico de los productos evaluados."
    combined_text = f"{text_1} {text_2}"
    note_combined = Paragraph(combined_text, style=custom_style_tight)

    text = "Verificación de Cumplimiento del Sistema de Gestión de la Calidad del Proveedor"
    a2 = Paragraph(text, style=custom_style)
    text2 = "Control en la Cadena de Suministro del Proveedor"
    a3 = Paragraph(text2, style=custom_style)
    text1 = "Evaluación de la Capacidad del Proveedor"
    a1 = Paragraph(text1, style=custom_style)

    if preevaluacion.verif_calidad_b is True:
        text = "Cumple con las condiciones"
    else:
        text = "No cumple con las condiciones"
    a7 = Paragraph(text, style=custom_style)

    if preevaluacion.control_cadena_b is True:
        text = "Cumple con las condiciones"
    else:
        text = "No cumple con las condiciones"
    a8 = Paragraph(text, style=custom_style)

    if preevaluacion.capacidad_proveedor_b is True:
        text = "Cumple con las condiciones"
    else:
        text = "No cumple con las condiciones"
    a9 = Paragraph(text, style=custom_style)

    calidad = preevaluacion.verif_calidad
    suministro = preevaluacion.control_cadena_suministro
    capacidad = preevaluacion.capacidad_proveedor

    if calidad is None:
        calidad = ' '
    if suministro is None:
        suministro = ' '
    if capacidad is None:
        capacidad = ' '
    a4 = Paragraph(calidad, style=custom_style)
    a5 = Paragraph(suministro, style=custom_style)
    a6 = Paragraph(capacidad, style=custom_style)
    # Datos de la tabla
    data = [
        [''' ''', '''Criterios para la Evaluación Inicial de proveedores - compras criticas ''', '''Cumple o no cumple'''], 
        [a2, a4, a7],
        [a3, a5, a8],
        [a1, a6, a9],
        ['',note_combined,''],
    ]

    text_3 = "Nota: Para los proveedores de productos y servicios no criticos, se requiere el cumplimiento de al menos uno de los requisitos para obtener la pre aprobación."
    text_4 = "La documentación de respaldo debe adjuntarse con la evaluación, incluyendo un detalle especifico de los productos evaluados."
    combined_text2 = f"{text_3} {text_4}"
    note_combined1 = Paragraph(combined_text2, style=custom_style_tight)

    text3 = "Verificación de Cumplimiento del Sistema de Gestión de la Calidad del Proveedor"
    b = Paragraph(text3, style=custom_style)
    text4 = "Evaluación del proveedor para que cumpla con los requisitos de compra de la organización"
    b1 = Paragraph(text4, style=custom_style)
    text5 = "Evaluación del producto luego de la entrega o actividades de una vez finalizadas"
    b2 = Paragraph(text5, style=custom_style)

    if preevaluacion.sgc_b is True:
        text = "Cumple con las condiciones"
    else:
        text = "No cumple con las condiciones"
    b6 = Paragraph(text, style=custom_style)

    if preevaluacion.eval_compra_b is True:
        text = "Cumple con las condiciones"
    else:
        text = "No cumple con las condiciones"
    b7 = Paragraph(text, style=custom_style)

    if preevaluacion.eval_actividades_b is True:
        text = "Cumple con las condiciones"
    else:
        text = "No cumple con las condiciones"
    b8 = Paragraph(text, style=custom_style)

    requisito_sgc = preevaluacion.requisitos_sgc_ver
    requisito_compra = preevaluacion.eval_compra
    producto_eval = preevaluacion.eval_actividades

    if requisito_sgc is None:
        requisito_sgc = ' '
    if requisito_compra is None:
        requisito_compra = ' '
    if producto_eval is None:
        producto_eval = ' '
    b3 = Paragraph(requisito_sgc, style=custom_style)
    b4 = Paragraph(requisito_compra, style=custom_style)
    b5 = Paragraph(producto_eval, style=custom_style)

    data1 = [
        [''' ''', '''Criterios para la Evaluación Inicial de proveedores - compras  no criticas ''', '''Cumple o no cumple'''], 
        [b, b3, b6],
        [b1, b4, b7],
        [b2, b5, b8],
        ['',note_combined1,''],
    ]

    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(50,145,200,15, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.rect(350,145,200,15, fill=True, stroke=False) #Barra azul superior Orden de Compra

    c.setFillColor(white)
    width, height = letter
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    c.drawCentredString(150,150,'Realizador de la Evaluación')
    c.drawCentredString(450,150,'Autorización del Gerente General')
    c.setFillColor(black)
    c.drawString(100,100,'Jefe de adquisiciones')
    c.drawCentredString(150,115,preevaluacion.creado_por.staff.first_name + ' ' + preevaluacion.creado_por.staff.last_name)
    c.line(60,110,240,110)
    c.drawString(420,100,'Gerente General')
    c.drawCentredString(450,115,str(gerente.staff.first_name + ' ' + gerente.staff.last_name))
    c.line(360,110,540,110)

    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)#Caja de hasta abajo
    c.setFillColor(white)

    table = Table(data, colWidths=[3 * cm, 14* cm, 3 * cm,])
    table_style = TableStyle([
        ('INNERGRID', (0, 0), (-1, -2), 0.25, colors.black),  # Líneas internas, excepto en la última fila
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Línea divisoria entre filas, excepto la última fila
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.black),
        ('LINEBELOW', (0, 1), (-1, 1), 0.5, colors.black),
        ('LINEBELOW', (0, 2), (-1, 2), 0.5, colors.black),
        ('LINEBELOW', (0, 3), (-1, 3), 0.5, colors.black),
        # ENCABEZADO
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fuente en negrita
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003153')),  # Usando un color azul prusia
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Centrar el texto del encabezado
        # CUERPO
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 5),
    ])

    table1 = Table(data1, colWidths=[3 * cm, 14* cm, 3 * cm,])
    table1_style = TableStyle([
        ('INNERGRID', (0, 0), (-1, -2), 0.25, colors.black),  # Líneas internas, excepto en la última fila
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Línea divisoria entre filas, excepto la última fila
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.black),
        ('LINEBELOW', (0, 1), (-1, 1), 0.5, colors.black),
        ('LINEBELOW', (0, 2), (-1, 2), 0.5, colors.black),
        ('LINEBELOW', (0, 3), (-1, 3), 0.5, colors.black),
        # ENCABEZADO
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fuente en negrita
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003153')),  # Usando un color azul prusia
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Centrar el texto del encabezado
        # CUERPO
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 5),
    ])
       
    table.setStyle(table_style)
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, 490) 
    table1.setStyle(table_style)
    table1.wrapOn(c, width, height)
    table1.drawOn(c, 20, 210) 

    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica', 12)
    c.setFillColor(prussian_blue)
   
    c.showPage()
    
    c.save()
    buf.seek(0)
    return buf

def generar_preevaluacion_pdf(preevaluacion):
    gerente = Profile.objects.get(tipo__nombre='Gerente')

    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',8)
    caja_iso = 760
    c.drawString(410,caja_iso + 10,'Preparado por:')
    c.drawString(510,caja_iso + 10,'Aprobación')
    c.drawString(20,caja_iso-20,'Número de documento')
    c.drawString(125,caja_iso-20,'Clasificación del documento')
    c.drawString(250,caja_iso-20,'Nivel del documento')
    c.drawString(345,caja_iso-20,'Revisión No.')
    c.drawString(410,caja_iso-20,'Fecha de Emisión')
    c.drawString(500,caja_iso-20,'Fecha de Modificación')
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)
    c.setFont('Helvetica',8)
    c.drawString(415,caja_iso,'Adquisiciones')
    c.drawString(485,caja_iso,'Subdirección Administrativa')
    c.drawString(30,caja_iso-30,'F-ADQ-N4-01.04')
    c.drawString(160,caja_iso-30,'Registro')
    c.drawString(280,caja_iso-30, 'N5')
    c.drawString(355,caja_iso-30,'000')
    c.drawString(425,caja_iso-30,'02/01/2024')
    c.drawString(520,caja_iso-30,'02/01/2024')

    caja_proveedor = caja_iso - 50
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(100,750,300,30, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.rect(20,caja_proveedor,565,10, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(250,760,'Reporte de evaluación inicial simplificada')
    c.setLineWidth(.3) #Grosor

    c.drawInlineImage('static/images/logo vordtec_documento.png',40,755, 1.5 * cm, 0.75 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',9)
    c.drawString(250,caja_proveedor+1,'Información General')

    c.setFillColor(black)
    c.setFont('Helvetica-Bold',8)
    c.drawString(40,caja_proveedor-15,'Fecha de Evaluación:')
    #c.line(133,caja_proveedor-16,330,caja_proveedor-16)
    c.drawString(40,caja_proveedor-35,'Nombre del Proveedor:')
    #c.line(133,caja_proveedor-36,330,caja_proveedor-36)
    c.drawString(40,caja_proveedor-55,'Número de identificación del proveedor:')
    #c.line(198,caja_proveedor-56,330,caja_proveedor-56)

    c.setFont('Helvetica',8)
    c.drawString(135,caja_proveedor-15, str(preevaluacion.creado_at.strftime("%d/%m/%Y %H:%M")))
    c.drawString(135,caja_proveedor-35, preevaluacion.nombre.razon_social)
    c.drawString(200,caja_proveedor-55, str(preevaluacion.nombre.id))

    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    #c.rect(20,caja_proveedor-90,565,10, fill=True, stroke=False) #Segunda caja

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',9)
    #c.drawString(90,caja_proveedor-89,'')


    data =[]
    high = 530
    item = 0
    

    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent = styles['Normal'],
        fontName = 'Helvetica',
        fontSize=6,
        leaging = 10,
        wordWrap='CJK',
        alignment=TA_CENTER,  # Centrar el texto
        )
    
    custom_style_tight = ParagraphStyle(
        'CustomTight',
        parent=styles['Normal'],
        leading=8,  # Ajusta este valor según el espaciado que desees
        fontSize=5,
        textColor=colors.black,
    )

    # Definición del texto para los párrafos
    text_1 = "Nota: Para los proveedores de productos y servicios clasificados como generales, es imperativo cumplir totalmente con los criterios establecidos para obtener la preaprobación. Se solicita que la documentación de respaldo se adjunte con la evaluación."
    text_2 = "proporcionando especifico de los productos evaluados, lo que incluye una comparativa de precios y evidencia que demuestre que el producto cumple con las especificaciones requeridas."
    combined_text = f"{text_1} {text_2}"
    note_combined = Paragraph(combined_text, style=custom_style_tight)

    text = "Verificar que el producto cumple con las especificaciones requeridas"
    a2 = Paragraph(text, style=custom_style)

    text2 = "Comparar los precios ofrecidos por el proveedor con al menos otras dos cotizaciones de diferentes proveedores, emitidas dentro del último año para asegurar que sean aceptables"
    a3 = Paragraph(text2, style=custom_style)
    if preevaluacion.especs_b is True:
        text4 = "Cumple con las condiciones"
    else:
        text4 = "No cumple con las condiciones"
    a4 = Paragraph(text4, style=custom_style)
    if preevaluacion.precios_b is True:
        text5 = "Cumple con las condiciones"
    else:
        text5 = "No cumple con las condiciones"
    a5 = Paragraph(text5, style=custom_style)
    especs = preevaluacion.especs_ver
    precios = preevaluacion.precios_ver
    if especs is None:
        especs = ' '
    if precios is None:
        precios = ' '
    a6 = Paragraph(especs, style=custom_style)
    a7 = Paragraph(precios, style=custom_style)
    # Datos de la tabla
    data = [
        [''' ''', '''Criterios para la Evaluación Inicial de proveedores - productos generales''', '''Cumple o no cumple'''], 
        [a2, a6, a4],
        [a3, a7, a5],
        ['',note_combined,''],
    ]

    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(50,195,200,15, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.rect(350,195,200,15, fill=True, stroke=False) #Barra azul superior Orden de Compra

    c.setFillColor(white)
    width, height = letter
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    c.drawCentredString(150,200,'Realizador de la Evaluación')
    c.drawCentredString(450,200,'Autorización del Gerente General')
    c.setFillColor(black)
    c.drawString(100,150,'Jefe de adquisiciones')
    c.drawCentredString(150,165,preevaluacion.creado_por.staff.first_name + ' ' + preevaluacion.creado_por.staff.last_name)
    c.line(60,160,240,160)
    c.drawString(420,150,'Gerente General')
    c.drawCentredString(450,165,str(gerente.staff.first_name + ' ' + gerente.staff.last_name))
    c.line(360,160,540,160)

    # Crear un marco (frame) en la posición específica
    frame = Frame(135, 0, width-145, height-648, id='normal')

    # Agregar el párrafo al marco
    #frame.addFromList([options_conditions_paragraph], c)
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)#Caja de hasta abajo
    c.setFillColor(white)

    table = Table(data, colWidths=[3 * cm, 14* cm, 3 * cm,])
    table_style = TableStyle([
        ('INNERGRID', (0, 0), (-1, -2), 0.25, colors.black),  # Líneas internas, excepto en la última fila
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Línea divisoria entre filas, excepto la última fila
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.black),
        ('LINEBELOW', (0, 1), (-1, 1), 0.5, colors.black),
        ('LINEBELOW', (0, 2), (-1, 2), 0.5, colors.black),  
        # ENCABEZADO
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fuente en negrita
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003153')),  # Usando un color azul prusia
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Centrar el texto del encabezado
        # CUERPO
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 5),
    ])
    
       
    table.setStyle(table_style)
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, 400) 
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica', 12)
    c.setFillColor(prussian_blue)
   
    c.showPage()
    
    c.save()
    buf.seek(0)
    return buf

def safe_string(value, default=''):
    return value if value is not None else default

def generar_oc_comparativa_pdf(request, pk):
    hoy = datetime.today().strftime("%d/%m/%Y %H:%M")

    compra = Compra.objects.get(id=pk)
    comparativo = Comparativo.objects.get(id=compra.comparativo_model.id)
    productos = Item_Comparativo.objects.get(comparativo = comparativo)

    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',8)
    caja_iso = 760
    #c.drawString(410,caja_iso + 10,'Preparado por:')
    c.drawString(510,caja_iso + 10,'Fecha reporte')
    c.setFont('Helvetica',8)
    c.drawString(505,caja_iso,str(hoy))

    caja_proveedor = caja_iso - 40
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(100,750,400,30, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.rect(20,caja_proveedor,565,14, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(310,760,'Registro Comparativo OC')
    c.setLineWidth(.3) #Grosor

    c.drawInlineImage('static/images/logo vordtec_documento.png',40,755, 1.5 * cm, 0.75 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',9)
    c.drawString(250,caja_proveedor+4,'Información General')
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',8)
    c.drawString(40,caja_proveedor-15,'Folio de la OC:')
    c.drawString(40,caja_proveedor-35,'Creada por:')
    c.drawString(40,caja_proveedor-55,'Creada el:')
    c.drawString(40,caja_proveedor-75,'Nombre comparativa:')
    c.drawString(40,caja_proveedor-95,'Superintendente:')
    c.drawString(40,caja_proveedor-115,'Autorizada el:')

    c.drawString(300,caja_proveedor-15,'Gerente:')
    c.drawString(300,caja_proveedor-35,'Autorizada el:')

    c.setFont('Helvetica',8)
    if compra.folio is not None:
        c.drawString(100,caja_proveedor-15,compra.folio)
    else:
        c.drawString(100,caja_proveedor-15,'N/A')

    if compra.creada_por is not None:
        c.drawString(100,caja_proveedor-35,compra.creada_por.staff.first_name + ' ' + compra.creada_por.staff.last_name)
    else:
        c.drawString(100,caja_proveedor-35,'N/A')
    c.drawString(100,caja_proveedor-55,compra.created_at.strftime("%d/%m/%Y %H:%M"))
    c.drawString(130,caja_proveedor-75,str(comparativo))
    if compra.oc_autorizada_por is not None:
        c.drawString(120,caja_proveedor-95,compra.oc_autorizada_por.staff.first_name + ' ' + compra.oc_autorizada_por.staff.last_name)
    else:
        c.setFillColor(red)
        c.drawString(120,caja_proveedor-95,'No autorizado aún')
        c.setFillColor(black)
    if compra.autorizado_date1 is not None:
        c.drawString(120,caja_proveedor-115,compra.autorizado_date1.strftime("%d/%m/%Y %H:%M"))
    else:
        c.drawString(120,caja_proveedor-115,'N/A')
    
    if compra.oc_autorizada_por2 is not None:
        c.drawString(360,caja_proveedor-15,compra.oc_autorizada_por2.staff.first_name + ' ' + compra.oc_autorizada_por2.staff.last_name)
    else:
        c.setFillColor(red)
        c.drawString(360,caja_proveedor-15,'No autorizado aún')
        c.setFillColor(black)
    if compra.autorizado_date2 is not None:
        c.drawString(360,caja_proveedor-35,compra.autorizado_date2.strftime("%d/%m/%Y %H:%M"))
    else:
        c.drawString(360,caja_proveedor-35,'N/A')

    data =[]
    styles = getSampleStyleSheet()
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent = styles['Normal'],
        fontName = 'Helvetica',
        fontSize=6,
        leaging = 10,
        wordWrap='CJK',
        alignment=TA_CENTER,  # Centrar el texto
        )
    custom_style_tight = ParagraphStyle(
        'CustomTight',
        parent=styles['Normal'],
        leading=8,  # Ajusta este valor según el espaciado que desees
        fontSize=5,
        textColor=colors.black,
    )

    if productos.comparativo.cotizacion is not None:
        coti1= 'Si'
    else:
        coti1 = 'Sin documento'
    if productos.comparativo.cotizacion is not None:
        coti2= 'Si'
    else:
        coti2 = 'Sin documento'
    if productos.comparativo.cotizacion is not None:
        coti3 = 'Si'
    else:
        coti3= 'Sin documento'

    # Usar la función safe_string para manejar None
    nombre = safe_string(productos.producto.producto.nombre)
    proveedor1 = safe_string(productos.comparativo.proveedor.nombre.razon_social)
    proveedor2 = safe_string(productos.comparativo.proveedor2.nombre.razon_social)
    proveedor3 = safe_string(productos.comparativo.proveedor3.nombre.razon_social)
    modelo1 = safe_string(productos.modelo)
    modelo2 = safe_string(productos.modelo2)
    modelo3 = safe_string(productos.modelo3)
    marca1 = safe_string(productos.marca)
    marca2 = safe_string(productos.marca2)
    marca3 = safe_string(productos.marca3)
    precio1 = safe_string(str(productos.precio))
    precio2 = safe_string(str(productos.precio2))
    precio3 = safe_string(str(productos.precio3))

    nombre = Paragraph(nombre, style=custom_style_tight)
    proveedor1 = Paragraph(proveedor1, style=custom_style_tight)
    proveedor2 = Paragraph(proveedor2, style=custom_style_tight)
    proveedor3 = Paragraph(proveedor3, style=custom_style_tight)
    modelo1 = Paragraph(modelo1, style=custom_style_tight)
    modelo2 = Paragraph(modelo2, style=custom_style_tight)
    modelo3 = Paragraph(modelo3, style=custom_style_tight)
    marca1 = Paragraph(marca1, style=custom_style_tight)
    marca2 = Paragraph(marca2, style=custom_style_tight)
    marca3 = Paragraph(marca3, style=custom_style_tight)
    precio1 = Paragraph(precio1, style=custom_style_tight)
    precio2 = Paragraph(precio2, style=custom_style_tight)
    precio3 = Paragraph(precio3, style=custom_style_tight)
    data = [
        ['''''','''''', 'Tabla comparativa', '''''','''''',''''''], 
        ['''Producto''','''Proveedor''', '''Modelo''', '''Marca''','''Precio''','''Cotización'''], 
        [nombre, proveedor1, modelo1, marca1, precio1, coti1],
        [nombre, proveedor2, modelo2, marca2, precio2, coti2],
        [nombre, proveedor3, modelo3, marca3, precio3, coti3, ],
    ]

    c.setFillColor(prussian_blue)
    c.setFillColor(white)
    width, height = letter

    table = Table(data, colWidths=[4 * cm, 4 * cm, 3.5 * cm, 3.5 * cm, 3.0 * cm, 2 * cm,])
    table_style = TableStyle([
        ('INNERGRID', (0, 1), (-1, -1), 0.25, colors.black),  # Líneas internas
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),  # Bordes de la tabla
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical
        
        # Estilo para el encabezado
        ('TEXTCOLOR', (0, 0), (-1, 1), colors.white),  # Texto blanco para la primera y segunda fila
        ('FONTSIZE', (0, 0), (-1, 1), 7),  # Tamaño de fuente para la primera y segunda fila
        ('FONT', (0, 0), (-1, 1), 'Helvetica-Bold'),  # Fuente en negrita para la primera y segunda fila
        ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#003153')),  # Fondo azul prusia para la primera y segunda fila
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),  # Centrar el texto del encabezado

        # Estilo para el cuerpo de la tabla
        ('TEXTCOLOR', (0, 2), (-1, -1), colors.black),  # Texto negro para el cuerpo de la tabla
        ('FONTSIZE', (0, 2), (-1, -1), 5),  # Tamaño de fuente para el cuerpo
    ])
    
       
    table.setStyle(table_style)
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, 400) 
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica', 12)
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)

    c.showPage()
    
    c.save()
    buf.seek(0)

    #return FileResponse(buf, as_attachment=True, filename='Comparativa_' + str(compras.folio) +'.pdf')
    return FileResponse(buf, as_attachment=True, filename='Comparativa_prueba'  +'.pdf')

def generar_oc_comparativas_pdf(request, pk):
    hoy = datetime.today().strftime("%d/%m/%Y %H:%M")

    comparativo = Comparativo.objects.get(id=pk)
    productos = Item_Comparativo.objects.filter(comparativo = comparativo)

    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',8)
    caja_iso = 760
    #c.drawString(410,caja_iso + 10,'Preparado por:')
    c.drawString(510,caja_iso + 10,'Fecha reporte')
    c.setFont('Helvetica',8)
    c.drawString(505,caja_iso,str(hoy))

    caja_proveedor = caja_iso - 40
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(100,750,400,30, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.rect(20,caja_proveedor,565,14, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(310,760,'Registro Comparativo OC')
    c.setFont('Helvetica-Bold',10)
    c.drawCentredString(310,caja_proveedor+4,str(comparativo))
    c.setLineWidth(.3) #Grosor

    c.drawInlineImage('static/images/logo vordtec_documento.png',40,755, 1.5 * cm, 0.75 * cm) #Imagen vortec

    width, height = letter
    data =[]
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        name="header_style",
        fontName="Helvetica-Bold",
        fontSize=7,
        textColor=colors.white,
        alignment=1,  # Centrado
    )
    custom_style_tight = ParagraphStyle(
        'CustomTight',
        parent=styles['Normal'],
        leading=8,  # Ajusta este valor según el espaciado que desees
        fontSize=5,
        textColor=colors.black,
        )

    header = [
        Paragraph('Nombre', style=header_style),
        Paragraph('Proveedor 1', style=header_style),
        #Paragraph('Modelo 1', style=header_style),
        #Paragraph('Marca 1', style=header_style),
        Paragraph('Precio 1', style=header_style),
        Paragraph('Cotización 1', style=header_style),
        Paragraph('Proveedor 2', style=header_style),
        #Paragraph('Modelo 2', style=header_style),
        #Paragraph('Marca 2', style=header_style),
        Paragraph('Precio 2', style=header_style),
        Paragraph('Cotización 2', style=header_style),
        Paragraph('Proveedor 3', style=header_style),
        #Paragraph('Modelo 3', style=header_style),
        #Paragraph('Marca 3', style=header_style),
        Paragraph('Precio 3', style=header_style),
        Paragraph('Cotización 3', style=header_style),
    ]

    row_height = 35  # Altura de cada fila
    max_items_per_page = 12  # Máximo de 12 elementos por página
    max_height_per_page = 600  # Altura máxima para datos en una página
    page_margin = 100  # Margen superior

    # Preparar datos para la tabla
    data = [header]
    current_height = max_height_per_page
    pages = []  # Lista para almacenar los datos de cada página
    current_item_count = 0  # Contador de elementos por página
    for i in range(1):
        for producto in productos:
            # Crear filas para la tabla
            nombre = Paragraph(safe_string(producto.producto.producto.nombre), style=custom_style_tight)
            proveedor1 = Paragraph(safe_string(producto.comparativo.proveedor.nombre.razon_social), style=custom_style_tight)
            precio1 = Paragraph('$' + safe_string(str(producto.precio)), style=custom_style_tight)
            coti1 = Paragraph('Si' if producto.comparativo.cotizacion else 'Sin documento', style=custom_style_tight)
            proveedor2 = Paragraph(safe_string(producto.comparativo.proveedor2.nombre.razon_social), style=custom_style_tight)
            precio2 = Paragraph('$' + safe_string(str(producto.precio2)), style=custom_style_tight)
            coti2 = Paragraph('Si' if producto.comparativo.cotizacion else 'Sin documento', style=custom_style_tight)
            proveedor3 = Paragraph(safe_string(producto.comparativo.proveedor3.nombre.razon_social), style=custom_style_tight)
            precio3 = Paragraph('$' + safe_string(str(producto.precio3)), style=custom_style_tight)
            coti3 = Paragraph('Si' if producto.comparativo.cotizacion else 'Sin documento', style=custom_style_tight)

            # Agregar fila al conjunto actual
            row = [nombre, proveedor1, precio1, coti1, proveedor2, precio2, coti2, proveedor3, precio3, coti3]
            data.append(row)
            current_item_count += 1
            current_height -= row_height  # Reducir la altura disponible

            # Si hemos alcanzado el límite de 12 elementos o no hay suficiente espacio
            if current_item_count >= max_items_per_page or current_height < page_margin:
                pages.append(data)  # Guardar las filas de la página actual
                data = [header]  # Reiniciar con el encabezado para la nueva página
                current_height = max_height_per_page  # Reiniciar la altura disponible
                current_item_count = 0  # Reiniciar el contador de elementos por página

    # Agregar la última página si quedan elementos
    if data:
        pages.append(data)

    # Dibujar la tabla
    for page_data in pages:
        table = Table(page_data, colWidths=[2 * cm] * 10)
        table.setStyle(TableStyle([
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003153')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 5),
        ]))

        # Calcular la altura de inicio para la página actual
        start_height = max_height_per_page - (row_height * current_item_count) - row_height  # Ajustar altura en función de los productos en la página

        table.wrapOn(c, width, height)
        table.drawOn(c, 20, start_height)  # Dibujar la tabla en la posición ajustada

        c.showPage()  # Crear nueva página

    # Guardar el archivo PDF
    c.save()
    buf.seek(0)


    return FileResponse(buf, as_attachment=True, filename='Comparativo_'+ str(comparativo.id)  + '.pdf')

@login_required(login_url='user-login')
def comparativo_historico(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    
    items_comparativos = Item_Comparativo.objects.filter(completo=True).select_related('comparativo').prefetch_related('comparativo__preevaluaciones_comparativo').order_by('-comparativo__id')
    myfilter = Item_ComparativoFilter(request.GET, queryset=items_comparativos)

    itemss = myfilter.qs

     #Set up pagination
    p = Paginator(itemss, 50)
    page = request.GET.get('page')
    item_list = p.get_page(page)
    
    context= {
        'myfilter':myfilter,
        'itemss':itemss,
        'item_list':item_list,
    }
    return render(request,'compras/comparativo_historico.html', context)