from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, DecimalField,  Sum, OuterRef, Subquery, DecimalField, Value, DecimalField
from django.db.models.functions import Concat
from django.http import JsonResponse, HttpResponse, Http404
from django.core.mail import EmailMessage, BadHeaderError
from .filters import EntradaArticuloFilter, No_ConformidadFilter, Reporte_CalidadFilter, EntradaCaducidadFilter, EntradaTerminadoFilter
from compras.models import Compra, ArticuloComprado
from compras.filters import CompraFilter
from compras.views import attach_oc_pdf
from dashboard.models import Inventario, Order, ArticulosparaSurtir, Producto_Calidad, Productos_Solicitud_Terminado
from requisiciones.models import Salidas, ArticulosRequisitados, Requis
from .models import Entrada, EntradaArticulo, Reporte_Calidad, No_Conformidad, NC_Articulo, Tipo_Nc
from .forms import EntradaArticuloForm, Reporte_CalidadForm, NoConformidadForm, NC_ArticuloForm, NC_Almacen_ArticuloForm, Cierre_NCForm
from user.models import Profile
from smtplib import SMTPException
import json

from datetime import date, datetime
import decimal
import socket
import os
from requisiciones.views import get_image_base64

from io import BytesIO
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import Color, black, white
from reportlab.lib.units import cm
from reportlab.lib import colors



@login_required(login_url='user-login')
def pendientes_recepcion(request):
    usuario = Profile.objects.get(staff__id=request.user.id)


    if usuario.tipo.comprador == True:
        compras = Compra.objects.filter(
            Q(cond_de_pago__nombre ='CREDITO') | Q(pagada = True), 
            recepcion_completa = False, 
            autorizado2= True).order_by('-autorizado_date2')
        for compra in compras:
            articulos_recepcion  = ArticuloComprado.objects.filter(oc=compra, recepcion_completa = False)
            servicios_pendientes = articulos_recepcion.filter(producto__producto__articulos__producto__producto__servicio=True)
            cant_entradas = articulos_recepcion.count()
            cant_servicios = servicios_pendientes.count()

            if cant_entradas == cant_servicios and cant_entradas > 0:
                compra.solo_servicios = True
                compra.save()
        compras = Compra.objects.filter(
            Q(cond_de_pago__nombre ='CREDITO') | Q(pagada = True),  
            Q(solo_servicios=False),
            recepcion_completa = False , 
            autorizado2= True).order_by('-autorizado_date2')
    else:
        compras = Compra.objects.filter(Q(cond_de_pago__nombre ='CREDITO') | Q(pagada = True), solo_servicios= True, entrada_completa = False, autorizado2= True, req__orden__staff = usuario).order_by('-autorizado_date2')

    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs

    #Set up pagination
    p = Paginator(compras, 25)
    page = request.GET.get('page')
    compras_list = p.get_page(page)

    context = {
        'compras':compras,
        'myfilter':myfilter,
        'compras_list':compras_list,
        }

    return render(request, 'entradas/pendientes_recepcion.html', context)

@login_required(login_url='user-login')
def recepcion_servicios(request):
    usuario = Profile.objects.get(staff__id=request.user.id)

    compras = Compra.objects.filter(
        Q(cond_de_pago__nombre ='CREDITO') | Q(pagada = True), 
        recepcion_completa = False,
        solo_servicios= True, 
        autorizado2= True, 
        req__orden__staff = usuario
    ).order_by('-autorizado_date2')


    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs

    #Set up pagination
    p = Paginator(compras, 50)
    page = request.GET.get('page')
    compras_list = p.get_page(page)

    context = {
        'compras':compras,
        'myfilter':myfilter,
        'compras_list':compras_list,
        }

    return render(request, 'entradas/pendientes_recepcion_servicios.html', context)

@login_required(login_url='user-login')
def devolucion_a_proveedor(request):

    articulos = Reporte_Calidad.objects.filter(completo = True, autorizado = False)

    context = {
        'articulos':articulos,
        }


def update_fecha(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        producto_id = data.get('productoId')
        nueva_fecha = data.get('nuevaFecha')

        # Realiza la lógica para actualizar la fecha del producto
        producto = EntradaArticulo.objects.get(id=producto_id)
        producto.fecha_caducidad = nueva_fecha
        producto.save()

        return JsonResponse({'nuevaFecha': nueva_fecha})
    return JsonResponse({'error': 'Método no permitido'}, status=405)

# Create your views here.
@login_required(login_url='user-login')
def pendientes_entrada(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    
    if usuario.tipo.almacen == True:
        # Subconsulta para total_recepcionado
        subquery_total_recepcionado = EntradaArticulo.objects.filter(
            articulo_comprado=OuterRef('articulo_comprado'),
            entrada__oc=OuterRef('entrada__oc'),
            recepcion=True
        ).values('articulo_comprado').annotate(
            total_recepcionado=Sum('cantidad')
        ).values('total_recepcionado')

        # Subconsulta para total_nc
        subquery_total_nc = NC_Articulo.objects.filter(
            articulo_comprado=OuterRef('articulo_comprado'), resuelto=False,
            nc__oc=OuterRef('entrada__oc')
        ).values('articulo_comprado').annotate(
            total_nc=Sum('cantidad')
        ).values('total_nc')

        # Consulta principal
        articulos_recepcionados = EntradaArticulo.objects.filter(
            recepcion=True,
            cantidad__gt=0,
            agotado = False,
            almacenado = False,
            articulo_comprado__producto__producto__articulos__producto__producto__servicio=False
        ).filter(
            # Filtro condicional usando Q
            Q(
                articulo_comprado__producto__producto__articulos__producto__producto__critico__in=[1, 2], 
                calidad=True
            ) | Q(
                ~Q(articulo_comprado__producto__producto__articulos__producto__producto__critico__in=[1, 2])  # Excluye critico en [1, 2]
            )
        ).annotate(
            total_recepcionado=Subquery(subquery_total_recepcionado, output_field=DecimalField(max_digits=14, decimal_places=2)),
            total_nc=Subquery(subquery_total_nc, output_field=DecimalField(max_digits=14, decimal_places=2))
        ).order_by('-id')

    myfilter = EntradaArticuloFilter(request.GET, queryset=articulos_recepcionados)
    articulos_recepcionados = myfilter.qs

    if request.method == "POST" and 'btnExcel' in request.POST:
        return reporte_recepcionados(articulos_recepcionados)

    if request.method == "POST" and 'entrada' in request.POST:
        pk = request.POST.get('entrada_articulo_id')
        entrada_item = EntradaArticulo.objects.get(id = pk)
        compra = Compra.objects.get(id = entrada_item.entrada.oc.id)

        productos_comprados = ArticuloComprado.objects.filter(oc=entrada_item.entrada.oc.id) #Esto son todos los productos de la OC

        producto_comprado = productos_comprados.get(id = entrada_item.articulo_comprado.id) #Este es el producto al que se le está dando entrada
        entrada = Entrada.objects.get(id = entrada_item.entrada.id)

        aggregation = EntradaArticulo.objects.filter(
            articulo_comprado = producto_comprado,
            entrada__completo = True, almacenado = True,
        ).aggregate(
            suma_cantidad = Sum('cantidad'),
            suma_cantidad_por_surtir = Sum('cantidad_por_surtir')
        )
        suma_cantidad = aggregation['suma_cantidad'] or 0   #Este es el resultado de la suma de las entradas

        nc_producto = NC_Articulo.objects.filter(resuelto=False, articulo_comprado = producto_comprado, nc__oc = compra).aggregate(Sum('cantidad')) 
        suma_nc_producto = nc_producto['cantidad__sum'] #Este es el resultado de la suma de las entradas por nc
        if suma_nc_producto is None:
            suma_nc_producto = 0
        suma_cantidad = suma_cantidad + suma_nc_producto
        #error
        pendientes_surtir = aggregation['suma_cantidad_por_surtir'] or 0 #La cantidad por surtir es la cantidad a la que no se le ha dado salida aún
        producto_inv = Inventario.objects.get(producto = producto_comprado.producto.producto.articulos.producto.producto)

        if entrada.oc.req.orden.tipo.tipo == 'resurtimiento': #si es resurtimiento
            try:
                producto_surtir = ArticulosparaSurtir.objects.filter(articulos__producto=producto_comprado.producto.producto.articulos.producto, requisitar=True, articulos__orden__tipo__tipo='normal')
                mismo_producto = ArticulosparaSurtir.objects.get(articulos = producto_comprado.producto.producto.articulos)
                print('producto_surtir:',producto_surtir)
                print('mismo_producto:',mismo_producto)
                # ...
            except ArticulosparaSurtir.MultipleObjectsReturned:
                # Maneja el caso en que se devuelven múltiples objetos
                print("Se encontraron múltiples objetos!")
            except ArticulosparaSurtir.DoesNotExist:
                # Maneja el caso en que no se encuentra ningún objeto
                print("No se encontró ningún objeto!")
        else:
            producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto_comprado.producto.producto.articulos)       
        tolerance = 0.01
        print(entrada_item.cantidad)
        print(entrada_item.cantidad_por_surtir)

        if abs(producto_comprado.cantidad_pendiente - entrada_item.cantidad) > tolerance: #Si la cantidad de las entradas es mayor a la cantidad de la compra se rechaza
            diferencia = entrada_item.cantidad_por_surtir - entrada_item.cantidad
            messages.error(request,f'La cantidad de entradas sobrepasa la cantidad entrada {entrada_item.cantidad} mayor que la comprada {producto_comprado.cantidad_pendiente} = {diferencia}') 
        
            #messages.error(request,f'La cantidad de entradas sobrepasa la cantidad comprada {suma_cantidad} > {entrada_item.cantidad}')
        else:   #En caso de que NO sea un RESURMIENTO
            #producto_comprado.cantidad_pendiente = producto_comprado.cantidad - suma_cantidad
            producto_comprado.cantidad_pendiente = producto_comprado.cantidad - suma_cantidad - entrada_item.cantidad
            #dato = producto_comprado.cantidad_pendiente
            #error
            if producto_inv.producto.servicio == False:     #Se sacan los cálculos de costeo en caso de NO sea un SERVICIO
                monto_inventario = producto_inv.cantidad * producto_inv.price + producto_inv.apartada * producto_inv.price
                cantidad_inventario = producto_inv.cantidad + producto_inv.apartada
                monto_total = monto_inventario + entrada_item.cantidad * producto_comprado.precio_unitario
                nueva_cantidad_inventario =  cantidad_inventario + entrada_item.cantidad
                if cantidad_inventario == 0:
                    precio_unit_promedio = producto_comprado.precio_unitario
                else:    
                    precio_unit_promedio = monto_total/nueva_cantidad_inventario
                producto_inv.price = precio_unit_promedio
            
                #Esta parte determina el comportamiento de todos las solicitudes que se tienen que activar cuando la entrada es de resurtimiento
            if entrada.oc.req.orden.tipo.tipo == 'resurtimiento':
                print('esto es un resurtimiento')
                entrada_item.almacenado = True
                producto_inv.cantidad_entradas = pendientes_surtir + entrada_item.cantidad
                producto_inv.cantidad = producto_inv.cantidad + entrada_item.cantidad 
                producto_inv._change_reason = 'Se modifica el inventario en view: update_entrada. Esto es una entrada para resurtimiento'
                producto_inv.save()
                entrada_item.save()
                if producto_surtir:
                    for producto in producto_surtir:  #producto surtir deben de ser todos los productos que estaban en espera de ser requisitados se itera sobre ellos
                        if entrada_item.agotado == False:
                            if (entrada_item.cantidad_por_surtir - producto.cantidad_requisitar) >= 0:       #si la entrada es mayor que la cantidad por surtir entonces                                                                                     #se evalua si la cantidad que queda de item es suficiente para cubrir el surtimiento
                                mismo_producto.cantidad = mismo_producto.cantidad + producto.cantidad_requisitar  #la cantidad de producto a surtir es igual a la cantidad por surtir mas la que se iba a requisitar (iba, ya no es necesario)
                                producto.cantidad = producto.cantidad + producto.cantidad_requisitar
                                entrada_item.cantidad_por_surtir = entrada_item.cantidad_por_surtir - producto.cantidad_requisitar #la cantidad disponible para surtir es igual a la que cantidad que había disponible menos la cantidad por requisitar (que ya se le sumo al surtir arriba por ello se resta para mantener el balance)
                                producto_inv.cantidad = producto_inv.cantidad - producto.cantidad_requisitar #a la cantidad de inventario también se le resta la cantidad por requisitar ya que originalmente aquí se suma la cantidad total de la entrada
                                producto.cantidad_requisitar = 0 #la cantidad a requisitar es 0 (ya no queda nada más por requisitar)
                                producto.requisitar = False #al no quedar cantidad por requisitar, el requisitar es falso
                                producto.precio = producto_comprado.precio_unitario 
                                producto_inv._change_reason = 'Se modifica el inventario en view: update_entrada. Esto es un apartado de un resurtimiento que estaba por requistarse'
                            else: #si la cantidad de entrada por surtir no es mayor que la necesidad del material que se quiere requisitar, entonces:
                                producto.cantidad = producto.cantidad + entrada_item.cantidad_por_surtir #a la cantidad por surtir se le suma lo que sea que quede en la entrada por surtir
                                producto_inv.cantidad = producto_inv.cantidad - entrada_item.cantidad_por_surtir # a la cantidad del inventario se le resta también lo que sea que quede en la entrada por surtir
                                entrada_item.cantidad_por_surtir = 0  #la cantidad por surtir se agota
                                mismo_producto.cantidad_requisitar = 0
                                mismo_producto.requisitar = False
                                entrada_item.agotado = True           #es decir, es producto se agota
                                producto_inv._change_reason = 'Aqui se acaba el resurtimiento'
                            producto.surtir = True
                            producto.save()
                            mismo_producto.save()
                            producto_inv.save()
                            entrada_item.save()
                            messages.success(request,'Haz agregado exitosamente un producto, desde un resurtimiento')
                else:
                    messages.success(request,'Haz agregado exitosamente un producto, desde un resurtimiento')
            else:
                print('esto no es resurtiminento') ############Creo es aqui
                if producto_surtir.articulos.producto.producto.especialista or producto_surtir.articulos.producto.producto.critico or producto_surtir.articulos.producto.producto.rev_calidad:
                    producto_surtir.surtir = False                           
                    entrada_item.liberado = False
                    archivo_oc = attach_oc_pdf(request, entrada_item.articulo_comprado.oc.id)
                    email = EmailMessage(
                            f'Compra Autorizada {compra.get_folio}',
                            f'Estimado *Inserte nombre de especialista*,\n Estás recibiendo este correo porque se ha recibido en almacén el producto código:{producto_surtir.articulos.producto.producto.codigo} descripción:{producto_surtir.articulos.producto.producto.nombre} el cual requiere la liberación de calidad\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                            'savia@vordtec.com',
                            ['ulises_huesc@hotmail.com'], #agregar correo de calidad
                            )
                    email.attach(f'OC_folio:{entrada_item.articulo_comprado.oc.folio}.pdf',archivo_oc,'application/pdf')
                    #email.send()
                if producto_surtir.articulos.producto.producto.activo == True:
                    producto_inv.cantidad += entrada_item.cantidad #Se suma al inventario como si fuera un resurtimiento ya que Activos lo asigna
                    producto_inv.cantidad_entradas += entrada_item.cantidad
                    producto_surtir.cantidad = 0                     #Al producto disponible para surtir
                    producto_surtir.cantidad_requisitar = 0   #Al producto pendiente por requisitar
                    producto_surtir.seleccionado = False
                    producto_surtir.surtir = False
                    producto_surtir.requisitar = False
                    producto_inv._change_reason = 'Se modifica el inventario en view: update_entrada. Esto es una entrada para solicitud normal con activo'
                    entrada.entrada_date = date.today()
                    entrada.entrada_hora = datetime.now().time()
                    entrada.save()
                    entrada_item.almacenado = True #Esto esta bien
                else:   
                    #Este código es el que tiene que suceder para hacer la entrada a almacén
                    producto_inv.cantidad_entradas = pendientes_surtir
                    producto_inv.cantidad_apartada = producto_inv.apartada_entradas
                    producto_surtir.cantidad = producto_surtir.cantidad + entrada_item.cantidad                       #Al producto disponible para surtir se le suma lo que entra
                    producto_surtir.cantidad_requisitar = producto_surtir.cantidad_requisitar - entrada_item.cantidad   #Al producto pendiente por requisitar se le resta lo que entra
                    producto_surtir.seleccionado = False
                    producto_surtir.surtir = True
                    producto_inv._change_reason = 'Se modifica el inventario en view: update_entrada. Esto es una entrada para solicitud normal'
                    entrada.entrada_date = date.today()
                    entrada.entrada_hora = datetime.now().time()
                    
                    entrada.save()
                    entrada_item.almacenado = True #Esto esta bien


                #Si cantidad de entradas es menor a cantidad total del producto
                #if suma_cantidad < producto_comprado.cantidad:
                    #producto_comprado.recepcion_completa = False
                    #producto_comprado.seleccionado = False
                    #compra.recepcion_completa = False
                #else:
                #    producto_comprado.entrada_completa = True
                
                if producto_comprado.cantidad_pendiente <= 0:
                    producto_comprado.entrada_completa = True
                    producto_comprado.recepcion_completa = True
                messages.success(request,'Haz agregado exitosamente un producto')
                entrada_item.save()
                producto_comprado.save()
                producto_inv.save()
                entrada.save()
                producto_surtir.save()
                #Se guardan todas las bases de datos
          
                #cantidad_entradas = entradas_producto.cantidad - entradas_producto.cantidad_por_surtir
            #messages.success(request,'Haz agregado exitosamente un producto')
            if producto_comprado.producto.producto.articulos.producto.producto.servicio == True:
                salida, created = Salidas.objects.get_or_create(producto = producto_surtir, salida_firmada=True, cantidad = entrada_item.cantidad)
                salida.comentario = 'Esta salida es un  servicio por lo tanto no pasa por almacén y no existe registro de la salida del mismo'
                producto_surtir.surtir = False
                salida.save()
            num_art_entregados = ArticuloComprado.objects.filter(oc=entrada_item.entrada.oc.id, entrada_completa=True).count()
            num_art_comprados = productos_comprados.count()
            #num_art_entregados = productos_comprados.filter(entrada_completa=True).count()
            if num_art_comprados == num_art_entregados:
                compra.entrada_completa = True
                compra.recepcion_completa = True
            compra.save()


            producto_comprado.save()
            producto_inv.save()
            static_path = settings.STATIC_ROOT
            img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
            img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
            image_base64 = get_image_base64(img_path)
            logo_v_base64 = get_image_base64(img_path2)
            html_message = f"""
            <html>
                <head>
                    <meta charset="UTF-8">
                </head>
                <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                    <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                        <tr>
                            <td align="center">
                                <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                    <tr>
                                        <td align="center">
                                            <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                        </td>
                                    </tr>
                                    <tr>
                                        <td style="padding: 20px;">
                                            <p style="font-size: 18px; text-align: justify;">
                                                Estimado {compra.req.orden.staff.staff.first_name} {compra.req.orden.staff.staff.last_name},
                                            </p>
                                            <p style="font-size: 16px; text-align: justify;">
                                                Estás recibiendo este correo porque tu OC <strong>{compra.get_folio}</strong> | RQ: <strong>{compra.req.folio}</strong> |Sol: <strong>{compra.req.orden.folio}</strong> ha sido recibida en el módulo de entrada por el 
                                                <strong>Almacen.</strong>
                                            </p>
                                            <p style="font-size: 16px; text-align: justify;">
                                                El siguiente paso del sistema: 'Salida' Puede ir por su material.
                                            </p>
                                            <p style="text-align: center; margin: 20px 0;">
                                                <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                            </p>
                                            <p style="font-size: 14px; color: #999; text-align: justify;">
                                                Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                            </p>
                                        </td>
                                    </tr>
                                </table>
                            </td>
                        </tr>
                    </table>
                </body>
            </html>
            """
            try:
                email = EmailMessage(
                    f'OC Autorizada {compra.get_folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                    body=html_message,
                    from_email = settings.DEFAULT_FROM_EMAIL,
                    to= [compra.req.orden.staff.staff.email,compra.creada_por.staff.email,'ulises_huesc@hotmail.com'],
                    headers={'Content-Type': 'text/html'}
                    )
                email.content_subtype = "html " # Importante para que se interprete como HTML
                email.send()
                messages.success(request, f'{usuario.staff.first_name} has autorizado la solicitud {compra.get_folio}')
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'{usuario.staff.first_name}, Has generado Recepción correctamente pero el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.warning(request, error_message)
            messages.success(request,'Haz agregado exitosamente un producto')
        return redirect('pendientes-entrada')
        



    #Set up pagination
    p = Paginator(articulos_recepcionados, 50)
    page = request.GET.get('page')
    articulos_recepcionados_list = p.get_page(page)
    

    context = {
        'articulos_recepcionados':articulos_recepcionados,
        'myfilter':myfilter,
        'articulos_recepcionados_list':articulos_recepcionados_list,
        }

    return render(request, 'entradas/pendientes_entrada.html', context)

@login_required(login_url='user-login')
def pendientes_calidad(request):
    articulos_entrada = EntradaArticulo.objects.filter(
        Q(articulo_comprado__producto__producto__articulos__producto__producto__especialista=True) |
        Q(articulo_comprado__producto__producto__articulos__producto__producto__critico=True) |
        Q(articulo_comprado__producto__producto__articulos__producto__producto__rev_calidad=True),
        liberado=False
        )
    
    print(articulos_entrada)
  
    context = {
        'articulos_entrada':articulos_entrada,
        }

    return render(request, 'entradas/pendientes_calidad.html', context)


@login_required(login_url='user-login')
def devolucion_a_proveedor(request):

    articulos = Reporte_Calidad.objects.filter(completo = True, autorizado = False)

    context = {
        'articulos':articulos,
        }

    return render(request, 'entradas/devolucion_a_proveedor.html', context)


#Esta es la vista que genera la recepción
@login_required(login_url='user-login')
def articulos_recepcion(request, pk):

    usuario = Profile.objects.get(staff=request.user.id)
    if usuario.tipo.compras == True:
        # Subconsulta para la suma de NC en artículos
        subquery_suma_nc_articulos = NC_Articulo.objects.filter(resuelto = False,
            articulo_comprado=OuterRef('pk')
        ).values('articulo_comprado').annotate(
            total_nc=Sum('cantidad')
        ).values('total_nc')

        # Subconsulta para la suma de entradas en artículos
        subquery_suma_entrada_articulos = EntradaArticulo.objects.filter(
            articulo_comprado=OuterRef('pk')
        ).values('articulo_comprado').annotate(
            total_entrada=Sum('cantidad')
        ).values('total_entrada')

        # Consulta principal
        articulos = ArticuloComprado.objects.filter(
            oc=pk,
            recepcion_completa=False,
            entrada_completa=False,
            seleccionado=False,
            producto__producto__articulos__producto__producto__servicio=False
        ).annotate(
            suma_nc_articulos=Subquery(subquery_suma_nc_articulos, output_field=DecimalField(max_digits=14, decimal_places=2)),
            suma_entrada_articulos=Subquery(subquery_suma_entrada_articulos, output_field=DecimalField(max_digits=14, decimal_places=2))
        )

    compra = Compra.objects.get(id=pk)

    entrada, created = Entrada.objects.get_or_create(oc=compra, almacenista= usuario, completo = False)
    articulos_entrada = EntradaArticulo.objects.filter(entrada = entrada)
    form = EntradaArticuloForm()

    for articulo in articulos:
        if articulo.cantidad_pendiente is None:
            articulo.cantidad_pendiente = articulo.cantidad

    if request.method == 'POST' and 'entrada' in request.POST:
        entrada.completo = True              
        entrada.entrada_date = date.today()
        entrada.entrada_hora = datetime.now().time()
        articulos_comprados = ArticuloComprado.objects.filter(oc=compra) #Se buscan los articulos comprados de la OC
        num_art_comprados = articulos_comprados.count() #Aqui sacamos el numero de articulos pedidos  
                
        for articulo in articulos_entrada:
            articulo_comprado = articulos_comprados.get(id = articulo.articulo_comprado.id)
            #Valen como entradas y recepciones
            nc_producto = NC_Articulo.objects.filter(resuelto = False, articulo_comprado = articulo_comprado, nc__oc = compra).aggregate(Sum('cantidad')) 
            #Valen como entradas
            entradas_producto = EntradaArticulo.objects.filter(articulo_comprado = articulo_comprado, entrada__oc = compra, almacenado = True).aggregate(Sum('cantidad'))#Busca la cantidad de entradas para ese producto añadiendo la cantidad para cada dato
            #Valen recepcionados
            recepcion_producto = EntradaArticulo.objects.filter(articulo_comprado = articulo_comprado, entrada__oc = compra, recepcion = True).aggregate(Sum('cantidad'))
            #Cantidad total de cada entrada nc y recepcion
            suma_entradas = entradas_producto['cantidad__sum']#Saca la suma de la cantidad total de las entradas de ese producto
            suma_nc_producto = nc_producto['cantidad__sum']
            suma_recepcion = recepcion_producto['cantidad__sum']
            #Formateo
            if suma_entradas is None:
                suma_entradas = 0
            if suma_nc_producto is None:
                suma_nc_producto = 0
            if suma_recepcion is None:
                suma_recepcion=0
            if articulo_comprado.cantidad_pendiente == None: 
                articulo_comprado.cantidad_pendiente = articulo_comprado.cantidad
            #Producto pendientes entradas 
            entrada_pendientes = articulo_comprado.cantidad - suma_entradas - suma_nc_producto
            articulo_comprado.cantidad_pendiente = entrada_pendientes
            #Producto pendientes recepcion
            recepcion_pendientes = articulo_comprado.cantidad - suma_recepcion - suma_nc_producto

            if recepcion_pendientes == 0:
                articulo_comprado.recepcion_completa = True
            if entrada_pendientes == 0:
                articulo_comprado.entrada_completa = True

            articulo_comprado.seleccionado = False
            articulo_comprado.save() #guarda el articulo comprado

        num_art_entregados = ArticuloComprado.objects.filter(oc=compra, entrada_completa=True).count() #Articulos completos
        articulos_recepcionados = ArticuloComprado.objects.filter(oc=compra,recepcion_completa=True)

        detalle = []
        for art in articulos_recepcionados:
            detalle.append({
                "codigo": art.producto.producto.articulos.producto.producto.codigo,
                "nombre": art.producto.producto.articulos.producto.producto.nombre,
                "cantidad": art.cantidad,   # <- ajusta si tu campo se llama distinto
            })

        
        filas_html = "".join(
            f"""
            <tr>
                <td style="border:1px solid #ddd; padding:8px;">{d['codigo']}</td>
                <td style="border:1px solid #ddd; padding:8px;">{d['nombre']}</td>
                <td style="border:1px solid #ddd; padding:8px; text-align:right;">{d['cantidad']}</td>
            </tr>
            """
            for d in detalle
        )

        num_art_recepcionados = articulos_recepcionados.count()
        if num_art_comprados == num_art_recepcionados:
            compra.recepcion_completa = True #Define la OC como recepcion completa
        if num_art_comprados == num_art_entregados: #Concuerda con el numero de pedidos
            compra.entrada_completa = True #Define en la OC entrada completa si numero de articulos entragados completos concuerda con los pedidos
        entrada.save()
        compra.save()

        static_path = settings.STATIC_ROOT
        img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
        img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
        image_base64 = get_image_base64(img_path)
        logo_v_base64 = get_image_base64(img_path2)
        html_message = f"""
        <html>
            <head>
                <meta charset="UTF-8">
            </head>
            <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                    <tr>
                        <td align="center">
                            <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                <tr>
                                    <td align="center">
                                        <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 20px;">
                                        <p style="font-size: 18px; text-align: justify;">
                                            Estimado {compra.req.orden.staff.staff.first_name} {compra.req.orden.staff.staff.last_name},
                                        </p>
                                        <p style="font-size: 16px; text-align: justify;">
                                            Estás recibiendo este correo porque tu OC <strong>{compra.get_folio}</strong> | RQ: <strong>{compra.req.folio}</strong> |Sol: <strong>{compra.req.orden.folio}</strong> ha sido recibida en el módulo de recepción por el 
                                            <strong>Departamento de Compras.</strong>
                                        </p>
                                        <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse; font-size:14px;">
                                            <thead>
                                                <tr style="background:#0B2D4A; color:#fff;">
                                                    <th style="border:1px solid #ddd; padding:8px; text-align:left;">Código</th>
                                                    <th style="border:1px solid #ddd; padding:8px; text-align:left;">Producto</th>
                                                    <th style="border:1px solid #ddd; padding:8px; text-align:right;">Cantidad</th>
                                                </tr>
                                            </thead>
                                            <tbody>
                                                {filas_html}
                                            </tbody>
                                        </table>
                                        <p style="font-size: 16px; text-align: justify;">
                                            El siguiente paso del sistema: Entrada por parte del Almacén.
                                        </p>
                                        <p style="text-align: center; margin: 20px 0;">
                                            <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                        </p>
                                        <p style="font-size: 14px; color: #999; text-align: justify;">
                                            Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                        </p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                </table>
            </body>
        </html>
        """
        # 1) Obtener correos de almacenistas
        almacenistas = Profile.objects.filter(tipo__almacenista=True, activo=True)

        correos_almacenistas = list(
            almacenistas.values_list('staff__email', flat=True)
        )

        # 2) Correos fijos
        correos_fijos = [
            compra.req.orden.staff.staff.email,
            compra.creada_por.staff.email,
            'ulises.huesca@grupovordcab.com',
        ]

        # 3) Unir todo
        destinatarios = correos_almacenistas + correos_fijos

        # (opcional pero recomendado) eliminar duplicados
        destinatarios = list(set(destinatarios))
        try:
            print(destinatarios)
            email = EmailMessage(
                f'Información Recepción| SAVIA | OC Autorizada {compra.get_folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                body=html_message,
                from_email = settings.DEFAULT_FROM_EMAIL,
                to= destinatarios,
                headers={'Content-Type': 'text/html'}
                )
            email.content_subtype = "html " # Importante para que se interprete como HTML
            email.send()
            messages.success(request, f'{usuario.staff.first_name} has autorizado la solicitud {compra.get_folio}')
        except (BadHeaderError, SMTPException, socket.gaierror) as e:
            error_message = f'{usuario.staff.first_name}, Has generado Recepción correctamente pero el correo de notificación no ha sido enviado debido a un error: {e}'
            messages.warning(request, error_message)
        messages.success(request, f'La entrada-recepcion {entrada.id} se ha realizado con éxito')
        return redirect('pendientes-recepcion')

    context = {
        'articulos':articulos,
        'entrada':entrada,
        'compra':compra,
        'form':form,
        'articulos_entrada':articulos_entrada,
        }

    return render(request, 'entradas/articulos_recepcion.html', context)

@login_required(login_url='user-login')
def articulos_recepcion_servicios(request, pk):

    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id=pk)
    articulos = ArticuloComprado.objects.filter(oc=compra, entrada_completa = False, recepcion_completa = False, seleccionado = False, producto__producto__articulos__producto__producto__servicio = True)


    entrada, created = Entrada.objects.get_or_create(oc=compra, almacenista= usuario, completo = False)
    articulos_entrada = EntradaArticulo.objects.filter(entrada = entrada)
    form = EntradaArticuloForm()
    productos_html = '<ul>'
    for articulo in articulos:
        if articulo.cantidad_pendiente == None:
            articulo.cantidad_pendiente = articulo.cantidad


    if request.method == 'POST' and 'entrada' in request.POST:
        entrada.completo = True              
        entrada.entrada_date = date.today()
        entrada.entrada_hora = datetime.now().time()
        articulos_comprados = ArticuloComprado.objects.filter(oc=pk)
        num_art_comprados = articulos_comprados.count()        
        

        for articulo in articulos_entrada:
            articulo_compra = articulos_comprados.get(id = articulo.articulo_comprado.id)
            aggregation = EntradaArticulo.objects.filter(
                articulo_comprado = articulo_compra,
                entrada__completo = True
            ).aggregate(
                suma_cantidad = Sum('cantidad'),
                suma_cantidad_por_surtir = Sum('cantidad_por_surtir')
            )
            suma_cantidad = aggregation['suma_cantidad'] or 0
            if suma_cantidad >=  articulo_compra.cantidad:
                articulo_compra.seleccionado = False
                articulo_compra.save()

            productos_html += f'<li>{articulo.articulo_comprado.producto.producto.articulos.producto.producto.nombre}: {articulo.cantidad}</li>'
        articulos_recepcionados = articulos_comprados.filter(recepcion_completa = True)
        num_art_recepcionados = articulos_recepcionados.count()
        if num_art_recepcionados >= num_art_comprados:
            compra.recepcion_completa = True
        
        entrada.save()
        compra.save()
        static_path = settings.STATIC_ROOT
        img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
        img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
        productos_html += '</ul>'
        image_base64 = get_image_base64(img_path)
        logo_v_base64 = get_image_base64(img_path2)
        # Crear el mensaje HTML
        html_message = f"""
        <html>
            <head>
                <meta charset="UTF-8">
            </head>
            <body>
                <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                <p>Estimado {compra.creada_por.staff.first_name} {compra.creada_por.staff.last_name},</p>
                <p>Estás recibiendo este correo porque se ha realizado la entrada para servicios: {entrada.id} de la OC: {compra.id}.</p>
                <p>Con los productos siguientes:</p>
                {productos_html}
                <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
            </body>
        </html>
        """
        try:
            email = EmailMessage(
                f'Entrada Servicio: {entrada.id} OC: {compra.id}',
                body=html_message,
                from_email= settings.DEFAULT_FROM_EMAIL,
                to=[compra.creada_por.staff.email],
                headers={'Content-Type': 'text/html'}
                )
            email.content_subtype = "html " # Importante para que se interprete como HTML
            email.send()
            messages.success(request, f'La recepcion del servicio {entrada.id} se ha realizado con éxito')
            return redirect('recepcion-servicios')
        except (BadHeaderError, SMTPException, socket.gaierror) as e:
            error_message = f'La recepcion del servicio {entrada.id} ha sido creada, pero el correo no ha sido enviado debido a un error: {e}'
            messages.success(request, error_message)
            return redirect('recepcion-servicios')

    context = {
        'articulos':articulos,
        'entrada':entrada,
        'compra':compra,
        'form':form,
        'articulos_entrada':articulos_entrada,
        }

    return render(request, 'entradas/articulos_recepcion.html', context)


@login_required(login_url='user-login')
def articulos_entrada(request, pk):

    usuario = Profile.objects.get(staff=request.user.id)
    if usuario.tipo.almacen == True:
        articulos = ArticuloComprado.objects.filter(oc=pk, entrada_completa = False, seleccionado = False, producto__producto__articulos__producto__producto__servicio = False)
       
    else:
        articulos = ArticuloComprado.objects.filter(oc=pk, entrada_completa = False, seleccionado = False, producto__producto__articulos__producto__producto__servicio = True)


    compra = Compra.objects.get(id=pk)
    conteo_de_articulos = articulos.count()
    entrada, created = Entrada.objects.get_or_create(oc=compra, almacenista= usuario, completo = False)
    articulos_entrada = EntradaArticulo.objects.filter(entrada = entrada)
    form = EntradaArticuloForm()

    for articulo in articulos:
        if articulo.cantidad_pendiente == None:
            articulo.cantidad_pendiente = articulo.cantidad


    if request.method == 'POST' and 'entrada' in request.POST:
        #entrada.completo = True                #Lo comenté porque esto ya está sucediendo en la recepción
        entrada.entrada_date = date.today()                                    #Se actualiza la fecha de la entrada
        entrada.entrada_hora = datetime.now().time()                           #Se actualiza la hora de la entrada
        entrada.almacenista = usuario                                          #Se actualiza almacenista que hizo la entrada
        articulos_comprados = ArticuloComprado.objects.filter(oc=pk)           #Traigo todos los articulos comprados 
        num_art_comprados = articulos_comprados.count()                         #Hago un conteo de los artículos comprados
        articulos_entregados = articulos_comprados.filter(entrada_completa=True) #Traigo todo los articulos que tienen la entrada completa de esa entrada
        articulos_seleccionados = articulos_entregados.filter(seleccionado = True) #De todos los entregados determino cuales están seleccionados
        num_art_entregados = articulos_entregados.count()        #cuento los articulos que tienen la entrada completa
        for elemento in articulos_seleccionados:                  # Con este ciclo les quito el seleccionado
            elemento.seleccionado = False
            elemento.save()

        for articulo in articulos_entrada:                        #Para cada de los articulos en la entrada
            producto_surtir = ArticulosparaSurtir.objects.get(articulos = articulo.articulo_comprado.producto.producto.articulos)
            producto_surtir.seleccionado = False                    #Se deselecciona el artículo para surtir relacionado
            if producto_surtir.articulos.producto.producto.especialista or producto_surtir.articulos.producto.producto.critico or producto_surtir.articulos.producto.producto.rev_calidad:
                producto_surtir.surtir = False                           
                articulo.liberado = False
                archivo_oc = attach_oc_pdf(request, articulo.articulo_comprado.oc.id)
                email = EmailMessage(
                        f'Compra Autorizada {compra.get_folio}',
                        f'Estimado *Inserte nombre de especialista*,\n Estás recibiendo este correo porque se ha recibido en almacén el producto código:{producto_surtir.articulos.producto.producto.codigo} descripción:{producto_surtir.articulos.producto.producto.nombre} el cual requiere la liberación de calidad\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                        'savia@vordtec.com',
                        ['ulises_huesc@hotmail.com'],
                        )
                email.attach(f'OC_folio:{articulo.articulo_comprado.oc.folio}.pdf',archivo_oc,'application/pdf')
                email.send()
            if entrada.oc.req.orden.tipo.tipo == 'resurtimiento':
                #Estas son todas las solicitudes pendientes por surtir que se podrían surtir con el resurtimiento
                productos_pendientes_surtir = ArticulosparaSurtir.objects.filter(
                    articulos__producto__producto = articulo.articulo_comprado.producto.producto.articulos.producto.producto,
                    salida = False, 
                    articulos__orden__tipo__tipo = 'normal',
                    cantidad_requisitar__gt=0
                    )
                inv_de_producto = Inventario.objects.get(producto = producto_surtir.articulos.producto.producto)
                for producto in productos_pendientes_surtir:    #Recorremos todas las solicitudes pendientes por surtir una por una
                    if producto_surtir.cantidad > 0:
                        inv_de_producto.cantidad = inv_de_producto.cantidad - producto.cantidad
                        if producto.cantidad_requisitar <= producto_surtir.cantidad:
                            producto_surtir.cantidad = producto_surtir.cantidad - producto.cantidad_requisitar
                            producto.cantidad = producto.cantidad + producto.cantidad_requisitar
                            producto.cantidad_requisitar = 0
                            producto.requisitar = False
                        else:
                            producto.cantidad_requisitar = producto.cantidad_requisitar - producto_surtir.cantidad
                            producto.cantidad = producto.cantidad + producto_surtir.cantidad
                            producto_surtir.cantidad = 0
                            producto.surtir = True
                            producto.save()
                            producto_surtir.save()
                            inv_de_producto.save()
                            solicitud = Order.objects.get(id = producto_surtir.articulos.orden.id)
                            productos_orden = ArticulosparaSurtir.objects.filter(articulos__orden = solicitud, requisitar=False).count()
                            if productos_orden == 0:
                                solicitud.requisitar = False
                                solicitud.save()
            if entrada.oc.req.orden.tipo.tipo == 'normal':
                if articulo.articulo_comprado.producto.producto.articulos.producto.producto.servicio == True:
                    producto_surtir.surtir = False
                else:
                    producto_surtir.surtir = True        #Si NO es un SERVICIO es surtir cambia a True
            producto_surtir.save()
        for articulo in articulos_comprados:
            #entradas_producto = EntradaArticulo.objects.filter(articulo_comprado = articulo, entrada__oc = articulo.oc, entrada__completo = True).aggregate(Sum('cantidad'))
            #suma_entradas = entradas_producto['cantidad__sum']
            if articulo.cantidad_pendiente == 0:  #Si la cantidad de la compra es igual a la cantida entonces la entrada está completamente entregada
                articulo.entrada_completa = True
            articulo.seleccionado = False
            articulo.save()
        #Se compara los articulos comprados contra los articulos que han entrado y que están totalmente entregados
        #En el bucle de arriba se redefine si la entrada de un articulo está completa o no por lo tanto debería de volver a calcular los artículos completos
        num_art_entregados = articulos_comprados.filter(entrada_completa=True).count()
        if num_art_comprados == num_art_entregados:
            compra.entrada_completa = True
        compra.save()
        entrada.save()
        messages.success(request, f'La entrada {entrada.id} se ha realizado con éxito')
        return redirect('pendientes_entrada')

    context = {
        'articulos':articulos,
        'entrada':entrada,
        'compra':compra,
        'form':form,
        'articulos_entrada':articulos_entrada,
        }

    return render(request, 'entradas/articulos_entradas.html', context)


#Esta es la vista que actualiza la cantidad en las entradas
def update_cantidad(request):
    data= json.loads(request.body)
    pk = data["solicitud_id"]
    dato = data["dato"]
    entrada = EntradaArticulo.objects.get(id=pk)
    articulo_comprado = ArticuloComprado.objects.get(id = entrada.articulo_comprado.id)
    anterior = entrada.cantidad
    entrada.cantidad = dato #Cantidad
    entrada.cantidad_por_surtir = dato
    oc = Compra.objects.get(id = articulo_comprado.oc.id)
    dif = dato - anterior  
            #Logica para establecer la recepción en el producto y la oc por si se regresa una cantidad a una OC ya recepcionada
    #if not producto_comprado.cantidad == entrada.cantidad:
    #    producto_comprado.recepcion_completa = False
    #    producto_comprado.seleccionado = False
    if anterior > dato:
        articulo_comprado.recepcion_completa = False
        oc.recepcion_completa = False
    elif anterior < dato:
        # Calcular el total_recepcionado
        total_recepcionado = EntradaArticulo.objects.filter(
            articulo_comprado=entrada.articulo_comprado,
            entrada__oc=entrada.entrada.oc,
            recepcion=True
        ).aggregate(total_recepcionado=Sum('cantidad'))['total_recepcionado'] or 0.00

        # Calcular el total_nc
        total_nc = NC_Articulo.objects.filter(
            articulo_comprado=entrada.articulo_comprado,
            nc__oc=entrada.entrada.oc, resuelto = False,
        ).aggregate(total_nc=Sum('cantidad'))['total_nc'] or 0.00
        total_recepcionado = decimal.Decimal(total_recepcionado)
        total_nc = decimal.Decimal(total_nc)
        if (total_recepcionado + total_nc + dif) == entrada.articulo_comprado.cantidad:
            articulo_comprado.recepcion_completa = True
    articulo_comprado.save()
    entrada.save()
    oc.save()
    # Construye un objeto de respuesta que incluya el dato y el tipo.
    response_data = {
        'dato': dato,
    }

    return JsonResponse(response_data, safe=False)
    

def update_recepcion_articulos(request):
    data = json.loads(request.body)
    cantidad = decimal.Decimal(data["cantidad_ingresada"])
    action = data["action"]
    producto_id = int(data["producto"])
    pk = int(data["entrada_id"])
    #referencia = data["referencia"]
    folio = ''

    producto_comprado = ArticuloComprado.objects.get(id = producto_id)
    productos_ids = ArticuloComprado.objects.filter(oc=producto_comprado.oc).order_by('id').values_list('id', flat=True)
    
    entrada = Entrada.objects.get(id = pk, completo = False)
    aggregation = EntradaArticulo.objects.filter(
        articulo_comprado = producto_comprado,
        entrada__completo = True
    ).aggregate(
        suma_cantidad = Sum('cantidad'),
        suma_cantidad_por_surtir = Sum('cantidad_por_surtir')
    )
    nc_producto = NC_Articulo.objects.filter(resuelto = False, articulo_comprado = producto_comprado, nc__oc = producto_comprado.oc, nc__completo = True).aggregate(Sum('cantidad'))
    suma_nc_producto = nc_producto['cantidad__sum'] or 0

    suma_cantidad = aggregation['suma_cantidad'] or 0
    #pendientes_surtir = aggregation['suma_cantidad_por_surtir'] or 0
    entrada_item, created = EntradaArticulo.objects.get_or_create(entrada = entrada, articulo_comprado = producto_comprado)

    if action == "add":
        total_entradas = suma_cantidad + cantidad + suma_nc_producto
        if producto_comprado.producto.producto.articulos.producto.producto.critico:
            if producto_comprado.producto.producto.articulos.producto.producto.critico.nombre == 'Crítico':
                try:
                    numero_partida = list(productos_ids).index(producto_comprado.id) + 1  # Agregar 1 para comenzar desde 1
                except ValueError:
                    numero_partida = None  # Manejar el caso donde el producto no esté en la lista
                folio = f"OC{producto_comprado.oc.id}-{numero_partida}-{total_entradas}"
        tolerance = decimal.Decimal('0.01')
        if total_entradas > (producto_comprado.cantidad + tolerance):
            messages.error(request, f'La cantidad recibida sobrepasa la cantidad comprada, Ya Ingresado: {suma_cantidad + suma_nc_producto} Ingresando: {cantidad} Comprado: {producto_comprado.cantidad}')
        else:
            entrada_item.cantidad = cantidad               #Se define por primera vez la variable cantidad de la entrada del producto
            entrada_item.cantidad_por_surtir = cantidad    #Se define por primera vez la variable cantidad_por_surtir de la entrada del producto
            entrada_item.referencia = folio        
            entrada_item.recepcion = True                  #Se define como recepcionado
            entrada_item.fecha_recepcion = datetime.now()  #Se captura la fecha de recepción
            entrada_item.save()                            #Se guarda la entrada
            total_entradas = suma_cantidad + entrada_item.cantidad      #Se determina el total de las entradas que puedan existir de ese mismo producto
            print(entrada_item.cantidad)                     
            print(producto_comprado.cantidad_pendiente)
            if producto_comprado.cantidad_pendiente is None:         #Se determina la cantidad pendiente, 
                producto_comprado.cantidad_pendiente = producto_comprado.cantidad         #Si no existe, la cantidad de la OC se convierte en el producto pendiente 
            messages.success(request,'Haz agregado exitosamente un producto')
            producto_comprado.seleccionado = True #Creé una variable booleana temporal para quitarlo del seleccionable
            if producto_comprado.cantidad == total_entradas: #Solo cuando el total de las entradas es igual a la cantidad comprada
                producto_comprado.recepcion_completa = True   #la recepción está completa
            producto_comprado.save()
            entrada_item.save()
            
    
    elif action == "remove":
        messages.success(request,'Has eliminado el artículo con éxito')
        producto_comprado.seleccionado = False
        producto_comprado.recepcion_completa = False
        producto_comprado.save()
        entrada_item.delete()
    mensaje ='Item was ' + action
    return JsonResponse(mensaje, safe=False)



    #elif action == "remove":
        #if producto_inv.producto.servicio == False:
        #    monto_total = monto_total - (entrada_item.cantidad * producto_comprado.precio_unitario)
        #else:
        #    monto_total = 0
        #if monto_total == 0:
        #    producto_inv.price = 0
        #else:
        #    producto_inv.price = monto_total/cantidad_inventario
        #cantidad_total = cantidad_inventario - entrada_item.cantidad
        #if entrada.oc.req.orden.tipo.tipo == 'resurtimiento':
            #producto_surtir.cantidad = producto_surtir.cantidad - entrada_item.cantidad
            #producto_surtir.cantidad_requisitar = producto_surtir.cantidad_requisitar + entrada_item.cantidad
            #producto_inv.cantidad = producto_inv.cantidad - entrada_item.cantidad
        #    producto_surtir.requisitar = True
            
        #    if producto_surtir.cantidad > entrada_item.cantidad:
        #        producto_surtir.cantidad = producto_surtir.cantidad - entrada_item.cantidad
        #    if producto_surtir.cantidad <= entrada_item.cantidad:
        #        producto_surtir.cantidad_requisitar = producto_surtir.cantidad
        #        producto_surtir.cantidad = 0
        #        producto_inv.cantidad = producto_inv.cantidad - entrada_item.cantidad + producto_surtir.cantidad
        #        producto_inv.cantidad_apartada = producto_inv.cantidad_apartada - producto_surtir.cantidad_requisitar
        #    producto_surtir.save()
        #else:
            #producto_inv.cantidad_apartada = producto_inv.cantidad_apartada - entrada_item.cantidad
        #    producto_surtir.cantidad_requisitar = producto_surtir.cantidad_requisitar + entrada_item.cantidad
        #    producto_surtir.cantidad = producto_surtir.cantidad - entrada_item.cantidad
        #    if producto_surtir == 0:
        #        producto_surtir.surtir = False
        #        producto_surtir.precio = 0
        #    producto_surtir.save()
        #producto_inv._change_reason = 'Se está borrando una entrada. view: update_entrada'
        #producto_inv.cantidad_entradas = producto_inv.cantidad_entradas - entrada_item.cantidad
        #if producto_comprado.cantidad_pendiente == None:
        #    producto_comprado.cantidad_pendiente = 0
        #producto_comprado.cantidad_pendiente = producto_comprado.cantidad_pendiente + entrada_item.cantidad
        #producto_comprado.entrada_completa = False
        #producto_comprado.seleccionado = False
        #messages.success(request,'Has eliminado el artículo con éxito')
        #Se borra el elemento de las entradas
        #Guardado de bases de datos
        #entrada_item.save()
        #producto_inv.save()
        #producto_comprado.save()
        #entrada_item.delete()
    #mensaje ='Item was ' + action
    #return JsonResponse(mensaje, safe=False)


def reporte_calidad(request, pk):
    perfil = Profile.objects.get(staff__id = request.user.id)
    articulo_entrada = EntradaArticulo.objects.get(id = pk, liberado = False)
    producto_calidad = Producto_Calidad.objects.get(producto = articulo_entrada.articulo_comprado.producto.producto.articulos.producto.producto)
    form = Reporte_CalidadForm()
    articulos_reportes = Reporte_Calidad.objects.filter(articulo = articulo_entrada, completo = True)
    reporte_actual, created = Reporte_Calidad.objects.get_or_create(articulo = articulo_entrada, completo = False)
    sum_articulos_reportes = 0

    for item in articulos_reportes:
        sum_articulos_reportes = item.cantidad + sum_articulos_reportes

    restantes_liberacion = articulo_entrada.cantidad - sum_articulos_reportes


    if request.method =='POST':
        form = Reporte_CalidadForm(request.POST, instance = reporte_actual)
        if decimal.Decimal(request.POST['cantidad']) <=  restantes_liberacion:
            if not request.POST['autorizado'] == None:
                if form.is_valid():
                    item = form.save()
                    item.articulo = articulo_entrada
                    item.reporte_date = date.today()
                    item.reporte_hora = datetime.now().time()
                    producto_surtir = ArticulosparaSurtir.objects.get(articulos = articulo_entrada.articulo_comprado.producto.producto.articulos)
                    articulos_restantes = articulo_entrada.cantidad - item.cantidad - sum_articulos_reportes
                    if item.autorizado == True:
                        if articulos_restantes == 0:
                            articulo_entrada.liberado = True
                        producto_surtir.cantidad = producto_surtir.cantidad + item.cantidad
                        producto_surtir.surtir = True
                        producto_surtir.save()
                    if item.autorizado == False:
                        if articulos_restantes == 0:
                            articulo_entrada.liberado = True
                    articulo_entrada.save()
                    item.completo = True
                    item.save()
                    messages.success(request, 'Has generado exitosamente tu reporte')
                    return HttpResponse(status=204)
            else:
                messages.error(request, 'Debes elegir un modo de liberación')
        else:
            messages.error(request, 'La cantidad liberada no puede ser mayor que cantidad de entradas restante')

    #else:
        #form = InventarioForm()

    context = {
        'form': form,
        'producto_calidad':producto_calidad,
        'articulo_entrada':articulo_entrada,
        'restantes_liberacion': restantes_liberacion,
        }

    return render(request,'entradas/calidad_entrada.html',context)

def productos(request, pk):
    compra = Compra.objects.get(id=pk)
    articulos_comprados = ArticuloComprado.objects.filter(oc=compra, entrada_completa=False)

    context = {
        'compra': compra,
        'articulos_comprados': articulos_comprados,
    }

    return render(request, 'entradas/productos.html', context)


def no_conformidad(request, pk):
    # Obtén la compra y el perfil asociado con la sesión actual
    compra = Compra.objects.get(id=pk)
    perfil = Profile.objects.get(staff__id = request.user.id)
    # Subconsulta para la suma de NC en artículos
    subquery_suma_nc_articulos = NC_Articulo.objects.filter(resuelto = False,
        articulo_comprado=OuterRef('pk')
    ).values('articulo_comprado').annotate(
        total_nc=Sum('cantidad')
    ).values('total_nc')

    # Subconsulta para la suma de entradas en artículos
    subquery_suma_entrada_articulos = EntradaArticulo.objects.filter(
        articulo_comprado=OuterRef('pk')
    ).values('articulo_comprado').annotate(
        total_entrada=Sum('cantidad')
    ).values('total_entrada')

    # Consulta principal
    articulos = ArticuloComprado.objects.filter(
        oc=pk,
        entrada_completa=False,
        recepcion_completa=False,
        seleccionado=False,
        producto__producto__articulos__producto__producto__servicio=False
    ).annotate(
        suma_nc_articulos=Subquery(subquery_suma_nc_articulos, output_field=DecimalField(max_digits=14, decimal_places=2)),
        suma_entrada_articulos=Subquery(subquery_suma_entrada_articulos, output_field=DecimalField(max_digits=14, decimal_places=2))
    )
        #Pasa por todos estableciendo la cantidad pendiente como la cantidad solicitada
    for articulo in articulos:
        if articulo.cantidad_pendiente == None:
            articulo.cantidad_pendiente = articulo.cantidad


    # Crear o obtener la instancia de No_Conformidad al momento de ingresar al form
    no_conformidad, created = No_Conformidad.objects.get_or_create(
        oc=compra,
        almacenista=perfil,
        completo = False,
    )

    articulos_nc = NC_Articulo.objects.filter(nc = no_conformidad,resuelto = False) #Aqui se buscan las NC_articulo para cada articulo
    form = NC_ArticuloForm()
    form2 = NoConformidadForm()

    productos_para_select2 = [
        {'id': producto.id,
         'text': str(producto.producto.producto.articulos.producto), 
         'cantidad': str(producto.cantidad), 
         'cantidad_pendiente': str(producto.cantidad_pendiente),
         'nc': str(producto.suma_nc_articulos),
         'entradas': str(producto.suma_entrada_articulos),
        } for producto in articulos]

    # Si el método de la petición es POST, procesar el formulario
    if request.method == "POST":
        #and 'BtnCrear' in request.POST:
        form2 = NoConformidadForm(request.POST, instance = no_conformidad)
        #Una vez se manda el form luego de el add y que se creara con este el NC_Articulo nuevo
        if form2.is_valid(): ################################## SI HAY UN PROBLEMA DE QUE DICE QUE YA QUEDO LA COMPRA Y AUN QUEDA por pulsar un NC O ENTRADA por complete solo quita este commit false en los dos ya que ya estas validando la cantidad en el update de cada uno y este evita que se guarde para que jale los complete
            no_conf = form2.save(commit=False)
            articulos_comprados = ArticuloComprado.objects.filter(oc=compra) #Se buscan los articulos comprados de la OC
            num_art_comprados = ArticuloComprado.objects.filter(oc=compra,).count() #Aqui sacamos el numero de articulos pedidos
            
            for articulo in articulos_nc: #Aqui se pasa por cada NC articulos de la orden NC
                articulo_comprado = articulos_comprados.get(producto=articulo.articulo_comprado.producto) #Se busca el articulo comprado para tener la cantidad y pendiente
                #Valen como entradas y recepciones
                nc_producto = NC_Articulo.objects.filter(resuelto = False, articulo_comprado = articulo_comprado, nc__oc = compra).aggregate(Sum('cantidad')) 
                #Valen como entradas                           #### Ojoooo no tiene complete entrada o nc creo se arregla con borrar el commit para que ya la guarde nada más mandar el form ya que no se hace más cuentas
                entradas_producto = EntradaArticulo.objects.filter(articulo_comprado = articulo_comprado, entrada__oc = compra, almacenado = True).aggregate(Sum('cantidad'))#Busca la cantidad de entradas para ese producto añadiendo la cantidad para cada dato
                #Valen recepcionados
                recepcion_producto = EntradaArticulo.objects.filter(articulo_comprado = articulo_comprado, entrada__oc = compra, recepcion = True).aggregate(Sum('cantidad'))
                #Cantidad total de cada entrada nc y recepcion
                suma_entradas = entradas_producto['cantidad__sum']#Saca la suma de la cantidad total de las entradas de ese producto
                suma_nc_producto = nc_producto['cantidad__sum']
                suma_recepcion = recepcion_producto['cantidad__sum']
                #Formateo
                if suma_entradas is None:
                    suma_entradas = 0
                if suma_nc_producto is None:
                    suma_nc_producto = 0
                if suma_recepcion is None:
                    suma_recepcion=0
                if articulo_comprado.cantidad_pendiente == None: 
                    articulo_comprado.cantidad_pendiente = articulo_comprado.cantidad
                #Producto pendientes entradas 
                entrada_pendientes = articulo_comprado.cantidad - suma_entradas - suma_nc_producto
                articulo_comprado.cantidad_pendiente = entrada_pendientes
                #Producto pendientes recepcion
                recepcion_pendientes = articulo_comprado.cantidad - suma_recepcion - suma_nc_producto

                if recepcion_pendientes == 0:
                    articulo_comprado.recepcion_completa = True
                if entrada_pendientes == 0:
                    articulo_comprado.entrada_completa = True

                articulo_comprado.seleccionado = False
                articulo_comprado.save() #guarda el articulo comprado

            static_path = settings.STATIC_ROOT
            #Generación de correo
            img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
            img_path2 = os.path.join(static_path,'images','logo vordtec_documento.png')
    
            image_base64 = get_image_base64(img_path)
            logo_v_base64 = get_image_base64(img_path2)

            #Luego de pasar por todos los articulos no recibidos
            #Se manda todos los articulos comprados de la oc--Se manda la cantidad (Count) de estos comprados-- se manda la OC
            num_art_entregados = ArticuloComprado.objects.filter(oc=compra, entrada_completa=True).count() #Articulos completos
            num_art_recepcionados = ArticuloComprado.objects.filter(oc=compra, recepcion_completa=True).count()
            if num_art_comprados == num_art_entregados: #Concuerda con el numero de pedidos
                compra.entrada_completa = True #Define en la OC entrada completa si numero de articulos entragados completos concuerda con los pedidos
                compra.recepcion_completa = True #Define la OC como recepcion completa
            elif num_art_comprados == num_art_recepcionados:
                compra.recepcion_completa = True #Define la OC como recepcion completa
            compra.save()
            no_conf.completo = True
            no_conf.nc_date = date.today()
            no_conf.nc_hora = datetime.now().time()
            no_conf.save()
            try:
                email = EmailMessage(
                    f'Compra| No conformidad {no_conf.id} OC {no_conf.oc.get_folio}',
                    f'Estimado {no_conf.oc.proveedor.nombre.razon_social},\n Estás recibiendo este correo porque se ha recibido en almacén el producto código:{articulo.articulo_comprado.producto.producto.articulos.producto.producto.codigo} descripción:{articulo.articulo_comprado.producto.producto.articulos.producto.producto.nombre} el cual no fue entregado al almacén\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                    'savia@vordtec.com',
                    ['ulises_huesc@hotmail.com',no_conf.oc.proveedor.email,no_conf.oc.creada_por.staff.email,],
                    )
                email.send()
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'Se ha dado de alta correctamente la NC el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.warning(request, error_message)
            messages.success(request,'Has completado la No Conformidad de manera exitosa')
            return redirect('pendientes-recepcion')
        else:
            messages.error(request,'No está validando')
    #else:
        #messages.error(request,'Está siguiendo de largo')


    context = {
        'productos_para_select2':productos_para_select2,
        'compra':compra,
        #'articulos':articulos,
        'articulos_nc':articulos_nc,
        'form': form,
        'form2':form2,
        'no_conformidad': no_conformidad,
    }

    return render(request, 'entradas/no_conformidad.html', context)

def update_no_conformidad(request):
    #Solo se evaluan las cantidades pero no se afectan
    data = json.loads(request.body)
    cantidad = decimal.Decimal(data["cantidad_ingresada"])
    action = data["action"]
    producto_id = int(data["producto"])
    pk = int(data["nc_id"])
    #referencia = data["referencia"]
    producto_comprado = ArticuloComprado.objects.get(id = producto_id) #Saca el producto
    nc = No_Conformidad.objects.get(id = pk, completo = False) #Saca el NC
    nc_producto = NC_Articulo.objects.filter(resuelto = False, articulo_comprado = producto_comprado, nc__oc = producto_comprado.oc, nc__completo = True).aggregate(Sum('cantidad')) #Saca el NC_Articulo con la cantidad del producto
    entradas_producto = EntradaArticulo.objects.filter(articulo_comprado = producto_comprado, entrada__oc = producto_comprado.oc, entrada__completo = True).aggregate(Sum('cantidad'))#Busca la cantidad de entradas para ese producto añadiendo la cantidad para cada dato
    suma_entradas = entradas_producto['cantidad__sum']#Saca la suma de la cantidad total de las entradas de ese producto
    suma_nc_producto = nc_producto['cantidad__sum']#Saca la cantidad del producto en la NC_Articulo
    #Saca ahora en las entradas la cantidad por surtir de ese producto
    entradas_producto = EntradaArticulo.objects.filter(articulo_comprado = producto_comprado, entrada__oc = producto_comprado.oc, entrada__completo = True).aggregate(Sum('cantidad_por_surtir'))
    #Saca el total de la cantidad por surtir de las entradas de ese producto
    pendientes_surtir = entradas_producto['cantidad_por_surtir__sum']
    #Formatea con 0 para evitar errores si es que no encuentra datos
    if pendientes_surtir == None:   #Esto sucede cuando no hay ningún producto en esos articulos
        pendientes_surtir = 0
    if suma_nc_producto == None:
        suma_nc_producto = 0
    if suma_entradas == None:
        suma_entradas = 0

    #Crea el dato NC_Articulo al que se le asignara la cantidad de articulos NC
    nc_item, created = NC_Articulo.objects.get_or_create(nc = nc, articulo_comprado = producto_comprado)
    nc_item.cantidad = cantidad #Se le añade la cantidad

    if action == "add":
        total_entradas_nc = pendientes_surtir + suma_nc_producto + nc_item.cantidad#Se suma los pendientes por surtir de las entradas
        #La suma total de la cantidad de los NC_articulo y la cantidad del nuevo item nc_articulo
        if total_entradas_nc > producto_comprado.cantidad: #Si la cantidad de las entradas es mayor a la cantidad de la compra se rechaza
            messages.error(request,f'Estas intentando ingresar mas productos de los comprados, Comprados: {producto_comprado.cantidad} Ya Recepcionados: {suma_entradas} Ingresados: {cantidad}')
        else:
            #producto_comprado.cantidad_pendiente = producto_comprado.cantidad - total_entradas_nc
            #Cree una variable booleana temporal para quitarlo del seleccionable
            producto_comprado.seleccionado = True
            messages.success(request,f'Has agregado el artículo con éxito {total_entradas_nc}')
            #Se guarda el nc_articulo una vez se comprobo que no sobrepasa la cantidad maxima pedida
            producto_comprado.save()
            nc_item.save()
            print("pendientes surtir")
            print(pendientes_surtir)
            print("suma_nc_producto")
            print(suma_nc_producto)
            print("nc_item.cantidad")
            print(nc_item.cantidad)
    elif action == "remove":
        #Se restablece la variable del seleccionable
        producto_comprado.seleccionado = False
        
        #Se borra el elemento de las entradas
        #Guardado de bases de datos
        nc_item.delete()
        producto_comprado.save()
        messages.success(request,'Has eliminado el artículo con éxito')
    return JsonResponse('Item was '+action, safe=False)

def no_conformidad_almacen(request, pk):
    # Obtén la entrada de artículo y el perfil asociado con la sesión actual
    objeto_nc = EntradaArticulo.objects.get(id=pk) #Entrada del articulo uno para cada producto que tiene su cantidad pedida y pendiente
    compra = objeto_nc.entrada.oc #La OC
    perfil = Profile.objects.get(staff__id=request.user.id)
    #articulos_comprados = ArticuloComprado.objects.filter(oc=pk) 
    articulo = objeto_nc.articulo_comprado #El articulo comprado
    # Crear o obtener la instancia de No_Conformidad
    no_conformidad, created = No_Conformidad.objects.get_or_create(
        oc=compra,
        almacenista=perfil,
        completo=False,
    )

    #articulo_comprado = articulo
    # Consulta para sumar la cantidad de NC_Articulo asociado a EntradaArticulo
    #sumatoria_cantidad_con_entrada = NC_Articulo.objects.filter(
    #    articulo_comprado=articulo_comprado,
    #    entrada_articulo__isnull=False  # Filtramos los que tienen una EntradaArticulo asociada
    #).aggregate(Sum('cantidad'))

    #sumatoria_cantidad_sin_entrada = NC_Articulo.objects.filter(
    #    articulo_comprado=articulo_comprado,
    #    entrada_articulo__isnull=True  # Filtramos los que NO tienen una EntradaArticulo asociada
    #).aggregate(Sum('cantidad'))

    #nc_con_entrada = sumatoria_cantidad_con_entrada['cantidad__sum'] or 0 #Saber cantidad NC en Almacen
    #nc_sin_entrada = sumatoria_cantidad_sin_entrada['cantidad__sum'] or 0 #Saber cantidad NC recepcion

    # Inicializar los formularios
    #articulos_nc = NC_Articulo.objects.filter(nc = no_conformidad, ) #Aqui se buscan las NC para cada articulo
    form = NC_Almacen_ArticuloForm()
    form2 = NoConformidadForm(instance=no_conformidad)
    # Si el método de la petición es POST, procesar el formulario
    if request.method == "POST":
        form2 = NoConformidadForm(request.POST, instance=no_conformidad)
        form = NC_Almacen_ArticuloForm(request.POST)
        if form2.is_valid() and form.is_valid():
            # Validación personalizada
            cantidad = form.cleaned_data['cantidad']
            if cantidad <= 0 or cantidad > objeto_nc.cantidad:
                messages.error(request, 'La cantidad ingresada es 0 o mayor a la cantidad máxima permitida.')
            else:
                # Guardar No_Conformidad sin guardar en la base de datos
                no_conf = form2.save(commit=False)
                # Guardar NC sin guardar en la base de datos
                nc_articulo = form.save(commit=False)
                nc_articulo.nc = no_conf
                nc_articulo.entrada_articulo = objeto_nc
                nc_articulo.articulo_comprado = articulo  # Asignar el artículo comprado
                #Actualizar #################Checar si es que surtir hace cosas raras
                objeto_nc.cantidad_por_surtir = objeto_nc.cantidad_por_surtir - cantidad
                objeto_nc.cantidad = objeto_nc.cantidad - cantidad
                # Guardar los formularios
                no_conf.save()
                nc_articulo.save()
                objeto_nc.save()

                #Logica update
                articulo_comprado = articulo
                nc_producto = NC_Articulo.objects.filter(resuelto = False, articulo_comprado = articulo_comprado, nc__oc = compra).aggregate(Sum('cantidad')) 
                #Valen como entradas                           #### Ojoooo no tiene complete entrada o nc creo se arregla con borrar el commit para que ya la guarde nada más mandar el form ya que no se hace más cuentas
                entradas_producto = EntradaArticulo.objects.filter(articulo_comprado = articulo_comprado, entrada__oc = compra, almacenado = True).aggregate(Sum('cantidad'))#Busca la cantidad de entradas para ese producto añadiendo la cantidad para cada dato
                #Valen recepcionados
                recepcion_producto = EntradaArticulo.objects.filter(articulo_comprado = articulo_comprado, entrada__oc = compra, recepcion = True).aggregate(Sum('cantidad'))
                #Cantidad total de cada entrada nc y recepcion
                suma_entradas = entradas_producto['cantidad__sum']#Saca la suma de la cantidad total de las entradas de ese producto
                suma_nc_producto = nc_producto['cantidad__sum']
                suma_recepcion = recepcion_producto['cantidad__sum']
                #Formateo
                if suma_entradas is None:
                    suma_entradas = 0
                if suma_nc_producto is None:
                    suma_nc_producto = 0
                if suma_recepcion is None:
                    suma_recepcion=0
                if articulo_comprado.cantidad_pendiente == None: 
                    articulo_comprado.cantidad_pendiente = articulo_comprado.cantidad
                #Producto pendientes entradas 
                entrada_pendientes = articulo_comprado.cantidad - suma_entradas - suma_nc_producto
                articulo_comprado.cantidad_pendiente = entrada_pendientes
                #Producto pendientes recepcion ###########################Este solo va aqui en almacen no poner en recepcion
                recepcion_pendientes = articulo_comprado.cantidad - suma_recepcion - suma_nc_producto

                if recepcion_pendientes == 0:
                    articulo_comprado.recepcion_completa = True
                if entrada_pendientes == 0:
                    articulo_comprado.entrada_completa = True

                articulo_comprado.seleccionado = False
                articulo_comprado.save() #guarda el articulo comprado

                # Generación de correo
                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path, 'images', 'SAVIA_Logo.png')
                img_path2 = os.path.join(static_path, 'images', 'logo vordtec_documento.png')

                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)

                #Luego de pasar por todos los articulos no recibidos
                #Se manda todos los articulos comprados de la oc--Se manda la cantidad (Count) de estos comprados-- se manda la OC
                num_art_entregados = ArticuloComprado.objects.filter(oc=compra, entrada_completa=True).count() #Articulos completos
                #num_art_comprados = articulos_comprados.count()
                num_art_comprados = ArticuloComprado.objects.filter(oc=compra,).count()
                num_art_recepcionados = ArticuloComprado.objects.filter(oc=compra, recepcion_completa=True).count()
                print('Objetos')
                print(num_art_comprados) #0
                print('Almacenados')
                print(num_art_entregados)
                print('Recepcionados')
                print(num_art_recepcionados)
                if num_art_comprados == num_art_entregados: #Concuerda con el numero de pedidos
                    compra.entrada_completa = True #Define en la OC entrada completa si numero de articulos entragados completos concuerda con los pedidos
                    compra.recepcion_completa = True #Define la OC como recepcion completa
                    print('Compra almacen completo')
                if num_art_comprados == num_art_recepcionados:
                    compra.recepcion_completa = True #Define la OC como recepcion completa
                compra.save()

                no_conf.completo = True
                no_conf.nc_date = date.today()
                no_conf.nc_hora = datetime.now().time()
                no_conf.save()
                try:
                    email = EmailMessage(
                        f'Compra| No conformidad {no_conf.id} OC {no_conf.oc.get_folio}',
                        f'Estimado {no_conf.oc.proveedor.nombre.razon_social},\n Estás recibiendo este correo porque se ha recibido en almacén el producto código:{articulo.producto.producto.articulos.producto.producto.codigo} descripción:{articulo.producto.producto.articulos.producto.producto.nombre} el cual no fue entregado al almacén\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                        'savia@vordtec.com',
                        ['ulises_huesc@hotmail.com',no_conf.oc.proveedor.email,no_conf.oc.creada_por.staff.email,],
                        )
                    email.send()
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'Se ha dado de alta correctamente la NC el correo de notificación no ha sido enviado debido a un error: {e}'
                    messages.warning(request, error_message)
                messages.success(request, 'Has completado la No Conformidad de manera exitosa')
                return redirect('pendientes-entrada')
        else:
            messages.error(request, 'No está validando')

    context = {
        'objeto_nc': objeto_nc,
        'form': form,
        'form2': form2,
        'compra': compra,
        'no_conformidad': no_conformidad,
    }

    return render(request, 'entradas/no_conformidad_almacen.html', context)

def productos_nc(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    articulos_nc = NC_Articulo.objects.filter(nc = pk)

    context = {
        'articulos_nc': articulos_nc,
    }

    return render(request, 'entradas/productos_nc.html', context)

def cierre_nc(request, pk):
    perfil = Profile.objects.get(staff__id=request.user.id)
    nc = No_Conformidad.objects.get(id = pk)
    #articulos_nc = NC_Articulo.objects.filter(nc = pk).first()
    articulos_nc = NC_Articulo.objects.filter(nc = pk)
    form = Cierre_NCForm(instance = nc)

    if request.method == "POST":
        #and 'BtnCrear' in request.POST:
        form = Cierre_NCForm(request.POST, request.FILES, instance = nc)

        if form.is_valid():
            nc = form.save(commit=False)
            nc.fecha_cierre = date.today()
            nc.save()
            oc = Compra.objects.get(id = nc.oc.id)
            if nc.cierre.id == 3:
                oc.entrada_completa = False
                oc.recepcion_completa = False
            for dato in articulos_nc:
                producto = ArticuloComprado.objects.get(id = dato.articulo_comprado.id)
                articulo = NC_Articulo.objects.get(id = dato.id)
                articulo.resuelto = True
                if nc.cierre.id == 3:
                    #Se debería de reactivas la OC, en la variable entrada_completa = False
                    producto.entrada_completa = False
                    producto.recepcion_completa = False
                    producto.cantidad_pendiente = producto.cantidad_pendiente + dato.cantidad
                    articulo.save()
                    producto.save()
            oc.save()
            if nc.cierre.id == 3:
                messages.success(request, 'Has reactivado correctamente la entrada del articulo')
            else:
                messages.success(request, 'Se a realizado correctamente el cierre del articulo')
            return redirect('entradas_nc')

    context = {
        'form': form,
        'nc': nc,
        'articulos_nc': articulos_nc,
    }

    return render(request, 'entradas/cierre_nc.html', context)


def matriz_reportes_calidad(request):
    perfil = Profile.objects.get(staff__id=request.user.id)
    reportes = Reporte_Calidad.objects.filter(completo = True).order_by('-reporte_date')
    form = Reporte_CalidadForm()


    context = {
        #'form': form,
        'reportes':reportes,
        #'restantes_liberacion': restantes_liberacion,
        }

    return render(request,'entradas/matriz_reportes_calidad.html',context)

def entradas_nc(request):
    perfil = Profile.objects.get(staff__id=request.user.id)
    ncs= No_Conformidad.objects.filter(completo = True, oc__req__orden__distrito = perfil.distrito)
    
    myfilter = No_ConformidadFilter(request.GET, queryset=ncs)
    ncs = myfilter.qs

    #Set up pagination
    p = Paginator(ncs, 50)
    page = request.GET.get('page')
    entradas_list = p.get_page(page)

    context = {
        'myfilter': myfilter,
        'ncs': ncs,
        'entradas_list': entradas_list,
        }

    return render(request,'entradas/entradas_nc.html',context)

@login_required(login_url='user-login')
def entradas_con_caducidad(request):
    perfil = Profile.objects.get(staff__id=request.user.id)
    entradas = EntradaArticulo.objects.filter(
        articulo_comprado__producto__producto__articulos__producto__producto__caducidad=True
    ).exclude(fecha_caducidad__isnull=True).exclude(cantidad_por_surtir=0).order_by('fecha_caducidad')
    
    myfilter = EntradaCaducidadFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs

    #Set up pagination
    p = Paginator(entradas, 50)
    page = request.GET.get('page')
    entradas_list = p.get_page(page)

    for e in entradas_list:
        e.dias_restantes = (e.fecha_caducidad - date.today()).days


    if request.method == "POST" and 'btnExcel' in request.POST:
        return convert_caducidad_to_xls2(entradas)
    
    context = {
        'myfilter':myfilter,
        'entradas': entradas,
        'entradas_list':entradas_list,
        }

    return render(request,'entradas/entradas_con_caducidad.html',context)

def convert_caducidad_to_xls2(entradas):
    # Crear un flujo en memoria para el archivo Excel
    output = io.BytesIO()

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=Entradas_Productos_Caducidad_{datetime.today().strftime("%Y-%m-%d")}.xlsx'
    response.set_cookie('descarga_iniciada', 'true', max_age=10)  # La cookie expira en 10 segundos

    # Crear un nuevo libro de trabajo y hoja de trabajo
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet('Entradas')
    #ws.title = 'Entradas'

    head_style = wb.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = wb.add_format({'font_name': 'Calibri', 'font_size': 10})
    number_style = wb.add_format({'num_format': '##0', 'font_name': 'Calibri', 'font_size': 10})
    date_style = wb.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = wb.add_format({'font_name': 'Arial Narrow', 'font_size': 11})

   

    # Configurar encabezados de columna
    columns = ['Entrada', 'OC', 'Producto','Proveedor', 'Cantidad', 'Cantidad por surtir', 'Fecha entrada',
               'Comentario', 'Fecha caducidad', 'Almacenista']
    
    for i, column in enumerate(columns):
        ws.write(0, i, column, head_style)
        ws.set_column(i, i, 20)  # Ancho de columna

    # Agregar mensajes
    max_col = len(columns) + 2
    ws.write(0, max_col - 1, 'Reporte Creado Automáticamente por SAVIA Vordcab. UH', messages_style)
    ws.write(1, max_col - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    ws.set_column(max_col - 1, max_col, 30)  # Ajusta el ancho de las columnas nuevas

    # Agregar datos de entradas
    row_num = 1
    for dev in entradas:
        if dev.entrada.comentario:
            comentario = dev.entrada.comentario
        else:
            comentario = ''
        row = [
            dev.id,
            dev.entrada.oc.id,
            str(dev.articulo_comprado.producto.producto.articulos.producto.producto.nombre),
            str(dev.entrada.oc.proveedor.nombre),###
            dev.cantidad,
            dev.cantidad_por_surtir,
            dev.entrada.entrada_date,
            #productos_str,  # Productos concatenados
            comentario,
            dev.fecha_caducidad,
            f"{dev.entrada.almacenista.staff.first_name} {dev.entrada.almacenista.staff.last_name}",
        ]

        for col_num, cell_value in enumerate(row):
            cell_format = number_style
            if col_num in [6, 8]:  # Columnas de fecha
                cell_format = date_style
            if col_num in [2, 3, 9]:  # Columnas numéricas
                cell_format = body_style
            ws.write(row_num, col_num, cell_value, cell_format)
        row_num += 1
            
    wb.close()
    
    # Eliminar la hoja predeterminada si existe y guardar el archivo en el objeto BytesIO
    #if 'Sheet' in wb.sheetnames:
    #    wb.remove(wb['Sheet'])

    #wb.save(output)
    output.seek(0)  # Asegurarse de que el puntero esté al principio del flujo de bytes

    # Establecer el contenido del archivo en la respuesta HTTP
    response.write(output.getvalue())
    output.close()

    return response


# Función para actualizar o crear el comentario
def update_comentario(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        entrada_articulo_id = data['entradaArticuloId']
        nuevo_comentario = data['nuevoComentario']

        entrada_articulo = get_object_or_404(EntradaArticulo, id=entrada_articulo_id)

        # Verifica si el Reporte_Calidad ya existe, si no lo crea
        reporte_calidad, created = Reporte_Calidad.objects.get_or_create(articulo=entrada_articulo)

        # Actualiza el comentario
        reporte_calidad.comentarios = nuevo_comentario
        reporte_calidad.completo = True
        reporte_calidad.cantidad = entrada_articulo.cantidad
        reporte_calidad.save()

        return JsonResponse({'status': 'success', 'nuevoComentario': nuevo_comentario})

def autorizar_calidad(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        entrada_articulo_id = data['entradaArticuloId']
        autorizado = data['autorizado']  # True para palomita, False para tache
        with transaction.atomic():
            entrada_articulo = get_object_or_404(EntradaArticulo, id=entrada_articulo_id)
            
            # Verifica si el Reporte_Calidad ya existe, si no lo crea
            reporte_calidad, created = Reporte_Calidad.objects.get_or_create(articulo=entrada_articulo)

            reporte_calidad.reporte_date = datetime.now().date()
            reporte_calidad.reporte_hora = datetime.now().time()
            # Actualiza los campos de EntradaArticulo y Reporte_Calidad
            if autorizado:
                print('autorizado')
                entrada_articulo.calidad = True
                reporte_calidad.autorizado = True
            if not autorizado:
                entrada_articulo.calidad = True
                reporte_calidad.autorizado = False
                tipo_nc = Tipo_Nc.objects.get(id=2)  

                nc = No_Conformidad.objects.create(
                    almacenista=getattr(request.user, 'profile', None),
                    oc=entrada_articulo.articulo_comprado.oc,
                    comentario=f'No conformidad por rechazo en calidad. Reporte:{reporte_calidad.id}',
                    nc_date=datetime.now().date(),
                    nc_hora=datetime.now().time(),
                    completo=True,
                    tipo_nc=tipo_nc, 
                )

                NC_Articulo.objects.create(
                    nc=nc,
                    cantidad=entrada_articulo.cantidad,
                    articulo_comprado=entrada_articulo.articulo_comprado,
                    entrada_articulo=entrada_articulo,
                    resuelto=False
                )

            # Asigna la fecha y hora actual
            reporte_calidad.reporte_date = datetime.now().date()  # Fecha actual
            reporte_calidad.reporte_hora = datetime.now().time()  # Hora actual
            reporte_calidad.cantidad = entrada_articulo.cantidad
            reporte_calidad.completo = True
            entrada_articulo.save()
            reporte_calidad.save()

            return JsonResponse({'status': 'success', 'autorizado': autorizado})
    
@login_required(login_url='user-login')
def calidad_entradas(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
  
    
    if usuario.tipo.calidad == True:
        # Subconsulta para total_recepcionado
        subquery_total_recepcionado = EntradaArticulo.objects.filter(
            articulo_comprado=OuterRef('articulo_comprado'),
            entrada__oc=OuterRef('entrada__oc'),
            recepcion=True,
        ).values('articulo_comprado').annotate(
            total_recepcionado=Sum('cantidad')
        ).values('total_recepcionado')

        # Subconsulta para total_nc
        subquery_total_nc = NC_Articulo.objects.filter(
            articulo_comprado=OuterRef('articulo_comprado'), resuelto=False,
            nc__oc=OuterRef('entrada__oc')
        ).values('articulo_comprado').annotate(
            total_nc=Sum('cantidad')
        ).values('total_nc')

        # Consulta principal
        articulos_recepcionados = EntradaArticulo.objects.filter(
            recepcion=True,
            cantidad__gt=0,
            almacenado = False,
            calidad = False,
            liberado = True,
            articulo_comprado__producto__producto__articulos__producto__producto__servicio=False,
            articulo_comprado__producto__producto__articulos__producto__producto__critico__in=[1, 2]  # Filtro para id 1 y 2 
        ).exclude(
            reportes_calidad__autorizado=True  # Excluye aquellos que tienen un Reporte_Calidad con autorizado=True
        ).annotate(
            total_recepcionado=Subquery(subquery_total_recepcionado, output_field=DecimalField(max_digits=14, decimal_places=2)),
            total_nc=Subquery(subquery_total_nc, output_field=DecimalField(max_digits=14, decimal_places=2))
        ).order_by('-id')

        myfilter = EntradaArticuloFilter(request.GET, queryset=articulos_recepcionados)
        articulos_recepcionados = myfilter.qs
        #Set up pagination
        p = Paginator(articulos_recepcionados, 50)
        page = request.GET.get('page')
        articulos_recepcionados_list = p.get_page(page)
    else:
        articulos_recepcionados = None
        articulos_recepcionados_list = None

    print(articulos_recepcionados)
    context = {
        'articulos_recepcionados':articulos_recepcionados,
        'myfilter':myfilter,
        'articulos_recepcionados_list':articulos_recepcionados_list,
        }

    return render(request, 'entradas/calidad_entradas.html', context)

@login_required(login_url='user-login')
def calidad_entradas_autorizadas(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    
    if usuario.tipo.calidad == True:
        # Consulta principal
        articulos_recepcionados = Reporte_Calidad.objects.filter(completo=True, autorizado=True).order_by('-id')

        myfilter = Reporte_CalidadFilter(request.GET, queryset=articulos_recepcionados)
        articulos_recepcionados = myfilter.qs
        #Set up pagination
        p = Paginator(articulos_recepcionados, 50)
        page = request.GET.get('page')
        articulos_recepcionados_list = p.get_page(page)
    else:
        articulos_recepcionados = None
        articulos_recepcionados_list = None

    context = {
        'articulos_recepcionados':articulos_recepcionados,
        'myfilter':myfilter,
        'articulos_recepcionados_list':articulos_recepcionados_list,
        }

    return render(request, 'entradas/calidad_entradas_autorizadas.html', context)

@login_required(login_url='user-login')
def productos_terminados_entrada(request):
    perfil = Profile.objects.get(staff__id=request.user.id)
     #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.almacen == True:
        entradas = EntradaArticulo.objects.filter(producto_terminado__isnull=False,producto_terminado__principal__isnull=True,recepcion=True,cantidad_por_surtir__gt=0,almacenado =False,producto_terminado__validado__isnull=True)
    else:
        entradas = EntradaArticulo.objects.none()

    myfilter=EntradaTerminadoFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs

    #Set up pagination
    p = Paginator(entradas, 25)
    page = request.GET.get('page')
    entradas_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:
        return convert_excel_salida_terminados(entradas)

    context= {
        'entradas_list':entradas_list,
        'myfilter':myfilter,
        }

    return render(request, 'entradas/productos_terminados_entradas.html',context)

@login_required(login_url='user-login')
def productos_terminados_salida(request):
    perfil = Profile.objects.get(staff__id=request.user.id)
     #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.almacen == True:
        entradas = EntradaArticulo.objects.filter(producto_terminado__isnull=False,producto_terminado__principal__isnull=True,almacenado =True, producto_terminado__validado=True,liberado=False)
    else:
        entradas = EntradaArticulo.objects.none()

    myfilter=EntradaTerminadoFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs

    #Set up pagination
    p = Paginator(entradas, 25)
    page = request.GET.get('page')
    entradas_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:
        return convert_excel_salida_terminados(entradas)

    context= {
        'entradas_list':entradas_list,
        'myfilter':myfilter,
        }

    return render(request, 'entradas/productos_terminados_salidas.html',context)

def terminado_salida_editar_cliente(request, pk):
    entrada = get_object_or_404(EntradaArticulo, id=pk)
    producto = entrada

    if request.method == 'POST':
        cliente = request.POST.get('cliente')
        destino = request.POST.get('destino')
        producto.cliente = cliente
        producto.destino = destino
        producto.save()
        messages.success(request, 'Datos del cliente agregados exitosamente.')
        return HttpResponse(status=204)
    
    context = {
        'producto': producto,
    }


    return render(request, 'entradas/terminado_salida_editar_cliente.html', context)

@login_required(login_url='user-login')
def validar_entrada_terminado(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    entrada_art = EntradaArticulo.objects.get(id = pk)
    entrada = Entrada.objects.get(id = entrada_art.entrada.id)
    solicitud = entrada.solicitud
    fecha = datetime.now()
    dia = date.today()
    hora = datetime.now().time()
    #Logica actual de que no hay devolucion o entradas por partes
    entrada_art.cantidad_por_surtir = 0
    entrada_art.entrada_date = fecha
    entrada_art.almacenado = True
    entrada_art.save()   
    
    producto = entrada_art.producto_terminado
    producto.validado = True
    producto.save()

    entrada.almacenista = usuario
    entrada.entrada_date = dia
    entrada.entrada_hora = hora
    entrada.save()
    #Parte donde se altera el inventario para la entrada
    inventario = entrada_art.producto_terminado.producto
    inventario.cantidad += entrada_art.cantidad
    inventario.comentario = 'Entrada en producto terminado'
    inventario.save()

    if entrada_art.producto_terminado.producto.producto.subfamilia:
        if entrada_art.producto_terminado.producto.producto.subfamilia.nombre == 'EQUIPO':
            # Obtener las entrada articulos (componentes) relacionados al producto principal
            componentes = EntradaArticulo.objects.filter(producto_terminado__principal = producto)
            print('Componentes')
            print(componentes)
            for componente in componentes:
                #Logica actual de que no hay devolucion o entradas por partes
                componente.cantidad_por_surtir = 0
                componente.entrada_date = fecha
                componente.almacenado = True
                componente.save()
                producto = componente.producto_terminado
                producto.validado = True
                producto.save()   
                entrada = componente.entrada
                entrada.almacenista = usuario
                entrada.entrada_date = dia
                entrada.entrada_hora = hora
                entrada.save()
    #Checar si ya fue completa la solicitud
    todos_productos = solicitud.productos.filter(complete=True)
    total_productos = todos_productos.count()
    validados = solicitud.productos.filter(validado__isnull=False,complete=True) 
    total_validados = validados.count() 
    if total_validados == total_productos:
        solicitud.concluida = True
        solicitud.save()
        print('Terminada la solicitud')
    return redirect('producto-terminado-entrada')

@login_required(login_url='user-login')
def no_validar_entrada_terminado(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    entrada_art = EntradaArticulo.objects.get(id = pk)
    entrada = Entrada.objects.get(id = entrada_art.entrada.id)
    solicitud = entrada.solicitud
    fecha = datetime.now()
    dia = date.today()
    hora = datetime.now().time()
    # Logica actual de que no hay devolucion o entradas por partes
    entrada_art.entrada_date = fecha
    entrada_art.almacenado = False
    entrada_art.save()
    
    producto = entrada_art.producto_terminado #ProductoSolicitud
    producto.validado = False
    producto.save()
    
    entrada.almacenista = usuario
    entrada.entrada_date = dia
    entrada.entrada_hora = hora
    entrada.save()
    if entrada_art.producto_terminado.producto.producto.subfamilia:
        if entrada_art.producto_terminado.producto.producto.subfamilia.nombre == 'EQUIPO':
            # Obtener las entrada articulos (componentes) relacionados al producto principal
            componentes = EntradaArticulo.objects.filter(producto_terminado__principal = producto)
            for componente in componentes:
                #Logica actual de que no hay devolucion o entradas por partes
                componente.entrada_date = fecha
                componente.almacenado = False
                componente.save()
                producto = componente.producto_terminado
                producto.validado = False
                producto.save()   
                entrada = componente.entrada
                entrada.almacenista = usuario
                entrada.entrada_date = dia
                entrada.entrada_hora = hora
                entrada.save()
    #Checar si ya fue completa la solicitud
    todos_productos = solicitud.productos.filter(complete=True)
    total_productos = todos_productos.count()
    validados = solicitud.productos.filter(validado__isnull=False,complete=True) 
    total_validados = validados.count() 
    if total_validados == total_productos:
        solicitud.concluida = True
        solicitud.save()
    return redirect('producto-terminado-entrada')

def producto_terminado_componente_ver(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    entrada = get_object_or_404(EntradaArticulo, id=pk)
    producto = entrada.producto_terminado
    componentes = Productos_Solicitud_Terminado.objects.filter(solicitud__id=producto.solicitud.id, principal = producto)
    
    return render(request, 'entradas/modal_producto_terminado_componentes.html', {'producto': producto,'componentes':componentes,'entrada':entrada,})
    
def convert_excel_salida_terminados(entradas):
    # Crear un flujo en memoria para el archivo Excel
    output = io.BytesIO()

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=Salida_Producto_Terminado_{date.today()}.xlsx'
    response.set_cookie('descarga_iniciada', 'true', max_age=10)

    # Crear un nuevo libro de trabajo y hoja de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Salida_Producto_Terminado'

    # Definir estilos
    styles = {
        'head_style': NamedStyle(name="head_style", font=Font(name='Arial', color='00FFFFFF', bold=True, size=11), fill=PatternFill("solid", fgColor='00003366')),
        'body_style': NamedStyle(name="body_style", font=Font(name='Calibri', size=10)),
        'messages_style': NamedStyle(name="messages_style", font=Font(name="Arial Narrow", size=11)),
        'date_style': NamedStyle(name='date_style', number_format='DD/MM/YYYY', font=Font(name='Calibri', size=10))
    }
    
    for style in styles.values():
        wb.add_named_style(style)
    
    # Configurar encabezados de columna
    columns = [
        'Solicitud', 'Producto terminado id', 'Entrada', 'Solicitó', 'Solicitado', 
        'Proyecto', 'Subproyecto', 'Producto', 'Tipo', '# Serie', 
        'Cantidad', 'Cliente', 'Destino','Entrada','Almacenista', 'Componentes','Vale de Salida', 'Fecha de Salida', 'Almacenista'
    ]
    
    for col_num, column_title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.style = styles['head_style']
        ws.column_dimensions[get_column_letter(col_num)].width = 25 if col_num == 6 else 16

    # Agregar mensajes
    max_col = len(columns) + 2
    ws.cell(row=1, column=max_col, value='{Reporte Creado Automáticamente por Savia X. UH}').style = styles['messages_style']
    ws.cell(row=2, column=max_col, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = styles['messages_style']
    ws.column_dimensions[get_column_letter(max_col)].width = 20

    # Agregar datos
    rows = entradas.values_list(
        'producto_terminado__solicitud__id',
        'producto_terminado__id',
        'entrada__id',
        Concat('producto_terminado__solicitud__staff__staff__first_name', Value(' '), 'producto_terminado__solicitud__staff__staff__last_name'),
        'producto_terminado__solicitud__created_at',
        'producto_terminado__solicitud__proyecto__nombre',
        'producto_terminado__solicitud__subproyecto__nombre',
        'producto_terminado__producto__producto__nombre',
        'producto_terminado__producto__producto__subfamilia__nombre',
        'producto_terminado__serie',
        'cantidad',
        'producto_terminado__cliente',
        'producto_terminado__destino',
        'entrada__entrada_date',
        Concat('entrada__almacenista__staff__first_name', Value(' '),'entrada__almacenista__staff__last_name'),
    )
    
    for row_num, row in enumerate(rows, start=2):
        # Escribir las columnas existentes
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=str(value))
            if col_num == 6:
                cell.style = styles['date_style']
            else:
                cell.style = styles['body_style']

        # Obtener componentes asociados
        id_producto_terminado = row[1]  # ID del producto terminado
        componentes = Productos_Solicitud_Terminado.objects.filter(principal__id=id_producto_terminado).values(
            'producto__producto__nombre', 'serie', 'cantidad'
        )
        componentes_texto = "\n".join([ 
            f"{comp['producto__producto__nombre']} (Serie: {comp['serie']}, Cantidad: {comp['cantidad']})"
            for comp in componentes
        ]) if componentes.exists() else "Sin componentes"

        # Agregar componentes a la columna correspondiente
        cell = ws.cell(row=row_num, column=len(columns) - 3, value=componentes_texto)
        cell.style = styles['body_style']

        # Obtener los datos de ValeSalidas y Salidas
        salida = Salidas.objects.filter(producto_terminado__id=id_producto_terminado).values_list('vale_salida__id', 'created_at', 'vale_salida__almacenista__staff__first_name', flat=False).first()

        # Verificar si se obtuvo algún dato
        if salida:
            vale_salida_id, fecha_salida, almacenista = salida
            # Ajustar la fecha de salida al formato adecuado
            fecha_salida = fecha_salida.strftime('%d/%m/%Y %H:%M:%S')  # Puedes ajustar el formato
        else:
            vale_salida_id, fecha_salida, almacenista = "Sin vale de salida", "Sin fecha", "Sin responsable"
        
        # Agregar el Vale de Salida ID
        ws.cell(row=row_num, column=len(columns) - 2, value=vale_salida_id).style = styles['body_style']

        # Agregar la Fecha de Salida
        ws.cell(row=row_num, column=len(columns) - 1, value=fecha_salida).style = styles['body_style']

        # Agregar el Responsable (Almacenista)
        ws.cell(row=row_num, column=len(columns), value=almacenista).style = styles['body_style']
    
    # Eliminar la hoja predeterminada y guardar el archivo en el objeto BytesIO
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    wb.save(output)
    output.seek(0)  # Asegurarse de que el puntero esté al principio del flujo de bytes
    
    # Establecer el contenido del archivo en la respuesta HTTP
    response.write(output.getvalue())
    output.close()
    
    return response



def reporte_recepcionados(entradas):
    
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet("Artículos Recepcionados")


    # Define los estilos
    head_style = wb.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = wb.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = wb.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = wb.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = wb.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = wb.add_format({'font_name':'Arial Narrow', 'font_size':11})


    # 🔹 Aquí defines los encabezados antes de usarlos
    columns = [
        "Compra", "Crítico", "Req.", "Sol.",
        "Solicitado por", "Proyecto", "Subproyecto", "Fecha Recepción",
        "Proveedor", "Concepto", "Cantidad","Precio"
    ]
    
    # Encabezados
    for i, column in enumerate(columns):
        ws.write(0, i, column, head_style)
        ws.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    row_num = 0
    for articulo in entradas:
        row_num += 1

        compra = articulo.articulo_comprado.oc.id if articulo.articulo_comprado.oc else 'ND'
        critico = "Sí" if articulo.articulo_comprado.producto.producto.articulos.producto.producto.critico in [1,2] else "No"
        #terminado = "Sí" if articulo.articulo_comprado.producto.producto.articulos.producto.producto.terminado else "No"
        req = articulo.articulo_comprado.oc.req.folio if articulo.articulo_comprado.oc.req else ''
        sol = articulo.articulo_comprado.oc.req.folio if articulo.articulo_comprado.oc.req else ''
        solicitado_por = f"{articulo.articulo_comprado.oc.req.orden.staff.staff.first_name} {articulo.articulo_comprado.oc.req.orden.staff.staff.last_name}" if articulo.articulo_comprado.oc.req and articulo.articulo_comprado.oc.req.orden.staff.staff else ''
        proyecto = articulo.articulo_comprado.oc.req.orden.proyecto.nombre if articulo.articulo_comprado.oc.req and articulo.articulo_comprado.oc.req.orden.proyecto else ''
        subproyecto = articulo.articulo_comprado.oc.req.orden.subproyecto.nombre if articulo.articulo_comprado.oc.req and articulo.articulo_comprado.oc.req.orden.subproyecto else ''
        fecha_recepcion = articulo.entrada.entrada_date
        proveedor = f"{articulo.articulo_comprado.oc.proveedor.nombre}" if articulo.articulo_comprado.oc and articulo.articulo_comprado.oc.proveedor else ''
        concepto = f"{articulo.articulo_comprado.producto.producto.articulos.producto.producto.codigo}|{articulo.articulo_comprado.producto.producto.articulos.producto.producto.id}|{articulo.articulo_comprado.producto.producto.articulos.producto.producto.nombre}"
        cantidad = articulo.cantidad
        precio = articulo.articulo_comprado.precio_unitario
        row = [
            compra, 
            critico, 
            req, 
            sol, 
            solicitado_por,
            proyecto, 
            subproyecto, 
            fecha_recepcion, 
            proveedor,
            concepto, 
            cantidad,
            precio,
        ]

        for col_num, value in enumerate(row):
            cell_format = body_style
            if col_num == 7:  # Columna de fecha
                cell_format = date_style
            ws.write(row_num, col_num, value, cell_format)

    wb.close()

    output.seek(0)  # Asegurarse de que el puntero esté al principio del flujo de bytes
    # Preparar respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    response['Content-Disposition'] = f'attachment; filename=articulos_recepcionados.xlsx'
    output.close()

    return response



PRUSSIAN_BLUE = Color(0.0859375, 0.1953125, 0.30859375)
def pdf_reporte_calidad(request, reporte_id):
    PRUSSIAN_BLUE = Color(0.0859375, 0.1953125, 0.30859375)

    try:
        reporte = Reporte_Calidad.objects.select_related(
            "articulo__articulo_comprado__oc",
            "articulo__articulo_comprado__producto__producto__articulos__orden__proyecto",
            "articulo__articulo_comprado__producto__producto__articulos__orden__subproyecto",
        ).get(id=reporte_id)
    except Reporte_Calidad.DoesNotExist:
        raise Http404("Reporte no encontrado")

    compra = reporte.articulo.articulo_comprado.oc
    producto = reporte.articulo.articulo_comprado.producto.producto.articulos.producto.producto
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    # --- Encabezado estilo Vordcab ---
    y = _draw_header(c, compra)

    # Sello de estado (usa reporte.autorizado)
    _draw_status_badge(c, reporte.autorizado)
    # (Opcional) Watermark diagonal suave
    _draw_watermark_if_needed(c, reporte.autorizado)


    # --- Título del reporte ---
    c.setFont("Helvetica-Bold", 14)
    c.setFillColor(PRUSSIAN_BLUE)
    c.drawCentredString(300, y - 20, "REPORTE DE CALIDAD")
    y -= 40

    # --- Datos generales ---
    c.setFont("Helvetica", 10)
    c.setFillColor(black)
    c.drawString(40, y, f"Folio OC: {compra.id}")
    c.drawString(250, y, f"Proveedor: {compra.proveedor.nombre if compra.proveedor else ''}")
    y -= 20
    c.drawString(40, y, f"Proyecto: {reporte.articulo.articulo_comprado.producto.producto.articulos.orden.proyecto.nombre}")
    c.drawString(40, y - 12, f"Subproyecto: {reporte.articulo.articulo_comprado.producto.producto.articulos.orden.subproyecto.nombre}")
    y -= 30

    # --- Información del producto ---
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(PRUSSIAN_BLUE)
    c.drawString(40, y, "Producto inspeccionado:")
    c.setFont("Helvetica", 10)
    c.setFillColor(black)
    y -= 15
    c.drawString(50, y, f"{producto.codigo} | {producto.nombre}")
    y -= 15
    c.drawString(50, y, f"Unidad: {producto.unidad.nombre if producto.unidad else ''}")
    y -= 25

    # --- Información del reporte ---
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(PRUSSIAN_BLUE)
    c.drawString(40, y, "Detalles del reporte:")
    c.setFont("Helvetica", 10)
    c.setFillColor(black)
    y -= 15
    c.drawString(50, y, f"Fecha: {reporte.reporte_date or 'N/A'}")
    c.drawString(250, y, f"Hora: {reporte.reporte_hora or 'N/A'}")
    y -= 15
    c.drawString(50, y, f"Cantidad revisada: {reporte.cantidad}")
    y -= 15
    c.drawString(50, y, f"Completo: {'Sí' if reporte.completo else 'No'}")
    c.drawString(250, y, f"Autorizado: {'Sí' if reporte.autorizado else 'Pendiente'}")
    y -= 15
    c.drawString(50, y, f"Comentarios: {reporte.comentarios or 'Sin comentarios'}")
    y -= 30

    # --- Requerimientos de criticidad ---
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(PRUSSIAN_BLUE)
    c.drawString(40, y, "Requisitos:")
    y -= 15
    c.setFont("Helvetica", 10)
    c.setFillColor(black)

    producto_calidad = getattr(producto, "producto_calidad", None)
    #requerimientos = producto_calidad.requerimientos_calidad.all() if producto_calidad else []
    if producto_calidad and producto_calidad.hecho:
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(PRUSSIAN_BLUE)
        c.drawString(40, y, "Características de calidad (hechas):")
        y -= 15

        c.setFont("Helvetica", 10)
        c.setFillColor(black)

        # Helpers para imprimir pares etiqueta/valor
        def _yn(v): 
            return "Sí" if v else "No"

        def _safe(v):
            return v if v not in (None, "") else "N/A"

        # Campos de texto
          # Requisitos con wrap (IMPORTANTE)
        y = draw_wrapped_text(
            c,
            f"Requisitos: {producto_calidad.requisitos or 'N/A'}",
            x=50,
            y=y,
            max_width=500,  # ajusta a tu layout
            font_name="Helvetica",
            font_size=10,
            line_height=12
        )
        y -= 6
      

        # Banderas
        c.drawString(50, y, f"Documental: {_yn(producto_calidad.documental)}")
        c.drawString(200, y, f"Inspección: {_yn(producto_calidad.inspeccion)}")
        c.drawString(350, y, f"Cumplimiento: {_yn(producto_calidad.cumplimiento)}")
        y -= 20

        # Grados (si aplican)
        if producto_calidad.g_documental_id:
            c.drawString(50, y, f"Grado documental: {producto_calidad.g_documental.nombre}")
            y -= 15
        if producto_calidad.g_inspeccion_id:
            c.drawString(50, y, f"Grado inspección: {producto_calidad.g_inspeccion.nombre}")
            y -= 15
        if producto_calidad.g_cumplimiento_id:
            c.drawString(50, y, f"Grado cumplimiento: {producto_calidad.g_cumplimiento.nombre}")
            y -= 15

        y -= 10
    else:
        # opcional: mostrar algo si no hay o no está hecho
        pass
  

    # --- Imagen del reporte ---
    if reporte.image:
        try:
            c.setFont("Helvetica-Bold", 11)
            c.setFillColor(PRUSSIAN_BLUE)
            c.drawString(40, y, "Evidencia fotográfica:")
            y -= 10
            c.drawImage(reporte.image.path, 60, y - 200, width=200, height=200, preserveAspectRatio=True, mask='auto')
            y -= 140
        except Exception:
            c.setFillColor(black)
            c.drawString(50, y, "(No se pudo cargar la imagen del reporte)")
            y -= 15

    # --- Pie de página ---
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(black)
    c.drawCentredString(300, 40, "Documento generado automáticamente por SAVIA 2.0")

    c.showPage()
    c.save()
    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="Reporte_Calidad_{reporte.id}.pdf"'
    response.write(pdf)
    return response

# Encabezado corporativo reutilizable
def _draw_header(c, compra):
    prussian_blue = PRUSSIAN_BLUE
    c.setFillColor(black)
    c.setFont('Helvetica', 8)
    caja_iso = 760

    c.drawString(430, caja_iso, 'Preparado por:')
    #c.drawString(405, caja_iso - 10, 'SUPT. DE ADQUISIONES')
    c.drawString(520, caja_iso, 'Aprobación')
    c.drawString(515, caja_iso - 10, 'SUBD ADTVO')
    c.drawString(150, caja_iso - 20, 'Número de documento')
    #c.drawString(160, caja_iso - 30, 'SEOV-ADQ-N4-01.02')
    c.drawString(245, caja_iso - 20, 'Clasificación del documento')
    c.drawString(275, caja_iso - 30, 'Controlado')
    c.drawString(355, caja_iso - 20, 'Nivel del documento')
    c.drawString(380, caja_iso - 30, 'N5')
    c.drawString(440, caja_iso - 20, 'Revisión No.')
    #c.drawString(452, caja_iso - 30, '003')
    c.drawString(510, caja_iso - 20, 'Fecha de Emisión')
    #c.drawString(525, caja_iso - 30, '13/11/2017')

    c.setFillColor(prussian_blue)
    c.rect(150, 750, 250, 20, fill=True, stroke=False)
    c.setFillColor(white)
    c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(280, 755, 'Reporte de Calidad')
    c.setFillColor(black)
    c.drawInlineImage('static/images/logo vordtec_documento.png', 45, 730, 3 * cm, 1.5 * cm)
    return 700

def _draw_status_badge(c, status, x=400, y=705):
    """
    Dibuja un badge de estado en la esquina superior derecha.
    status: True -> AUTORIZADO (verde)
            False -> NO AUTORIZADO (rojo)
            None -> PENDIENTE (ámbar)
    (x,y) es la esquina inferior-izquierda del badge.
    """
    if status is True:
        txt, fill = "LIBERADO POR CALIDAD", colors.green
    elif status is False:
        txt, fill = "NO LIBERADO POR CALIDAD", colors.red
    else:
        txt, fill = "PENDIENTE", colors.orange

    # Caja
    c.setFillColor(fill)
    c.setStrokeColor(fill)
    c.rect(x- 10, y, 150, 22, fill=True, stroke=False)

    # Texto
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 11)
    # Centrar horizontalmente en la caja (130 de ancho)
    c.drawCentredString(x + 65, y + 6, txt)

def _draw_watermark_if_needed(c, status):
    """
    Watermark diagonal suave según el estado.
    Solo para NO AUTORIZADO (rojo) o PENDIENTE (ámbar).
    """
    if status is True:
        return
    c.saveState()
    c.setFont("Helvetica-Bold", 60)
    c.setFillColor(colors.lightcoral if status is False else colors.lightgoldenrodyellow)
    # Girar y posicionar
    c.translate(120, 200)
    c.rotate(30)
    c.drawString(0, 0, "NO LIBERADO" if status is False else "PENDIENTE")
    c.restoreState()

from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.colors import black

def draw_wrapped_text(c, text, x, y, max_width, font_name="Helvetica", font_size=10, line_height=12):
    """
    Dibuja texto con word-wrap dentro de max_width.
    Retorna el nuevo 'y' después de dibujar todas las líneas.
    """
    c.setFont(font_name, font_size)
    c.setFillColor(black)

    words = (text or "").split()
    if not words:
        c.drawString(x, y, "N/A")
        return y - line_height

    line = ""
    for w in words:
        test = f"{line} {w}".strip()
        if stringWidth(test, font_name, font_size) <= max_width:
            line = test
        else:
            c.drawString(x, y, line)
            y -= line_height
            line = w

    if line:
        c.drawString(x, y, line)
        y -= line_height

    return y