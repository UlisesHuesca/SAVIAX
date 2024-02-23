from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from compras.models import ArticuloComprado, Compra
from compras.forms import CompraForm
from compras.filters import CompraFilter
from compras.views import dof, attach_oc_pdf, convert_excel_matriz_compras
from dashboard.models import Subproyecto
from .models import Pago, Cuenta, Facturas
from gastos.models import Solicitud_Gasto
from viaticos.models import Solicitud_Viatico
from .forms import PagoForm, Facturas_Form, Facturas_Completas_Form, Saldo_Form, ComprobanteForm
from .filters import PagoFilter, Matriz_Pago_Filter
from viaticos.filters import Solicitud_Viatico_Filter
from gastos.filters import Solicitud_Gasto_Filter
from gastos.models import Articulo_Gasto
from user.models import Profile
from django.contrib import messages
from django.db.models import Sum, Prefetch
from datetime import date, datetime
import decimal
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
#Excel stuff
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt
from django.db.models.functions import Concat



# Create your views here.
@login_required(login_url='user-login')
def compras_autorizadas(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    if usuario.tipo.tesoreria == True:
        compras = Compra.objects.filter(autorizado2=True, pagada=False).order_by('-folio')
    else:
        compras = Compra.objects.filter(flete=True,costo_fletes='1')
    compras = Compra.objects.filter(autorizado2=True, pagada=False).order_by('-folio')
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    if request.method == 'POST' and 'btnReporte' in request.POST:
        return convert_excel_matriz_compras(compras)

    context= {
        'compras':compras,
        'myfilter':myfilter,
        }

    return render(request, 'tesoreria/compras_autorizadas.html',context)

@login_required(login_url='user-login')
def compras_pagos(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id=pk)
    productos = ArticuloComprado.objects.filter(oc=pk)
    pagos = Pago.objects.filter(oc=compra.id, hecho=True) #.aggregate(Sum('monto'))
    sub = Subproyecto.objects.get(id=compra.req.orden.subproyecto.id)
    pagos_alt = Pago.objects.filter(oc=compra.id, hecho=True)
    #Esta es otra forma de sacar lo que hice con los pagos, me parece mas legible
    #compra_pagos = compra.pago_set.aggregate(Sum('monto'))

    #if pagos['monto__sum'] == None:
        #    monto_anterior = 0
    #else:
        #if compra.moneda.nombre == 'PESOS':
       #     monto_anterior = pagos['monto__sum']
        #cero = Money(0, 'MXN')
       # if compra.moneda.nombre == 'DOLARES':
       #     monto_anterior = pagos['monto__sum']

    suma_pago = 0

    for pago in pagos:
        if pago.oc.moneda.nombre == "DOLARES":
            if pago.cuenta.moneda.nombre == "PESOS":
                monto_pago = pago.monto/pago.tipo_de_cambio
                suma_pago = suma_pago + monto_pago
            else:
                suma_pago = suma_pago + pago.monto
        else:
            suma_pago = suma_pago + pago.monto


    if compra.moneda.nombre == 'PESOS':
        cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
    if compra.moneda.nombre == 'DOLARES':
        cuentas = Cuenta.objects.all()


    pago, created = Pago.objects.get_or_create(tesorero = usuario, distrito = usuario.distrito, oc=compra, hecho=False)
    form = PagoForm(instance=pago)
    remanente = compra.costo_plus_adicionales - suma_pago


    if request.method == 'POST':
        form = PagoForm(request.POST, request.FILES or None, instance = pago)
        if form.is_valid():
            pago = form.save(commit = False)
            pago.pagado_date = date.today()
            pago.pagado_hora = datetime.now().time()
            pago.hecho = True
            #Traigo la cuenta que se capturo en el form
            cuenta = Cuenta.objects.get(cuenta = pago.cuenta.cuenta)
            #La utilizo para sacar la información de todos los pagos relacionados con esa cuenta y sumarlos

            # Actualizo el saldo de la cuenta
            monto_actual = pago.monto #request.POST['monto_0']
            if compra.moneda.nombre == "PESOS":
                sub.gastado = sub.gastado + monto_actual
            #    cuenta.saldo = cuenta_pagos['monto__sum'] + monto_actual
            if compra.moneda.nombre == "DOLARES":
                if pago.cuenta.moneda.nombre == "PESOS": #Si la cuenta es en pesos
                    sub.gastado = sub.gastado + monto_actual * pago.tipo_de_cambio
                    monto_actual = monto_actual/pago.tipo_de_cambio
            #        cuenta.saldo = cuenta_pagos['monto__sum'] + monto_actual * decimal.Decimal(request.POST['tipo_de_cambio'])
                if pago.cuenta.moneda.nombre == "DOLARES":
                    tipo_de_cambio = decimal.Decimal(dof())
                    sub.gastado = sub.gastado + monto_actual * tipo_de_cambio
                    #cuenta.saldo = cuenta_pagos['monto__sum'] + monto_actual
            #actualizar la cuenta de la que se paga
            monto_total= monto_actual + suma_pago
            compra.monto_pagado = monto_total
            costo_oc = compra.costo_plus_adicionales
            if monto_actual <= 0:
                messages.error(request,f'El pago {monto_actual} debe ser mayor a 0')
            elif round(monto_total,2) <= round(costo_oc,2):
                if round(monto_total,2) == round(costo_oc,2):
                    compra.pagada= True
                    if compra.cond_de_pago.nombre == "CONTADO":
                        pagos = Pago.objects.filter(oc=compra, hecho=True)
                        archivo_oc = attach_oc_pdf(request, compra.id)
                        if compra.referencia:
                            email = EmailMessage(
                                f'Compra Autorizada {compra.get_folio}',
                                f'Estimado(a) {compra.proveedor.contacto} | Proveedor {compra.proveedor.nombre}:\n\nEstás recibiendo este correo porque has sido seleccionado para surtirnos la OC adjunta con folio: {compra.get_folio} y referencia: {compra.referencia}.\n\n Atte. {compra.creada_por.staff.first_name} {compra.creada_por.staff.last_name} \nVordtec de México S.A. de C.V.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                                'savia@vordtec.com',
                                ['ulises_huesc@hotmail.com',compra.proveedor.email,'lizeth.ojeda@vordtec.com','carlos.ramon@vordtec.com'],
                                )
                        else:
                            email = EmailMessage(
                                f'Compra Autorizada {compra.get_folio}',
                                f'Estimado(a) {compra.proveedor.contacto} | Proveedor {compra.proveedor.nombre}:\n\nEstás recibiendo este correo porque has sido seleccionado para surtirnos la OC adjunta con folio: {compra.get_folio}.\n\n Atte. {compra.creada_por.staff.first_name} {compra.creada_por.staff.last_name} \nVordtec de México S.A. de C.V.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                                'savia@vordtec.com',
                                ['ulises_huesc@hotmail.com',compra.proveedor.email,'lizeth.ojeda@vordtec.com','carlos.ramon@vordtec.com'],
                                )
                        email.attach(f'OC_folio_{compra.get_folio}.pdf',archivo_oc,'application/pdf')
                        email.attach('Pago.pdf',request.FILES['comprobante_pago'].read(),'application/pdf')
                        if pagos.count() > 0:
                            for pago in pagos:
                                email.attach(f'Pago_folio_{pago.id}.pdf',pago.comprobante_pago.path,'application/pdf')
                        email.send()
                        for producto in productos:
                            if producto.producto.producto.articulos.producto.producto.especialista == True:
                                archivo_oc = attach_oc_pdf(request, compra.id)
                                email = EmailMessage(
                                f'Compra Autorizada {compra.get_folio}',
                                f'Estimado Especialista,\n Estás recibiendo este correo porque ha sido pagada una OC que contiene el producto código:{producto.producto.producto.articulos.producto.producto.codigo} descripción:{producto.producto.producto.articulos.producto.producto.codigo} el cual requiere la liberación de calidad\n Este mensaje ha sido automáticamente generado por SAVIA X',
                                'savia@vordtec.com',
                                ['ulises_huesc@hotmail.com'],
                                )
                                email.attach(f'folio:{compra.get_folio}.pdf',archivo_oc,'application/pdf')
                                email.send()
                pago.save()
                compra.save()
                form.save()
                sub.save()
                cuenta.save()

                messages.success(request,f'Gracias por registrar tu pago, {usuario.staff.first_name}')
                return redirect('compras-autorizadas') 
            elif monto_total > compra.costo_plus_adicionales:
                messages.error(request,'El monto total pagado es mayor que el costo de la compra')
            else:
                form = PagoForm()
                messages.error(request,f'{usuario.staff.first_name}, No se pudo subir tu documento')
        else:
            messages.error(request,f'{usuario.staff.first_name}, No está validando')

    context= {
        'compra':compra,
        'pago':pago,
        'form':form,
        'monto':suma_pago,
        'suma_pagos': suma_pago,
        'pagos_alt':pagos_alt,
        'cuentas':cuentas,
        'remanente':remanente,
    }

    return render(request, 'tesoreria/compras_pagos.html',context)

@login_required(login_url='user-login')
def edit_comprobante_pago(request, pk):
    pago = Pago.objects.get(id = pk)
    #print(pago.id)
    form = ComprobanteForm(instance = pago)

    if request.method == 'POST':
        form = ComprobanteForm(request.POST, request.FILES, instance=pago)
        if form.is_valid():
            form.save()
            return HttpResponse(status=204) #No content to render nothing and send a "signal" to javascript in order to close window
    
    context = {
        'pago':pago,
        'form':form, 
    }
    
    return render(request, 'tesoreria/edit_comprobante_pago.html',context)

def edit_pago(request, pk):
    # Obtener el objeto Pago basado en el pk
    usuario = Profile.objects.get(staff__id=request.user.id)
    pago = Pago.objects.get(id=pk)
    print('pago_id',pk)
     
    compra = Compra.objects.get(id = pago.oc.id)
    productos = ArticuloComprado.objects.filter(oc=compra.id)
    pagos = Pago.objects.filter(oc=compra.id, hecho=True) #.aggregate(Sum('monto'))
    sub = Subproyecto.objects.get(id=compra.req.orden.subproyecto.id)
    pagos_alt = Pago.objects.filter(oc=compra.id, hecho=True)
    suma_pago = 0

    for item in pagos:
        if item.oc.moneda.nombre == "DOLARES":
            if item.cuenta.moneda.nombre == "PESOS":
                monto_pago = item.monto/item.tipo_de_cambio
                suma_pago = suma_pago + monto_pago
            else:
                suma_pago = suma_pago + item.monto
        else:
            suma_pago = suma_pago + item.monto


    if compra.moneda.nombre == 'PESOS':
        cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
    if compra.moneda.nombre == 'DOLARES':
        cuentas = Cuenta.objects.all()

    remanente = compra.costo_plus_adicionales - suma_pago
    # Verificar si es un POST para guardar los cambios
    print(pago)
    if request.method == "POST":
        form = PagoForm(request.POST, request.FILES or None, instance=pago)
        if "btn_actualizar" in request.POST:
            if form.is_valid():
                form.save()
                messages.success(request,f'Has actualizado el pago {pago.id} de manera satisfactoria')
                # Redirigir al usuario a donde quieras luego de guardar los cambios
                return redirect('compras-autorizadas')
        if "btn_eliminar" in request.POST:
            print('ya estoy aquí')
            if compra.moneda.nombre == "PESOS":
                sub.gastado = sub.gastado - pago.monto
            if compra.moneda.nombre == "DOLARES":
                if pago.cuenta.moneda.nombre == "PESOS": #Si la cuenta es en pesos
                    sub.gastado = sub.gastado - pago.monto * pago.tipo_de_cambio
                if pago.cuenta.moneda.nombre == "DOLARES":
                    tipo_de_cambio = decimal.Decimal(dof())
                    sub.gastado = sub.gastado - pago.monto * tipo_de_cambio
                messages.success(request,f'Has eliminado el pago {pago.id} de manera satisfactoria')
            pago.delete()
            return redirect('compras-autorizadas')
    else:
        # Si no es un POST, simplemente carga el formulario con el objeto Pago
        form = PagoForm(instance=pago)


    context= {
        'compra':compra,
        'pago':pago,
        'form':form,
        'monto':suma_pago,
        'suma_pagos': suma_pago,
        'pagos_alt':pagos_alt,
        'cuentas':cuentas,
        'remanente':remanente,
    }
    return render(request, 'tesoreria/compras_pagos.html', context)

@login_required(login_url='user-login')
def saldo_a_favor(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id=pk)
    form = Saldo_Form(instance = compra)
    total_pagado = Pago.objects.filter(oc=compra).aggregate(Sum('monto'))['monto__sum']
    total_pagado = total_pagado if total_pagado is not None else 0
    remanente = compra.costo_plus_adicionales - total_pagado


    if request.method == 'POST':
        form = Saldo_Form(request.POST, instance = compra)
        if form.is_valid():
            compra =form.save()
            total = total_pagado + compra.saldo_a_favor
            if compra.costo_plus_adicionales == total:
                compra.pagada = True
            compra.save()
            messages.success(request,f'El saldo se ha registrado correctamente, {usuario.staff.first_name}')
            return HttpResponse(status=204) 

    context= {
        'compra':compra,
        'form':form,
        'remanente': remanente,
        'total_pagado': total_pagado,
    }

    return render(request, 'tesoreria/saldo_a_favor.html',context)

# Create your views here.
@login_required(login_url='user-login')
def matriz_pagos(request):
    pagos = Pago.objects.filter(hecho=True)
    myfilter = Matriz_Pago_Filter(request.GET, queryset=pagos) 

    pagos = myfilter.qs
    for pago in pagos:
        articulos_gasto = Articulo_Gasto.objects.filter(gasto=pago.gasto)

        proyectos = set()
        subproyectos = set()

        for articulo in articulos_gasto:
            if articulo.proyecto:
                proyectos.add(str(articulo.proyecto.nombre))
            if articulo.subproyecto:
                subproyectos.add(str(articulo.subproyecto.nombre))

        pago.proyectos = ', '.join(proyectos)
        pago.subproyectos = ', '.join(subproyectos)

    #Set up pagination
    p = Paginator(pagos, 50)
    page = request.GET.get('page')
    pagos_list = p.get_page(page)

    if request.method == 'POST' and 'btnReporte' in request.POST:
        return convert_excel_matriz_pagos(pagos)

    context= {
        'pagos_list':pagos_list,
        'pagos':pagos,
        'myfilter':myfilter,
        }

    return render(request, 'tesoreria/matriz_pagos.html',context)


@login_required(login_url='user-login')
def matriz_facturas(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    facturas = Facturas.objects.filter(oc = compra, hecho=True)
    factura, created = Facturas.objects.get_or_create(oc=compra, hecho=False)

    form = Facturas_Form()

    if request.method == 'POST':
        if "btn_factura" in request.POST:
            form = Facturas_Form(request.POST or None, request.FILES or None, instance = factura)
            if form.is_valid():
                factura = form.save(commit = False)
                factura.fecha_subido = date.today()
                factura.hora_subido = datetime.now().time()
                factura.hecho = True
                factura.subido_por = usuario
                factura.save()
                form.save()
                messages.success(request,'Haz registrado tu factura')
                return HttpResponse(status=204) #No content to render nothing and send a "signal" to javascript in order to close window
            else:
                messages.error(request,'No está validando')
        if "btn_editar" in request.POST:
            form

    context={
        'form':form,
        'facturas':facturas,
        'compra':compra,
        }

    return render(request, 'tesoreria/matriz_facturas.html', context)

@login_required(login_url='user-login')
def matriz_facturas_nomodal(request, pk):
    compra = Compra.objects.get(id = pk)
    facturas = Facturas.objects.filter(oc = compra, hecho=True)
    pagos = Pago.objects.filter(oc=compra)
    form = Facturas_Completas_Form(instance=compra)

    if request.method == 'POST':
        form = Facturas_Completas_Form(request.POST, instance=compra)
        if "btn_factura_completa" in request.POST:
            if form.is_valid():
                form.save()
                messages.success(request,'Haz cambiado el status de facturas completas')
                return redirect('matriz-pagos')
            else:
                messages.error(request,'No está validando')

    context={
        'pagos':pagos,
        'form':form,
        'facturas':facturas,
        'compra':compra,
        }

    return render(request, 'tesoreria/matriz_factura_no_modal.html', context)

def factura_nueva(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    compra = Compra.objects.get(id = pk)
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)
    factura, created = Facturas.objects.get_or_create(oc=compra, hecho=False)
    form = Facturas_Form()

    if request.method == 'POST':
        if 'btn_registrar' in request.POST:
            form = Facturas_Form(request.POST or None, request.FILES or None, instance = factura)
            if form.is_valid():
                factura = form.save(commit=False)
                factura.hecho=True
                factura.fecha_subido =date.today()
                factura.hora_subido = datetime.now().time()
                factura.subido_por =  usuario
                form.save()
                factura.save()
                messages.success(request,'Las factura se registró de manera exitosa')
            else:
                messages.error(request,'No se pudo subir tu documento')


    context={
        'form':form,
        }

    return render(request, 'tesoreria/registrar_nueva_factura.html', context)

def factura_compra_edicion(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    factura = Facturas.objects.get(id = pk)
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)
    #factura, created = Facturas.objects.get_or_create(pago=pago, hecho=False)
    form = Facturas_Form(instance= factura)

    if request.method == 'POST':
        if 'btn_edicion' in request.POST:
            form = Facturas_Form(request.POST or None, request.FILES or None, instance = factura)
            if form.is_valid():
                factura = form.save(commit = False)
                factura.subido_por = usuario
                factura.save()
                form.save()
                messages.success(request,'Las facturas se subieron de manera exitosa')
            else:
                messages.error(request,'No se pudo subir tu documento')


    context={
        'factura':factura,
        'form':form,
        }

    return render(request, 'tesoreria/factura_compra_edicion.html', context)

def factura_eliminar(request, pk):
    factura = Facturas.objects.get(id = pk)
    compra = factura.oc
    messages.success(request,f'La factura {factura.id} ha sido eliminado exitosamente')
    factura.delete()

    return redirect('matriz-facturas-nomodal',pk= compra.id)

def mis_gastos(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    gastos = Solicitud_Gasto.objects.filter(complete=True, staff = usuario)
    myfilter = Solicitud_Gasto_Filter(request.GET, queryset=gastos)
    gastos = myfilter.qs

    context= {
        'gastos':gastos,
        'myfilter':myfilter,
        }

    return render(request, 'tesoreria/mis_gastos.html',context)

def mis_viaticos(request):
    usuario = Profile.objects.get(staff__id=request.user.id)
    viaticos = Solicitud_Viatico.objects.filter(complete=True, staff = usuario)
    myfilter = Solicitud_Viatico_Filter(request.GET, queryset=viaticos)
    viaticos = myfilter.qs

    context= {
        'viaticos':viaticos,
        'myfilter':myfilter,
        }

    return render(request, 'tesoreria/mis_viaticos.html',context)

def convert_excel_matriz_pagos(pagos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Matriz_pagos_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Pagos')
    #Comenzar en la fila 1
    row_num = 1

    # Función para manejar los IDs de las compras, gastos o viáticos
    def get_transaction_id(pago):
        if pago.oc:
            return 'OC'+str(pago.oc.id)
        elif pago.gasto:
            return 'G'+str(pago.gasto.id)
        elif pago.viatico:
            return 'V'+str(pago.viatico.id)
        else:
            return None

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Id','Compra/Gasto','Solicitado','Proyecto','Subproyecto','Proveedor/Colaborador','Factuas_Completas',
               'Importe','Fecha', 'Moneda','Tipo de cambio', 'Total en Pesos']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordtec. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de Pagos").style = head_style
    ws.cell(row=4, column = columna_max, value="Sumatoria de Pagos").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=4, column=columna_max + 1, value=f"=SUM(L:L)").style = money_resumen_style
  

   # Aquí debes extraer el conjunto completo de pagos en lugar de solo ciertos valores
    
    for pago in pagos:
        row_num = row_num + 1
        # Define los valores de las columnas basándote en el tipo de pago
        if pago.oc:
            proveedor = pago.oc.proveedor
            facturas_completas = pago.oc.facturas_completas
            cuenta_moneda = pago.cuenta.moneda.nombre if pago.cuenta else None
            solicitado = pago.oc.req.orden.staff.staff.first_name + ' ' + pago.oc.req.orden.staff.staff.last_name
            if cuenta_moneda == 'PESOS':
                tipo_de_cambio = ''
            elif cuenta_moneda == 'DOLARES':
                 tipo_de_cambio = pago.tipo_de_cambio or pago.oc.tipo_de_cambio or 17
            else:
                tipo_de_cambio = ''  # default si no se cumplen las condiciones anteriores
            moneda = pago.oc.moneda.nombre, 
        elif pago.gasto:
            solicitado = pago.gasto.staff.staff.first_name + ' ' + pago.gasto.staff.staff.last_name
            proveedor = pago.gasto.staff.staff.first_name
            facturas_completas = pago.gasto.facturas_completas
            tipo_de_cambio = '' # Asume que no se requiere tipo de cambio para gastos
            moneda = pago.gasto.moneda.nombre
        elif pago.viatico:
            proveedor = pago.viatico.staff.staff.first_name
            solicitado = pago.viatico.staff.staff.first_name + ' ' + pago.viatico.staff.staff.last_name
            facturas_completas = pago.viatico.facturas_completas
            tipo_de_cambio = '' # Asume que no se requiere tipo de cambio para viáticos
            moneda = pago.viatico.moneda.nombre,
        else:
            proveedor = None
            facturas_completas = None
            tipo_de_cambio = ''


        row = [
            pago.id,
            get_transaction_id(pago),
            solicitado,
            pago.oc.req.orden.proyecto.nombre if pago.oc else '',
            pago.oc.req.orden.subproyecto.nombre if pago.oc else '',
            proveedor,
            'Verdadero' if facturas_completas else 'Falso',
            pago.monto,
            pago.pagado_date.strftime('%d/%m/%Y') if pago.pagado_date else '',
            moneda,# Modificación aquí
            tipo_de_cambio,
            f'=IF(K{row_num}="",H{row_num},H{row_num}*K{row_num})'  # Calcula total en pesos usando la fórmula de Excel
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 7 or col_num == 10 or col_num == 11:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
       
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)
